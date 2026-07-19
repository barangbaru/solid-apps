from flask import (Flask, render_template, request, redirect, url_for,
                   flash, g, session, jsonify, Response, abort)
from werkzeug.security import generate_password_hash, check_password_hash
from functools import wraps
import sqlite3, os, smtplib, json, secrets, requests as req_lib, io, base64, signal, subprocess
try:
    from cryptography.fernet import Fernet, InvalidToken as _FernetInvalidToken
    _CRYPTO_OK = True
except ImportError:
    _CRYPTO_OK = False
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from datetime import datetime, date, timedelta
from seed_data import ALL_DIVISIONS, ABILITY_ITEMS
from version import VERSION, RELEASE_DATE, RELEASE_NOTES
import pyotp, qrcode

# Load .env file manually if it exists
_env_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), '.env')
if os.path.exists(_env_path):
    with open(_env_path, 'r', encoding='utf-8') as _f:
        for _line in _f:
            _line = _line.strip()
            if _line and not _line.startswith('#') and '=' in _line:
                _k, _v = _line.split('=', 1)
                _k = _k.strip()
                _v = _v.strip().strip('\'"').strip("'\"")
                if _k:
                    os.environ[_k] = _v

app = Flask(__name__)
_default_secret = os.environ.get('SECRET_KEY', '')
if not _default_secret:
    import warnings
    warnings.warn("SECRET_KEY env var tidak diset! Gunakan nilai acak yang kuat di production.")
    _default_secret = 'evalkey-2024-superadmin-secure!'
app.secret_key = _default_secret

@app.template_filter('from_json')
def _tpl_from_json(s):
    import json as _j
    try:
        return _j.loads(s) if s else []
    except Exception:
        return []

# Google OAuth config (set via env vars or portal settings)
GOOGLE_CLIENT_ID     = os.environ.get('GOOGLE_CLIENT_ID', '')
GOOGLE_CLIENT_SECRET = os.environ.get('GOOGLE_CLIENT_SECRET', '')
GOOGLE_AUTH_URL      = 'https://accounts.google.com/o/oauth2/v2/auth'
GOOGLE_TOKEN_URL     = 'https://oauth2.googleapis.com/token'
GOOGLE_USERINFO_URL  = 'https://www.googleapis.com/oauth2/v3/userinfo'

# Agar request.host_url benar di balik nginx (ProxyFix)
from werkzeug.middleware.proxy_fix import ProxyFix
app.wsgi_app = ProxyFix(app.wsgi_app, x_for=1, x_proto=1, x_host=1, x_prefix=1)

DB_TYPE = 'postgresql'
DIVISI_LIST = list(ALL_DIVISIONS.keys())

UPLOAD_FOLDER = os.path.join(os.path.dirname(__file__), 'static', 'uploads')
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
ALLOWED_IMAGE_EXT = {'jpg', 'jpeg', 'png', 'webp', 'gif'}
ALLOWED_ATTACHMENT_EXT = ALLOWED_IMAGE_EXT | {'pdf', 'doc', 'docx', 'xls', 'xlsx', 'txt', 'eml', 'msg'}

def _save_upload_file(file_obj, subfolder='', allowed_ext=None):
    import uuid
    import shutil
    if not file_obj or not file_obj.filename or '.' not in file_obj.filename:
        return None
    ext = file_obj.filename.rsplit('.', 1)[-1].lower()
    
    DANGEROUS_EXT = {
        'php', 'sh', 'bat', 'exe', 'py', 'cmd', 'ps1', 'js', 'vbs', 'pl', 'msi', 'cgi', 
        'asp', 'aspx', 'jsp', 'jar', 'com', 'scr', 'pif', 'vbe', 'jse', 'reg'
    }
    actual_allowed = allowed_ext if allowed_ext is not None else ALLOWED_ATTACHMENT_EXT
    
    if ext in DANGEROUS_EXT or ext not in actual_allowed:
        try:
            from flask import flash
            flash(f'File "{file_obj.filename}" ditolak. Hanya dokumen dan gambar yang diperbolehkan.', 'danger')
        except Exception:
            pass
        return None
    fname = uuid.uuid4().hex + '.' + ext

    # Check media storage type from settings
    storage_type = 'local'
    try:
        db = get_db()
        cfg = get_settings(db)
        storage_type = cfg.get('media_storage_type', 'local')
    except Exception:
        pass

    if storage_type == 's3':
        temp_path = None
        try:
            endpoint = cfg.get('backup_dest_s3_endpoint', '').strip()
            access_key = cfg.get('backup_dest_s3_access_key', '').strip()
            secret_key = cfg.get('backup_dest_s3_secret_key', '').strip()
            bucket = cfg.get('backup_dest_s3_bucket', '').strip()
            region = cfg.get('backup_dest_s3_region', '').strip()

            # Only attempt S3 upload if required credentials are provided
            if access_key and secret_key and bucket:
                s3_key = f"upload/media/{subfolder}/{fname}" if subfolder else f"upload/media/{fname}"

                # Temporary local save to upload
                temp_path = os.path.join(UPLOAD_FOLDER, 'temp_' + fname)
                file_obj.save(temp_path)

                import boto3
                from botocore.config import Config
                config = Config(
                    region_name=region or 'us-east-1',
                    signature_version='s3v4',
                    connect_timeout=3,
                    read_timeout=5,
                    retries={'max_attempts': 1}
                )
                s3 = boto3.client(
                    's3',
                    endpoint_url=endpoint or None,
                    aws_access_key_id=access_key,
                    aws_secret_access_key=secret_key,
                    config=config
                )
                # Upload (try with public-read ACL, fallback if forbidden)
                try:
                    s3.upload_file(temp_path, bucket, s3_key, ExtraArgs={'ACL': 'public-read'})
                except Exception:
                    s3.upload_file(temp_path, bucket, s3_key)

                try:
                    os.remove(temp_path)
                except Exception:
                    pass

                return f"/media/proxy/{s3_key}"
        except Exception:
            # Fallback: if we already saved the file to temp_path, move it to local destination
            if temp_path and os.path.exists(temp_path):
                folder = os.path.join(UPLOAD_FOLDER, subfolder)
                os.makedirs(folder, exist_ok=True)
                try:
                    shutil.move(temp_path, os.path.join(folder, fname))
                    return f'/static/uploads/{subfolder}/{fname}' if subfolder else f'/static/uploads/{fname}'
                except Exception:
                    pass

    # Local Storage fallback
    try:
        file_obj.seek(0)
    except Exception:
        pass
    folder = os.path.join(UPLOAD_FOLDER, subfolder)
    os.makedirs(folder, exist_ok=True)
    file_obj.save(os.path.join(folder, fname))
    return f'/static/uploads/{subfolder}/{fname}' if subfolder else f'/static/uploads/{fname}'

def _save_upload(file_obj, subfolder=''):
    return _save_upload_file(file_obj, subfolder, ALLOWED_IMAGE_EXT)

def get_divisi_list(db):
    try:
        rows = db.execute('SELECT name FROM divisions WHERE is_active=1 ORDER BY sort_order, name').fetchall()
        return [r['name'] for r in rows] if rows else DIVISI_LIST
    except Exception:
        return DIVISI_LIST


def _is_db_integrity_error(exc):
    """True untuk unique/constraint violation dari psycopg2."""
    try:
        import psycopg2
        return isinstance(exc, psycopg2.IntegrityError)
    except ImportError:
        return any(
            cls.__name__ == 'IntegrityError' and cls.__module__.startswith('psycopg2')
            for cls in type(exc).__mro__
        )

# ─── DB Helpers ────────────────────────────────────────────────────────────────
# Thin wrapper sehingga kode SQLite (placeholder "?", row['col']) tetap bekerja
# tanpa perubahan saat menggunakan PostgreSQL (placeholder "%s", RealDictCursor).

class _DBWrapper:
    """Wrap koneksi SQLite atau psycopg2 agar keduanya terlihat sama dari luar."""
    def __init__(self, conn, is_pg=False):
        self._conn = conn
        self._is_pg = is_pg

    def _fix(self, sql):
        """Konversi SQL dialek SQLite → PostgreSQL."""
        if not self._is_pg:
            return sql
        import re
        sql_stripped = sql.strip()
        # INSERT OR REPLACE → ON CONFLICT DO UPDATE SET col=EXCLUDED.col, ...
        m_replace = re.match(
            r'INSERT\s+OR\s+REPLACE\s+INTO\s+(\w+)\s*\(([^)]+)\)',
            sql_stripped, re.IGNORECASE | re.DOTALL)
        if m_replace:
            cols = [c.strip() for c in m_replace.group(2).split(',')]
            # Kolom pertama dijadikan conflict target; kolom lainnya di-update
            conflict_col = cols[0]
            update_cols  = cols[1:] if len(cols) > 1 else cols
            update_set   = ', '.join(f'{c}=EXCLUDED.{c}' for c in update_cols)
            sql = re.sub(r'\bINSERT\s+OR\s+REPLACE\s+INTO\b', 'INSERT INTO', sql, flags=re.IGNORECASE)
            sql = re.sub(r'\?', '%s', sql)
            sql = sql.rstrip().rstrip(';') + f' ON CONFLICT ({conflict_col}) DO UPDATE SET {update_set}'
            return sql
        # INSERT OR IGNORE → ON CONFLICT DO NOTHING (tandai agar execute tahu tidak perlu RETURNING)
        is_or_ignore = bool(re.search(r'\bINSERT\s+OR\s+IGNORE\b', sql, re.IGNORECASE))
        sql = re.sub(r'\bINSERT\s+OR\s+IGNORE\s+INTO\b', 'INSERT INTO', sql, flags=re.IGNORECASE)
        # Placeholder ? → %s
        sql = re.sub(r'\?', '%s', sql)
        if is_or_ignore:
            sql = sql.rstrip().rstrip(';') + ' ON CONFLICT DO NOTHING'
        # ── Konversi fungsi SQLite → PostgreSQL ───────────────────────────────
        # julianday(col) - julianday('now') → (NULLIF(col,'')::date - CURRENT_DATE)
        # pakai NULLIF agar empty string tidak crash saat di-cast ke date
        sql = re.sub(
            r"julianday\(([^)]+)\)\s*-\s*julianday\('now'\)",
            lambda m: f"(NULLIF({m.group(1).strip()},'')" r"::date - CURRENT_DATE)",
            sql, flags=re.IGNORECASE)
        # julianday('now') - julianday(col) → (CURRENT_DATE - NULLIF(col,'')::date)
        sql = re.sub(
            r"julianday\('now'\)\s*-\s*julianday\(([^)]+)\)",
            lambda m: r"(CURRENT_DATE - NULLIF(" + m.group(1).strip() + ",'')::date)",
            sql, flags=re.IGNORECASE)
        # julianday(col) sisa (standalone)
        sql = re.sub(
            r"julianday\(([^)]+)\)",
            lambda m: f"NULLIF({m.group(1).strip()},'')" r"::date",
            sql, flags=re.IGNORECASE)
        # GROUP_CONCAT(col, sep) → STRING_AGG(col::text, sep)
        sql = re.sub(
            r'\bGROUP_CONCAT\s*\(([^,)]+),\s*([^)]+)\)',
            lambda m: f"STRING_AGG({m.group(1).strip()}::text, {m.group(2).strip()})",
            sql, flags=re.IGNORECASE)
        # GROUP_CONCAT(col) tanpa separator → STRING_AGG(col::text, ',')
        sql = re.sub(
            r'\bGROUP_CONCAT\s*\(([^)]+)\)',
            lambda m: f"STRING_AGG({m.group(1).strip()}::text, ',')",
            sql, flags=re.IGNORECASE)
        # last_insert_rowid() → lastval()
        sql = re.sub(r'\blast_insert_rowid\s*\(\s*\)', 'lastval()', sql, flags=re.IGNORECASE)
        # date('now') / date("now") → CURRENT_DATE  (harus sebelum pola date(col) di bawah)
        sql = re.sub(r"date\(\s*['\"]now['\"]\s*\)", "CURRENT_DATE", sql, flags=re.IGNORECASE)
        # date(col_expr) → (col_expr)::date  — konversi SQLite date() ke PostgreSQL cast
        sql = re.sub(r"\bdate\(([^)]+)\)", r"(\1)::date", sql, flags=re.IGNORECASE)
        # datetime('now','localtime') / datetime("now","localtime") → NOW()
        sql = re.sub(r"datetime\(\s*['\"]now['\"]\s*,\s*['\"]localtime['\"]\s*\)", "NOW()", sql, flags=re.IGNORECASE)
        # datetime('now') / datetime("now") → NOW()
        sql = re.sub(r"datetime\(\s*['\"]now['\"]\s*\)", "NOW()", sql, flags=re.IGNORECASE)
        # strftime('%Y-%m-%d', col) → TO_CHAR(col::date, 'YYYY-MM-DD')
        sql = re.sub(
            r"strftime\('%Y-%m-%d',\s*([^)]+)\)",
            r"TO_CHAR(\1::date, 'YYYY-MM-DD')",
            sql, flags=re.IGNORECASE)
        return sql

    @property
    def _is_or_ignore(self):
        return False  # placeholder, dipakai di execute

    def execute(self, sql, params=()):
        if self._is_pg:
            import re
            # Cek apakah ini INSERT OR IGNORE (sebelum _fix mengubahnya)
            is_or_ignore = bool(re.search(r'\bINSERT\s+OR\s+IGNORE\b', sql, re.IGNORECASE))
            is_or_replace = bool(re.search(r'\bINSERT\s+OR\s+REPLACE\b', sql, re.IGNORECASE))
            fixed = self._fix(sql)
            cur = self._conn.cursor()
            # Untuk INSERT biasa, tambahkan RETURNING id agar lastrowid tersedia
            is_insert = bool(re.match(r'\s*INSERT\b', fixed, re.IGNORECASE))
            needs_returning = is_insert and not is_or_ignore and 'RETURNING' not in fixed.upper()
            fixed_exec = fixed
            if needs_returning:
                fixed_exec = fixed.rstrip().rstrip(';') + ' RETURNING id'
            try:
                cur.execute(fixed_exec, params if params else None)
            except Exception as e:
                if needs_returning and 'column "id" does not exist' in str(e):
                    # Tabel ini tidak punya kolom id (misal app_settings pakai key)
                    self._conn.rollback()
                    cur = self._conn.cursor()
                    cur.execute(fixed, params if params else None)
                    needs_returning = False
                else:
                    raise
            wrapper = _CursorWrapper(cur, is_pg=True)
            if needs_returning:
                row = cur.fetchone()
                wrapper._last_id = row[0] if row else None
            elif is_or_replace and 'RETURNING' not in fixed.upper():
                # ON CONFLICT DO UPDATE SET sudah di fixed, tambahkan RETURNING
                pass  # lastrowid tidak kritis untuk OR REPLACE
            return wrapper
        else:
            return self._conn.execute(sql, params)

    def executescript(self, sql):
        if self._is_pg:
            # PostgreSQL: eksekusi per-statement dengan retry multi-pass
            # agar FK ordering di schema tidak masalah (misal sc_ticket_history
            # di-CREATE sebelum sc_tickets karena SQLite tidak enforce FK order)
            stmts = [s.strip() for s in sql.split(';') if s.strip()]
            pending = stmts
            cur = self._conn.cursor()
            max_passes = len(stmts) + 1
            for _ in range(max_passes):
                if not pending:
                    break
                still_pending = []
                for stmt in pending:
                    try:
                        cur.execute(stmt)
                        self._conn.commit()
                    except Exception:
                        self._conn.rollback()
                        cur = self._conn.cursor()
                        still_pending.append(stmt)
                if len(still_pending) == len(pending):
                    # Tidak ada kemajuan — statement benar-benar error
                    # Eksekusi sekali lagi agar exception muncul dengan jelas
                    cur.execute(still_pending[0])
                pending = still_pending
            return cur
        return self._conn.executescript(sql)

    def commit(self):
        self._conn.commit()

    def rollback(self):
        self._conn.rollback()

    def close(self):
        self._conn.close()


class _CursorWrapper:
    """Wrap psycopg2 cursor agar fetchone()/fetchall() mengembalikan RealDictRow
    yang dapat diakses dengan row['col'] DAN row[0] seperti sqlite3.Row."""
    def __init__(self, cur, is_pg=False):
        self._cur = cur
        self._is_pg = is_pg

    def fetchone(self):
        row = self._cur.fetchone()
        if row is None:
            return None
        if self._is_pg:
            return _DictRow(row, self._cur.description)
        return row

    def fetchall(self):
        rows = self._cur.fetchall()
        if self._is_pg:
            desc = self._cur.description
            return [_DictRow(r, desc) for r in rows]
        return rows

    @property
    def lastrowid(self):
        if self._is_pg:
            return getattr(self, '_last_id', None)
        return self._cur.lastrowid

    # Proxy agar sqlite3.Cursor attrs seperti rowcount dll bisa diakses
    def __getattr__(self, name):
        return getattr(self._cur, name)

    def __iter__(self):
        for row in self._cur:
            if self._is_pg:
                yield _DictRow(row, self._cur.description)
            else:
                yield row


class _DictRow:
    """Row psycopg2 yang bisa diakses via row['col'] atau row[0]."""
    __slots__ = ('_data', '_keys')

    def __init__(self, row, description):
        self._keys = [d[0] for d in description]
        self._data = list(row)

    def __getitem__(self, key):
        if isinstance(key, int):
            return self._data[key]
        return self._data[self._keys.index(key)]

    def __iter__(self):
        return iter(self._data)

    def keys(self):
        return self._keys

    def get(self, key, default=None):
        try:
            return self[key]
        except (KeyError, ValueError):
            return default


def _pg_connect():
    """Buat koneksi psycopg2 dari env vars PG_*."""
    import psycopg2
    return psycopg2.connect(
        host=os.environ.get('PG_HOST', 'localhost'),
        port=int(os.environ.get('PG_PORT', 5432)),
        dbname=os.environ.get('PG_NAME', 'hive_db'),
        user=os.environ.get('PG_USER', 'hive'),
        password=os.environ.get('PG_PASS', ''),
        options='-c search_path=public',
    )


def _get_raw_db():
    """Buat koneksi DB baru (untuk background tasks / scheduler di luar request context)."""
    conn = _pg_connect()
    conn.autocommit = False
    return _DBWrapper(conn, is_pg=True)


def get_db():
    if 'db' not in g:
        conn = _pg_connect()
        conn.autocommit = False
        g.db = _DBWrapper(conn, is_pg=True)
    return g.db

@app.teardown_appcontext
def close_db(e=None):
    db = g.pop('db', None)
    if db: db.close()

# ─── Schema ────────────────────────────────────────────────────────────────────

SCHEMA = """
CREATE TABLE IF NOT EXISTS users (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    username TEXT UNIQUE NOT NULL,
    password_hash TEXT NOT NULL,
    full_name TEXT DEFAULT '',
    role TEXT DEFAULT 'admin',
    is_active INTEGER DEFAULT 1,
    last_login TEXT DEFAULT '',
    email TEXT DEFAULT '',
    google_id TEXT DEFAULT '',
    created_at TEXT DEFAULT (datetime('now','localtime'))
);
CREATE TABLE IF NOT EXISTS employees (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    name TEXT NOT NULL,
    jabatan TEXT DEFAULT '',
    divisi TEXT NOT NULL,
    level TEXT DEFAULT 'Staff',
    employment_type TEXT DEFAULT 'tetap',
    contract_start TEXT DEFAULT '',
    contract_end TEXT DEFAULT '',
    rate_mandays REAL DEFAULT NULL,
    email TEXT DEFAULT '',
    phone TEXT DEFAULT '',
    telegram_id TEXT DEFAULT '',
    is_active INTEGER DEFAULT 1,
    notes TEXT DEFAULT '',
    created_at TEXT DEFAULT (datetime('now','localtime'))
);
CREATE TABLE IF NOT EXISTS evaluations (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    employee_id INTEGER NOT NULL,
    periode TEXT NOT NULL,
    status TEXT DEFAULT 'draft',
    pp_score INTEGER DEFAULT 0,
    pp_total REAL DEFAULT 0,
    hs_total REAL DEFAULT 0,
    ability_score REAL DEFAULT 0,
    competency_total REAL DEFAULT 0,
    final_total REAL DEFAULT 0,
    evaluator TEXT DEFAULT '',
    overall_assessment TEXT DEFAULT '',
    development_plan TEXT DEFAULT '',
    created_at TEXT DEFAULT (datetime('now','localtime')),
    FOREIGN KEY(employee_id) REFERENCES employees(id) ON DELETE CASCADE
);
CREATE TABLE IF NOT EXISTS project_entries (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    eval_id INTEGER NOT NULL,
    entry_type TEXT NOT NULL,
    project_name TEXT DEFAULT '',
    detail_task TEXT DEFAULT '',
    status TEXT DEFAULT 'DONE',
    notes TEXT DEFAULT '',
    sort_order INTEGER DEFAULT 0,
    FOREIGN KEY(eval_id) REFERENCES evaluations(id) ON DELETE CASCADE
);
CREATE TABLE IF NOT EXISTS skill_categories (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    divisi TEXT NOT NULL,
    name TEXT NOT NULL,
    sort_order INTEGER DEFAULT 0
);
CREATE TABLE IF NOT EXISTS skill_items (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    category_id INTEGER NOT NULL,
    name TEXT NOT NULL,
    description TEXT DEFAULT '',
    bobot REAL DEFAULT 1.0,
    sort_order INTEGER DEFAULT 0,
    FOREIGN KEY(category_id) REFERENCES skill_categories(id) ON DELETE CASCADE
);
CREATE TABLE IF NOT EXISTS skill_scores (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    eval_id INTEGER NOT NULL,
    skill_item_id INTEGER NOT NULL,
    score INTEGER DEFAULT 0,
    UNIQUE(eval_id, skill_item_id),
    FOREIGN KEY(eval_id) REFERENCES evaluations(id) ON DELETE CASCADE
);
CREATE TABLE IF NOT EXISTS ability_items (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    divisi TEXT NOT NULL,
    name TEXT NOT NULL,
    desc_a TEXT DEFAULT '',
    desc_b TEXT DEFAULT '',
    desc_c TEXT DEFAULT '',
    desc_d TEXT DEFAULT '',
    sort_order INTEGER DEFAULT 0
);
CREATE TABLE IF NOT EXISTS ability_scores (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    eval_id INTEGER NOT NULL,
    ability_item_id INTEGER NOT NULL,
    level TEXT DEFAULT '',
    UNIQUE(eval_id, ability_item_id),
    FOREIGN KEY(eval_id) REFERENCES evaluations(id) ON DELETE CASCADE
);
CREATE TABLE IF NOT EXISTS competency_items (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    divisi TEXT NOT NULL,
    point_measurement TEXT NOT NULL,
    bobot REAL DEFAULT 0.0,
    sort_order INTEGER DEFAULT 0,
    is_hardskill INTEGER DEFAULT 0
);
CREATE TABLE IF NOT EXISTS competency_scores (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    eval_id INTEGER NOT NULL,
    competency_item_id INTEGER NOT NULL,
    rating INTEGER DEFAULT 0,
    UNIQUE(eval_id, competency_item_id),
    FOREIGN KEY(eval_id) REFERENCES evaluations(id) ON DELETE CASCADE
);
CREATE TABLE IF NOT EXISTS peer_reviews (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    eval_id INTEGER NOT NULL,
    slot INTEGER DEFAULT 1,
    reviewer_name TEXT DEFAULT '',
    feedback TEXT DEFAULT '',
    FOREIGN KEY(eval_id) REFERENCES evaluations(id) ON DELETE CASCADE
);
CREATE TABLE IF NOT EXISTS app_settings (
    key TEXT PRIMARY KEY,
    value TEXT DEFAULT ''
);
CREATE TABLE IF NOT EXISTS reminder_logs (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    employee_id INTEGER,
    channel TEXT DEFAULT '',
    subject TEXT DEFAULT '',
    message TEXT DEFAULT '',
    status TEXT DEFAULT 'pending',
    error_msg TEXT DEFAULT '',
    triggered_by TEXT DEFAULT 'auto',
    created_at TEXT DEFAULT (datetime('now','localtime')),
    FOREIGN KEY(employee_id) REFERENCES employees(id) ON DELETE SET NULL
);
CREATE TABLE IF NOT EXISTS password_reset_tokens (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    user_id INTEGER NOT NULL,
    token TEXT UNIQUE NOT NULL,
    expires_at TEXT NOT NULL,
    used INTEGER DEFAULT 0,
    created_at TEXT DEFAULT (datetime('now','localtime')),
    FOREIGN KEY(user_id) REFERENCES users(id) ON DELETE CASCADE
);
CREATE TABLE IF NOT EXISTS eval_tokens (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    eval_id INTEGER NOT NULL UNIQUE,
    token TEXT UNIQUE NOT NULL,
    email_sent_to TEXT DEFAULT '',
    sent_at TEXT DEFAULT '',
    accessed_at TEXT DEFAULT '',
    submitted_at TEXT DEFAULT '',
    status TEXT DEFAULT 'active',
    FOREIGN KEY(eval_id) REFERENCES evaluations(id) ON DELETE CASCADE
);
CREATE TABLE IF NOT EXISTS eval_reviews (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    eval_id INTEGER NOT NULL,
    reviewer_user_id INTEGER NOT NULL,
    reviewer_role TEXT DEFAULT '',
    notes TEXT DEFAULT '',
    status TEXT DEFAULT 'pending',
    submitted_at TEXT DEFAULT '',
    FOREIGN KEY(eval_id) REFERENCES evaluations(id) ON DELETE CASCADE,
    FOREIGN KEY(reviewer_user_id) REFERENCES users(id),
    UNIQUE(eval_id, reviewer_user_id)
);
CREATE TABLE IF NOT EXISTS divisions (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    name TEXT UNIQUE NOT NULL,
    description TEXT DEFAULT '',
    is_active INTEGER DEFAULT 1,
    sort_order INTEGER DEFAULT 0,
    created_at TEXT DEFAULT (datetime('now','localtime'))
);
CREATE TABLE IF NOT EXISTS roles (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    name TEXT UNIQUE NOT NULL,
    description TEXT DEFAULT '',
    is_system INTEGER DEFAULT 0,
    created_at TEXT DEFAULT (datetime('now','localtime'))
);
CREATE TABLE IF NOT EXISTS role_permissions (
    role_name TEXT NOT NULL,
    permission TEXT NOT NULL,
    PRIMARY KEY (role_name, permission)
);
CREATE TABLE IF NOT EXISTS employee_salary (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    employee_id INTEGER NOT NULL,
    year INTEGER NOT NULL,
    base_salary TEXT DEFAULT '',
    al_001 TEXT DEFAULT '',
    al_002 TEXT DEFAULT '',
    al_003 TEXT DEFAULT '',
    al_004 TEXT DEFAULT '',
    increase_pct REAL DEFAULT 0,
    notes TEXT DEFAULT '',
    created_at TEXT DEFAULT (datetime('now','localtime')),
    updated_at TEXT DEFAULT (datetime('now','localtime')),
    UNIQUE(employee_id, year),
    FOREIGN KEY(employee_id) REFERENCES employees(id) ON DELETE CASCADE
);
CREATE TABLE IF NOT EXISTS superapp_apps (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    slug TEXT UNIQUE NOT NULL,
    name TEXT NOT NULL,
    description TEXT DEFAULT '',
    icon TEXT DEFAULT 'grid',
    color TEXT DEFAULT '#4da8da',
    bg_color TEXT DEFAULT '#e8f4fd',
    url TEXT DEFAULT '/',
    is_active INTEGER DEFAULT 1,
    is_coming_soon INTEGER DEFAULT 0,
    sort_order INTEGER DEFAULT 0,
    required_permission TEXT DEFAULT ''
);
CREATE TABLE IF NOT EXISTS user_app_access (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    user_id INTEGER NOT NULL,
    app_slug TEXT NOT NULL,
    app_role TEXT DEFAULT 'user',
    is_active INTEGER DEFAULT 1,
    granted_at TEXT DEFAULT (datetime('now','localtime')),
    UNIQUE(user_id, app_slug),
    FOREIGN KEY(user_id) REFERENCES users(id) ON DELETE CASCADE
);
CREATE TABLE IF NOT EXISTS app_menus (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    app_slug TEXT NOT NULL,
    parent_id INTEGER,
    title TEXT NOT NULL,
    url TEXT NOT NULL,
    icon TEXT DEFAULT 'circle',
    required_permission TEXT DEFAULT '',
    sort_order INTEGER DEFAULT 0,
    is_active INTEGER DEFAULT 1,
    FOREIGN KEY(parent_id) REFERENCES app_menus(id) ON DELETE CASCADE
);
CREATE TABLE IF NOT EXISTS role_menus (
    role_name TEXT NOT NULL,
    menu_id INTEGER NOT NULL,
    PRIMARY KEY (role_name, menu_id),
    FOREIGN KEY(menu_id) REFERENCES app_menus(id) ON DELETE CASCADE
);


-- ─── SupportCore: Ticket History & Multi-Assignee ────────────────────────────
CREATE TABLE IF NOT EXISTS sc_ticket_attachments (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    ticket_id INTEGER NOT NULL,
    section TEXT NOT NULL DEFAULT 'description',
    filename TEXT NOT NULL,
    original_name TEXT DEFAULT '',
    uploaded_by INTEGER,
    uploaded_by_name TEXT DEFAULT '',
    created_at TEXT DEFAULT (datetime('now','localtime')),
    FOREIGN KEY(ticket_id) REFERENCES sc_tickets(id) ON DELETE CASCADE
);
CREATE TABLE IF NOT EXISTS sc_ticket_history (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    ticket_id INTEGER NOT NULL,
    changed_by INTEGER,
    changed_by_name TEXT DEFAULT '',
    action TEXT DEFAULT '',
    field_name TEXT DEFAULT '',
    old_value TEXT DEFAULT '',
    new_value TEXT DEFAULT '',
    notes TEXT DEFAULT '',
    created_at TEXT DEFAULT (datetime('now','localtime')),
    FOREIGN KEY(ticket_id) REFERENCES sc_tickets(id) ON DELETE CASCADE
);
CREATE TABLE IF NOT EXISTS sc_ticket_external_assignees (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    ticket_id INTEGER NOT NULL,
    name TEXT NOT NULL,
    role_note TEXT DEFAULT '',
    added_by INTEGER,
    created_at TEXT DEFAULT (datetime('now','localtime')),
    FOREIGN KEY(ticket_id) REFERENCES sc_tickets(id) ON DELETE CASCADE
);
CREATE TABLE IF NOT EXISTS pc_task_external_assignees (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    task_id INTEGER NOT NULL,
    name TEXT NOT NULL,
    added_by INTEGER,
    created_at TEXT DEFAULT (datetime('now','localtime')),
    FOREIGN KEY(task_id) REFERENCES pc_tasks(id) ON DELETE CASCADE
);
CREATE TABLE IF NOT EXISTS sc_ticket_assignees (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    ticket_id INTEGER NOT NULL,
    employee_id INTEGER NOT NULL,
    divisi TEXT DEFAULT '',
    role_note TEXT DEFAULT '',
    assigned_at TEXT DEFAULT (datetime('now','localtime')),
    UNIQUE(ticket_id, employee_id),
    FOREIGN KEY(ticket_id) REFERENCES sc_tickets(id) ON DELETE CASCADE,
    FOREIGN KEY(employee_id) REFERENCES employees(id) ON DELETE CASCADE
);
-- ─── SupportCore: Presales & POC ─────────────────────────────────────────────
CREATE TABLE IF NOT EXISTS sc_presales_requests (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    req_no TEXT NOT NULL UNIQUE,
    customer_id INTEGER NOT NULL,
    request_type TEXT DEFAULT 'presales',
    subject TEXT NOT NULL,
    description TEXT DEFAULT '',
    status TEXT DEFAULT 'new',
    status_note TEXT DEFAULT '',
    created_by INTEGER,
    created_at TEXT DEFAULT (datetime('now','localtime')),
    FOREIGN KEY(customer_id) REFERENCES sc_customers(id)
);
CREATE TABLE IF NOT EXISTS sc_presales_assignees (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    request_id INTEGER NOT NULL,
    employee_id INTEGER NOT NULL,
    divisi TEXT DEFAULT '',
    role_note TEXT DEFAULT '',
    UNIQUE(request_id, employee_id),
    FOREIGN KEY(request_id) REFERENCES sc_presales_requests(id) ON DELETE CASCADE,
    FOREIGN KEY(employee_id) REFERENCES employees(id) ON DELETE CASCADE
);
CREATE TABLE IF NOT EXISTS sc_presales_history (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    request_id INTEGER NOT NULL,
    changed_by INTEGER,
    changed_by_name TEXT DEFAULT '',
    action TEXT DEFAULT '',
    notes TEXT DEFAULT '',
    created_at TEXT DEFAULT (datetime('now','localtime')),
    FOREIGN KEY(request_id) REFERENCES sc_presales_requests(id) ON DELETE CASCADE
);

-- ─── Audit Trail ─────────────────────────────────────────────────────────────
CREATE TABLE IF NOT EXISTS audit_activity (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    app_slug TEXT NOT NULL DEFAULT 'portal',
    user_id INTEGER,
    username TEXT DEFAULT '',
    action TEXT NOT NULL,
    resource TEXT DEFAULT '',
    resource_id TEXT DEFAULT '',
    detail TEXT DEFAULT '',
    ip TEXT DEFAULT '',
    user_agent TEXT DEFAULT '',
    created_at TEXT DEFAULT (datetime('now','localtime')),
    FOREIGN KEY(user_id) REFERENCES users(id) ON DELETE SET NULL
);
CREATE TABLE IF NOT EXISTS audit_errors (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    app_slug TEXT NOT NULL DEFAULT 'portal',
    user_id INTEGER,
    username TEXT DEFAULT '',
    url TEXT DEFAULT '',
    method TEXT DEFAULT '',
    error_code INTEGER DEFAULT 500,
    error_type TEXT DEFAULT '',
    error_msg TEXT DEFAULT '',
    traceback TEXT DEFAULT '',
    ip TEXT DEFAULT '',
    created_at TEXT DEFAULT (datetime('now','localtime')),
    FOREIGN KEY(user_id) REFERENCES users(id) ON DELETE SET NULL
);
CREATE TABLE IF NOT EXISTS audit_notifications (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    app_slug TEXT NOT NULL DEFAULT 'portal',
    channel TEXT DEFAULT '',
    recipient TEXT DEFAULT '',
    subject TEXT DEFAULT '',
    message TEXT DEFAULT '',
    status TEXT DEFAULT 'sent',
    error_msg TEXT DEFAULT '',
    triggered_by TEXT DEFAULT '',
    created_at TEXT DEFAULT (datetime('now','localtime'))
);

-- ─── SupportCore ─────────────────────────────────────────────────────────────
CREATE TABLE IF NOT EXISTS sc_apps (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    name TEXT NOT NULL UNIQUE,
    description TEXT DEFAULT '',
    is_active INTEGER DEFAULT 1,
    created_at TEXT DEFAULT (datetime('now','localtime'))
);
CREATE TABLE IF NOT EXISTS sc_modules (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    app_id INTEGER NOT NULL,
    name TEXT NOT NULL,
    description TEXT DEFAULT '',
    is_active INTEGER DEFAULT 1,
    created_at TEXT DEFAULT (datetime('now','localtime')),
    FOREIGN KEY(app_id) REFERENCES sc_apps(id) ON DELETE CASCADE,
    UNIQUE(app_id, name)
);
CREATE TABLE IF NOT EXISTS sc_customers (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    code TEXT NOT NULL UNIQUE,
    name TEXT NOT NULL,
    address TEXT DEFAULT '',
    contact_person TEXT DEFAULT '',
    phone TEXT DEFAULT '',
    email TEXT DEFAULT '',
    notes TEXT DEFAULT '',
    is_active INTEGER DEFAULT 1,
    created_at TEXT DEFAULT (datetime('now','localtime'))
);
CREATE TABLE IF NOT EXISTS sc_services (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    code TEXT NOT NULL UNIQUE,
    name TEXT NOT NULL,
    description TEXT DEFAULT '',
    is_active INTEGER DEFAULT 1,
    created_at TEXT DEFAULT (datetime('now','localtime'))
);
CREATE TABLE IF NOT EXISTS sc_support_types (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    code TEXT NOT NULL UNIQUE,
    name TEXT NOT NULL,
    description TEXT DEFAULT '',
    is_active INTEGER DEFAULT 1,
    created_at TEXT DEFAULT (datetime('now','localtime'))
);
CREATE TABLE IF NOT EXISTS sc_sla_categories (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    code TEXT NOT NULL UNIQUE,
    name TEXT NOT NULL,
    priority TEXT DEFAULT 'Medium',
    response_time_hours REAL DEFAULT 4,
    resolution_time_hours REAL DEFAULT 24,
    description TEXT DEFAULT '',
    is_active INTEGER DEFAULT 1,
    created_at TEXT DEFAULT (datetime('now','localtime'))
);
CREATE TABLE IF NOT EXISTS sc_contracts (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    code TEXT NOT NULL UNIQUE,
    customer_id INTEGER NOT NULL,
    title TEXT NOT NULL,
    description TEXT DEFAULT '',
    start_date TEXT NOT NULL,
    end_date TEXT NOT NULL,
    contract_value REAL DEFAULT 0,
    status TEXT DEFAULT 'active',
    notes TEXT DEFAULT '',
    created_by INTEGER,
    created_at TEXT DEFAULT (datetime('now','localtime')),
    FOREIGN KEY(customer_id) REFERENCES sc_customers(id)
);
CREATE TABLE IF NOT EXISTS sc_contract_services (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    contract_id INTEGER NOT NULL,
    service_id INTEGER NOT NULL,
    UNIQUE(contract_id, service_id),
    FOREIGN KEY(contract_id) REFERENCES sc_contracts(id) ON DELETE CASCADE,
    FOREIGN KEY(service_id) REFERENCES sc_services(id)
);
CREATE TABLE IF NOT EXISTS sc_contract_support_types (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    contract_id INTEGER NOT NULL,
    support_type_id INTEGER NOT NULL,
    UNIQUE(contract_id, support_type_id),
    FOREIGN KEY(contract_id) REFERENCES sc_contracts(id) ON DELETE CASCADE,
    FOREIGN KEY(support_type_id) REFERENCES sc_support_types(id)
);
CREATE TABLE IF NOT EXISTS sc_tickets (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    ticket_no TEXT NOT NULL UNIQUE,
    contract_id INTEGER DEFAULT NULL,
    customer_id INTEGER NOT NULL,
    support_type_id INTEGER NOT NULL,
    sla_category_id INTEGER DEFAULT NULL,
    subject TEXT NOT NULL,
    description TEXT DEFAULT '',
    status TEXT DEFAULT 'open',
    reported_by TEXT DEFAULT '',
    reported_at TEXT NOT NULL DEFAULT (datetime('now','localtime')),
    responded_at TEXT DEFAULT NULL,
    resolved_at TEXT DEFAULT NULL,
    closed_at TEXT DEFAULT NULL,
    assigned_to INTEGER DEFAULT NULL,
    created_by INTEGER DEFAULT NULL,
    notes TEXT DEFAULT '',
    created_at TEXT DEFAULT (datetime('now','localtime')),
    FOREIGN KEY(customer_id) REFERENCES sc_customers(id),
    FOREIGN KEY(contract_id) REFERENCES sc_contracts(id),
    FOREIGN KEY(support_type_id) REFERENCES sc_support_types(id),
    FOREIGN KEY(sla_category_id) REFERENCES sc_sla_categories(id)
);
CREATE TABLE IF NOT EXISTS bk_resources (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    name TEXT NOT NULL,
    type TEXT DEFAULT 'room',
    subtype TEXT DEFAULT '',
    capacity INTEGER DEFAULT 0,
    description TEXT DEFAULT '',
    location TEXT DEFAULT '',
    color TEXT DEFAULT '#d97706',
    icon TEXT DEFAULT 'door-open',
    is_active INTEGER DEFAULT 1,
    sort_order INTEGER DEFAULT 0,
    created_at TEXT DEFAULT (datetime('now','localtime'))
);
CREATE TABLE IF NOT EXISTS bk_bookings (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    resource_id INTEGER NOT NULL,
    title TEXT NOT NULL,
    purpose TEXT DEFAULT '',
    booked_by INTEGER NOT NULL,
    attendees TEXT DEFAULT '',
    attendee_count INTEGER DEFAULT 0,
    start_dt TEXT NOT NULL,
    end_dt TEXT NOT NULL,
    status TEXT DEFAULT 'confirmed',
    notes TEXT DEFAULT '',
    is_recurring INTEGER DEFAULT 0,
    recurring_type TEXT DEFAULT '',
    recurring_days TEXT DEFAULT '',
    recurring_until TEXT DEFAULT '',
    parent_id INTEGER DEFAULT NULL,
    destination TEXT DEFAULT '',
    cancelled_at TEXT DEFAULT '',
    cancelled_by INTEGER DEFAULT NULL,
    created_at TEXT DEFAULT (datetime('now','localtime'))
);
CREATE TABLE IF NOT EXISTS bk_resource_images (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    resource_id INTEGER NOT NULL REFERENCES bk_resources(id) ON DELETE CASCADE,
    image TEXT NOT NULL,
    caption TEXT DEFAULT '',
    sort_order INTEGER DEFAULT 0,
    created_at TEXT DEFAULT (datetime('now','localtime'))
);
CREATE TABLE IF NOT EXISTS bk_items (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    name TEXT NOT NULL,
    category TEXT DEFAULT 'other',
    description TEXT DEFAULT '',
    icon TEXT DEFAULT 'box',
    is_active INTEGER DEFAULT 1,
    sort_order INTEGER DEFAULT 0
);
CREATE TABLE IF NOT EXISTS bk_booking_items (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    booking_id INTEGER NOT NULL REFERENCES bk_bookings(id) ON DELETE CASCADE,
    item_id INTEGER NOT NULL REFERENCES bk_items(id),
    qty INTEGER DEFAULT 1,
    notes TEXT DEFAULT ''
);
CREATE TABLE IF NOT EXISTS ac_assets (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    employee_id INTEGER REFERENCES employees(id) ON DELETE SET NULL,
    device_type TEXT DEFAULT 'Laptop',
    brand TEXT DEFAULT '',
    os TEXT DEFAULT '',
    os_license_type TEXT DEFAULT '',
    processor TEXT DEFAULT '',
    ram TEXT DEFAULT '',
    disk TEXT DEFAULT '',
    office_version TEXT DEFAULT '',
    asset_tag TEXT DEFAULT '',
    serial_number TEXT DEFAULT '',
    purchase_date TEXT DEFAULT '',
    condition TEXT DEFAULT 'Baik',
    notes TEXT DEFAULT '',
    created_at TEXT DEFAULT (datetime('now','localtime')),
    updated_at TEXT DEFAULT (datetime('now','localtime'))
);
CREATE TABLE IF NOT EXISTS ac_asset_software (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    asset_id INTEGER REFERENCES ac_assets(id) ON DELETE CASCADE,
    software_name TEXT NOT NULL
);
CREATE TABLE IF NOT EXISTS ac_asset_history (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    asset_id INTEGER NOT NULL REFERENCES ac_assets(id) ON DELETE CASCADE,
    employee_id INTEGER REFERENCES employees(id) ON DELETE SET NULL,
    manual_employee_name TEXT DEFAULT '',
    started_at TEXT DEFAULT '',
    ended_at TEXT DEFAULT '',
    reason TEXT DEFAULT '',
    notes TEXT DEFAULT '',
    created_at TEXT DEFAULT (datetime('now','localtime'))
);
CREATE TABLE IF NOT EXISTS ac_infrastructure (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    device_type TEXT NOT NULL,
    brand TEXT DEFAULT '',
    model TEXT DEFAULT '',
    description TEXT DEFAULT '',
    serial_number TEXT DEFAULT '',
    nickname TEXT DEFAULT '',
    ups_group TEXT DEFAULT '',
    location TEXT DEFAULT '',
    status TEXT DEFAULT 'Aktif',
    condition_notes TEXT DEFAULT '',
    created_at TEXT DEFAULT (datetime('now','localtime'))
);
CREATE TABLE IF NOT EXISTS ac_licenses (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    software_name TEXT NOT NULL,
    license_key TEXT DEFAULT '',
    license_type TEXT DEFAULT 'Perpetual',
    version TEXT DEFAULT '',
    year INTEGER DEFAULT NULL,
    max_seats INTEGER DEFAULT 1,
    is_active INTEGER DEFAULT 1,
    notes TEXT DEFAULT '',
    created_at TEXT DEFAULT (datetime('now','localtime'))
);
CREATE TABLE IF NOT EXISTS ac_license_assignments (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    license_id INTEGER REFERENCES ac_licenses(id) ON DELETE CASCADE,
    employee_id INTEGER REFERENCES employees(id) ON DELETE SET NULL,
    seat_number INTEGER DEFAULT 1,
    assigned_at TEXT DEFAULT (datetime('now','localtime'))
);
CREATE TABLE IF NOT EXISTS ac_subscriptions (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    provider TEXT NOT NULL,
    category TEXT DEFAULT 'SaaS',
    billing_cycle TEXT DEFAULT 'Monthly',
    start_date TEXT DEFAULT '',
    end_date TEXT DEFAULT '',
    username TEXT DEFAULT '',
    password TEXT DEFAULT '',
    access_url TEXT DEFAULT '',
    notes TEXT DEFAULT '',
    is_active INTEGER DEFAULT 1,
    created_at TEXT DEFAULT (datetime('now','localtime'))
);
CREATE TABLE IF NOT EXISTS ac_software_requests (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    employee_id INTEGER REFERENCES employees(id) ON DELETE SET NULL,
    software_name TEXT NOT NULL,
    version TEXT DEFAULT '',
    reason TEXT DEFAULT '',
    status TEXT DEFAULT 'Pending',
    requested_at TEXT DEFAULT (datetime('now','localtime')),
    resolved_at TEXT DEFAULT '',
    resolved_by TEXT DEFAULT '',
    notes TEXT DEFAULT ''
);
CREATE TABLE IF NOT EXISTS ac_tool_requests (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    employee_id INTEGER REFERENCES employees(id) ON DELETE SET NULL,
    manual_user_name TEXT DEFAULT '',
    requestor_name TEXT DEFAULT '',
    item_name TEXT NOT NULL,
    item_category TEXT DEFAULT 'Laptop',
    request_channel TEXT DEFAULT 'Email',
    request_channel_other TEXT DEFAULT '',
    reason TEXT DEFAULT '',
    status TEXT DEFAULT 'Pending',
    admin_item_type TEXT DEFAULT '',
    admin_specs TEXT DEFAULT '',
    admin_url TEXT DEFAULT '',
    admin_price REAL DEFAULT 0.0,
    request_date TEXT DEFAULT '',
    purchase_date TEXT DEFAULT '',
    received_date TEXT DEFAULT '',
    receipt_date TEXT DEFAULT '',
    pic_support TEXT DEFAULT '',
    ket TEXT DEFAULT '',
    spec_cpu_type TEXT DEFAULT '',
    spec_ram TEXT DEFAULT '',
    spec_disk TEXT DEFAULT '',
    spec_gpu TEXT DEFAULT '',
    spec_screen TEXT DEFAULT '',
    spec_os TEXT DEFAULT '',
    spec_office TEXT DEFAULT '',
    asset_tag TEXT DEFAULT '',
    serial_number TEXT DEFAULT '',
    asset_id INTEGER REFERENCES ac_assets(id) ON DELETE SET NULL,
    requested_at TEXT DEFAULT (datetime('now','localtime')),
    resolved_at TEXT DEFAULT '',
    resolved_by TEXT DEFAULT '',
    notes TEXT DEFAULT '',
    updated_at TEXT DEFAULT ''
);
CREATE TABLE IF NOT EXISTS ac_tool_request_attachments (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    request_id INTEGER NOT NULL REFERENCES ac_tool_requests(id) ON DELETE CASCADE,
    section TEXT NOT NULL DEFAULT 'request_capture',
    filename TEXT NOT NULL,
    original_name TEXT DEFAULT '',
    uploaded_by INTEGER,
    uploaded_by_name TEXT DEFAULT '',
    created_at TEXT DEFAULT (datetime('now','localtime'))
);
CREATE TABLE IF NOT EXISTS ac_maintenance (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    title TEXT NOT NULL,
    category TEXT DEFAULT 'Lainnya',
    location TEXT DEFAULT '',
    description TEXT DEFAULT '',
    frequency TEXT DEFAULT 'Bulanan',
    last_maintenance TEXT DEFAULT '',
    next_maintenance TEXT DEFAULT '',
    vendor TEXT DEFAULT '',
    pic TEXT DEFAULT '',
    cost_estimate REAL DEFAULT NULL,
    is_active INTEGER DEFAULT 1,
    notes TEXT DEFAULT '',
    created_at TEXT DEFAULT (datetime('now','localtime')),
    updated_at TEXT DEFAULT (datetime('now','localtime'))
);
CREATE TABLE IF NOT EXISTS ac_maintenance_log (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    maintenance_id INTEGER NOT NULL REFERENCES ac_maintenance(id) ON DELETE CASCADE,
    done_at TEXT DEFAULT '',
    done_by TEXT DEFAULT '',
    cost REAL DEFAULT NULL,
    result TEXT DEFAULT 'OK',
    notes TEXT DEFAULT '',
    created_at TEXT DEFAULT (datetime('now','localtime'))
);
CREATE TABLE IF NOT EXISTS pc_projects (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    code TEXT NOT NULL UNIQUE,
    name TEXT NOT NULL,
    client TEXT DEFAULT '',
    customer_id INTEGER DEFAULT NULL REFERENCES sc_customers(id) ON DELETE SET NULL,
    description TEXT DEFAULT '',
    status TEXT DEFAULT 'active',
    start_date TEXT DEFAULT NULL,
    end_date TEXT DEFAULT NULL,
    pic_id INTEGER DEFAULT NULL REFERENCES employees(id) ON DELETE SET NULL,
    implementor_id INTEGER DEFAULT NULL REFERENCES employees(id) ON DELETE SET NULL,
    co_leader_id INTEGER DEFAULT NULL REFERENCES employees(id) ON DELETE SET NULL,
    pic_ext TEXT DEFAULT '',
    implementor_ext TEXT DEFAULT '',
    co_leader_ext TEXT DEFAULT '',
    color TEXT DEFAULT '#0ea5e9',
    created_by INTEGER DEFAULT NULL REFERENCES users(id),
    created_at TEXT DEFAULT (datetime('now','localtime'))
);
CREATE TABLE IF NOT EXISTS pc_members (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    project_id INTEGER NOT NULL REFERENCES pc_projects(id) ON DELETE CASCADE,
    employee_id INTEGER DEFAULT NULL REFERENCES employees(id) ON DELETE CASCADE,
    name_ext TEXT DEFAULT '',
    role TEXT DEFAULT 'developer'
);
CREATE TABLE IF NOT EXISTS pc_issues (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    project_id INTEGER NOT NULL REFERENCES pc_projects(id) ON DELETE CASCADE,
    issue_no TEXT NOT NULL,
    title TEXT NOT NULL,
    description TEXT DEFAULT '',
    role TEXT DEFAULT '',
    menu TEXT DEFAULT '',
    stage TEXT DEFAULT '',
    solution_type TEXT DEFAULT '',
    priority TEXT DEFAULT 'Medium',
    severity TEXT DEFAULT 'Medium',
    difficulty TEXT DEFAULT 'Normal',
    issued_type TEXT DEFAULT 'Bugs',
    issued_date TEXT DEFAULT NULL,
    issued_by TEXT DEFAULT '',
    pic_programmer_id INTEGER DEFAULT NULL REFERENCES employees(id) ON DELETE SET NULL,
    pic_tester_id INTEGER DEFAULT NULL REFERENCES employees(id) ON DELETE SET NULL,
    md_days REAL DEFAULT NULL,
    plan_hours REAL DEFAULT NULL,
    bobot_plan REAL DEFAULT 0,
    bobot_actual REAL DEFAULT 0,
    status_programmer TEXT DEFAULT 'New',
    status_testing TEXT DEFAULT '',
    testing_date TEXT DEFAULT NULL,
    resolved_date TEXT DEFAULT NULL,
    notes TEXT DEFAULT '',
    redmine TEXT DEFAULT '',
    created_at TEXT DEFAULT (datetime('now','localtime'))
);
CREATE TABLE IF NOT EXISTS pc_issue_history (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    issue_id INTEGER NOT NULL REFERENCES pc_issues(id) ON DELETE CASCADE,
    action TEXT NOT NULL,
    old_value TEXT DEFAULT '',
    new_value TEXT DEFAULT '',
    changed_by INTEGER DEFAULT NULL REFERENCES users(id),
    notes TEXT DEFAULT '',
    created_at TEXT DEFAULT (datetime('now','localtime'))
);
CREATE TABLE IF NOT EXISTS pc_milestones (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    project_id INTEGER NOT NULL REFERENCES pc_projects(id) ON DELETE CASCADE,
    title TEXT NOT NULL,
    description TEXT DEFAULT '',
    due_date TEXT NOT NULL,
    status TEXT DEFAULT 'upcoming',
    sort_order INTEGER DEFAULT 0,
    created_at TEXT DEFAULT (datetime('now','localtime'))
);
CREATE TABLE IF NOT EXISTS pc_tasks (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    project_id INTEGER NOT NULL REFERENCES pc_projects(id) ON DELETE CASCADE,
    milestone_id INTEGER DEFAULT NULL REFERENCES pc_milestones(id) ON DELETE SET NULL,
    title TEXT NOT NULL,
    description TEXT DEFAULT '',
    status TEXT DEFAULT 'backlog',
    priority TEXT DEFAULT 'Medium',
    difficulty TEXT DEFAULT 'Normal',
    due_date TEXT DEFAULT NULL,
    sort_order INTEGER DEFAULT 0,
    created_at TEXT DEFAULT (datetime('now','localtime'))
);
CREATE TABLE IF NOT EXISTS pc_task_assignees (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    task_id INTEGER NOT NULL REFERENCES pc_tasks(id) ON DELETE CASCADE,
    employee_id INTEGER NOT NULL REFERENCES employees(id) ON DELETE CASCADE,
    self_assigned INTEGER DEFAULT 0,
    UNIQUE(task_id, employee_id)
);
CREATE TABLE IF NOT EXISTS peer_review_dimensions (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    eval_id INTEGER NOT NULL REFERENCES evaluations(id) ON DELETE CASCADE,
    slot INTEGER DEFAULT 1,
    reviewer_name TEXT DEFAULT '',
    dim_kerjasama INTEGER DEFAULT NULL,
    dim_komunikasi INTEGER DEFAULT NULL,
    dim_keandalan INTEGER DEFAULT NULL,
    dim_inisiatif INTEGER DEFAULT NULL,
    dim_kualitas INTEGER DEFAULT NULL,
    feedback TEXT DEFAULT '',
    created_at TEXT DEFAULT (datetime('now','localtime'))
);
CREATE TABLE IF NOT EXISTS grade_benchmarks (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    divisi TEXT NOT NULL,
    level TEXT NOT NULL,
    benchmark_per_month REAL DEFAULT 100,
    notes TEXT DEFAULT '',
    UNIQUE(divisi, level)
);
CREATE TABLE IF NOT EXISTS pc_proposed_changes (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    project_id INTEGER NOT NULL REFERENCES pc_projects(id) ON DELETE CASCADE,
    title TEXT NOT NULL,
    description TEXT DEFAULT '',
    module TEXT DEFAULT '',
    pekerjaan TEXT DEFAULT '',
    impact TEXT DEFAULT 'Medium',
    difficulty TEXT DEFAULT 'Normal',
    status TEXT DEFAULT 'proposed',
    tester TEXT DEFAULT '',
    notes TEXT DEFAULT '',
    created_at TEXT DEFAULT (datetime('now','localtime'))
);
CREATE TABLE IF NOT EXISTS pc_phases (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    project_id INTEGER NOT NULL REFERENCES pc_projects(id) ON DELETE CASCADE,
    phase_type TEXT DEFAULT 'custom',
    name TEXT NOT NULL,
    start_date TEXT DEFAULT NULL,
    end_date TEXT DEFAULT NULL,
    status TEXT DEFAULT 'planned',
    sort_order INTEGER DEFAULT 0,
    pic_id INTEGER DEFAULT NULL REFERENCES employees(id) ON DELETE SET NULL,
    pic_ext TEXT DEFAULT '',
    sign_off_date TEXT DEFAULT NULL,
    notes TEXT DEFAULT '',
    created_at TEXT DEFAULT (datetime('now','localtime'))
);

-- ─── Task Performance Config ─────────────────────────────────────────────────
CREATE TABLE IF NOT EXISTS task_perf_config (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    task_type TEXT NOT NULL UNIQUE,
    label TEXT NOT NULL,
    base_points REAL DEFAULT 0,
    priority_critical REAL DEFAULT 2.0,
    priority_high REAL DEFAULT 1.5,
    priority_medium REAL DEFAULT 1.0,
    priority_low REAL DEFAULT 0.7,
    ontime_mult REAL DEFAULT 1.1,
    late_mult REAL DEFAULT 0.9,
    sort_order INTEGER DEFAULT 0
);

-- ─── Notification Settings ───────────────────────────────────────────────────
CREATE TABLE IF NOT EXISTS notif_type_settings (
    slug TEXT PRIMARY KEY,
    is_active INTEGER DEFAULT 1,
    label TEXT NOT NULL DEFAULT '',
    description TEXT DEFAULT ''
);
CREATE TABLE IF NOT EXISTS notif_recipients (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    name TEXT NOT NULL DEFAULT '',
    channel TEXT NOT NULL DEFAULT 'email',
    address TEXT NOT NULL DEFAULT '',
    notif_types TEXT NOT NULL DEFAULT '["*"]',
    is_active INTEGER DEFAULT 1,
    created_at TEXT DEFAULT (datetime('now','localtime'))
);

CREATE TABLE IF NOT EXISTS attendance (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    user_id INTEGER NOT NULL,
    date TEXT NOT NULL,
    clock_in TEXT,
    clock_out TEXT,
    location_in TEXT,
    location_out TEXT,
    notes_in TEXT,
    notes_out TEXT,
    status TEXT,
    plan TEXT DEFAULT '',
    progress TEXT DEFAULT '',
    created_at TEXT DEFAULT (datetime('now','localtime'))
);
CREATE TABLE IF NOT EXISTS attendance_leaves (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    user_id INTEGER NOT NULL,
    leave_type TEXT NOT NULL,
    start_date TEXT NOT NULL,
    end_date TEXT NOT NULL,
    reason TEXT,
    status TEXT DEFAULT 'pending',
    approved_by INTEGER,
    created_at TEXT DEFAULT (datetime('now','localtime'))
);
CREATE TABLE IF NOT EXISTS attendance_overtime (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    user_id INTEGER NOT NULL,
    date TEXT NOT NULL,
    start_time TEXT,
    end_time TEXT,
    hours REAL DEFAULT 0,
    reason TEXT,
    status TEXT DEFAULT 'pending',
    approved_by INTEGER,
    created_at TEXT DEFAULT (datetime('now','localtime'))
);
CREATE TABLE IF NOT EXISTS ac_masters (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    category TEXT NOT NULL,
    name TEXT NOT NULL,
    UNIQUE(category, name)
);
CREATE TABLE IF NOT EXISTS attendance_corrections (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    user_id INTEGER NOT NULL,
    date TEXT NOT NULL,
    requested_clock_in TEXT,
    requested_clock_out TEXT,
    reason TEXT,
    status TEXT DEFAULT 'pending',
    approved_by INTEGER,
    created_at TEXT DEFAULT (datetime('now','localtime'))
);
"""

# ALTER existing column types (PostgreSQL only) — run before MIGRATIONS ADD COLUMN
COLUMN_TYPE_MIGRATIONS = [
    # employee_salary columns were originally REAL but now store Fernet-encrypted TEXT
    ('employee_salary', 'base_salary', 'TEXT', "''"),
    ('employee_salary', 'al_001',      'TEXT', "''"),
    ('employee_salary', 'al_002',      'TEXT', "''"),
    ('employee_salary', 'al_003',      'TEXT', "''"),
    ('employee_salary', 'al_004',      'TEXT', "''"),
]

MIGRATIONS = [
    ('users', 'email',     "TEXT DEFAULT ''"),
    ('users', 'google_id', "TEXT DEFAULT ''"),
    ('employees', 'birthday',        "TEXT DEFAULT ''"),
    ('employees', 'level',           "TEXT DEFAULT 'Staff'"),
    ('employees', 'employment_type', "TEXT DEFAULT 'tetap'"),
    ('employees', 'contract_start',  "TEXT DEFAULT ''"),
    ('employees', 'contract_end',    "TEXT DEFAULT ''"),
    ('employees', 'rate_mandays',    'REAL DEFAULT NULL'),
    ('employees', 'email',           "TEXT DEFAULT ''"),
    ('employees', 'phone',           "TEXT DEFAULT ''"),
    ('employees', 'telegram_id',     "TEXT DEFAULT ''"),
    ('employees', 'is_active',       "INTEGER DEFAULT 1"),
    ('employees', 'notes',           "TEXT DEFAULT ''"),
    ('employees', 'supervisor_id',   'INTEGER DEFAULT NULL'),
    ('employees', 'leader_id',       'INTEGER DEFAULT NULL'),
    ('employees', 'manager_id',      'INTEGER DEFAULT NULL'),
    ('employees', 'user_id',         'INTEGER DEFAULT NULL'),
    ('evaluations', 'self_notes',        "TEXT DEFAULT ''"),
    ('evaluations', 'self_achievements', "TEXT DEFAULT ''"),
    ('evaluations', 'self_improvements', "TEXT DEFAULT ''"),
    ('evaluations', 'self_assessment_json', "TEXT DEFAULT ''"),
    ('evaluations', 'review_status',     "TEXT DEFAULT 'draft'"),
    ('evaluations', 'reviewed_by',       'INTEGER DEFAULT NULL'),
    ('evaluations', 'reviewed_at',       "TEXT DEFAULT ''"),
    ('evaluations', 'review_notes',      "TEXT DEFAULT ''"),
    ('users',       'totp_secret',       "TEXT DEFAULT ''"),
    ('users',       'mfa_enabled',       "INTEGER DEFAULT 0"),
    ('users',       'email',             "TEXT DEFAULT ''"),
    ('users',       'phone',             "TEXT DEFAULT ''"),
    ('users',       'telegram_id',       "TEXT DEFAULT ''"),
    ('employees',   'salary',            "TEXT DEFAULT ''"),
    ('employee_salary', 'increase_date', "TEXT DEFAULT ''"),
    ('roles',              'app_slug',               "TEXT DEFAULT 'evaluasi'"),
    ('sc_contracts',       'pic_helpdesk_id',         'INTEGER DEFAULT NULL'),
    ('sc_customers',       'pic_helpdesk_id',         'INTEGER DEFAULT NULL'),
    ('sc_customers',       'pic_helpdesk_backup_id',  'INTEGER DEFAULT NULL'),
    ('sc_customers',       'pic_implementor_id',      'INTEGER DEFAULT NULL'),
    ('sc_customers',       'pic_coleader_id',         'INTEGER DEFAULT NULL'),
    ('sc_customers',       'telegram_group_id',       "TEXT DEFAULT ''"),
    ('sc_tickets',         'module_id',               'INTEGER DEFAULT NULL'),
    ('sc_tickets',         'assignee_id',             'INTEGER DEFAULT NULL'),
    ('sc_tickets',         'status_note',             "TEXT DEFAULT ''"),
    ('sc_tickets',         'mandays',                 'REAL DEFAULT NULL'),
    ('sc_tickets',         'pct_done',                'INTEGER DEFAULT 0'),
    ('sc_tickets',         'solution_type',           "TEXT DEFAULT ''"),
    ('sc_tickets',         'solution_note',           "TEXT DEFAULT ''"),
    ('sc_tickets',         'due_date',                "TEXT DEFAULT NULL"),
    ('sc_tickets',         'work_start_date',         "TEXT DEFAULT NULL"),
    ('sc_tickets',         'media_lapor',             "TEXT DEFAULT ''"),
    ('sc_tickets',         'priority',                "TEXT DEFAULT 'Medium'"),
    ('sc_customers',       'customer_type',           "TEXT DEFAULT 'aktif'"),
    ('sc_customers',       'pic_sales_id',            'INTEGER DEFAULT NULL'),
    ('sc_sla_categories',  'workaround_time_hours',   'REAL DEFAULT NULL'),
    ('sc_sla_categories',  'maintenance_type',        "TEXT DEFAULT 'corrective'"),
    ('sc_sla_categories',  'priority',                "TEXT DEFAULT 'Medium'"),
    ('ac_assets',            'manual_employee_name',    "TEXT DEFAULT ''"),
    ('ac_assets',            'status',                  "TEXT DEFAULT 'Aktif'"),
    ('ac_assets',            'started_using',           "TEXT DEFAULT ''"),
    ('ac_subscriptions',     'last_reminder_sent',      "TEXT DEFAULT ''"),
    ('ac_infrastructure',    'updated_at',              "TEXT DEFAULT ''"),
    ('ac_licenses',          'updated_at',              "TEXT DEFAULT ''"),
    ('ac_subscriptions',     'updated_at',              "TEXT DEFAULT ''"),
    ('ac_software_requests', 'updated_at',              "TEXT DEFAULT ''"),
    ('ac_tool_requests',     'manual_user_name',        "TEXT DEFAULT ''"),
    ('ac_tool_requests',     'item_category',           "TEXT DEFAULT 'Laptop'"),
    ('ac_tool_requests',     'request_channel',         "TEXT DEFAULT 'Email'"),
    ('ac_tool_requests',     'request_channel_other',   "TEXT DEFAULT ''"),
    ('ac_tool_requests',     'request_date',            "TEXT DEFAULT ''"),
    ('ac_tool_requests',     'purchase_date',           "TEXT DEFAULT ''"),
    ('ac_tool_requests',     'requestor_name',          "TEXT DEFAULT ''"),
    ('ac_tool_requests',     'received_date',           "TEXT DEFAULT ''"),
    ('ac_tool_requests',     'receipt_date',            "TEXT DEFAULT ''"),
    ('ac_tool_requests',     'pic_support',             "TEXT DEFAULT ''"),
    ('ac_tool_requests',     'ket',                     "TEXT DEFAULT ''"),
    ('ac_tool_requests',     'spec_cpu_type',           "TEXT DEFAULT ''"),
    ('ac_tool_requests',     'spec_ram',                "TEXT DEFAULT ''"),
    ('ac_tool_requests',     'spec_disk',               "TEXT DEFAULT ''"),
    ('ac_tool_requests',     'spec_gpu',                "TEXT DEFAULT ''"),
    ('ac_tool_requests',     'spec_screen',             "TEXT DEFAULT ''"),
    ('ac_tool_requests',     'spec_os',                 "TEXT DEFAULT ''"),
    ('ac_tool_requests',     'spec_office',             "TEXT DEFAULT ''"),
    ('ac_tool_requests',     'asset_tag',               "TEXT DEFAULT ''"),
    ('ac_tool_requests',     'serial_number',           "TEXT DEFAULT ''"),
    ('ac_tool_requests',     'asset_id',                'INTEGER DEFAULT NULL'),
    ('bk_resources',         'image',                   "TEXT DEFAULT ''"),
    ('bk_resources',         'facilities',              "TEXT DEFAULT ''"),
    ('bk_resources',         'notes',                   "TEXT DEFAULT ''"),
    ('pc_projects',          'customer_id',             'INTEGER DEFAULT NULL'),
    ('pc_projects',          'implementor_id',          'INTEGER DEFAULT NULL'),
    ('pc_projects',          'co_leader_id',            'INTEGER DEFAULT NULL'),
    ('pc_projects',          'pic_ext',                 "TEXT DEFAULT ''"),
    ('pc_projects',          'implementor_ext',         "TEXT DEFAULT ''"),
    ('pc_projects',          'co_leader_ext',           "TEXT DEFAULT ''"),
    ('pc_members',           'name_ext',                "TEXT DEFAULT ''"),
    ('pc_projects',          'deleted_at',              'TEXT DEFAULT NULL'),
    ('pc_phases',            'pic_ext',                 "TEXT DEFAULT ''"),
    ('pc_phases',            'sign_off_date',           "TEXT DEFAULT NULL"),
    ('pc_phases',            'app_id',                  'INTEGER DEFAULT NULL'),
    ('pc_phases',            'module_id',               'INTEGER DEFAULT NULL'),
    ('evaluations',          'task_score',              'REAL DEFAULT NULL'),
    ('evaluations',          'task_date_from',          "TEXT DEFAULT ''"),
    ('evaluations',          'task_date_to',            "TEXT DEFAULT ''"),
    ('evaluations',          'task_benchmark',          'REAL DEFAULT 100'),
    ('notif_type_settings',  'label',                   "TEXT NOT NULL DEFAULT ''"),
    ('notif_type_settings',  'description',             "TEXT DEFAULT ''"),
    ('notif_type_settings',  'is_active',               'INTEGER DEFAULT 1'),
    ('notif_recipients',     'name',                    "TEXT NOT NULL DEFAULT ''"),
    ('notif_recipients',     'channel',                 "TEXT NOT NULL DEFAULT 'email'"),
    ('notif_recipients',     'address',                 "TEXT NOT NULL DEFAULT ''"),
    ('notif_recipients',     'notif_types',             'TEXT NOT NULL DEFAULT \'["*"]\''),
    ('notif_recipients',     'is_active',               'INTEGER DEFAULT 1'),
    ('notif_recipients',     'created_at',              "TEXT DEFAULT (datetime('now','localtime'))"),
    # Task difficulty & self_assigned
    ('pc_tasks',             'difficulty',              "TEXT DEFAULT 'Normal'"),
    ('pc_task_assignees',    'self_assigned',           'INTEGER DEFAULT 0'),
    # Peer review structured dimensions
    ('peer_reviews',         'dim_kerjasama',           'INTEGER DEFAULT NULL'),
    ('peer_reviews',         'dim_komunikasi',          'INTEGER DEFAULT NULL'),
    ('peer_reviews',         'dim_keandalan',           'INTEGER DEFAULT NULL'),
    ('peer_reviews',         'dim_inisiatif',           'INTEGER DEFAULT NULL'),
    ('peer_reviews',         'dim_kualitas',            'INTEGER DEFAULT NULL'),
    # Evaluasi AI narrative
    ('evaluations',          'ai_summary',              "TEXT DEFAULT ''"),
    ('evaluations',          'ai_recommendation',       "TEXT DEFAULT ''"),
    ('evaluations',          'ai_generated_at',         "TEXT DEFAULT ''"),
    # Attendance Plan & Progress
    ('attendance',           'plan',                    "TEXT DEFAULT ''"),
    ('attendance',           'progress',                "TEXT DEFAULT ''"),
    ('attendance',           'checkout_reminder_sent',  "INTEGER DEFAULT 0"),
]

SC_TICKET_PRIORITIES = [
    ('Critical', 'Critical', 'danger'),
    ('High',     'High',     'warning'),
    ('Medium',   'Medium',   'primary'),
    ('Low',      'Low',      'secondary'),
]

SC_TICKET_STATUSES = [
    ('new',         'Baru',        'secondary'),
    ('in_progress', 'In Progress', 'primary'),
    ('hold',        'Hold',        'warning'),
    ('resolved',    'Resolved',    'success'),
    ('feedback',    'Feedback',    'info'),
    ('closed',      'Closed',      'dark'),
    ('rejected',    'Rejected',    'danger'),
]

SC_SOLUTION_TYPES = [
    ('workaround_restart',    'Workaround — Restart Aplikasi'),
    ('workaround_db',         'Workaround — DB'),
    ('workaround_coding',     'Workaround — Coding'),
    ('workaround_suggestion', 'Workaround — Suggestion'),
    ('final_db',              'Final — DB'),
    ('final_coding',          'Final — Coding'),
    ('final_suggestion',      'Final — Suggestion'),
]

SC_MEDIA_LAPOR = [
    ('wa_helpdesk', 'WA ke Nomor Helpdesk'),
    ('email',       'Email'),
    ('telegram',    'Telegram'),
    ('wa_pic',      'WA Pribadi PIC Helpdesk'),
]

PC_ISSUE_STATUSES_PRG = [
    ('New',           'New',            'secondary'),
    ('In Progress',   'In Progress',    'primary'),
    ('Done',          'Done',           'success'),
    ('Ready to Test', 'Ready to Test',  'info'),
    ('Need Deploy',   'Need Deploy',    'warning'),
    ('Hold',          'Hold',           'dark'),
    ('Feedback',      'Feedback',       'danger'),
]
PC_ISSUE_STATUSES_TEST = [
    ('',         '—',        'light'),
    ('Testing',  'Testing',  'primary'),
    ('Done',     'Done',     'success'),
    ('Feedback', 'Feedback', 'warning'),
    ('Reject',   'Reject',   'danger'),
]
PC_PRIORITIES  = ['Low', 'Medium', 'High', 'Critical']
PC_SEVERITIES  = ['Low', 'Medium', 'High', 'Critical']
PC_DIFFICULTIES= ['Easy', 'Normal', 'Hard', 'Very Hard']
PC_ISSUED_TYPES= ['Bugs', 'New Feature', 'Enhancement', 'Change Request']
PC_SOLUTION_TYPES = ['Coding', 'Config', 'DB', 'Design', 'Documentation', 'Other']
PC_TASK_STATUSES = [
    ('backlog',     'Backlog',     '#6b7280'),
    ('todo',        'To Do',       '#3b82f6'),
    ('in_progress', 'In Progress', '#f59e0b'),
    ('review',      'Review',      '#8b5cf6'),
    ('done',        'Done',        '#10b981'),
]
PC_MILESTONE_STATUSES = [
    ('upcoming',     'Upcoming',     'secondary'),
    ('in_progress',  'In Progress',  'primary'),
    ('completed',    'Completed',    'success'),
    ('delayed',      'Delayed',      'danger'),
]
PC_PROPOSED_STATUSES = [
    ('proposed',    'Proposed',     'secondary'),
    ('approved',    'Approved',     'success'),
    ('in_progress', 'In Progress',  'primary'),
    ('done',        'Done',         'success'),
    ('rejected',    'Rejected',     'danger'),
    ('hold',        'Hold',         'warning'),
]
PC_PHASE_TYPES = [
    ('sit',     'SIT',                   '#3b82f6'),
    ('uat',     'UAT',                   '#8b5cf6'),
    ('bast',    'BAST',                  '#f59e0b'),
    ('promote', 'Promote to Production', '#ef4444'),
    ('golive',  'Go Live',               '#10b981'),
    ('custom',  'Custom',                '#6b7280'),
]
PC_PHASE_STATUSES = [
    ('planned',     'Planned',     'secondary'),
    ('in_progress', 'In Progress', 'primary'),
    ('done',        'Done',        'success'),
    ('skipped',     'Skipped',     'dark'),
]
PC_PROJECT_COLORS = [
    '#0ea5e9','#10b981','#f59e0b','#ef4444','#8b5cf6','#ec4899','#0d9488','#6366f1',
]

DEFAULT_SETTINGS = {
    'smtp_host': '',
    'smtp_port': '587',
    'smtp_user': '',
    'smtp_password': '',
    'smtp_from': '',
    'smtp_ssl': '0',
    'telegram_bot_token': '',
    'telegram_default_chat_id': '',
    'reminder_days': '30,14,7,1',
    'reminder_enabled': '1',
    'app_name': 'TalentCore',
    'notification_emails': '',
    'notification_telegram_ids': '',
    'openwa_url': '',
    'openwa_api_key': '',
    'openwa_session_id': 'default',      # session fallback jika app tidak punya session sendiri
    'openwa_session_evaluasi': '',        # TalentCore
    'openwa_session_support': '',         # SupportCore
    'openwa_session_booking': '',         # BookingCore
    'openwa_session_aset': '',            # AssetCore
    'openwa_enabled': '0',
    'openwa_extra_phones': '',
    'app_url': '',  # URL publik aplikasi mis. https://evaluasi.perusahaan.com (kosong = auto-detect)
    'ac_sub_reminder_enabled': '1',
    'ac_sub_reminder_days': '30,14,7,1',
    'ac_notification_emails': '',
    'ac_notification_telegram_ids': '',
    'ac_notification_wa_phones': '',
    # Update Center
    'update_check_enabled':  '1',           # cek update otomatis
    'update_notify_roles':   'superadmin,admin',  # role yang dapat notifikasi
    'update_trigger_roles':  'superadmin',  # role yang bisa trigger update
    'update_available':      '0',
    'update_latest_version': '',
    'update_latest_tag':     '',
    'update_release_notes':  '',
    'update_check_last':     '',
    'github_repo':           'barangbaru/solid-apps',
    # AI Chatbot
    'chatbot_enabled':       '0',
    'ai_provider':           'anthropic',      # anthropic | openai | openai_compat
    'ai_api_key':            '',
    'ai_model':              '',               # kosong = pakai default per provider
    'ai_base_url':           '',               # hanya untuk openai_compat
    # backward compat
    'anthropic_api_key':     '',
    'backup_sched_enabled': '1',
    'backup_sched_interval': 'daily',
    'backup_sched_time': '02:00',
    'backup_target_app': '1',
    'backup_target_uploads': '1',
    'backup_target_db': '1',
    'backup_dest_email_enabled': '0',
    'backup_dest_email_recipient': '',
    'backup_dest_s3_enabled': '1',
    'backup_dest_s3_endpoint': '',
    'backup_dest_s3_access_key': '',
    'backup_dest_s3_secret_key': '',
    'backup_dest_s3_bucket': '',
    'backup_dest_s3_region': '',
    'backup_sched_last_run': '',
    'backup_last_status': '',
    'backup_last_log': '',
    'backup_retention_days': '30',
    'media_storage_type': 'local',
}

LEVEL_CHOICES = ['Staff', 'Senior Staff', 'Co-Leader', 'Leader', 'Manager', 'Senior Manager', 'General Manager', 'Director']

# Permissions per-app. Portal-level (manage_users, manage_roles) dikelola via superadmin.
APP_PERMISSIONS = {
    'evaluasi': {
        'manage_settings':    'Pengaturan notifikasi sistem',
        'manage_template':    'Edit template evaluasi (skill/kompetensi/ability)',
        'manage_employees':   'Tambah/edit/hapus data karyawan',
        'manage_divisions':   'Kelola divisi',
        'manage_evaluations': 'Buat/hapus evaluasi karyawan',
        'view_evaluations':   'Lihat hasil evaluasi',
        'send_reminders':     'Kirim reminder kontrak',
        'view_salary':        'Lihat data gaji karyawan',
        'manage_salary':      'Edit data gaji karyawan',
    },
    'aset': {
        'ac_view':            'Lihat data AssetCore',
        'ac_manage_assets':   'Kelola inventaris laptop/PC',
        'ac_manage_infra':    'Kelola inventaris infrastruktur',
        'ac_manage_licenses': 'Kelola lisensi software',
        'ac_manage_subs':     'Kelola subscription & ISP',
        'ac_manage_requests': 'Kelola request software',
    },
    'support': {
        'sc_view':             'Lihat data SupportCore',
        'sc_manage_customers': 'Kelola master customer',
        'sc_manage_services':  'Kelola master layanan/jasa',
        'sc_manage_types':     'Kelola master tipe support',
        'sc_manage_sla':       'Kelola master kategori SLA',
        'sc_manage_apps':      'Kelola master aplikasi & modul',
        'sc_manage_contracts': 'Kelola kontrak support tahunan',
        'sc_manage_tickets':   'Buat/update tiket support',
        'sc_manage_presales':  'Kelola request Presales & POC',
        'sc_view_reports':     'Lihat laporan & monitoring SLA',
    },
    'attendance': {
        'at_view':   'Lihat data AttendanceCore (Kehadiran Saya)',
        'at_manage': 'Kelola & Approve Kehadiran, Cuti, Lembur, dan Koreksi Karyawan',
    },
}
# Backward-compat: ALL_PERMISSIONS = gabungan semua app + portal perms
ALL_PERMISSIONS = {
    'manage_users':  'Kelola pengguna (tambah/edit/hapus)',
    'manage_roles':  'Kelola role dan permission',
}
for _perms in APP_PERMISSIONS.values():
    ALL_PERMISSIONS.update(_perms)

# Permission yang hanya bisa diset oleh superadmin (admin tidak boleh assign ke role)
CRITICAL_PERMISSIONS = {'view_salary', 'manage_salary', 'manage_users', 'manage_roles'}

SYSTEM_ROLE_DEFAULTS = {
    'superadmin': list(ALL_PERMISSIONS.keys()),
    'admin': ['manage_employees','manage_divisions','manage_evaluations',
              'view_evaluations','send_reminders','manage_template',
              'sc_view','sc_manage_customers','sc_manage_services','sc_manage_types','sc_manage_sla',
              'sc_manage_apps','sc_manage_contracts','sc_manage_tickets','sc_manage_presales','sc_view_reports',
              'ac_view','ac_manage_assets','ac_manage_infra','ac_manage_licenses','ac_manage_subs','ac_manage_requests',
              'at_view', 'at_manage'],
    'viewer': ['view_evaluations','sc_view','sc_view_reports','ac_view', 'at_view'],
    'user': ['at_view'],
}

def _pg_adapt_schema(schema):
    """Konversi DDL SQLite → PostgreSQL."""
    import re
    s = schema
    # AUTOINCREMENT → SERIAL (untuk kolom id)
    s = re.sub(r'\bINTEGER PRIMARY KEY AUTOINCREMENT\b', 'SERIAL PRIMARY KEY', s, flags=re.IGNORECASE)
    # datetime('now','localtime') → NOW()
    s = re.sub(r"datetime\('now',\s*'localtime'\)", 'NOW()', s, flags=re.IGNORECASE)
    # SQLite CREATE INDEX IF NOT EXISTS sudah kompatibel dengan PG
    return s


def _pg_column_exists(db, table, col):
    """Cek apakah kolom ada di PostgreSQL via information_schema."""
    row = db.execute(
        "SELECT 1 FROM information_schema.columns WHERE table_name=%s AND column_name=%s",
        (table, col)).fetchone()
    return row is not None


def _seed_menus(db):
    cnt = db.execute('SELECT COUNT(*) as c FROM app_menus').fetchone()
    if not cnt or cnt['c'] == 0:
        default_menus = [
            # Portal
            {"app_slug": "portal", "parent_title": None, "title": "Semua Aplikasi", "url": "/portal", "icon": "grid-3x3-gap", "required_permission": "", "sort_order": 1},
            {"app_slug": "portal", "parent_title": None, "title": "Kelola User", "url": "/portal/users", "icon": "people", "required_permission": "manage_users", "sort_order": 2},
            {"app_slug": "portal", "parent_title": None, "title": "Role & Permission", "url": "/portal/roles", "icon": "person-badge", "required_permission": "manage_roles", "sort_order": 3},
            {"app_slug": "portal", "parent_title": None, "title": "Akses Aplikasi", "url": "/portal/settings", "icon": "shield-lock", "required_permission": "manage_roles", "sort_order": 4},
            {"app_slug": "portal", "parent_title": None, "title": "Pengaturan Sistem", "url": "/portal/system-settings", "icon": "gear", "required_permission": "manage_roles", "sort_order": 5},
            {"app_slug": "portal", "parent_title": None, "title": "Pengaturan Notifikasi", "url": "/portal/notifications", "icon": "bell-fill", "required_permission": "manage_roles", "sort_order": 6},
            {"app_slug": "portal", "parent_title": None, "title": "Audit Trail", "url": "/portal/audit", "icon": "shield-check", "required_permission": "manage_roles", "sort_order": 7},
            {"app_slug": "portal", "parent_title": None, "title": "Update Center", "url": "/portal/update", "icon": "arrow-up-circle", "required_permission": "manage_roles", "sort_order": 8},
            {"app_slug": "portal", "parent_title": None, "title": "Backup Center", "url": "/portal/backup", "icon": "database-fill-gear", "required_permission": "manage_roles", "sort_order": 9},

            # TalentCore (evaluasi)
            {"app_slug": "evaluasi", "parent_title": None, "title": "Menu Utama", "url": "#", "icon": "", "required_permission": "", "sort_order": 1},
            {"app_slug": "evaluasi", "parent_title": "Menu Utama", "title": "Dashboard", "url": "/", "icon": "house-door", "required_permission": "", "sort_order": 1},
            {"app_slug": "evaluasi", "parent_title": "Menu Utama", "title": "Karyawan", "url": "/karyawan", "icon": "people", "required_permission": "manage_employees", "sort_order": 2},
            {"app_slug": "evaluasi", "parent_title": "Menu Utama", "title": "Review Evaluasi", "url": "/reviews", "icon": "clipboard2-check", "required_permission": "", "sort_order": 3},
            {"app_slug": "evaluasi", "parent_title": "Menu Utama", "title": "Kinerja Task", "url": "/kinerja/tim", "icon": "graph-up-arrow", "required_permission": "", "sort_order": 4},
            {"app_slug": "evaluasi", "parent_title": "Menu Utama", "title": "Analitik Divisi", "url": "/kinerja/analitik", "icon": "diagram-3", "required_permission": "", "sort_order": 5},
            {"app_slug": "evaluasi", "parent_title": "Menu Utama", "title": "Performa Divisi", "url": "/kinerja/divisi", "icon": "people-fill", "required_permission": "", "sort_order": 6},
            {"app_slug": "evaluasi", "parent_title": "Menu Utama", "title": "Log Reminder", "url": "/reminders", "icon": "bell", "required_permission": "send_reminders", "sort_order": 7},
            {"app_slug": "evaluasi", "parent_title": "Menu Utama", "title": "Tabel Gaji", "url": "/salary", "icon": "cash-stack", "required_permission": "view_salary", "sort_order": 8},
            
            {"app_slug": "evaluasi", "parent_title": None, "title": "Konfigurasi", "url": "#", "icon": "", "required_permission": "", "sort_order": 2},
            {"app_slug": "evaluasi", "parent_title": "Konfigurasi", "title": "Template Evaluasi", "url": "/admin", "icon": "sliders", "required_permission": "manage_template", "sort_order": 1},
            {"app_slug": "evaluasi", "parent_title": "Konfigurasi", "title": "Manajemen Divisi", "url": "/admin/divisions", "icon": "diagram-3", "required_permission": "manage_divisions", "sort_order": 2},
            {"app_slug": "evaluasi", "parent_title": "Konfigurasi", "title": "Pengaturan Notifikasi", "url": "/settings", "icon": "gear", "required_permission": "manage_settings", "sort_order": 3},

            # AssetCore (aset)
            {"app_slug": "aset", "parent_title": None, "title": "Menu Utama", "url": "#", "icon": "", "required_permission": "ac_view", "sort_order": 1},
            {"app_slug": "aset", "parent_title": "Menu Utama", "title": "Dashboard", "url": "/aset/", "icon": "grid-1x2", "required_permission": "ac_view", "sort_order": 1},
            
            {"app_slug": "aset", "parent_title": None, "title": "Inventaris", "url": "#", "icon": "", "required_permission": "ac_view", "sort_order": 2},
            {"app_slug": "aset", "parent_title": "Inventaris", "title": "Laptop / PC", "url": "/aset/assets", "icon": "laptop", "required_permission": "ac_manage_assets", "sort_order": 1},
            {"app_slug": "aset", "parent_title": "Inventaris", "title": "Infrastruktur", "url": "/aset/infra", "icon": "hdd-network", "required_permission": "ac_manage_infra", "sort_order": 2},
            {"app_slug": "aset", "parent_title": "Inventaris", "title": "Maintenance", "url": "/aset/maintenance", "icon": "tools", "required_permission": "ac_manage_assets", "sort_order": 3},
            
            {"app_slug": "aset", "parent_title": None, "title": "Permintaan & Lisensi", "url": "#", "icon": "", "required_permission": "ac_view", "sort_order": 3},
            {"app_slug": "aset", "parent_title": "Permintaan & Lisensi", "title": "Lisensi", "url": "/aset/licenses", "icon": "key", "required_permission": "ac_manage_licenses", "sort_order": 1},
            {"app_slug": "aset", "parent_title": "Permintaan & Lisensi", "title": "Subscription", "url": "/aset/subscriptions", "icon": "cloud-check", "required_permission": "ac_manage_subs", "sort_order": 2},
            {"app_slug": "aset", "parent_title": "Permintaan & Lisensi", "title": "Request Software", "url": "/aset/requests", "icon": "inbox", "required_permission": "ac_manage_requests", "sort_order": 3},
            {"app_slug": "aset", "parent_title": "Permintaan & Lisensi", "title": "Request Alat Kerja", "url": "/aset/tool-requests", "icon": "laptop", "required_permission": "ac_manage_requests", "sort_order": 4},
            
            {"app_slug": "aset", "parent_title": None, "title": "Pengaturan", "url": "#", "icon": "", "required_permission": "ac_view", "sort_order": 4},
            {"app_slug": "aset", "parent_title": "Pengaturan", "title": "Notifikasi", "url": "/aset/settings", "icon": "bell-fill", "required_permission": "ac_manage_assets", "sort_order": 1},
            {"app_slug": "aset", "parent_title": "Pengaturan", "title": "Master Data Spec", "url": "/aset/masters", "icon": "database", "required_permission": "ac_manage_assets", "sort_order": 2},

            # SupportCore (support)
            {"app_slug": "support", "parent_title": None, "title": "Menu Utama", "url": "#", "icon": "", "required_permission": "sc_view", "sort_order": 1},
            {"app_slug": "support", "parent_title": "Menu Utama", "title": "Dashboard", "url": "/support/", "icon": "grid-1x2", "required_permission": "sc_view", "sort_order": 1},
            
            {"app_slug": "support", "parent_title": None, "title": "Master Data", "url": "#", "icon": "", "required_permission": "sc_view", "sort_order": 2},
            {"app_slug": "support", "parent_title": "Master Data", "title": "Customer", "url": "/support/customers", "icon": "building", "required_permission": "sc_manage_customers", "sort_order": 1},
            {"app_slug": "support", "parent_title": "Master Data", "title": "Apps & Modul", "url": "/support/apps", "icon": "grid-3x3-gap", "required_permission": "sc_manage_apps", "sort_order": 2},
            {"app_slug": "support", "parent_title": "Master Data", "title": "Layanan/Jasa", "url": "/support/services", "icon": "wrench-adjustable", "required_permission": "sc_manage_services", "sort_order": 3},
            {"app_slug": "support", "parent_title": "Master Data", "title": "Tipe Support", "url": "/support/support-types", "icon": "tools", "required_permission": "sc_manage_types", "sort_order": 4},
            {"app_slug": "support", "parent_title": "Master Data", "title": "Kategori SLA", "url": "/support/sla-categories", "icon": "speedometer2", "required_permission": "sc_manage_sla", "sort_order": 5},
            
            {"app_slug": "support", "parent_title": None, "title": "Operasional", "url": "#", "icon": "", "required_permission": "sc_view", "sort_order": 3},
            {"app_slug": "support", "parent_title": "Operasional", "title": "Kontrak", "url": "/support/contracts", "icon": "file-earmark-text", "required_permission": "sc_manage_contracts", "sort_order": 1},
            {"app_slug": "support", "parent_title": "Operasional", "title": "Presales & POC", "url": "/support/presales", "icon": "person-plus", "required_permission": "sc_manage_presales", "sort_order": 2},
            {"app_slug": "support", "parent_title": "Operasional", "title": "Tiket Support", "url": "/support/tickets", "icon": "ticket-detailed", "required_permission": "sc_manage_tickets", "sort_order": 3},
            
            {"app_slug": "support", "parent_title": None, "title": "Analitik", "url": "#", "icon": "", "required_permission": "sc_view", "sort_order": 4},
            {"app_slug": "support", "parent_title": "Analitik", "title": "Monitoring SLA", "url": "/support/sla-monitor", "icon": "graph-up", "required_permission": "sc_view_reports", "sort_order": 1},
            {"app_slug": "support", "parent_title": "Analitik", "title": "Laporan", "url": "/support/reports", "icon": "bar-chart-line", "required_permission": "sc_view_reports", "sort_order": 2},

            # BookingCore (booking)
            {"app_slug": "booking", "parent_title": None, "title": "Booking", "url": "#", "icon": "", "required_permission": "", "sort_order": 1},
            {"app_slug": "booking", "parent_title": "Booking", "title": "Semua Booking", "url": "/booking/", "icon": "calendar2-check", "required_permission": "", "sort_order": 1},
            {"app_slug": "booking", "parent_title": "Booking", "title": "Buat Booking", "url": "/booking/new", "icon": "plus-circle", "required_permission": "", "sort_order": 2},
            {"app_slug": "booking", "parent_title": "Booking", "title": "Tambah Resource", "url": "/booking/resource/add", "icon": "building-add", "required_permission": "", "sort_order": 3},

            # ProjectCore (project)
            {"app_slug": "project", "parent_title": None, "title": "Menu Utama", "url": "#", "icon": "", "required_permission": "", "sort_order": 1},
            {"app_slug": "project", "parent_title": "Menu Utama", "title": "Dashboard", "url": "/project/", "icon": "grid-1x2", "required_permission": "", "sort_order": 1},
            {"app_slug": "project", "parent_title": "Menu Utama", "title": "Semua Proyek", "url": "/project/projects", "icon": "kanban", "required_permission": "", "sort_order": 2},
            {"app_slug": "project", "parent_title": "Menu Utama", "title": "Tambah Proyek", "url": "/project/projects/add", "icon": "plus-circle", "required_permission": "", "sort_order": 3},
            
            {"app_slug": "project", "parent_title": None, "title": "Master Data", "url": "#", "icon": "", "required_permission": "", "sort_order": 2},
            {"app_slug": "project", "parent_title": "Master Data", "title": "Customer", "url": "/support/customers", "icon": "building", "required_permission": "", "sort_order": 1},

            # AttendanceCore (attendance)
            {"app_slug": "attendance", "parent_title": None, "title": "Kehadiran Saya", "url": "#", "icon": "", "required_permission": "at_view", "sort_order": 1},
            {"app_slug": "attendance", "parent_title": "Kehadiran Saya", "title": "Presensi Harian", "url": "/attendance/", "icon": "clock", "required_permission": "at_view", "sort_order": 1},
            {"app_slug": "attendance", "parent_title": "Kehadiran Saya", "title": "Pengajuan Cuti", "url": "/attendance/leave", "icon": "calendar-event", "required_permission": "at_view", "sort_order": 2},
            {"app_slug": "attendance", "parent_title": "Kehadiran Saya", "title": "Pengajuan Lembur", "url": "/attendance/overtime", "icon": "clock-history", "required_permission": "at_view", "sort_order": 3},
            {"app_slug": "attendance", "parent_title": "Kehadiran Saya", "title": "Koreksi Absen", "url": "/attendance/correction", "icon": "patch-exclamation", "required_permission": "at_view", "sort_order": 4},
            {"app_slug": "attendance", "parent_title": "Kehadiran Saya", "title": "Laporan Kehadiran", "url": "/attendance/report", "icon": "file-earmark-bar-graph", "required_permission": "at_view", "sort_order": 5},
            
            {"app_slug": "attendance", "parent_title": None, "title": "Manajemen", "url": "#", "icon": "", "required_permission": "at_manage", "sort_order": 2},
            {"app_slug": "attendance", "parent_title": "Manajemen", "title": "Approval Console", "url": "/attendance/admin/approvals", "icon": "check2-square", "required_permission": "at_manage", "sort_order": 1},
        ]

        parent_map = {}
        for m in default_menus:
            if m['parent_title'] is None:
                cur = db.execute('''INSERT INTO app_menus (app_slug, parent_id, title, url, icon, required_permission, sort_order, is_active)
                                    VALUES (?, NULL, ?, ?, ?, ?, ?, 1)''',
                                 (m['app_slug'], m['title'], m['url'], m['icon'], m['required_permission'], m['sort_order']))
                pid = cur.lastrowid
                parent_map[(m['app_slug'], m['title'])] = pid

        for m in default_menus:
            if m['parent_title'] is not None:
                pid = parent_map.get((m['app_slug'], m['parent_title']))
                if pid:
                    db.execute('''INSERT INTO app_menus (app_slug, parent_id, title, url, icon, required_permission, sort_order, is_active)
                                        VALUES (?, ?, ?, ?, ?, ?, ?, 1)''',
                                     (m['app_slug'], pid, m['title'], m['url'], m['icon'], m['required_permission'], m['sort_order']))

    # Dynamic insertion for existing databases to add Laporan Kehadiran menu
    chk = db.execute("SELECT id FROM app_menus WHERE app_slug='attendance' AND url='/attendance/report'").fetchone()
    if not chk:
        parent = db.execute("SELECT id FROM app_menus WHERE app_slug='attendance' AND title='Kehadiran Saya' AND parent_id IS NULL").fetchone()
        if parent:
            db.execute('''
                INSERT INTO app_menus (app_slug, parent_id, title, url, icon, required_permission, sort_order, is_active)
                VALUES ('attendance', ?, 'Laporan Kehadiran', '/attendance/report', 'file-earmark-bar-graph', 'at_view', 5, 1)
            ''', (parent['id'],))

    # Re-seed role defaults for any missing relationships (always run this to sync defaults!)
    all_db_menus = db.execute('SELECT id, required_permission, app_slug FROM app_menus').fetchall()
    
    for rname, rperms in SYSTEM_ROLE_DEFAULTS.items():
        db.execute("INSERT OR IGNORE INTO roles (name, description, app_slug, is_system) VALUES (?, ?, '', 1)",
                   (rname, f"System default {rname} role"))
                   
        for row in all_db_menus:
            menu_id = row['id']
            req_perm = row['required_permission']
            if rname == 'superadmin':
                db.execute('INSERT OR IGNORE INTO role_menus (role_name, menu_id) VALUES (?, ?)', (rname, menu_id))
            else:
                if not req_perm or req_perm in rperms:
                    db.execute('INSERT OR IGNORE INTO role_menus (role_name, menu_id) VALUES (?, ?)', (rname, menu_id))
                    
        # Re-derive permissions
        if rname == 'superadmin':
            for perm in ALL_PERMISSIONS:
                db.execute('INSERT OR IGNORE INTO role_permissions(role_name,permission) VALUES(?,?)', (rname, perm))
        else:
            for perm in rperms:
                db.execute('INSERT OR IGNORE INTO role_permissions(role_name,permission) VALUES(?,?)', (rname, perm))
    db.commit()


def init_db():
    conn = _pg_connect()
    conn.autocommit = False
    db = _DBWrapper(conn, is_pg=True)

    schema_sql = _pg_adapt_schema(SCHEMA)
    db.executescript(schema_sql)
    db.commit()

    _seed_menus(db)


    # Migrations
    import re as _re
    _valid_ident = _re.compile(r'^[a-zA-Z_][a-zA-Z0-9_]*$')
    # ALTER column types — PostgreSQL only
    for table, col, new_type, default_val in COLUMN_TYPE_MIGRATIONS:
        if not _valid_ident.match(table) or not _valid_ident.match(col):
            raise ValueError(f"Invalid identifier in type migration: {table!r}.{col!r}")
        row = db.execute(
            "SELECT data_type FROM information_schema.columns "
            "WHERE table_name=%s AND column_name=%s", (table, col)
        ).fetchone()
        if row and row[0].lower() not in ('text', 'character varying'):
            db.execute(
                f'ALTER TABLE {table} ALTER COLUMN {col} '
                f'TYPE {new_type} USING COALESCE({col}::text, {default_val})'
            )
    db.commit()
    for table, col, col_def in MIGRATIONS:
        if not _valid_ident.match(table) or not _valid_ident.match(col):
            raise ValueError(f"Invalid identifier in migration: table={table!r}, col={col!r}")
        # Konversi tipe kolom SQLite → PostgreSQL jika perlu
        col_def_pg = col_def.replace('INTEGER', 'INTEGER').replace(
            "datetime('now','localtime')", 'NOW()')
        if not _pg_column_exists(db, table, col):
            db.execute(f'ALTER TABLE {table} ADD COLUMN {col} {col_def_pg}')
    db.commit()
    # Seed evaluation data
    cnt_row = db.execute('SELECT COUNT(*) as c FROM skill_categories').fetchone()
    if not cnt_row or cnt_row['c'] == 0:
        seed_db(db)

    # Seed default ac_masters data
    try:
        cnt_masters = db.execute('SELECT COUNT(*) as c FROM ac_masters').fetchone()
        if not cnt_masters or cnt_masters['c'] == 0:
            default_masters = [
                ('cpu', 'Intel Core i3'), ('cpu', 'Intel Core i5'), ('cpu', 'Intel Core i7'), ('cpu', 'Intel Core i9'),
                ('cpu', 'AMD Ryzen 3'), ('cpu', 'AMD Ryzen 5'), ('cpu', 'AMD Ryzen 7'), ('cpu', 'AMD Ryzen 9'),
                ('cpu', 'Apple M1'), ('cpu', 'Apple M2'), ('cpu', 'Apple M3'),
                ('ram', '4 GB'), ('ram', '8 GB'), ('ram', '16 GB'), ('ram', '32 GB'), ('ram', '64 GB'),
                ('disk', '256 GB SSD'), ('disk', '512 GB SSD'), ('disk', '1 TB SSD'), ('disk', '2 TB SSD'), ('disk', '1 TB HDD'),
                ('gpu', 'Intel Iris Xe Graphics'), ('gpu', 'NVIDIA GeForce RTX 3050'), ('gpu', 'NVIDIA GeForce RTX 4050'),
                ('gpu', 'NVIDIA GeForce RTX 4060'), ('gpu', 'AMD Radeon Graphics'), ('gpu', 'Integrated'),
                ('screen', '13.3"'), ('screen', '14"'), ('screen', '15.6"'), ('screen', '16"'), ('screen', '24"'), ('screen', '27"'),
                ('os', 'Windows 10 Pro'), ('os', 'Windows 11 Pro'), ('os', 'macOS Sonoma'), ('os', 'macOS Sequoia'), ('os', 'Ubuntu 22.04 LTS'),
                ('office', 'Microsoft Office 2019'), ('office', 'Microsoft Office 2021'), ('office', 'Microsoft 365 Business'), ('office', 'None'),
                ('software', 'VS Code'), ('software', 'DBeaver'), ('software', 'Postman'), ('software', 'Docker Desktop'),
                ('software', 'Slack'), ('software', 'Google Chrome'), ('software', 'Zoom'), ('software', 'FortiClient'),
                ('software', 'GlobalProtect'), ('software', 'AnyDesk'), ('software', 'Adobe Acrobat Reader'),
            ]
            for category, name in default_masters:
                db.execute('INSERT OR IGNORE INTO ac_masters(category, name) VALUES(?,?)', (category, name))
            db.commit()
    except Exception:
        pass
    # Default settings
    for k, v in DEFAULT_SETTINGS.items():
        db.execute('INSERT OR IGNORE INTO app_settings(key, value) VALUES(?,?)', (k, v))
    db.commit()
    # Default superadmin
    if db.execute('SELECT COUNT(*) FROM users').fetchone()[0] == 0:
        db.execute('''INSERT INTO users(username, password_hash, full_name, role)
                      VALUES(?,?,?,?)''',
                   ('superadmin', generate_password_hash('Admin@123'), 'Super Administrator', 'superadmin'))
        db.commit()
        print("=" * 55)
        print(" SUPERADMIN DIBUAT:")
        print("   Username : superadmin")
        print("   Password : Admin@123")
        print(" !! Segera ganti password setelah login !!")
        print("=" * 55)
    # Migrate old level name
    db.execute("UPDATE employees SET level='Leader' WHERE level='Team Lead'")
    # Seed divisions
    if db.execute('SELECT COUNT(*) FROM divisions').fetchone()[0] == 0:
        for i, name in enumerate(ALL_DIVISIONS.keys()):
            db.execute('INSERT OR IGNORE INTO divisions(name, sort_order) VALUES(?,?)', (name, i))
    # Seed superapp apps registry
    _apps = [
        ('evaluasi', 'TalentCore', 'Penilaian & review kinerja karyawan',
         'clipboard2-check', '#4da8da', '#e8f4fd', '/karyawan', 1, 0, 0, ''),
        ('aset', 'AssetCore', 'Pencatatan & tracking aset perusahaan',
         'box-seam', '#6f42c1', '#f0ecff', '/aset/', 1, 0, 1, 'ac_view'),
        ('support', 'SupportCore', 'Monitoring technical support, SLA & presales',
         'headset', '#0d9488', '#e6faf8', '/support/', 1, 0, 2, ''),
        ('booking', 'BookingCore', 'Pemesanan & penjadwalan ruangan, kendaraan & aset',
         'calendar2-check', '#d97706', '#fff8e1', '/booking/', 1, 0, 3, ''),
        ('project', 'ProjectCore', 'Manajemen proyek, task & timeline tim',
         'kanban', '#0ea5e9', '#e0f2fe', '/project/', 1, 0, 4, ''),
        ('docs', 'DocsCore', 'Pengelolaan dokumen, SOP & knowledge base perusahaan',
         'file-earmark-richtext', '#10b981', '#d1fae5', '/docs/', 1, 1, 5, ''),
        ('finance', 'FinanceCore', 'Pencatatan keuangan, anggaran & laporan finansial',
         'cash-coin', '#f59e0b', '#fef3c7', '/finance/', 1, 1, 6, ''),
        ('attendance', 'AttendanceCore', 'Live attendance, leave, overtime, dan correction',
         'clock', '#8b5cf6', '#ede9fe', '/attendance/', 1, 0, 7, ''),
    ]
    for slug, name, desc, icon, color, bg, url, active, soon, sort, perm in _apps:
        db.execute('''INSERT INTO superapp_apps
            (slug,name,description,icon,color,bg_color,url,is_active,is_coming_soon,sort_order,required_permission)
            VALUES(?,?,?,?,?,?,?,?,?,?,?)
            ON CONFLICT(slug) DO UPDATE SET
                name=excluded.name,
                description=excluded.description,
                icon=excluded.icon,
                color=excluded.color,
                bg_color=excluded.bg_color,
                url=excluded.url,
                is_active=excluded.is_active,
                is_coming_soon=excluded.is_coming_soon,
                sort_order=excluded.sort_order,
                required_permission=excluded.required_permission''',
            (slug, name, desc, icon, color, bg, url, active, soon, sort, perm))
    # Hapus aplikasi obsolete (HelpdeskCore) yang sudah digantikan oleh AttendanceCore
    db.execute("DELETE FROM superapp_apps WHERE slug='helpdesk'")
    db.commit()
    # Seed system roles sebagai global (app_slug='')
    for rname, rdesc, rsys in [('superadmin','Super Administrator',1),('admin','Administrator',1),('viewer','Viewer Read-Only',1),('user','User / Staff',1)]:
        db.execute('INSERT OR IGNORE INTO roles(name,description,is_system,app_slug) VALUES(?,?,?,?)', (rname, rdesc, rsys, ''))
    # Migrate existing system roles agar global
    db.execute("UPDATE roles SET app_slug='' WHERE is_system=1 AND (app_slug='evaluasi' OR app_slug IS NULL)")
    for rname, perms in SYSTEM_ROLE_DEFAULTS.items():
        for perm in perms:
            db.execute('INSERT OR IGNORE INTO role_permissions(role_name,permission) VALUES(?,?)', (rname, perm))
    # Seed default support types
    for st_code, st_name, st_desc in [
        ('CORRECTIVE', 'Corrective Support', 'Penanganan gangguan/kerusakan sistem yang sudah terjadi'),
        ('PREVENTIVE', 'Preventive Support', 'Pemeliharaan rutin untuk mencegah kerusakan/gangguan'),
        ('ONSITE',     'Onsite Support',     'Kunjungan langsung ke lokasi customer'),
    ]:
        db.execute('INSERT OR IGNORE INTO sc_support_types(code,name,description) VALUES(?,?,?)',
                   (st_code, st_name, st_desc))
    db.commit()
    # Seed SLA Categories best practice (INSERT OR IGNORE — aman dijalankan ulang)
    _sla_seed = [
        # code, name, priority, response_h, workaround_h, resolution_h, mtype, description
        # ── CORRECTIVE ──────────────────────────────────────────────────────────────────
        ('COR-P1', 'Corrective P1 — Critical (System Down)',
         'Critical', 1, 4, 24, 'corrective',
         'Sistem tidak dapat diakses sama sekali / data corrupt. Berdampak ke seluruh user & operasional. '
         'Workaround wajib dalam 4 jam, solusi final dalam 1 hari kerja. Eskalasi immediate ke manajemen.'),
        ('COR-P2', 'Corrective P2 — High (Major Feature Broken)',
         'High', 2, 8, 72, 'corrective',
         'Fitur utama tidak berfungsi, operasional terganggu signifikan namun sistem masih bisa diakses sebagian. '
         'Workaround dalam 8 jam, solusi final dalam 3 hari kerja.'),
        ('COR-P3', 'Corrective P3 — Medium (Minor Feature Impaired)',
         'Medium', 4, 24, 120, 'corrective',
         'Fitur minor tidak berfungsi namun ada workaround manual yang dapat digunakan user. '
         'Workaround dalam 1 hari kerja, solusi final dalam 5 hari kerja.'),
        ('COR-P4', 'Corrective P4 — Low (Cosmetic / Minor Bug)',
         'Low', 8, 48, 240, 'corrective',
         'Bug minor / kosmetik, tidak mengganggu operasional. Antrian normal sesuai sprint. '
         'Workaround dalam 2 hari kerja, solusi final dalam 10 hari kerja.'),
        # ── PREVENTIVE ──────────────────────────────────────────────────────────────────
        ('PREV-CRIT', 'Preventive — Emergency Security Patch (Zero-Day)',
         'Critical', 2, None, 24, 'preventive',
         'Patch darurat untuk kerentanan zero-day atau exploit aktif yang mengancam keamanan data. '
         'Respons dalam 2 jam, patch diterapkan dalam 1 hari kerja meskipun di luar jam operasional.'),
        ('PREV-PATCH', 'Preventive — Security Patch & Update (Terjadwal)',
         'High', 8, None, 48, 'preventive',
         'Patch keamanan rutin, update library, atau upgrade versi minor. Dilakukan terjadwal dengan persetujuan customer. '
         'Konfirmasi jadwal dalam 8 jam, selesai dalam 2 hari kerja.'),
        ('PREV-STD', 'Preventive — Pemeliharaan Rutin',
         'Medium', 24, None, 72, 'preventive',
         'Pemeliharaan terjadwal: optimasi database, pembersihan log, cek performa, backup verification, monitoring review. '
         'Terjadwal bulanan/triwulan. Konfirmasi dalam 1 hari, selesai dalam 3 hari kerja.'),
        ('PREV-OPT', 'Preventive — Optimasi & Refactor Minor',
         'Low', 48, None, 168, 'preventive',
         'Optimasi performa non-urgent, refactor kode minor, update dokumentasi teknis, atau peningkatan kecil. '
         'Direncanakan dalam sprint berikutnya. Selesai dalam 7 hari kerja.'),
        # ── ONSITE ──────────────────────────────────────────────────────────────────────
        ('ONS-CRIT', 'Onsite — Critical Emergency',
         'Critical', 2, None, 4, 'onsite',
         'Kunjungan onsite darurat untuk gangguan kritikal infrastruktur / hardware yang tidak dapat diselesaikan remote sama sekali. '
         'Tim berangkat dalam 2 jam, target selesai dalam 4 jam sejak tiba di lokasi.'),
        ('ONS-URG', 'Onsite — Urgent',
         'High', 4, None, 8, 'onsite',
         'Kunjungan onsite urgen untuk masalah yang berdampak signifikan dan tidak dapat diselesaikan remote. '
         'Tim onsite di lokasi dalam 4 jam, penyelesaian dalam 8 jam sejak tiba.'),
        ('ONS-STD', 'Onsite — Standard',
         'Medium', 8, None, 24, 'onsite',
         'Kunjungan onsite terjadwal: training user, setup perangkat, instalasi, atau pendampingan rutin. '
         'Konfirmasi jadwal dalam 8 jam, pelaksanaan selesai dalam 1 hari kerja.'),
        ('ONS-PLAN', 'Onsite — Planned / Scheduled Visit',
         'Low', 24, None, 72, 'onsite',
         'Kunjungan onsite terencana: audit sistem, demo fitur baru, workshop, atau review berkala. '
         'Penjadwalan dalam 1 hari kerja, pelaksanaan disesuaikan kalender customer, selesai dalam 3 hari kerja.'),
    ]
    for code, name, prio, resp, wta, reso, mtype, desc in _sla_seed:
        db.execute('''INSERT OR IGNORE INTO sc_sla_categories
            (code,name,priority,response_time_hours,workaround_time_hours,resolution_time_hours,maintenance_type,description)
            VALUES(?,?,?,?,?,?,?,?)''', (code, name, prio, resp, wta, reso, mtype, desc))
    db.commit()
    # Seed default task_perf_config
    _tpc_seed = [
        # (task_type, label, base_pts, crit, high, med, low, ontime, late, sort)
        ('project_lead',      'Project (PIC/Lead)',       15, 2.0, 1.5, 1.0, 0.7, 1.1, 0.9, 1),
        ('project_impl',      'Project (Implementor)',    10, 2.0, 1.5, 1.0, 0.7, 1.1, 0.9, 2),
        ('project_member',    'Project (Member Tim)',      7, 2.0, 1.5, 1.0, 0.7, 1.1, 0.9, 3),
        ('project_issue',     'Issue Project (Programmer)',3, 2.0, 1.5, 1.0, 0.5, 1.1, 0.9, 4),
        ('project_task',      'Task ProjectCore (Done)',   2, 2.0, 1.5, 1.0, 0.7, 1.1, 0.9, 5),
        ('poc_presales',      'POC / Presales',            5, 2.0, 1.5, 1.0, 0.7, 1.1, 0.9, 6),
        ('support_ticket',    'Tiket Support (Closed)',    2, 2.0, 1.5, 1.0, 0.7, 1.1, 0.9, 7),
    ]
    for row in _tpc_seed:
        db.execute('''INSERT OR IGNORE INTO task_perf_config
            (task_type,label,base_points,priority_critical,priority_high,priority_medium,priority_low,ontime_mult,late_mult,sort_order)
            VALUES(?,?,?,?,?,?,?,?,?,?)''', row)
    db.commit()

    # Pastikan superadmin punya akses ke semua app (hanya jika belum ada)
    sa = db.execute("SELECT id FROM users WHERE role='superadmin' LIMIT 1").fetchone()
    if sa:
        for _slug in ['evaluasi', 'aset', 'support', 'booking']:
            db.execute('''INSERT OR IGNORE INTO user_app_access(user_id,app_slug,app_role,is_active)
                VALUES(?,?,?,1)''', (sa[0], _slug, 'superadmin'))
    db.commit()
    # Database indexes untuk kolom yang sering di-WHERE/JOIN
    _indexes = [
        ('idx_employees_is_active',       'employees',              'is_active'),
        ('idx_evaluations_employee_id',   'evaluations',            'employee_id'),
        ('idx_evaluations_review_status', 'evaluations',            'review_status'),
        ('idx_eval_tokens_eval_id',       'eval_tokens',            'eval_id'),
        ('idx_reminder_logs_employee',    'reminder_logs',          'employee_id'),
        ('idx_audit_created_at',          'audit_activity',         'created_at'),
        ('idx_audit_app_slug',            'audit_activity',         'app_slug'),
        ('idx_user_app_access_user',      'user_app_access',        'user_id, app_slug'),
        ('idx_sc_tickets_customer',       'sc_tickets',             'customer_id'),
        ('idx_sc_tickets_status',         'sc_tickets',             'status'),
        ('idx_sc_tickets_reported_at',    'sc_tickets',             'reported_at'),
        ('idx_sc_assignees_ticket',       'sc_ticket_assignees',    'ticket_id'),
        ('idx_sc_attachments_ticket',     'sc_ticket_attachments',  'ticket_id'),
        ('idx_ac_tool_req_attach_request','ac_tool_request_attachments', 'request_id'),
        ('idx_sc_ext_assignees_ticket',   'sc_ticket_external_assignees', 'ticket_id'),
        ('idx_pc_ext_assignees_task',     'pc_task_external_assignees',   'task_id'),
        ('idx_sc_presales_assignees',     'sc_presales_assignees',  'request_id'),
        ('idx_bk_bookings_resource',      'bk_bookings',            'resource_id, start_dt'),
        ('idx_bk_bookings_booked_by',     'bk_bookings',            'booked_by'),
        ('idx_bk_bookings_parent',        'bk_bookings',            'parent_id'),
    ]
    for idx_name, tbl, cols in _indexes:
        db.execute(f'CREATE INDEX IF NOT EXISTS {idx_name} ON {tbl}({cols})')
    db.commit()
    # Hapus duplikat bk_resources — pertahankan baris dengan id terkecil per nama
    db.execute('''DELETE FROM bk_resources WHERE id NOT IN (
        SELECT MIN(id) FROM bk_resources GROUP BY name)''')
    # Seed booking resources hanya jika belum ada data
    if db.execute('SELECT COUNT(*) FROM bk_resources').fetchone()[0] == 0:
        _resources = [
            ('Big Meeting Room', 'room', 'meeting', 20, 'Ruang rapat utama kapasitas besar', 'Lantai 3', '#0d6efd', 'people-fill', 1),
            ('Small Meeting Room A', 'room', 'meeting', 8, 'Ruang rapat kecil A', 'Lantai 2', '#6f42c1', 'door-open', 2),
            ('Small Meeting Room B', 'room', 'meeting', 8, 'Ruang rapat kecil B', 'Lantai 2', '#20c997', 'door-open', 3),
            ('Lounge Room', 'room', 'lounge', 15, 'Area lounge untuk diskusi santai', 'Lantai 1', '#fd7e14', 'cup-hot-fill', 4),
            ('Mobil Operasional', 'vehicle', 'car', 6, 'Kendaraan operasional perusahaan', 'Parkir Basement', '#d97706', 'car-front-fill', 5),
        ]
        for name, rtype, subtype, cap, desc, loc, color, icon, sort in _resources:
            db.execute('''INSERT INTO bk_resources(name,type,subtype,capacity,description,location,color,icon,sort_order)
                VALUES(?,?,?,?,?,?,?,?,?)''', (name, rtype, subtype, cap, desc, loc, color, icon, sort))
    # Seed booking items (minuman, makanan) hanya jika belum ada
    if db.execute('SELECT COUNT(*) FROM bk_items').fetchone()[0] == 0:
        _items = [
            ('Air Mineral', 'minuman', 'Botol air mineral 600ml', 'droplet-fill', 1),
            ('Kopi', 'minuman', 'Kopi hitam / kopi susu', 'cup-hot-fill', 2),
            ('Teh', 'minuman', 'Teh hangat / teh manis', 'cup-hot', 3),
            ('Jus Buah', 'minuman', 'Jus buah segar', 'cup-straw', 4),
            ('Makanan Ringan', 'makanan', 'Snack / kue ringan', 'bag', 5),
            ('Lunch Box', 'makanan', 'Makan siang kotak', 'box2', 6),
            ('Nasi Box', 'makanan', 'Nasi kotak lengkap', 'grid-3x2', 7),
            ('Buah Potong', 'makanan', 'Buah segar potong', 'basket', 8),
        ]
        for name, cat, desc, icon, sort in _items:
            db.execute('INSERT INTO bk_items(name,category,description,icon,sort_order) VALUES(?,?,?,?,?)',
                       (name, cat, desc, icon, sort))
    db.commit()
    
    # Seed employee birthdays
    try:
        _seed_employee_birthdays(db)
    except Exception as e:
        print(f"[Birthday Seed Error] {e}")

    # Seed grade_benchmarks default (hanya jika belum ada data)
    if db.execute('SELECT COUNT(*) FROM grade_benchmarks').fetchone()[0] == 0:
        _grade_bm = [
            ('Programmer',      'Senior',  130, 'Senior dev — project lead capable'),
            ('Programmer',      'Staff',   100, 'Mid-level dev'),
            ('Programmer',      'Junior',   80, 'Junior dev — supervised'),
            ('Implementor/BPS', 'Senior',  120, 'Senior implementor'),
            ('Implementor/BPS', 'Staff',    90, 'Implementor'),
            ('Implementor/BPS', 'Junior',   70, 'Junior implementor'),
            ('Helpdesk Support','Senior',  110, 'Senior HD — SLA keeper'),
            ('Helpdesk Support','Staff',    90, 'Helpdesk'),
            ('Helpdesk Support','Junior',   70, 'Junior HD'),
            ('Tester',          'Senior',  110, 'QA lead'),
            ('Tester',          'Staff',    90, 'Tester'),
            ('Tester',          'Junior',   70, 'Junior tester'),
            ('Management',      'Manager', 120, 'Manager / Project Manager'),
            ('Management',      'Senior',  110, 'Senior management'),
            ('Management',      'Staff',    90, 'Staff management'),
        ]
        for divisi, level, bm, notes in _grade_bm:
            db.execute(
                "INSERT OR IGNORE INTO grade_benchmarks(divisi,level,benchmark_per_month,notes) VALUES(?,?,?,?)",
                (divisi, level, bm, notes)
            )
    db.commit()
    db.close()

def _seed_employee_birthdays(db):
    birthdays_data = [
        ("muhammad thoriq zihni", "07-18"),
        ("Amalia", "08-04"),
        ("Aliyah", "08-06"),
        ("Khonza Izzati", "08-18"),
        ("Riska Agustia", "08-23"),
        ("dioharvandy", "08-27"),
        ("R Rizky Aria Putra", "09-13"),
        ("Okta", "10-05"),
        ("Muhammad Muhsin", "10-14"),
        ("M Iskandar Adi pratama", "10-20"),
        ("Riski", "10-22"),
        ("day", "10-25"),
        ("Andhy Ardhianto", "10-25"),
        ("RiRi", "11-07"),
        ("Ismed Iqbal", "11-24"),
        ("Ryno Andryano", "11-26"),
        ("vika andini", "11-29"),
        ("Fathur Zhafran", "12-28"),
        ("SirNewton", "01-07"),
        ("Gita", "01-09"),
        ("Nurul", "01-19"),
        ("Deti", "01-24"),
        ("Farah Fitriah", "02-05"),
        ("Windi Rahma", "02-17"),
        ("Feri Widiyanto", "02-25"),
        ("Yuhyi Wahyudin", "03-15"),
        ("Bamara Anugrah A.P", "03-19"),
        ("Faiz Muhammad", "03-31"),
        ("Rimayani", "03-31"),
        ("Vita Apriliana", "04-02"),
        ("Royandi Nicolas Naibaho", "04-03"),
        ("Yoshfia A.Z", "04-09"),
        ("Ahmad Sauki", "04-13"),
        ("Tri Hermawan", "04-20"),
        ("Ismail", "05-03"),
        ("Reinhard Meinard", "05-13"),
        ("JundiAufá", "05-26"),
        ("Fabian Dewantara Santonie", "06-04"),
        ("Devina Amanda", "06-10"),
        ("nita", "06-25"),
        ("Dwi Wahyuni", "07-07")
    ]
    for name_part, mmdd in birthdays_data:
        bday_val = f"1990-{mmdd}"
        emp = db.execute('SELECT id FROM employees WHERE LOWER(name) = LOWER(?)', (name_part.strip(),)).fetchone()
        if not emp:
            emp = db.execute('SELECT id FROM employees WHERE LOWER(name) LIKE ?', (f"%{name_part.strip()}%",)).fetchone()
        if emp:
            db.execute('UPDATE employees SET birthday=? WHERE id=?', (bday_val, emp['id']))
        else:
            db.execute('''
                INSERT INTO employees (name, divisi, birthday, is_active)
                VALUES (?, ?, ?, 1)
            ''', (name_part.strip(), 'Telegram Core', bday_val))

def seed_db(db):
    for divisi, (skill_cats, competency_items) in ALL_DIVISIONS.items():
        for order, (cat_name, items) in enumerate(skill_cats):
            cur = db.execute(
                'INSERT INTO skill_categories(divisi, name, sort_order) VALUES(?,?,?)',
                (divisi, cat_name, order))
            cat_id = cur.lastrowid
            for iorder, (name, desc, bobot) in enumerate(items):
                db.execute(
                    'INSERT INTO skill_items(category_id, name, description, bobot, sort_order) VALUES(?,?,?,?,?)',
                    (cat_id, name, desc, bobot, iorder))
        for order, (measurement, bobot, is_hs) in enumerate(competency_items):
            db.execute(
                'INSERT INTO competency_items(divisi, point_measurement, bobot, sort_order, is_hardskill) VALUES(?,?,?,?,?)',
                (divisi, measurement, bobot, order, is_hs))
        for order, (name, a, b, c, d) in enumerate(ABILITY_ITEMS):
            db.execute(
                'INSERT INTO ability_items(divisi, name, desc_a, desc_b, desc_c, desc_d, sort_order) VALUES(?,?,?,?,?,?,?)',
                (divisi, name, a, b, c, d, order))
    db.commit()

# ─── Settings Helpers ──────────────────────────────────────────────────────────

def get_settings(db):
    rows = db.execute('SELECT key, value FROM app_settings').fetchall()
    return {r['key']: r['value'] for r in rows}

def save_setting(db, key, value):
    if getattr(db, '_is_pg', False):
        # Bypass _fix() — langsung pakai PG syntax agar tidak terkena RETURNING id
        cur = db._conn.cursor()
        cur.execute(
            'INSERT INTO app_settings(key,value) VALUES(%s,%s) ON CONFLICT (key) DO UPDATE SET value=EXCLUDED.value',
            (key, value)
        )
    else:
        db.execute('INSERT OR REPLACE INTO app_settings(key, value) VALUES(?,?)', (key, value))

def get_notification_emails(settings, emp_email=''):
    """Gabungkan email karyawan + daftar email statis dari pengaturan."""
    recipients = []
    if emp_email and emp_email.strip():
        recipients.append(emp_email.strip())
    for e in settings.get('notification_emails', '').split(','):
        e = e.strip()
        if e and e not in recipients:
            recipients.append(e)
    return recipients

def normalize_telegram_id(value):
    """Normalisasi chat_id Telegram.
    - Angka (misal 123456789 atau -100123456789) → tetap
    - Username tanpa @ (misal 'grupku') → '@grupku'
    - Sudah ada @ → tetap
    """
    v = value.strip()
    if not v:
        return v
    if v.startswith('@'):
        return v
    if v.lstrip('-').isdigit():
        return v
    return '@' + v

def get_notification_telegram_ids(settings, emp_tg_id='', default_chat=''):
    """Gabungkan telegram_id karyawan + daftar statis + default chat, semua dinormalisasi."""
    ids = []
    if emp_tg_id and emp_tg_id.strip():
        ids.append(normalize_telegram_id(emp_tg_id))
    for t in settings.get('notification_telegram_ids', '').split(','):
        t = normalize_telegram_id(t)
        if t and t not in ids:
            ids.append(t)
    if not ids and default_chat:
        ids.append(normalize_telegram_id(default_chat))
    return ids

def get_notification_wa_phones(settings, emp_phone=''):
    """Gabungkan nomor HP karyawan + daftar extra phones dari pengaturan."""
    phones = []
    if emp_phone and emp_phone.strip():
        phones.append(emp_phone.strip())
    for p in settings.get('openwa_extra_phones', '').split(','):
        p = p.strip()
        if p and p not in phones:
            phones.append(p)
    return phones

_NOTIF_TYPES = [
    ('contract_reminder', 'Reminder Kontrak Karyawan',    'Notifikasi kontrak karyawan akan berakhir'),
    ('daily_report',      'Laporan Harian',               'Laporan otomatis harian jam 22:00 WIB'),
    ('eval_request',      'Permintaan Self-Assessment',   'Notifikasi pengisian self-assessment karyawan'),
    ('asset_change',      'Perubahan Aset IT',            'Notifikasi assign/release/perubahan aset'),
    ('ticket_new',        'Tiket Support Baru',           'Notifikasi tiket baru dibuat'),
    ('ticket_update',     'Update Status Tiket',          'Notifikasi perubahan status tiket support'),
    ('new_user',          'User Baru Terdaftar',          'Notifikasi saat user baru login via Google'),
]

def _init_notif_types(db):
    """Seed notif_type_settings jika belum ada."""
    for slug, label, desc in _NOTIF_TYPES:
        db.execute(
            'INSERT OR IGNORE INTO notif_type_settings(slug, label, description, is_active) VALUES(?,?,?,1)',
            (slug, label, desc)
        )
    db.commit()

def _migrate_notif_from_settings(db):
    """Impor penerima lama dari app_settings ke notif_recipients (sekali saja)."""
    import json as _j
    if db.execute('SELECT COUNT(*) as c FROM notif_recipients').fetchone()['c'] > 0:
        return
    cfg = get_settings(db)
    def _ins(name, channel, addr, types):
        addr = addr.strip()
        if addr:
            db.execute(
                'INSERT INTO notif_recipients(name,channel,address,notif_types,is_active) VALUES(?,?,?,?,1)',
                (name, channel, addr, _j.dumps(types))
            )
    for e in _parse_list(cfg.get('notification_emails', '')):
        _ins(f'Email: {e}', 'email', e, ['contract_reminder', 'eval_request'])
    for t in _parse_list(cfg.get('notification_telegram_ids', '')):
        _ins(f'Telegram: {t}', 'telegram', normalize_telegram_id(t), ['contract_reminder', 'eval_request'])
    for p in _parse_list(cfg.get('openwa_extra_phones', '')):
        _ins(f'WA: {p}', 'wa', p, ['contract_reminder'])
    for e in _parse_list(cfg.get('ac_notification_emails', '')):
        _ins(f'Email Aset: {e}', 'email', e, ['asset_change'])
    for t in _parse_list(cfg.get('ac_notification_telegram_ids', '')):
        _ins(f'Telegram Aset: {t}', 'telegram', normalize_telegram_id(t), ['asset_change'])
    for p in _parse_list(cfg.get('ac_notification_wa_phones', '')):
        _ins(f'WA Aset: {p}', 'wa', p, ['asset_change'])
    db.commit()

def get_notif_recipients(db, notif_slug, channel=None):
    """Return list address aktif untuk notif_slug & channel tertentu."""
    import json as _j
    q = 'SELECT address, notif_types FROM notif_recipients WHERE is_active=1'
    p = []
    if channel:
        q += ' AND channel=?'
        p.append(channel)
    result = []
    for r in db.execute(q, p).fetchall():
        try:
            types = _j.loads(r['notif_types'] or '["*"]')
        except Exception:
            types = ['*']
        if '*' in types or notif_slug in types:
            addr = r['address'].strip()
            if addr and addr not in result:
                result.append(addr)
    return result

def is_notif_enabled(db, notif_slug):
    """True jika tipe notifikasi ini aktif."""
    row = db.execute('SELECT is_active FROM notif_type_settings WHERE slug=?', (notif_slug,)).fetchone()
    return bool(row and row['is_active'])

def get_openwa_session(settings, app_slug=''):
    """Dapatkan OpenWA session ID untuk app tertentu.
    Prioritas: per-app session → global openwa_session_id → 'default'."""
    if app_slug:
        specific = settings.get(f'openwa_session_{app_slug}', '').strip()
        if specific:
            return specific
    return settings.get('openwa_session_id', 'default').strip() or 'default'

def _parse_list(raw):
    """Parse newline/comma-separated string jadi list string non-kosong."""
    return [x.strip() for x in raw.replace('\n', ',').split(',') if x.strip()]

def get_ac_notification_emails(settings):
    return _parse_list(settings.get('ac_notification_emails', ''))

def get_ac_notification_telegram_ids(settings):
    return [normalize_telegram_id(t) for t in _parse_list(settings.get('ac_notification_telegram_ids', ''))]

def get_ac_notification_wa_phones(settings):
    return _parse_list(settings.get('ac_notification_wa_phones', ''))

# ─── RBAC Helpers ─────────────────────────────────────────────────────────────

def get_role_permissions(db, role_name):
    rows = db.execute('SELECT permission FROM role_permissions WHERE role_name=?', (role_name,)).fetchall()
    return {r['permission'] for r in rows}

def has_permission(role_name, permission, db=None):
    if role_name == 'superadmin':
        return True
    if db is None:
        return False
    from flask import session
    try:
        if 'user_id' in session:
            app_slug = None
            for slug, perms in APP_PERMISSIONS.items():
                if permission in perms:
                    app_slug = slug
                    break
            if app_slug:
                row = db.execute(
                    'SELECT app_role FROM user_app_access WHERE user_id=? AND app_slug=? AND is_active=1',
                    (session['user_id'], app_slug)
                ).fetchone()
                if row:
                    role_name = row['app_role']
    except Exception:
        pass

    if role_name == 'superadmin':
        return True
    return permission in get_role_permissions(db, role_name)

def permission_required(perm):
    def decorator(f):
        @wraps(f)
        def decorated(*args, **kwargs):
            if 'user_id' not in session:
                return redirect(url_for('login'))
            db = get_db()
            if not has_permission(session.get('user_role', ''), perm, db):
                flash(f'Akses ditolak — permission "{perm}" diperlukan', 'danger')
                return redirect(url_for('index'))
            return f(*args, **kwargs)
        return decorated
    return decorator

# ─── MFA Helpers ──────────────────────────────────────────────────────────────

MFA_CHALLENGE_TTL = 900  # 15 menit

def generate_totp_secret():
    return pyotp.random_base32()

def get_totp_uri(secret, username, issuer='TalentCore'):
    return pyotp.totp.TOTP(secret).provisioning_uri(name=username, issuer_name=issuer)

def verify_totp(secret, code):
    return pyotp.TOTP(secret).verify((code or '').strip(), valid_window=1)

def qr_png_base64(uri):
    img = qrcode.make(uri)
    buf = io.BytesIO()
    img.save(buf, format='PNG')
    return base64.b64encode(buf.getvalue()).decode()

def mfa_session_valid(session_key):
    ts = session.get(session_key, 0)
    return (datetime.now().timestamp() - ts) < MFA_CHALLENGE_TTL

def mfa_challenge_required(session_key='mfa_verified'):
    def decorator(f):
        @wraps(f)
        def decorated(*args, **kwargs):
            if 'user_id' not in session:
                return redirect(url_for('login'))
            db = get_db()
            user = db.execute('SELECT mfa_enabled FROM users WHERE id=?', (session['user_id'],)).fetchone()
            if user and user['mfa_enabled'] and not mfa_session_valid(session_key):
                qs = request.query_string.decode()
                session['mfa_return_to']      = request.path + ('?' + qs if qs else '')
                session['mfa_session_key']    = session_key
                session['mfa_return_app_slug'] = session.get('active_app', 'portal')
                return redirect(url_for('mfa_challenge'))
            return f(*args, **kwargs)
        return decorated
    return decorator

def get_chain_contacts(db, emp):
    """Ambil data co-leader (supervisor), leader, dan manager dari employee record.
    Returns list of sqlite3.Row, masing-masing unik dan tidak null."""
    chain_ids = []
    for col in ('supervisor_id', 'leader_id', 'manager_id'):
        try:
            val = emp[col]
        except (IndexError, KeyError):
            val = None
        if val and val not in chain_ids and val != emp['id']:
            chain_ids.append(val)
    if not chain_ids:
        return []
    placeholders = ','.join('?' * len(chain_ids))
    rows = db.execute(
        f'SELECT * FROM employees WHERE id IN ({placeholders}) AND is_active=1',
        chain_ids
    ).fetchall()
    seen, result = set(), []
    for r in rows:
        if r['id'] not in seen:
            seen.add(r['id'])
            result.append(r)
    return result

# ─── Review / Self-Assessment Helpers ─────────────────────────────────────────

def get_or_create_eval_token(db, eval_id):
    row = db.execute('SELECT token FROM eval_tokens WHERE eval_id=?', (eval_id,)).fetchone()
    if row:
        return row['token']
    token = secrets.token_urlsafe(32)
    db.execute('INSERT INTO eval_tokens(eval_id, token) VALUES(?,?)', (eval_id, token))
    db.commit()
    return token

def get_user_emp_id(db, user_id):
    """Return the employee.id linked to this user, or None."""
    row = db.execute('SELECT id FROM employees WHERE user_id=?', (user_id,)).fetchone()
    return row['id'] if row else None

def get_user_contacts(db, user):
    """Return merged contact dict: user fields take priority, fallback ke linked employee."""
    uid = user['id'] if hasattr(user, 'keys') else user.get('id')
    emp = db.execute(
        'SELECT email, phone, telegram_id, name FROM employees WHERE user_id=? AND is_active=1',
        (uid,)
    ).fetchone()
    email    = (user['email']       or '').strip() or (emp['email']       if emp else '')
    phone    = (user['phone']       or '').strip() or (emp['phone']       if emp else '')
    telegram = (user['telegram_id'] or '').strip() or (emp['telegram_id'] if emp else '')
    return {
        'email':       email,
        'phone':       phone,
        'telegram_id': telegram,
        'emp_name':    emp['name'] if emp else '',
        'linked_emp':  emp is not None,
    }

def can_review_eval(db, user_id, user_role, eval_id):
    if user_role == 'superadmin':
        return True
    eid = get_user_emp_id(db, user_id)
    if not eid:
        return False
    row = db.execute('''SELECT emp.supervisor_id, emp.leader_id, emp.manager_id
                        FROM evaluations e JOIN employees emp ON emp.id = e.employee_id
                        WHERE e.id=?''', (eval_id,)).fetchone()
    if not row:
        return False
    return eid in [row['supervisor_id'], row['leader_id'], row['manager_id']]

def get_pending_review_count(db, user_id, user_role):
    if user_role == 'superadmin':
        return db.execute(
            "SELECT COUNT(*) FROM evaluations WHERE review_status='self_filled'"
        ).fetchone()[0]
    eid = get_user_emp_id(db, user_id)
    if not eid:
        return 0
    return db.execute('''
        SELECT COUNT(*) FROM evaluations e
        JOIN employees emp ON emp.id = e.employee_id
        WHERE e.review_status='self_filled'
        AND (emp.supervisor_id=? OR emp.leader_id=? OR emp.manager_id=?)
    ''', (eid, eid, eid)).fetchone()[0]

# ─── Auth Helpers ──────────────────────────────────────────────────────────────

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            flash('Silakan login terlebih dahulu', 'warning')
            return redirect(url_for('login', next=request.path))
        db = get_db()
        user = db.execute('SELECT is_active FROM users WHERE id=?', (session['user_id'],)).fetchone()
        if not user or not user['is_active']:
            session.clear()
            flash('Sesi Anda tidak valid atau pengguna tidak ditemukan', 'danger')
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated

def superadmin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        db = get_db()
        user = db.execute('SELECT is_active, role FROM users WHERE id=?', (session['user_id'],)).fetchone()
        if not user or not user['is_active']:
            session.clear()
            flash('Sesi Anda tidak valid atau pengguna tidak ditemukan', 'danger')
            return redirect(url_for('login'))
        if user['role'] != 'superadmin':
            flash('Akses ditolak — fitur ini hanya untuk Superadmin', 'danger')
            return redirect(url_for('index'))
        return f(*args, **kwargs)
    return decorated

@app.route('/media/proxy/<path:key>')
def media_proxy(key):
    # Public access is allowed for bookingcore/resources images. Others require login.
    if 'bookingcore' not in key and 'resources' not in key:
        if not session.get('user_id'):
            abort(401)
    from flask import send_file
    import io
    
    # Retrieve S3 configurations
    db = get_db()
    cfg = get_settings(db)
    storage_type = cfg.get('media_storage_type', 'local')
    
    if storage_type != 's3':
        # If media storage type is local, but they requested via proxy, try to serve from local uploads
        safe_path = os.path.join(UPLOAD_FOLDER, key.replace('upload/media/', '', 1))
        if os.path.exists(safe_path):
            return send_file(safe_path)
        abort(404)
        
    # S3 mode: fetch from S3 and stream back
    endpoint = cfg.get('backup_dest_s3_endpoint', '').strip()
    access_key = cfg.get('backup_dest_s3_access_key', '').strip()
    secret_key = cfg.get('backup_dest_s3_secret_key', '').strip()
    bucket = cfg.get('backup_dest_s3_bucket', '').strip()
    region = cfg.get('backup_dest_s3_region', '').strip()
    
    if not (access_key and secret_key and bucket):
        # Fallback to local
        safe_path = os.path.join(UPLOAD_FOLDER, key.replace('upload/media/', '', 1))
        if os.path.exists(safe_path):
            return send_file(safe_path)
        abort(404)
        
    try:
        import boto3
        from botocore.config import Config
        
        config = Config(
            region_name=region or 'us-east-1',
            signature_version='s3v4',
            connect_timeout=3,
            read_timeout=5,
            retries={'max_attempts': 1}
        )
        s3 = boto3.client(
            's3',
            endpoint_url=endpoint or None,
            aws_access_key_id=access_key,
            aws_secret_access_key=secret_key,
            config=config
        )
        
        # Get S3 object
        response = s3.get_object(Bucket=bucket, Key=key)
        content_type = response.get('ContentType', 'application/octet-stream')
        data = response['Body'].read()
        
        return send_file(
            io.BytesIO(data),
            mimetype=content_type,
            download_name=os.path.basename(key)
        )
    except Exception:
        # Fallback to local
        safe_path = os.path.join(UPLOAD_FOLDER, key.replace('upload/media/', '', 1))
        if os.path.exists(safe_path):
            return send_file(safe_path)
        abort(404)

def _get_update_badge(sess):
    """Return info update jika tersedia dan user berhak lihat, else None."""
    role = sess.get('user_role', '')
    if not role:
        return None
    try:
        db = get_db()
        settings = get_settings(db)
        notify_roles = [r.strip() for r in settings.get('update_notify_roles', 'superadmin,admin').split(',')]
        if role not in notify_roles:
            return None
        if settings.get('update_available', '0') != '1':
            return None
        return {
            'latest_version': settings.get('update_latest_version', ''),
            'latest_tag':     settings.get('update_latest_tag', ''),
        }
    except Exception:
        return None


@app.context_processor
def inject_globals():
    pending    = 0
    divisi_list = DIVISI_LIST
    user_perms  = set()
    portal_apps = []
    if 'user_id' in session:
        try:
            db = get_db()
            pending     = get_pending_review_count(db, session['user_id'], session.get('user_role', ''))
            divisi_list = get_divisi_list(db)
            user_perms  = get_role_permissions(db, session.get('user_role', ''))
            try:
                app_access_rows = db.execute(
                    'SELECT app_slug, app_role FROM user_app_access WHERE user_id=? AND is_active=1',
                    (session['user_id'],)
                ).fetchall()
                for row in app_access_rows:
                    slug = row['app_slug']
                    role = row['app_role']
                    if role == 'superadmin':
                        user_perms.update(ALL_PERMISSIONS.keys())
                    else:
                        app_allowed_perms = APP_PERMISSIONS.get(slug, {})
                        role_perms = get_role_permissions(db, role)
                        user_perms.update(p for p in role_perms if p in app_allowed_perms)
            except Exception:
                pass
            if session.get('user_role') == 'superadmin':
                user_perms = set(ALL_PERMISSIONS.keys())
            portal_apps = db.execute(
                'SELECT * FROM superapp_apps WHERE is_active=1 ORDER BY sort_order, name'
            ).fetchall()
            bk_resources = db.execute(
                'SELECT * FROM bk_resources WHERE is_active=1 ORDER BY sort_order, name'
            ).fetchall()
            try:
                _chatbot_on = get_settings(db).get('chatbot_enabled','0') == '1'
            except Exception:
                _chatbot_on = False
            
            # Fetch hierarchical menus for user role
            user_menus = []
            try:
                current_app = session.get('active_app') or 'portal'
                active_role = session.get('user_role', '')
                if current_app != 'portal':
                    app_access_rows = db.execute(
                        'SELECT app_slug, app_role FROM user_app_access WHERE user_id=? AND is_active=1',
                        (session['user_id'],)
                    ).fetchall()
                    for row in app_access_rows:
                        if row['app_slug'] == current_app:
                            active_role = row['app_role']
                            break
                
                if active_role == 'superadmin' or session.get('user_role') == 'superadmin':
                    menu_rows = db.execute(
                        'SELECT * FROM app_menus WHERE app_slug=? AND is_active=1 ORDER BY sort_order, id',
                        (current_app,)
                    ).fetchall()
                else:
                    menu_rows = db.execute(
                        '''SELECT m.* FROM app_menus m
                           JOIN role_menus rm ON m.id = rm.menu_id
                           WHERE m.app_slug=? AND rm.role_name=? AND m.is_active=1
                           ORDER BY m.sort_order, m.id''',
                        (current_app, active_role)
                    ).fetchall()
                
                parents = []
                children_by_parent = {}
                for row in menu_rows:
                    m = dict(row)
                    if m['parent_id'] is None:
                        parents.append(m)
                    else:
                        children_by_parent.setdefault(m['parent_id'], []).append(m)
                for p in parents:
                    p['children'] = children_by_parent.get(p['id'], [])
                user_menus = parents
            except Exception:
                pass
        except Exception:
            bk_resources = []
            _chatbot_on = False
            user_menus = []
    else:
        bk_resources = []
        _chatbot_on = False
        user_menus = []
    return {
        'current_user': {
            'id':       session.get('user_id'),
            'username': session.get('username'),
            'name':     session.get('user_name'),
            'role':     session.get('user_role'),
        } if 'user_id' in session else None,
        'now_year':        date.today().year,
        'today_date':      date.today().isoformat(),
        'pending_reviews': pending,
        'divisi_list':     divisi_list,
        'level_choices':   LEVEL_CHOICES,
        'user_perms':      user_perms,
        'ALL_PERMISSIONS': ALL_PERMISSIONS,
        'portal_apps':      portal_apps,
        'bk_resources':     bk_resources,
        'user_menus':       user_menus,
        'current_app_slug':    session.get('active_app') or 'portal',
        'app_version':         VERSION,
        'app_release_date':    RELEASE_DATE,
        'update_badge':        _get_update_badge(session),
        'chatbot_enabled':     _chatbot_on,
    }

# ─── Auto-set active_app dari URL path ────────────────────────────────────────

_SCANNER_PATHS = (
    '/wp-', '/wordpress', '/wp-login', '/wp-admin', '/xmlrpc',
    '/.env', '/.git', '/.htaccess', '/.htpasswd', '/config',
    '/admin/config', '/phpmyadmin', '/pma', '/mysql', '/myadmin',
    '/manager/', '/administrator', '/shell', '/cmd',
    '/cgi-bin', '/cgi/', '/api/v1/version', '/actuator', '/debug',
    '/vendor/', '/composer', '/package.json', '/package-lock',
    '/node_modules', '/proc/', '/etc/', '/usr/', '/var/log',
    '/.aws', '/.ssh', '/id_rsa', '/credentials',
    '/setup.php', '/install.php', '/update.php', '/upgrade.php',
    '/backup', '/dump', '/db.sql', '/database.sql',
)
_SCANNER_EXTENSIONS = (
    '.php', '.asp', '.aspx', '.jsp', '.cgi', '.pl', '.rb',
    '.bak', '.sql', '.tar', '.gz', '.zip', '.rar', '.7z',
    '.log', '.cfg', '.conf', '.ini', '.yaml', '.yml', '.toml',
    '.pem', '.key', '.crt', '.p12', '.pfx',
)

@app.before_request
def block_scanner():
    path = request.path.lower()
    if request.path.startswith('/static/'):
        return None
    if any(path.startswith(p) for p in _SCANNER_PATHS):
        return ('', 404)
    if any(path.endswith(ext) for ext in _SCANNER_EXTENSIONS):
        return ('', 404)
    if '../' in path or '%2e%2e' in path.lower():
        return ('', 404)


@app.before_request
def add_security_headers():
    pass  # headers ditambahkan via after_request


@app.after_request
def set_security_headers(response):
    response.headers['X-Content-Type-Options']  = 'nosniff'
    response.headers['X-Frame-Options']          = 'SAMEORIGIN'
    response.headers['X-XSS-Protection']         = '1; mode=block'
    response.headers['Referrer-Policy']           = 'strict-origin-when-cross-origin'
    response.headers['Permissions-Policy']        = 'geolocation=(), microphone=(), camera=()'
    response.headers['Server']                    = 'Hive'
    return response


_EVALUASI_PREFIXES = (
    '/emp', '/eval', '/salary', '/contracts', '/karyawan',
    '/reminders', '/reviews', '/assess', '/settings', '/admin',
    '/kinerja',
)

@app.before_request
def auto_set_active_app():
    path = request.path
    skip = ('/login', '/logout', '/static', '/mfa', '/portal/open')
    if any(path.startswith(p) for p in skip):
        # MFA: pertahankan app context asal (misal 'evaluasi' saat redirect dari /salary)
        # Hanya set portal jika belum ada context
        if path.startswith('/mfa') and not session.get('active_app'):
            session['active_app'] = 'portal'
        return
    if path.startswith('/support'):
        session['active_app'] = 'support'
    elif path.startswith('/portal'):
        session['active_app'] = 'portal'
    elif path.startswith('/booking'):
        session['active_app'] = 'booking'
    elif path.startswith('/aset'):
        session['active_app'] = 'aset'
    elif path.startswith('/project'):
        session['active_app'] = 'project'
    elif any(path.startswith(p) for p in _EVALUASI_PREFIXES):
        session['active_app'] = 'evaluasi'
    elif path == '/':
        # Dashboard TalentCore — jaga active_app yang sudah di-set oleh portal_open
        if session.get('active_app') not in ('evaluasi',):
            session['active_app'] = 'evaluasi'
    else:
        # Default: portal — untuk /profile, /users, /chatbot, dll
        session['active_app'] = 'portal'

# ─── Enforce app-level access dari user_app_access ──────────────────────────────

# Prefix path → app_slug. Urutan penting: lebih spesifik di atas.
_APP_PATH_MAP = [
    ('/support',     'support'),
    ('/portal',      'portal'),
    ('/booking',     'booking'),
    ('/aset',        'aset'),
    ('/attendance',  'attendance'),
    ('/project',     'project'),
    ('/docs',        'docs'),
    ('/finance',     'finance'),
]
# Path yang bebas diakses tanpa cek app_access
_APP_ACCESS_EXEMPT = {
    '/login', '/logout', '/static', '/mfa', '/portal/open',
    '/reset-password', '/set-password', '/media/proxy',
}

@app.before_request
def enforce_app_access():
    """Pastikan user hanya bisa akses app yang ada di user_app_access mereka.
    Superadmin selalu boleh. Evaluasi (TalentCore) juga dicek.
    """
    if 'user_id' not in session:
        return
    path = request.path
    if any(path.startswith(p) for p in _APP_ACCESS_EXEMPT):
        return
    if request.path.startswith('/static'):
        return
    # Tentukan slug app berdasarkan path
    app_slug = 'evaluasi'
    for prefix, slug in _APP_PATH_MAP:
        if path.startswith(prefix):
            app_slug = slug
            break
    # Superadmin bebas
    if session.get('user_role') == 'superadmin':
        return
    # Portal management hanya untuk admin/superadmin (sudah dijaga is_portal_admin)
    if app_slug == 'portal':
        return
    try:
        db  = get_db()
        row = db.execute(
            'SELECT id FROM user_app_access WHERE user_id=? AND app_slug=? AND is_active=1',
            (session['user_id'], app_slug)
        ).fetchone()
        if not row:
            flash(f'Anda tidak memiliki akses ke aplikasi ini.', 'danger')
            return redirect(url_for('portal'))
    except Exception:
        pass

# ─── Force MFA setup untuk user non-Google ─────────────────────────────────────

@app.before_request
def enforce_mfa_setup():
    # MFA tidak mandatory — hanya set flag reminder untuk banner di base.html
    pass

# ─── Score Helpers ──────────────────────────────────────────────────────────────

def calc_hs_total(db, eval_id, divisi):
    rows = db.execute('''
        SELECT si.bobot, COALESCE(ss.score, 0) AS score
        FROM skill_categories sc
        JOIN skill_items si ON si.category_id = sc.id
        LEFT JOIN skill_scores ss ON ss.skill_item_id = si.id AND ss.eval_id = ?
        WHERE sc.divisi = ?
    ''', (eval_id, divisi)).fetchall()
    if not rows: return 0.0
    sum_bobot = sum(r['bobot'] for r in rows)
    if sum_bobot == 0: return 0.0
    return round(sum(((r['score'] or 0) / 4.0) * (r['bobot'] or 0) for r in rows) / sum_bobot * 100, 4)

def calc_competency_total(db, eval_id, divisi, hs_total):
    items = db.execute('''
        SELECT ci.bobot, ci.is_hardskill, COALESCE(cs.rating, 0) AS rating
        FROM competency_items ci
        LEFT JOIN competency_scores cs ON cs.competency_item_id = ci.id AND cs.eval_id = ?
        WHERE ci.divisi = ? ORDER BY ci.sort_order
    ''', (eval_id, divisi)).fetchall()
    return round(sum(
        hs_total * (i['bobot'] or 0) if i['is_hardskill'] else (i['rating'] or 0) * 20.0 * (i['bobot'] or 0)
        for i in items
    ), 4)

def calc_ability_score(db, eval_id, divisi):
    rows = db.execute('''
        SELECT COALESCE(abs.level, '') AS level FROM ability_items ai
        LEFT JOIN ability_scores abs ON abs.ability_item_id = ai.id AND abs.eval_id = ?
        WHERE ai.divisi = ?
    ''', (eval_id, divisi)).fetchall()
    if not rows: return 0.0
    lv = {'A':1,'B':2,'C':3,'D':4}
    scored = [lv[r['level']] for r in rows if r['level'] in lv]
    return round(sum(scored) / len(scored) * 25, 2) if scored else 0.0

def recalc(db, eval_id):
    ev = db.execute('''SELECT e.*, emp.divisi FROM evaluations e
                       JOIN employees emp ON emp.id = e.employee_id WHERE e.id=?''', (eval_id,)).fetchone()
    if not ev: return
    pp_total = ev['pp_score'] * 20.0
    hs_total = calc_hs_total(db, eval_id, ev['divisi'])
    ability  = calc_ability_score(db, eval_id, ev['divisi'])
    comp     = calc_competency_total(db, eval_id, ev['divisi'], hs_total)
    final    = round(pp_total * 0.3 + comp * 0.7, 4)
    db.execute('''UPDATE evaluations SET pp_total=?,hs_total=?,ability_score=?,
                  competency_total=?,final_total=? WHERE id=?''',
               (pp_total, hs_total, ability, comp, final, eval_id))
    db.commit()

def calc_accumulated_metrics(db, ev, emp):
    from datetime import date, timedelta, datetime as _dt
    # 1. Fallback date range logic (3 months)
    created_date_str = ev['created_at'][:10] if ev.get('created_at') else ''
    if not created_date_str:
        created_date_str = date.today().strftime('%Y-%m-%d')
    
    date_to = ev.get('task_date_to') or created_date_str
    try:
        d_to = _dt.strptime(date_to, '%Y-%m-%d')
        d_from = d_to - timedelta(days=90)
        default_from_str = d_from.strftime('%Y-%m-%d')
    except:
        default_from_str = ''
    date_from = ev.get('task_date_from') or default_from_str
    
    bm = get_benchmark_for_emp(db, emp)
    perf = calc_task_perf(db, emp['id'], date_from, date_to, bm)
    task_score = perf.get('task_score', 0.0)
    
    # 2. Performance (Project Performance)
    pp_score = ev.get('pp_score') or 0
    pp_total = pp_score * 20.0
    
    # 3. Training / Self-Improvement
    proj_rows = db.execute('SELECT * FROM project_entries WHERE eval_id=? AND entry_type=?', (ev['id'], 'improvement')).fetchall()
    if proj_rows:
        done_count = sum(1 for p in proj_rows if p['status'] == 'DONE')
        training_score = round((done_count / len(proj_rows)) * 100.0, 2)
    else:
        training_score = 100.0
        
    # 4. Soft Skill
    comp_rows = db.execute('''
        SELECT ci.bobot, ci.is_hardskill, COALESCE(cs.rating, 0) AS rating
        FROM competency_items ci
        LEFT JOIN competency_scores cs ON cs.competency_item_id = ci.id AND cs.eval_id = ?
        WHERE ci.divisi = ? AND ci.is_hardskill = 0
    ''', (ev['id'], emp['divisi'])).fetchall()
    if comp_rows:
        sum_bobot = sum(c['bobot'] for c in comp_rows)
        if sum_bobot > 0:
            soft_skill_score = round(sum(c['rating'] * 20.0 * c['bobot'] for c in comp_rows) / sum_bobot, 2)
        else:
            soft_skill_score = 0.0
    else:
        soft_skill_score = 0.0
        
    # 5. Hard Skill
    hs_total = ev.get('hs_total') or 0.0
    
    overall = round((task_score + pp_total + training_score + soft_skill_score + hs_total) / 5.0, 2)
    
    return {
        'task_score': task_score,
        'pp_total': pp_total,
        'training_score': training_score,
        'soft_skill_score': soft_skill_score,
        'hs_total': hs_total,
        'overall': overall,
        'date_from': date_from,
        'date_to': date_to
    }

def get_eval_or_404(db, eval_id):
    return db.execute('''SELECT e.*, emp.name AS emp_name, emp.jabatan, emp.divisi
                         FROM evaluations e JOIN employees emp ON emp.id = e.employee_id
                         WHERE e.id=?''', (eval_id,)).fetchone()

# ─── Notification Helpers ───────────────────────────────────────────────────────

def send_email(settings, to_email, subject, html_body):
    try:
        msg = MIMEMultipart('alternative')
        msg['Subject'] = subject
        msg['From']    = settings.get('smtp_from') or settings.get('smtp_user', '')
        msg['To']      = to_email
        msg.attach(MIMEText(html_body, 'html', 'utf-8'))
        use_ssl = settings.get('smtp_ssl', '0') == '1'
        host, port = settings.get('smtp_host',''), int(settings.get('smtp_port', 587))
        server = smtplib.SMTP_SSL(host, port) if use_ssl else smtplib.SMTP(host, port)
        if not use_ssl:
            server.starttls()
        server.login(settings.get('smtp_user',''), settings.get('smtp_password',''))
        server.send_message(msg)
        server.quit()
        return True, None
    except Exception as e:
        return False, str(e)

def normalize_phone_wa(phone):
    """Konversi nomor HP ke format WhatsApp chat ID: 628xxx@c.us"""
    p = ''.join(c for c in (phone or '') if c.isdigit())
    if not p:
        return ''
    if p.startswith('0') and len(p) > 1:
        p = '62' + p[1:]
    elif not p.startswith('62'):
        p = '62' + p
    return f'{p}@c.us'

def send_whatsapp(openwa_url, api_key, session_id, phone, message):
    """Kirim pesan WhatsApp via OpenWA REST API (rmyndharis/OpenWA)."""
    try:
        chat_id = normalize_phone_wa(phone)
        if not chat_id:
            return False, 'Nomor HP tidak valid'
        sid = (session_id or 'default').strip()
        url = openwa_url.rstrip('/') + f'/api/sessions/{sid}/messages/send-text'
        headers = {'Content-Type': 'application/json'}
        if api_key:
            headers['X-API-Key'] = api_key
        r = req_lib.post(url, json={'chatId': chat_id, 'text': message},
                         headers=headers, timeout=15)
        r.raise_for_status()
        data = r.json() if r.text else {}
        if isinstance(data, dict) and data.get('error'):
            return False, str(data['error'])
        return True, None
    except Exception as e:
        return False, str(e)

def send_telegram(bot_token, chat_id, message, _log_subject='', _app_slug=None):
    try:
        r = req_lib.post(
            f'https://api.telegram.org/bot{bot_token}/sendMessage',
            json={'chat_id': chat_id, 'text': message, 'parse_mode': 'HTML'},
            timeout=5)
        r.raise_for_status()
        audit_notif('telegram', chat_id, _log_subject or message[:60], message, True, app_slug=_app_slug)
        return True, None
    except Exception as e:
        audit_notif('telegram', chat_id, _log_subject or message[:60], message, False, str(e), app_slug=_app_slug)
        return False, str(e)

def send_telegram_bg(bot_token, chat_id, message, _log_subject='', _app_slug=None):
    """Kirim notifikasi Telegram di background thread (fire-and-forget).
    Tidak memblokir request handler — timeout / kegagalan dicatat di audit_notifications.
    """
    import threading
    def _send():
        send_telegram(bot_token, chat_id, message, _log_subject=_log_subject, _app_slug=_app_slug)
    t = threading.Thread(target=_send, daemon=True)
    t.start()

def _real_ip():
    """Ambil IP asli client — baca X-Forwarded-For / X-Real-IP dari reverse proxy."""
    xff = request.headers.get('X-Forwarded-For', '')
    if xff:
        return xff.split(',')[0].strip()
    return request.headers.get('X-Real-IP', '') or request.remote_addr or ''

def audit_log(action, resource='', resource_id='', detail='', app_slug=None):
    """Catat aktivitas user ke audit_activity. Dipanggil dari route mana saja."""
    try:
        db = get_db()
        slug = app_slug or session.get('active_app') or 'portal'
        db.execute('''INSERT INTO audit_activity(app_slug,user_id,username,action,resource,resource_id,detail,ip,user_agent)
                      VALUES(?,?,?,?,?,?,?,?,?)''',
                   (slug, session.get('user_id'), session.get('user_name',''),
                    action, resource, str(resource_id), detail,
                    _real_ip(), request.user_agent.string[:200] if request.user_agent else ''))
        db.commit()
    except Exception:
        pass

def audit_notif(channel, recipient, subject, message, ok, err='', triggered_by='', app_slug=None):
    """Catat log notifikasi (email/telegram/wa)."""
    try:
        db = get_db()
        slug = app_slug or session.get('active_app') or 'portal'
        db.execute('''INSERT INTO audit_notifications(app_slug,channel,recipient,subject,message,status,error_msg,triggered_by)
                      VALUES(?,?,?,?,?,?,?,?)''',
                   (slug, channel, str(recipient), subject, message[:1000],
                    'sent' if ok else 'failed', err or '', triggered_by))
        db.commit()
    except Exception:
        pass

def log_reminder(db, emp_id, channel, subject, message, ok, err, triggered_by='auto'):
    db.execute('''INSERT INTO reminder_logs(employee_id, channel, subject, message, status, error_msg, triggered_by)
                  VALUES(?,?,?,?,?,?,?)''',
               (emp_id, channel, subject, message,
                'sent' if ok else 'failed', err or '', triggered_by))

def compose_contract_message(emp, days_left):
    from html import escape as _esc
    status = 'berakhir hari ini' if days_left == 0 else \
             f'berakhir dalam <b>{days_left} hari</b>' if days_left > 0 else \
             f'<b>sudah berakhir {abs(days_left)} hari lalu</b>'
    return f"""<h3>⚠️ Reminder Kontrak Karyawan</h3>
<table>
  <tr><td><b>Nama</b></td><td>: {_esc(emp['name'] or '')}</td></tr>
  <tr><td><b>Jabatan</b></td><td>: {_esc(emp['jabatan'] or '-')}</td></tr>
  <tr><td><b>Divisi</b></td><td>: {_esc(emp['divisi'] or '')}</td></tr>
  <tr><td><b>Tipe Kontrak</b></td><td>: Kontrak</td></tr>
  <tr><td><b>Akhir Kontrak</b></td><td>: {_esc(emp['contract_end'] or '')}</td></tr>
  <tr><td><b>Status</b></td><td>: Kontrak {status}</td></tr>
</table>
<p>Mohon segera tindak lanjut perpanjangan atau pemutusan kontrak.</p>"""

def compose_telegram_message(emp, days_left):
    from html import escape as _esc
    icon = '🔴' if days_left <= 7 else '🟡' if days_left <= 30 else '🟢'
    status = 'Berakhir HARI INI!' if days_left == 0 else \
             f'Berakhir dalam {days_left} hari' if days_left > 0 else \
             f'SUDAH BERAKHIR {abs(days_left)} hari lalu!'
    return (f"{icon} <b>Reminder Kontrak Karyawan</b>\n\n"
            f"👤 <b>{_esc(emp['name'] or '')}</b>\n"
            f"🏢 {_esc(emp['divisi'] or '')} — {_esc(emp['jabatan'] or '-')}\n"
            f"📅 Akhir kontrak: <b>{_esc(emp['contract_end'] or '')}</b>\n"
            f"⏳ {status}\n\n"
            f"<i>Segera tindak lanjut perpanjangan / pemutusan kontrak.</i>")

def compose_contract_wa_message(emp, days_left):
    icon = '🔴' if days_left <= 7 else '🟡' if days_left <= 30 else '🟢'
    status = 'Berakhir HARI INI!' if days_left == 0 else \
             f'Berakhir dalam {days_left} hari' if days_left > 0 else \
             f'SUDAH BERAKHIR {abs(days_left)} hari lalu!'
    return (f"{icon} *Reminder Kontrak Karyawan*\n\n"
            f"👤 *{emp['name']}*\n"
            f"🏢 {emp['divisi']} — {emp['jabatan'] or '-'}\n"
            f"📅 Akhir kontrak: *{emp['contract_end']}*\n"
            f"⏳ {status}\n\n"
            f"_Segera tindak lanjut perpanjangan / pemutusan kontrak._")

def compose_contract_chain_email(emp, days_left, recipient_role=''):
    status = 'berakhir hari ini' if days_left == 0 else \
             f'berakhir dalam <b>{days_left} hari</b>' if days_left > 0 else \
             f'<b>sudah berakhir {abs(days_left)} hari lalu</b>'
    role_note = f' (sebagai {recipient_role})' if recipient_role else ''
    return (f"<h3>⚠️ Info Kontrak Karyawan — Notifikasi Atasan{role_note}</h3>"
            f"<table>"
            f"<tr><td><b>Nama</b></td><td>: {emp['name']}</td></tr>"
            f"<tr><td><b>Jabatan</b></td><td>: {emp['jabatan'] or '-'}</td></tr>"
            f"<tr><td><b>Divisi</b></td><td>: {emp['divisi']}</td></tr>"
            f"<tr><td><b>Akhir Kontrak</b></td><td>: {emp['contract_end']}</td></tr>"
            f"<tr><td><b>Status</b></td><td>: Kontrak {status}</td></tr>"
            f"</table>"
            f"<p>Mohon segera tindak lanjut perpanjangan atau pemutusan kontrak karyawan di atas.</p>")

def compose_contract_chain_tg(emp, days_left, recipient_role=''):
    icon = '🔴' if days_left <= 7 else '🟡' if days_left <= 30 else '🟢'
    status = 'Berakhir HARI INI!' if days_left == 0 else \
             f'Berakhir dalam {days_left} hari' if days_left > 0 else \
             f'SUDAH BERAKHIR {abs(days_left)} hari lalu!'
    role_note = f' [{recipient_role}]' if recipient_role else ''
    return (f"{icon} <b>Info Kontrak — Notifikasi Atasan{role_note}</b>\n\n"
            f"👤 <b>{emp['name']}</b>\n"
            f"🏢 {emp['divisi']} — {emp['jabatan'] or '-'}\n"
            f"📅 Akhir kontrak: <b>{emp['contract_end']}</b>\n"
            f"⏳ {status}\n\n"
            f"<i>Mohon segera tindak lanjut perpanjangan / pemutusan kontrak.</i>")

def compose_contract_chain_wa(emp, days_left, recipient_role=''):
    icon = '🔴' if days_left <= 7 else '🟡' if days_left <= 30 else '🟢'
    status = 'Berakhir HARI INI!' if days_left == 0 else \
             f'Berakhir dalam {days_left} hari' if days_left > 0 else \
             f'SUDAH BERAKHIR {abs(days_left)} hari lalu!'
    role_note = f' [{recipient_role}]' if recipient_role else ''
    return (f"{icon} *Info Kontrak — Notifikasi Atasan{role_note}*\n\n"
            f"👤 *{emp['name']}*\n"
            f"🏢 {emp['divisi']} — {emp['jabatan'] or '-'}\n"
            f"📅 Akhir kontrak: *{emp['contract_end']}*\n"
            f"⏳ {status}\n\n"
            f"_Mohon segera tindak lanjut perpanjangan / pemutusan kontrak._")

def _send_contract_notification(db, emp, days_left, settings, triggered_by='auto'):
    """Kirim notifikasi kontrak ke staff + chain (co-leader/leader/manager) + extra WA.
    Returns (sent, failed)."""
    sent = failed = 0
    subject   = f"[Reminder] Kontrak {emp['name']} — {days_left} hari lagi"
    html      = compose_contract_message(emp, days_left)
    tg_msg    = compose_telegram_message(emp, days_left)
    wa_msg    = compose_contract_wa_message(emp, days_left)

    bot_token    = settings.get('telegram_bot_token', '').strip()
    default_chat = settings.get('telegram_default_chat_id', '').strip()
    wa_url       = settings.get('openwa_url', '').strip()
    wa_key       = settings.get('openwa_api_key', '').strip()
    wa_session   = get_openwa_session(settings, 'evaluasi')
    wa_enabled   = settings.get('openwa_enabled', '0') == '1'

    # ── Staff: Email ──
    if settings.get('smtp_host', '').strip():
        for to_email in get_notification_emails(settings, emp['email'] or ''):
            ok, err = send_email(settings, to_email, subject, html)
            log_reminder(db, emp['id'], 'email', subject, html, ok, err, triggered_by)
            if ok: sent += 1
            else:  failed += 1

    # ── Staff: Telegram ──
    if bot_token:
        for chat_id in get_notification_telegram_ids(settings, emp['telegram_id'] or '', default_chat):
            ok, err = send_telegram(bot_token, chat_id, tg_msg)
            log_reminder(db, emp['id'], 'telegram', subject, tg_msg, ok, err, triggered_by)
            if ok: sent += 1
            else:  failed += 1

    # ── Staff: WhatsApp ──
    if wa_enabled and wa_url and emp['phone']:
        ok, err = send_whatsapp(wa_url, wa_key, wa_session, emp['phone'], wa_msg)
        log_reminder(db, emp['id'], 'whatsapp', subject, wa_msg, ok, err, triggered_by)
        if ok: sent += 1
        else:  failed += 1

    # ── Chain: Manajerial / Leader / Manager ──
    col_role_map = [('supervisor_id', 'Manajerial'), ('leader_id', 'Leader'), ('manager_id', 'Manager')]
    notified_chain_ids = set()
    for col, role_label in col_role_map:
        try:
            chain_emp_id = emp[col]
        except (IndexError, KeyError):
            chain_emp_id = None
        if not chain_emp_id or chain_emp_id in notified_chain_ids:
            continue
        chain_emp = db.execute('SELECT * FROM employees WHERE id=? AND is_active=1',
                               (chain_emp_id,)).fetchone()
        if not chain_emp:
            continue
        notified_chain_ids.add(chain_emp_id)
        c_html = compose_contract_chain_email(emp, days_left, role_label)
        c_tg   = compose_contract_chain_tg(emp, days_left, role_label)
        c_wa   = compose_contract_chain_wa(emp, days_left, role_label)
        c_subj = f"[Info Kontrak] {emp['name']} — {days_left} hari lagi ({role_label})"

        if settings.get('smtp_host', '').strip() and chain_emp['email']:
            ok, err = send_email(settings, chain_emp['email'], c_subj, c_html)
            log_reminder(db, emp['id'], 'email', c_subj, c_html, ok, err, triggered_by)
            if ok: sent += 1
            else:  failed += 1

        if bot_token and chain_emp['telegram_id']:
            tg_id = normalize_telegram_id(chain_emp['telegram_id'])
            ok, err = send_telegram(bot_token, tg_id, c_tg)
            log_reminder(db, emp['id'], 'telegram', c_subj, c_tg, ok, err, triggered_by)
            if ok: sent += 1
            else:  failed += 1

        if wa_enabled and wa_url and chain_emp['phone']:
            ok, err = send_whatsapp(wa_url, wa_key, wa_session, chain_emp['phone'], c_wa)
            log_reminder(db, emp['id'], 'whatsapp', c_subj, c_wa, ok, err, triggered_by)
            if ok: sent += 1
            else:  failed += 1

    # ── Extra WA phones (diluar orang terkait) ──
    if wa_enabled and wa_url:
        extra_wa_msg = compose_contract_wa_message(emp, days_left)
        for phone in settings.get('openwa_extra_phones', '').split(','):
            phone = phone.strip()
            if not phone:
                continue
            ok, err = send_whatsapp(wa_url, wa_key, wa_session, phone, extra_wa_msg)
            log_reminder(db, emp['id'], 'whatsapp', subject, extra_wa_msg, ok, err, triggered_by)
            if ok: sent += 1
            else:  failed += 1

    return sent, failed

def run_contract_reminders(triggered_by='auto'):
    """Check contracts and send reminders. Call directly (not in request context)."""
    db = _get_raw_db()
    try:
        settings = get_settings(db)
        if settings.get('reminder_enabled', '1') != '1' and triggered_by == 'auto':
            return 0, 0
        reminder_days = [int(d.strip()) for d in settings.get('reminder_days', '30,14,7,1').split(',')
                         if d.strip().isdigit()]
        sent = failed = 0
        today = date.today()
        emps = db.execute('''SELECT * FROM employees
                             WHERE employment_type IN ('kontrak','staff_worker') AND contract_end != ''
                             AND contract_end IS NOT NULL AND is_active=1''').fetchall()
        for emp in emps:
            try:
                end_date = date.fromisoformat(emp['contract_end'])
            except Exception:
                continue
            days_left = (end_date - today).days
            if days_left not in reminder_days and triggered_by == 'auto':
                continue
            s, f = _send_contract_notification(db, emp, days_left, settings, triggered_by)
            sent += s
            failed += f
        db.commit()
        return sent, failed
    finally:
        db.close()

# ─── AssetCore: Subscription Reminders ────────────────────────────────────────

def compose_sub_email(sub, days_left):
    icon = '🔴' if days_left <= 7 else ('🟡' if days_left <= 14 else '🟢')
    cat  = sub['category'] or 'Subscription'
    status = 'berakhir <b>HARI INI</b>' if days_left == 0 else \
             f'berakhir dalam <b>{days_left} hari</b>' if days_left > 0 else \
             f'<b>sudah berakhir {abs(days_left)} hari lalu</b>'
    return (f"<h3>{icon} Reminder {cat}: {sub['provider']}</h3>"
            f"<table>"
            f"<tr><td><b>Layanan</b></td><td>: {sub['provider']}</td></tr>"
            f"<tr><td><b>Kategori</b></td><td>: {cat}</td></tr>"
            f"<tr><td><b>Billing</b></td><td>: {sub['billing_cycle'] or '-'}</td></tr>"
            f"<tr><td><b>Tgl Berakhir</b></td><td>: {sub['end_date']}</td></tr>"
            f"<tr><td><b>Status</b></td><td>: Subscription {status}</td></tr>"
            f"</table>"
            f"<p>Mohon segera lakukan perpanjangan atau evaluasi layanan ini.</p>")

def compose_sub_telegram(sub, days_left):
    icon = '🔴' if days_left <= 7 else ('🟡' if days_left <= 14 else '🟢')
    cat  = sub['category'] or 'Subscription'
    status = 'Berakhir HARI INI!' if days_left == 0 else \
             f'Berakhir dalam {days_left} hari' if days_left > 0 else \
             f'SUDAH BERAKHIR {abs(days_left)} hari lalu!'
    return (f"{icon} <b>Reminder {cat}</b>\n\n"
            f"📦 <b>{sub['provider']}</b>\n"
            f"💳 Billing: {sub['billing_cycle'] or '-'}\n"
            f"📅 Berakhir: <b>{sub['end_date']}</b>\n"
            f"⏳ {status}\n\n"
            f"<i>Segera lakukan perpanjangan atau evaluasi layanan ini.</i>")

def compose_sub_wa(sub, days_left):
    icon = '🔴' if days_left <= 7 else ('🟡' if days_left <= 14 else '🟢')
    cat  = sub['category'] or 'Subscription'
    status = 'Berakhir HARI INI!' if days_left == 0 else \
             f'Berakhir dalam {days_left} hari' if days_left > 0 else \
             f'SUDAH BERAKHIR {abs(days_left)} hari lalu!'
    return (f"{icon} *Reminder {cat}*\n\n"
            f"📦 *{sub['provider']}*\n"
            f"💳 Billing: {sub['billing_cycle'] or '-'}\n"
            f"📅 Berakhir: *{sub['end_date']}*\n"
            f"⏳ {status}\n\n"
            f"_Segera lakukan perpanjangan atau evaluasi layanan ini._")

# ─── AssetCore: Asset User Change Notifications ───────────────────────────────

def _asset_user_display(db, asset):
    """Ambil nama pengguna saat ini dari asset (linked / manual / kosong)."""
    if asset['employee_id']:
        row = db.execute('SELECT name FROM employees WHERE id=?', (asset['employee_id'],)).fetchone()
        return row['name'] if row else f'ID#{asset["employee_id"]}'
    return (asset['manual_employee_name'] or '').strip()

def _compose_asset_notif(asset, event, old_user, new_user, reason, fmt='email'):
    """
    event: 'assigned' | 'released' | 'end'
    fmt:   'email' | 'telegram' | 'wa'
    """
    name  = f"{asset['brand'] or ''} {asset['device_type'] or ''}".strip()
    tag   = asset['asset_tag'] or '—'
    today = date.today().isoformat()

    if event == 'assigned':
        icon, title = '🔄', 'Perubahan Pengguna Asset'
        line = f'Dari: {old_user or "—"}  →  Ke: {new_user or "—"}'
    elif event == 'released':
        icon, title = '📦', 'Asset Menjadi Available (Stok)'
        line = f'Pengguna sebelumnya: {old_user or "—"}'
    else:
        icon, title = '❌', 'Asset Ditandai End'
        line = f'Pengguna terakhir: {old_user or "—"}'

    if fmt == 'email':
        return (f"<h3>{icon} {title}</h3>"
                f"<table>"
                f"<tr><td><b>Asset</b></td><td>: {name}</td></tr>"
                f"<tr><td><b>Asset Tag</b></td><td>: {tag}</td></tr>"
                f"<tr><td><b>S/N</b></td><td>: {asset['serial_number'] or '—'}</td></tr>"
                f"<tr><td><b>Info</b></td><td>: {line}</td></tr>"
                f"<tr><td><b>Alasan</b></td><td>: {reason or '—'}</td></tr>"
                f"<tr><td><b>Tanggal</b></td><td>: {today}</td></tr>"
                f"</table>")
    elif fmt == 'telegram':
        return (f"{icon} <b>{title}</b>\n\n"
                f"💻 <b>{name}</b>  |  Tag: <code>{tag}</code>\n"
                f"👤 {line}\n"
                f"📋 Alasan: {reason or '—'}\n"
                f"📅 {today}")
    else:
        return (f"{icon} *{title}*\n\n"
                f"💻 *{name}*  |  Tag: {tag}\n"
                f"👤 {line}\n"
                f"📋 Alasan: {reason or '—'}\n"
                f"📅 {today}")

def _notify_asset_change(db, asset, event, old_user, new_user, reason):
    """Kirim notifikasi perubahan pengguna/status asset ke semua channel AC."""
    try:
        settings   = get_settings(db)
        ac_emails  = get_ac_notification_emails(settings)
        ac_tg_ids  = get_ac_notification_telegram_ids(settings)
        ac_phones  = get_ac_notification_wa_phones(settings)
        if not (ac_emails or ac_tg_ids or ac_phones):
            return

        name = f"{asset['brand'] or ''} {asset['device_type'] or ''}".strip()
        tag  = asset['asset_tag'] or ''
        ev_label = ('Perubahan Pengguna' if event == 'assigned'
                    else ('Asset Available' if event == 'released' else 'Asset End'))
        subj = f"[AssetCore] {ev_label}: {name} {tag}".strip()

        email_body = _compose_asset_notif(asset, event, old_user, new_user, reason, 'email')
        tg_msg     = _compose_asset_notif(asset, event, old_user, new_user, reason, 'telegram')
        wa_msg     = _compose_asset_notif(asset, event, old_user, new_user, reason, 'wa')

        bot_token  = settings.get('telegram_bot_token', '').strip()
        wa_url     = settings.get('openwa_url', '').strip()
        wa_key     = settings.get('openwa_api_key', '').strip()
        wa_session = get_openwa_session(settings, 'aset')
        wa_enabled = settings.get('openwa_enabled', '0') == '1'

        if settings.get('smtp_host', '').strip() and ac_emails:
            for to_email in ac_emails:
                ok, err = send_email(settings, to_email, subj, email_body)
                audit_notif('email', to_email, subj, email_body, ok, err, 'manual', app_slug='aset')

        if bot_token and ac_tg_ids:
            for chat_id in ac_tg_ids:
                send_telegram_bg(bot_token, chat_id, tg_msg, _log_subject=subj, _app_slug='aset')

        if wa_enabled and wa_url and ac_phones:
            for phone in ac_phones:
                ok, err = send_whatsapp(wa_url, wa_key, wa_session, phone, wa_msg)
                audit_notif('whatsapp', phone, subj, wa_msg, ok, err, 'manual', app_slug='aset')
    except Exception:
        pass  # Jangan ganggu flow utama jika notifikasi gagal


def run_subscription_reminders(triggered_by='auto'):
    """Cek ac_subscriptions yang mendekati expired dan kirim notifikasi.
    Dipanggil dari scheduler (auto) atau route manual (manual)."""
    db = _get_raw_db()
    try:
        settings = get_settings(db)
        if settings.get('ac_sub_reminder_enabled', '1') != '1' and triggered_by == 'auto':
            return 0, 0

        reminder_days_raw = settings.get('ac_sub_reminder_days', '30,14,7,1')
        reminder_days = [int(d.strip()) for d in reminder_days_raw.split(',') if d.strip().isdigit()]

        bot_token  = settings.get('telegram_bot_token', '').strip()
        wa_url     = settings.get('openwa_url', '').strip()
        wa_key     = settings.get('openwa_api_key', '').strip()
        wa_session = get_openwa_session(settings, 'aset')
        wa_enabled = settings.get('openwa_enabled', '0') == '1'

        # Gunakan daftar penerima khusus AssetCore
        ac_emails  = get_ac_notification_emails(settings)
        ac_tg_ids  = get_ac_notification_telegram_ids(settings)
        ac_phones  = get_ac_notification_wa_phones(settings)

        sent = failed = 0
        today = date.today()
        today_str = today.isoformat()

        subs = db.execute(
            "SELECT * FROM ac_subscriptions WHERE is_active=1 AND end_date!='' AND end_date IS NOT NULL"
        ).fetchall()

        for sub in subs:
            try:
                end_date = date.fromisoformat(sub['end_date'])
            except Exception:
                continue
            days_left = (end_date - today).days

            if triggered_by == 'auto':
                if days_left not in reminder_days:
                    continue
                if sub['last_reminder_sent'] == today_str:
                    continue

            subj   = f"[AssetCore] Reminder {sub['provider']} — {'berakhir dalam '+str(days_left)+' hari' if days_left > 0 else 'BERAKHIR HARI INI'}"
            html   = compose_sub_email(sub, days_left)
            tg_msg = compose_sub_telegram(sub, days_left)
            wa_msg = compose_sub_wa(sub, days_left)

            # Email — daftar AC penerima
            if settings.get('smtp_host', '').strip() and ac_emails:
                for to_email in ac_emails:
                    ok, err = send_email(settings, to_email, subj, html)
                    audit_notif('email', to_email, subj, html, ok, err, triggered_by, app_slug='aset')
                    if ok: sent += 1
                    else:  failed += 1

            # Telegram — daftar AC penerima
            if bot_token and ac_tg_ids:
                for chat_id in ac_tg_ids:
                    ok, err = send_telegram(bot_token, chat_id, tg_msg,
                                            _log_subject=subj, _app_slug='aset')
                    if ok: sent += 1
                    else:  failed += 1

            # WhatsApp — daftar AC penerima
            if wa_enabled and wa_url and ac_phones:
                for phone in ac_phones:
                    ok, err = send_whatsapp(wa_url, wa_key, wa_session, phone, wa_msg)
                    audit_notif('whatsapp', phone, subj, wa_msg, ok, err, triggered_by, app_slug='aset')
                    if ok: sent += 1
                    else:  failed += 1

            db.execute("UPDATE ac_subscriptions SET last_reminder_sent=? WHERE id=?",
                       (today_str, sub['id']))

        db.commit()
        return sent, failed
    finally:
        db.close()

# ─── APScheduler ───────────────────────────────────────────────────────────────

# ─── Update Center ─────────────────────────────────────────────────────────────

UPDATE_TRIGGER_FILE = '/tmp/hive_update_trigger'
UPDATE_LOG_FILE     = '/tmp/hive_update.log'
UPDATE_DONE_MARKER  = 'HIVE_DEPLOY_DONE'
UPDATE_FAIL_MARKER  = 'HIVE_DEPLOY_FAILED'


def check_for_updates():
    """Cek GitHub untuk release terbaru. Dipanggil dari scheduler."""
    db = _get_raw_db()
    try:
        settings = get_settings(db)
        if settings.get('update_check_enabled', '1') != '1':
            return
        repo = settings.get('github_repo', 'barangbaru/solid-apps')
        resp = req_lib.get(
            f'https://api.github.com/repos/{repo}/tags',
            headers={'Accept': 'application/vnd.github.v3+json'},
            timeout=10)
        if resp.status_code != 200:
            return
        tags = resp.json()
        if not tags:
            return
        # Ambil tag terbaru (urutan dari API sudah descending by creation)
        latest_tag  = tags[0].get('name', '')
        latest_ver  = latest_tag.lstrip('v')
        now_str     = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

        # Fetch semua releases sekaligus (lebih efisien dari per-tag)
        import json
        rel_all_resp = req_lib.get(
            f'https://api.github.com/repos/{repo}/releases?per_page=50',
            headers={'Accept': 'application/vnd.github.v3+json'},
            timeout=10)
        releases_by_tag = {}
        if rel_all_resp.status_code == 200:
            for r in rel_all_resp.json():
                releases_by_tag[r.get('tag_name', '')] = r.get('body', '') or ''

        release_notes = releases_by_tag.get(latest_tag, '')

        # Gabungkan tags + release notes jadi satu struktur JSON
        all_tags_data = []
        for t in tags:
            tname = t.get('name', '')
            if tname.startswith('v'):
                all_tags_data.append({
                    'tag':   tname,
                    'notes': releases_by_tag.get(tname, ''),
                })

        is_newer = _version_gt(latest_ver, VERSION)
        for key, val in [
            ('update_check_last',     now_str),
            ('update_latest_version', latest_ver),
            ('update_latest_tag',     latest_tag),
            ('update_release_notes',  release_notes),
            ('update_available',      '1' if is_newer else '0'),
            ('update_all_tags',       json.dumps(all_tags_data)),
        ]:
            db.execute("INSERT INTO app_settings(key,value) VALUES(?,?) ON CONFLICT(key) DO UPDATE SET value=excluded.value", (key, val))
        db.commit()
        print(f" Cek update: {'update tersedia ' + latest_tag if is_newer else 'sudah terbaru'} (terpasang v{VERSION})")
    except Exception as e:
        print(f" Cek update gagal: {e}")
    finally:
        db.close()


def _version_gt(a, b):
    """Return True jika versi a > b (semantic versioning)."""
    def parts(v):
        try:
            return [int(x) for x in v.strip().split('.')]
        except Exception:
            return [0]
    return parts(a) > parts(b)


_GRADE_BENCHMARK_DEFAULTS = {
    # (divisi_substr, level) : benchmark_per_month
    ('Programmer',      'Senior'): 130,
    ('Programmer',      'Staff'):  100,
    ('Programmer',      'Junior'):  80,
    ('Implementor',     'Senior'): 120,
    ('Implementor',     'Staff'):   90,
    ('Implementor',     'Junior'):  70,
    ('Helpdesk',        'Senior'): 110,
    ('Helpdesk',        'Staff'):   90,
    ('Helpdesk',        'Junior'):  70,
    ('Tester',          'Senior'): 110,
    ('Tester',          'Staff'):   90,
    ('Tester',          'Junior'):  70,
    ('Management',      'Manager'): 120,
    ('Management',      'Senior'): 110,
    ('Management',      'Staff'):   90,
}

def get_benchmark_for_emp(db, emp):
    """Ambil benchmark per-bulan sesuai divisi+level. Cek DB dulu, fallback ke default."""
    divisi = emp['divisi'] or ''
    level  = emp['level'] or 'Staff'
    row = db.execute(
        "SELECT benchmark_per_month FROM grade_benchmarks WHERE divisi=? AND level=?",
        (divisi, level)
    ).fetchone()
    if row:
        return float(row['benchmark_per_month'])
    # Fuzzy match ke defaults
    for (div_substr, lvl), bm in _GRADE_BENCHMARK_DEFAULTS.items():
        if div_substr.lower() in divisi.lower() and lvl.lower() == level.lower():
            return float(bm)
    return 100.0


def calc_task_perf(db, emp_id, date_from='', date_to='', benchmark_per_month=100.0):
    """Hitung skor kinerja task otomatis dari semua sumber data.

    Return dict:
      breakdown: list of {type, label, count, raw_pts}
      total_raw: total raw points
      task_score: normalized 0-100
      detail: list of individual tasks (untuk tabel detail)
      months: durasi periode dalam bulan
    """
    # Load config bobot
    cfg_rows = db.execute('SELECT * FROM task_perf_config ORDER BY sort_order').fetchall()
    cfg = {r['task_type']: r for r in cfg_rows}

    def _mult_priority(c, priority):
        p = (priority or 'Medium').lower()
        if p in ('critical', 'blocker', 'kritis'):   return float(c['priority_critical'])
        if p in ('high', 'tinggi'):                  return float(c['priority_high'])
        if p in ('low', 'rendah'):                   return float(c['priority_low'])
        return float(c['priority_medium'])

    def _mult_ontime(c, due_date, done_date):
        if not due_date or not done_date:
            return 1.0
        return float(c['ontime_mult']) if done_date <= due_date else float(c['late_mult'])

    # Helper: base WHERE clause untuk filter tanggal
    def _date_where(col_done, params, df=date_from, dt=date_to):
        clauses = []
        if df:
            clauses.append(f"{col_done} >= ?")
            params.append(df)
        if dt:
            clauses.append(f"{col_done} <= ?")
            params.append(dt)
        return (' AND ' + ' AND '.join(clauses)) if clauses else ''

    # Hitung durasi bulan (min 1)
    months = 1.0
    if date_from and date_to:
        try:
            from datetime import datetime as _dt
            d0 = _dt.strptime(date_from, '%Y-%m-%d')
            d1 = _dt.strptime(date_to,   '%Y-%m-%d')
            months = max(1.0, (d1 - d0).days / 30.44)
        except Exception:
            months = 1.0

    detail = []
    summary = {}

    def _add(task_type, label, source, pts, count=1):
        summary.setdefault(task_type, {'type': task_type, 'label': label, 'count': 0, 'raw_pts': 0.0})
        summary[task_type]['count']   += count
        summary[task_type]['raw_pts'] += pts
        detail.append({'type': task_type, 'label': label, 'source': source, 'pts': pts})

    # ── 1. Project (sebagai PIC/Lead) ──────────────────────────────────────────
    if 'project_lead' in cfg:
        c = cfg['project_lead']
        p = [emp_id]
        dw = _date_where('p.end_date', p)
        rows = db.execute(f'''
            SELECT p.code, p.name, p.end_date, p.status
            FROM pc_projects p
            WHERE p.pic_id=? AND p.deleted_at IS NULL
              AND p.status IN ('done','closed','selesai'){dw}
        ''', p).fetchall()
        for r in rows:
            pts = round(float(c['base_points']) * _mult_ontime(c, r['end_date'], r['end_date']), 2)
            _add('project_lead', c['label'], f"Project: {r['name']}", pts)

    # ── 2. Project (sebagai Implementor) ────────────────────────────────────────
    if 'project_impl' in cfg:
        c = cfg['project_impl']
        p = [emp_id, emp_id, emp_id]
        dw = _date_where('p.end_date', p)
        rows = db.execute(f'''
            SELECT p.code, p.name, p.end_date
            FROM pc_projects p
            WHERE (p.implementor_id=? OR p.co_leader_id=?)
              AND p.pic_id != ? AND p.deleted_at IS NULL
              AND p.status IN ('done','closed','selesai'){dw}
        ''', p).fetchall()
        for r in rows:
            pts = round(float(c['base_points']) * _mult_ontime(c, r['end_date'], r['end_date']), 2)
            _add('project_impl', c['label'], f"Project: {r['name']}", pts)

    # ── 3. Project (sebagai Member Tim) ─────────────────────────────────────────
    if 'project_member' in cfg:
        c = cfg['project_member']
        p = [emp_id, emp_id, emp_id, emp_id]
        dw = _date_where('p.end_date', p)
        rows = db.execute(f'''
            SELECT p.code, p.name, p.end_date
            FROM pc_members m
            JOIN pc_projects p ON p.id=m.project_id
            WHERE m.employee_id=? AND p.deleted_at IS NULL
              AND p.status IN ('done','closed','selesai')
              AND p.pic_id != ? AND p.implementor_id != ?
              AND (p.co_leader_id IS NULL OR p.co_leader_id != ?){dw}
        ''', p).fetchall()
        for r in rows:
            pts = round(float(c['base_points']) * _mult_ontime(c, r['end_date'], r['end_date']), 2)
            _add('project_member', c['label'], f"Project: {r['name']}", pts)

    # ── 4. Project Issues (sebagai Programmer) ──────────────────────────────────
    if 'project_issue' in cfg:
        c = cfg['project_issue']
        difficulty_map = {'hard': 2.0, 'sulit': 2.0, 'normal': 1.0, 'easy': 0.5, 'mudah': 0.5}
        p = [emp_id]
        dw = _date_where('i.resolved_date', p)
        rows = db.execute(f'''
            SELECT i.issue_no, i.title, i.difficulty, i.priority, i.resolved_date, p.name AS proj_name
            FROM pc_issues i
            JOIN pc_projects p ON p.id=i.project_id
            WHERE i.pic_programmer_id=?
              AND i.status_programmer IN ('done','closed','resolved'){dw}
        ''', p).fetchall()
        for r in rows:
            diff_mult = difficulty_map.get((r['difficulty'] or 'Normal').lower(), 1.0)
            pts = round(float(c.get('base_points') or 0) * diff_mult *
                        (_mult_ontime(c, r['resolved_date'], r['resolved_date']) or 1.0), 2)
            _add('project_issue', c['label'], f"Issue {r['issue_no']}: {r['title'][:40]} ({r['proj_name']})", pts)

    # ── 5. Project Tasks (done, assignee) — dengan difficulty & self_assigned ─────
    _diff_map = {'hard': 2.0, 'sulit': 2.0, 'normal': 1.0, 'easy': 0.5, 'mudah': 0.5}
    _self_bonus = 1.15  # inisiatif self-assign +15%
    if 'project_task' in cfg:
        c = cfg['project_task']
        p = [emp_id]
        dw = _date_where('t.due_date', p)
        rows = db.execute(f'''
            SELECT t.title, t.priority, t.due_date, t.status, t.difficulty,
                   p.name AS proj_name, COALESCE(ta.self_assigned,0) AS self_assigned
            FROM pc_task_assignees ta
            JOIN pc_tasks t ON t.id=ta.task_id
            JOIN pc_projects p ON p.id=t.project_id
            WHERE ta.employee_id=? AND t.status IN ('done','closed'){dw}
        ''', p).fetchall()
        for r in rows:
            diff_m = _diff_map.get((r['difficulty'] or 'Normal').lower(), 1.0)
            init_m = _self_bonus if r['self_assigned'] else 1.0
            pts = round(
                float(c.get('base_points') or 0)
                * (_mult_priority(c, r['priority'] or 'Medium') or 1.0)
                * diff_m * init_m, 2
            )
            label_sfx = ' [Self]' if r['self_assigned'] else ''
            _add('project_task', c['label'],
                 f"Task: {r['title'][:35]} ({r['proj_name']}) [{r['difficulty'] or 'Normal'}]{label_sfx}", pts)

    # ── 6. POC / Presales ───────────────────────────────────────────────────────
    if 'poc_presales' in cfg:
        c = cfg['poc_presales']
        p = [emp_id]
        dw = _date_where('r.created_at', p)
        rows = db.execute(f'''
            SELECT r.req_no, r.subject, r.request_type, r.status, r.created_at
            FROM sc_presales_assignees pa
            JOIN sc_presales_requests r ON r.id=pa.request_id
            WHERE pa.employee_id=? AND r.status IN ('done','closed','approved'){dw}
        ''', p).fetchall()
        for r in rows:
            pts = round(float(c['base_points']), 2)
            _add('poc_presales', c['label'], f"{r['request_type'].upper()} {r['req_no']}: {r['subject'][:40]}", pts)

    # ── 7. Support Tickets (assignee) ───────────────────────────────────────────
    if 'support_ticket' in cfg:
        c = cfg['support_ticket']
        p = [emp_id]
        # Gunakan work_start_date sebagai proxy tanggal penyelesaian jika resolved_at tidak ada
        dw = _date_where('t.reported_at', p)
        rows = db.execute(f'''
            SELECT t.ticket_no, t.subject, t.priority, t.due_date, t.reported_at, t.status
            FROM sc_ticket_assignees ta
            JOIN sc_tickets t ON t.id=ta.ticket_id
            WHERE ta.employee_id=? AND t.status IN ('resolved','closed'){dw}
        ''', p).fetchall()
        for r in rows:
            pts = round(float(c.get('base_points') or 0) * (_mult_priority(c, r['priority'] or 'Medium') or 1.0)
                        * (_mult_ontime(c, r['due_date'], r['reported_at']) or 1.0), 2)
            _add('support_ticket', c['label'], f"Tiket {r['ticket_no']}: {r['subject'][:40]}", pts)

    total_raw = round(sum(v['raw_pts'] for v in summary.values()), 2)
    task_score = round(min(total_raw / (benchmark_per_month * months) * 100, 100), 1)

    return {
        'breakdown': list(summary.values()),
        'total_raw': total_raw,
        'task_score': task_score,
        'months': round(months, 1),
        'benchmark': benchmark_per_month,
        'detail': detail,
    }


def calc_task_analytics(db, emp_id, date_from='', date_to=''):
    """Analitik detail per karyawan: timeliness, concurrency, breakdown tipe.

    Return dict:
      tasks_all      : list semua task (done & open) dengan info lengkap
      done_ontime    : task selesai tepat/sebelum due date
      done_delay     : task selesai setelah due date
      done_no_due    : task selesai tanpa due date
      open_ontime    : masih open, due date belum lewat (atau tanpa due)
      open_overtime  : masih open, due date sudah lewat!
      concurrent_max : maks task aktif bersamaan di satu titik waktu
      concurrent_avg : rata-rata task aktif bersamaan per hari
      by_type        : {type: {done, delay, overtime, ontime, open}} count
      total_done     : jumlah task selesai
      total_open     : jumlah task masih open
      ontime_rate    : pct ontime dari yang punya due_date
    """
    from datetime import date as _date, datetime as _dt, timedelta as _td

    today = _date.today().isoformat()

    def _in_period(d):
        if not d:
            return True
        if date_from and d < date_from:
            return False
        if date_to and d > date_to:
            return False
        return True

    def _timeliness(due, done_date):
        """Return: 'ontime'|'delay'|'no_due'|'open_ontime'|'open_overtime'"""
        is_done = bool(done_date)
        if not due:
            return 'done_no_due' if is_done else 'open_ontime'
        if is_done:
            return 'done_ontime' if done_date <= due else 'done_delay'
        return 'open_ontime' if today <= due else 'open_overtime'

    tasks_all = []

    def _add(task_type, label, source, due, start, done, pts, priority='Medium', project=''):
        tl = _timeliness(due, done)
        tasks_all.append({
            'type': task_type, 'label': label, 'source': source,
            'due': due, 'start': start, 'done': done,
            'timeliness': tl, 'pts': pts,
            'priority': priority, 'project': project,
            'is_done': bool(done),
        })

    # Load config bobot
    cfg_rows = db.execute('SELECT * FROM task_perf_config ORDER BY sort_order').fetchall()
    cfg = {r['task_type']: r for r in cfg_rows}

    def _bpts(ctype):
        return float(cfg[ctype]['base_points']) if ctype in cfg else 0

    def _pmult(ctype, priority):
        if ctype not in cfg:
            return 1.0
        c = cfg[ctype]
        p = (priority or 'Medium').lower()
        if p in ('critical', 'blocker', 'kritis'): return float(c['priority_critical'])
        if p in ('high', 'tinggi'):                return float(c['priority_high'])
        if p in ('low', 'rendah'):                 return float(c['priority_low'])
        return float(c['priority_medium'])

    def _omult(ctype, due, done):
        if ctype not in cfg or not due or not done:
            return 1.0
        c = cfg[ctype]
        return float(c['ontime_mult']) if done <= due else float(c['late_mult'])

    def _dw(col, params):
        cl = []
        if date_from: cl.append(f'{col} >= ?'); params.append(date_from)
        if date_to:   cl.append(f'{col} <= ?'); params.append(date_to)
        return (' AND ' + ' AND '.join(cl)) if cl else ''

    # ── Project tasks sebagai PIC ───────────────────────────────────────────
    p = [emp_id]
    rows = db.execute(f'''
        SELECT p.name, p.start_date, p.end_date, p.status, p.code
        FROM pc_projects p WHERE p.pic_id=? AND p.deleted_at IS NULL{_dw("p.start_date", p)}
    ''', p).fetchall()
    for r in rows:
        done_date = r['end_date'] if r['status'] in ('done','closed','selesai') else None
        pts = round(_bpts('project_lead') * _omult('project_lead', r['end_date'], done_date), 2)
        _add('project_lead', 'Project PIC/Lead', f"Project: {r['name']}",
             r['end_date'], r['start_date'], done_date, pts, project=r['name'])

    # ── Project sebagai Implementor ─────────────────────────────────────────
    p = [emp_id, emp_id]
    rows = db.execute(f'''
        SELECT p.name, p.start_date, p.end_date, p.status
        FROM pc_projects p
        WHERE (p.implementor_id=? OR p.co_leader_id=?)
          AND p.pic_id != ? AND p.deleted_at IS NULL{_dw("p.start_date", p)}
    ''', p + [emp_id]).fetchall()
    for r in rows:
        done_date = r['end_date'] if r['status'] in ('done','closed','selesai') else None
        pts = round(_bpts('project_impl') * _omult('project_impl', r['end_date'], done_date), 2)
        _add('project_impl', 'Project Implementor', f"Project: {r['name']}",
             r['end_date'], r['start_date'], done_date, pts, project=r['name'])

    # ── Project Member ──────────────────────────────────────────────────────
    p = [emp_id]
    rows = db.execute(f'''
        SELECT p.name, p.start_date, p.end_date, p.status
        FROM pc_members m JOIN pc_projects p ON p.id=m.project_id
        WHERE m.employee_id=? AND p.deleted_at IS NULL
          AND p.pic_id != ? AND p.implementor_id != ?
          AND (p.co_leader_id IS NULL OR p.co_leader_id != ?){_dw("p.start_date", p)}
    ''', p + [emp_id, emp_id, emp_id]).fetchall()
    for r in rows:
        done_date = r['end_date'] if r['status'] in ('done','closed','selesai') else None
        pts = round(_bpts('project_member') * _omult('project_member', r['end_date'], done_date), 2)
        _add('project_member', 'Project Member', f"Project: {r['name']}",
             r['end_date'], r['start_date'], done_date, pts, project=r['name'])

    # ── Project Issues ──────────────────────────────────────────────────────
    difficulty_map = {'hard':2.0,'sulit':2.0,'normal':1.0,'easy':0.5,'mudah':0.5}
    p = [emp_id]
    rows = db.execute(f'''
        SELECT i.issue_no, i.title, i.difficulty, i.priority,
               i.resolved_date, i.created_at, i.status_programmer, p.name AS proj
        FROM pc_issues i JOIN pc_projects p ON p.id=i.project_id
        WHERE i.pic_programmer_id=?{_dw("i.created_at", p)}
    ''', p).fetchall()
    for r in rows:
        dm = difficulty_map.get((r['difficulty'] or 'Normal').lower(), 1.0)
        is_done = r['status_programmer'] in ('done','closed','resolved')
        done_date = r['resolved_date'] if is_done else None
        pts = round(_bpts('project_issue') * dm
                    * _pmult('project_issue', r['priority'])
                    * _omult('project_issue', done_date, done_date), 2)
        _add('project_issue', f"Issue ({r['difficulty'] or 'Normal'})",
             f"#{r['issue_no']} {r['title'][:35]} ({r['proj']})",
             done_date, r['created_at'][:10] if r['created_at'] else None,
             done_date, pts, r['priority'] or 'Medium', r['proj'])

    # ── Project Tasks ───────────────────────────────────────────────────────
    _diff_map_a = {'hard': 2.0, 'sulit': 2.0, 'normal': 1.0, 'easy': 0.5, 'mudah': 0.5}
    p = [emp_id]
    rows = db.execute(f'''
        SELECT t.title, t.priority, t.due_date, t.status, t.created_at, t.difficulty,
               p.name AS proj, COALESCE(ta.self_assigned,0) AS self_assigned
        FROM pc_task_assignees ta JOIN pc_tasks t ON t.id=ta.task_id
        JOIN pc_projects p ON p.id=t.project_id
        WHERE ta.employee_id=?{_dw("t.created_at", p)}
    ''', p).fetchall()
    for r in rows:
        is_done  = r['status'] in ('done','closed')
        done_date = r['due_date'] if is_done else None
        diff_m   = _diff_map_a.get((r['difficulty'] or 'Normal').lower(), 1.0)
        init_m   = 1.15 if r['self_assigned'] else 1.0
        pts      = round(_bpts('project_task') * _pmult('project_task', r['priority']) * diff_m * init_m, 2)
        label    = f"Task [{r['difficulty'] or 'Normal'}]" + (' [Self]' if r['self_assigned'] else '')
        _add(label, 'Task Project', f"{r['title'][:40]} ({r['proj']})",
             r['due_date'], r['created_at'][:10] if r['created_at'] else None,
             done_date if is_done else None, pts, r['priority'] or 'Medium', r['proj'])

    # ── POC / Presales ──────────────────────────────────────────────────────
    p = [emp_id]
    rows = db.execute(f'''
        SELECT r.req_no, r.subject, r.status, r.request_type, r.created_at
        FROM sc_presales_assignees pa JOIN sc_presales_requests r ON r.id=pa.request_id
        WHERE pa.employee_id=?{_dw("r.created_at", p)}
    ''', p).fetchall()
    for r in rows:
        is_done = r['status'] in ('done','closed','approved')
        done_date = r['created_at'][:10] if (is_done and r['created_at']) else None
        pts = round(_bpts('poc_presales'), 2)
        _add('poc_presales', f"POC/Presales",
             f"{r['request_type'].upper()} {r['req_no']}: {r['subject'][:35]}",
             None, r['created_at'][:10] if r['created_at'] else None,
             done_date, pts, 'Medium')

    # ── Support Tickets ─────────────────────────────────────────────────────
    p = [emp_id]
    rows = db.execute(f'''
        SELECT t.ticket_no, t.subject, t.priority, t.due_date, t.reported_at, t.status
        FROM sc_ticket_assignees ta JOIN sc_tickets t ON t.id=ta.ticket_id
        WHERE ta.employee_id=?{_dw("t.reported_at", p)}
    ''', p).fetchall()
    for r in rows:
        is_done = r['status'] in ('resolved','closed')
        done_date = r['due_date'] if is_done else None  # proxy
        pts = round(_bpts('support_ticket') * _pmult('support_ticket', r['priority'] or 'Medium')
                    * _omult('support_ticket', r['due_date'], done_date), 2)
        _add('support_ticket', 'Tiket Support',
             f"#{r['ticket_no']}: {r['subject'][:40]}",
             r['due_date'],
             r['reported_at'][:10] if r['reported_at'] else None,
             done_date, pts, r['priority'] or 'Medium')

    # ── Hitung timeliness ───────────────────────────────────────────────────
    def _cnt(tl): return sum(1 for t in tasks_all if t['timeliness'] == tl)
    done_ontime   = _cnt('done_ontime')
    done_delay    = _cnt('done_delay')
    done_no_due   = _cnt('done_no_due')
    open_ontime   = _cnt('open_ontime')
    open_overtime = _cnt('open_overtime')
    total_done    = done_ontime + done_delay + done_no_due
    total_open    = open_ontime + open_overtime
    total_with_due = done_ontime + done_delay
    ontime_rate   = round(done_ontime / total_with_due * 100, 1) if total_with_due else None

    # ── Concurrency (max & avg task aktif bersamaan) ─────────────────────────
    # Pakai event scan: setiap task punya [start, end]. Scan timeline, hitung max tumpang tindih.
    events = []
    for t in tasks_all:
        s = t['start'] or t['due'] or today
        e = t['done'] or (today if not t['is_done'] else t['done']) or today
        if s and e and s <= e:
            events.append((s, +1))
            events.append((e, -1))
    events.sort(key=lambda x: (x[0], x[1]))
    cur = mx = 0
    for _, d in events:
        cur += d
        if cur > mx: mx = cur
    concurrent_max = mx

    # Avg concurrent: total task-days / period days
    total_task_days = 0
    for t in tasks_all:
        s = t['start'] or t['due'] or today
        e = t['done'] or today
        try:
            d0 = _dt.strptime(s[:10], '%Y-%m-%d').date()
            d1 = _dt.strptime(e[:10], '%Y-%m-%d').date()
            total_task_days += max(1, (d1 - d0).days + 1)
        except Exception:
            total_task_days += 1
    period_days = 1
    if date_from and date_to:
        try:
            period_days = max(1, (_dt.strptime(date_to, '%Y-%m-%d').date()
                                  - _dt.strptime(date_from, '%Y-%m-%d').date()).days + 1)
        except Exception:
            pass
    concurrent_avg = round(total_task_days / period_days, 1) if period_days else 0

    # ── By type summary ──────────────────────────────────────────────────────
    by_type = {}
    for t in tasks_all:
        tp = t['type']
        if tp not in by_type:
            by_type[tp] = {'label': t['label'], 'done': 0, 'ontime': 0,
                           'delay': 0, 'overtime': 0, 'open': 0}
        tl = t['timeliness']
        if 'done' in tl:
            by_type[tp]['done'] += 1
            if tl == 'done_ontime': by_type[tp]['ontime'] += 1
            elif tl == 'done_delay': by_type[tp]['delay'] += 1
        else:
            by_type[tp]['open'] += 1
            if tl == 'open_overtime': by_type[tp]['overtime'] += 1

    total_raw = round(sum(t['pts'] for t in tasks_all if t['is_done']), 2)

    return {
        'tasks_all':      tasks_all,
        'done_ontime':    done_ontime,
        'done_delay':     done_delay,
        'done_no_due':    done_no_due,
        'open_ontime':    open_ontime,
        'open_overtime':  open_overtime,
        'total_done':     total_done,
        'total_open':     total_open,
        'ontime_rate':    ontime_rate,
        'concurrent_max': concurrent_max,
        'concurrent_avg': concurrent_avg,
        'by_type':        by_type,
        'total_raw':      total_raw,
    }


def run_birthday_reminders():
    try:
        db = _get_raw_db()
        settings = get_settings(db)
        bot_token = settings.get('telegram_bot_token', '').strip()
        if not bot_token:
            bot_token = os.environ.get('TELEGRAM_BOT_TOKEN', '1095530966:AAFkSV9puxmT2z7cvpsbBQy_TWqj9-MCvbM').strip()
        if not bot_token:
            db.close()
            return

        from datetime import datetime
        today_mmdd = datetime.now().strftime('%m-%d')
        
        # Get employees having birthday today
        employees = db.execute("SELECT id, name, telegram_id, user_id FROM employees WHERE is_active=1 AND birthday LIKE ?", (f"%-{today_mmdd}",)).fetchall()
        
        for emp in employees:
            # 1. Broadcast to groups where this employee has ever clocked in/out
            att_notes = db.execute("SELECT DISTINCT notes_in, notes_out FROM attendance WHERE user_id = ?", (emp['user_id'],)).fetchall()
            
            group_ids = set()
            import re
            for row in att_notes:
                for note in [row['notes_in'], row['notes_out']]:
                    if note:
                        m = re.search(r'\(ID:\s*(-?\d+)\)', note)
                        if m:
                            group_ids.add(m.group(1))
            
            # Template wishes
            wishes = [
                f"🎉 Selamat Ulang Tahun untuk <b>{emp['name']}</b>! Semoga sehat selalu, sukses dalam karir, dan panjang umur! 🎂🎈",
                f"🎁 Happy Birthday <b>{emp['name']}</b>! Semoga hari ini luar biasa dan tahun depan dipenuhi berkah & kebahagiaan! 🎉🍰",
                f"🎂 Selamat Hari Lahir <b>{emp['name']}</b>! Semoga segala cita-cita tercapai, sukses selalu, dan diberkahi kebahagiaan! 🎁✨"
            ]
            import random
            wish = random.choice(wishes)
            
            # Send to groups
            for gid in group_ids:
                try:
                    url = f"https://api.telegram.org/bot{bot_token}/sendMessage"
                    payload = {
                        "chat_id": gid,
                        "text": wish,
                        "parse_mode": "HTML"
                    }
                    req_lib.post(url, json=payload, timeout=5)
                except Exception as ex:
                    print(f"[Birthday Broadcast Group Error] {ex}")
            
            # 2. Send personal chat greeting
            if emp['telegram_id']:
                try:
                    url = f"https://api.telegram.org/bot{bot_token}/sendMessage"
                    payload = {
                        "chat_id": emp['telegram_id'],
                        "text": f"🎉 Halo <b>{emp['name']}</b>,\n\nKami segenap keluarga besar HIVE mengucapkan: Selamat Ulang Tahun! 🎂🎁\nSemoga sehat, bahagia, dan sukses selalu di tahun yang baru ini! ✨🎈",
                        "parse_mode": "HTML"
                    }
                    req_lib.post(url, json=payload, timeout=5)
                except Exception as ex:
                    print(f"[Birthday Broadcast Personal Error] {ex}")
        
        db.close()
    except Exception as e:
        print(f"[run_birthday_reminders Error] {e}")


def check_checkout_reminders():
    try:
        db = _get_raw_db()
        settings = get_settings(db)
        bot_token = settings.get('telegram_bot_token', '').strip()
        if not bot_token:
            bot_token = os.environ.get('TELEGRAM_BOT_TOKEN', '1095530966:AAFkSV9puxmT2z7cvpsbBQy_TWqj9-MCvbM').strip()
        if not bot_token:
            db.close()
            return

        from datetime import datetime
        today = datetime.now().strftime('%Y-%m-%d')
        
        # Get active attendance today without clock out and checkout_reminder_sent = 0
        rows = db.execute('''
            SELECT a.id, a.user_id, a.clock_in, a.notes_in, u.full_name, e.telegram_id, u.username
            FROM attendance a
            JOIN users u ON a.user_id = u.id
            JOIN employees e ON e.user_id = u.id
            WHERE a.date = ? AND a.clock_in IS NOT NULL AND a.clock_out IS NULL AND (a.checkout_reminder_sent IS NULL OR a.checkout_reminder_sent = 0)
        ''', (today,)).fetchall()
        
        for r in rows:
            try:
                in_dt = datetime.strptime(f"{today} {r['clock_in']}", "%Y-%m-%d %H:%M:%S")
                curr_dt = datetime.now()
                diff = curr_dt - in_dt
                total_seconds = diff.total_seconds()
                
                # 9 hours = 32400 seconds
                if total_seconds >= 9 * 3600:
                    import re
                    import html
                    
                    user_display = f"@{html.escape(r['username'])}" if r['username'] else html.escape(r['full_name'])
                    msg_text = (
                        f"🔔 <b>[REMINDER]</b> {user_display}\n"
                        f"Waktu kerja Anda hari ini telah mencapai 9 jam (sejak {r['clock_in']}).\n"
                        f"Anda sudah dapat melakukan <b>Clock Out</b>.\n\n"
                        f"⚠️ Jangan lupa untuk mengisi <code>#PLAN</code> dan <code>#PROGRESS</code> sebelum melakukan Clock Out."
                    )
                    
                    # 1. Send to the group where the user checked in (if any)
                    group_id = None
                    if r['notes_in']:
                        m = re.search(r'\(ID:\s*(-?\d+)\)', r['notes_in'])
                        if m:
                            group_id = m.group(1)
                    
                    if group_id:
                        try:
                            url = f"https://api.telegram.org/bot{bot_token}/sendMessage"
                            payload = {
                                "chat_id": group_id,
                                "text": msg_text,
                                "parse_mode": "HTML"
                            }
                            resp = req_lib.post(url, json=payload, timeout=5)
                            if resp.status_code == 200:
                                resp_json = resp.json()
                                if resp_json.get('ok'):
                                    bot_msg_id = resp_json['result']['message_id']
                                    # Auto delete group reminder after 60 seconds
                                    def delete_msg(gid, mid):
                                        try:
                                            req_lib.post(f"https://api.telegram.org/bot{bot_token}/deleteMessage", json={
                                                'chat_id': gid,
                                                'message_id': mid
                                            }, timeout=5)
                                        except Exception:
                                            pass
                                    import threading
                                    threading.Timer(60.0, lambda: delete_msg(group_id, bot_msg_id)).start()
                        except Exception as ex:
                            print(f"[Checkout Reminder Group Error] {ex}")
                            
                    # 2. Send private message
                    if r['telegram_id']:
                        try:
                            url = f"https://api.telegram.org/bot{bot_token}/sendMessage"
                            payload = {
                                "chat_id": r['telegram_id'],
                                "text": msg_text,
                                "parse_mode": "HTML"
                            }
                            req_lib.post(url, json=payload, timeout=5)
                        except Exception as ex:
                            print(f"[Checkout Reminder Personal Error] {ex}")
                            
                    # 3. Mark as sent
                    db.execute('UPDATE attendance SET checkout_reminder_sent = 1 WHERE id = ?', (r['id'],))
                    db.commit()
            except Exception as e:
                print(f"[check_checkout_reminders Row Error] {e}")
                
        db.close()
    except Exception as e:
        print(f"[check_checkout_reminders Error] {e}")


def start_scheduler():
    try:
        from apscheduler.schedulers.background import BackgroundScheduler
        scheduler = BackgroundScheduler(daemon=True)
        scheduler.add_job(lambda: run_contract_reminders('auto'),
                          'cron', hour=8, minute=0,
                          id='contract_reminder', replace_existing=True)
        scheduler.add_job(lambda: run_subscription_reminders('auto'),
                          'cron', hour=8, minute=5,
                          id='sub_reminder', replace_existing=True)
        scheduler.add_job(run_birthday_reminders,
                          'cron', hour=9, minute=0,
                          id='birthday_reminder', replace_existing=True)
        scheduler.add_job(check_checkout_reminders,
                          'interval', minutes=5,
                          id='checkout_reminders', replace_existing=True)
        scheduler.add_job(check_for_updates,
                          'interval', hours=6,
                          id='update_check', replace_existing=True)
        scheduler.add_job(check_and_run_scheduled_backup,
                          'interval', minutes=1,
                          id='backup_scheduler', replace_existing=True)
        scheduler.start()
        import atexit
        atexit.register(scheduler.shutdown)
        print(" Scheduler aktif: cek kontrak 08:00, cek subscription 08:05, reminder ultah 09:00, cek update setiap 6 jam")
        # Cek update saat startup (non-blocking, delay 30 detik)
        import threading
        threading.Timer(30, check_for_updates).start()
    except Exception as e:
        print(f" Scheduler gagal: {e}")

# ─── Auth Routes ───────────────────────────────────────────────────────────────

def _verify_recaptcha(token, secret_key):
    """Verify reCAPTCHA v3 token, return score (0.0–1.0) or None on failure."""
    try:
        resp = req_lib.post('https://www.google.com/recaptcha/api/siteverify',
                            data={'secret': secret_key, 'response': token}, timeout=5)
        data = resp.json()
        if data.get('success'):
            return data.get('score', 0.0)
    except Exception:
        pass
    return None


@app.route('/login', methods=['GET', 'POST'])
def login():
    if 'user_id' in session:
        return redirect(url_for('index'))
    db = get_db()
    cfg = get_settings(db)
    recaptcha_enabled  = cfg.get('recaptcha_enabled') == '1'
    recaptcha_site_key = cfg.get('recaptcha_site_key', '').strip()
    recaptcha_secret   = cfg.get('recaptcha_secret_key', '').strip()
    if request.method == 'POST':
        username = request.form.get('username','').strip()
        password = request.form.get('password','')
        # reCAPTCHA v3 check
        if recaptcha_enabled and recaptcha_secret:
            rc_token = request.form.get('g-recaptcha-response', '')
            score = _verify_recaptcha(rc_token, recaptcha_secret)
            if score is None or score < 0.5:
                flash('Verifikasi bot gagal. Coba lagi.', 'danger')
                return render_template('login.html',
                    google_oauth_enabled=(cfg.get('google_oauth_enabled') == '1' and bool(cfg.get('google_client_id') or GOOGLE_CLIENT_ID)),
                    recaptcha_enabled=recaptcha_enabled,
                    recaptcha_site_key=recaptcha_site_key)
        user = db.execute('SELECT * FROM users WHERE (username=? OR LOWER(email)=?) AND is_active=1',
                          (username, username.lower())).fetchone()
        if user and check_password_hash(user['password_hash'], password):
            if user['mfa_enabled'] and user['totp_secret']:
                session['pending_mfa_user_id']   = user['id']
                _next_mfa = request.args.get('next') or ''
                if _next_mfa and (not _next_mfa.startswith('/') or _next_mfa.startswith('//')):
                    _next_mfa = ''
                session['pending_mfa_next'] = _next_mfa or url_for('portal')
                return redirect(url_for('login_mfa'))
            session['user_id']   = user['id']
            session['username']  = user['username']
            session['user_name'] = user['full_name'] or user['username']
            session['user_role'] = user['role']
            session['show_mfa_prompt'] = not bool(user['mfa_enabled'])
            db.execute('UPDATE users SET last_login=? WHERE id=?',
                       (datetime.now().strftime('%Y-%m-%d %H:%M:%S'), user['id']))
            db.commit()
            audit_log('login', 'users', user['id'], f"Login berhasil via form", app_slug='portal')
            flash(f'Selamat datang, {session["user_name"]}!', 'success')
            _next = request.args.get('next') or ''
            if _next and (not _next.startswith('/') or _next.startswith('//')):
                _next = ''
            return redirect(_next or url_for('portal'))
        flash('Username atau password salah', 'danger')
    google_on = (cfg.get('google_oauth_enabled') == '1' and bool(cfg.get('google_client_id') or GOOGLE_CLIENT_ID))
    return render_template('login.html',
        google_oauth_enabled=google_on,
        recaptcha_enabled=recaptcha_enabled,
        recaptcha_site_key=recaptcha_site_key)

@app.route('/login/mfa', methods=['GET', 'POST'])
def login_mfa():
    pending_id = session.get('pending_mfa_user_id')
    if not pending_id:
        return redirect(url_for('login'))
    if request.method == 'POST':
        code = request.form.get('code', '').strip()
        db   = get_db()
        user = db.execute('SELECT * FROM users WHERE id=?', (pending_id,)).fetchone()
        if user and verify_totp(user['totp_secret'], code):
            session.pop('pending_mfa_user_id', None)
            next_url = session.pop('pending_mfa_next', url_for('portal'))
            session['user_id']   = user['id']
            session['username']  = user['username']
            session['user_name'] = user['full_name'] or user['username']
            session['user_role'] = user['role']
            session['show_mfa_prompt'] = False  # sudah MFA, tidak perlu prompt
            db.execute('UPDATE users SET last_login=? WHERE id=?',
                       (datetime.now().strftime('%Y-%m-%d %H:%M:%S'), user['id']))
            db.commit()
            audit_log('login', 'users', user['id'], 'Login berhasil via MFA', app_slug='portal')
            flash(f'Selamat datang, {session["user_name"]}!', 'success')
            return redirect(next_url)
        flash('Kode MFA salah atau kadaluarsa. Coba lagi.', 'danger')
    return render_template('mfa_login.html')

def _google_callback_url(settings):
    """Bangun redirect URI yang konsisten: utamakan app_url dari settings."""
    base = (settings.get('app_url') or '').rstrip('/')
    if base:
        return base + '/login/google/callback'
    return url_for('login_google_callback', _external=True)

@app.route('/login/google')
def login_google():
    db = get_db()
    settings = get_settings(db)
    client_id = settings.get('google_client_id') or GOOGLE_CLIENT_ID
    oauth_on  = settings.get('google_oauth_enabled', '0') == '1' or GOOGLE_CLIENT_ID
    if not client_id or not oauth_on:
        flash('Login Google belum dikonfigurasi. Hubungi administrator.', 'warning')
        return redirect(url_for('login'))
    state = secrets.token_urlsafe(16)
    session['oauth_state'] = state
    callback_url = _google_callback_url(settings)
    params = {
        'client_id':     client_id,
        'redirect_uri':  callback_url,
        'response_type': 'code',
        'scope':         'openid email profile',
        'state':         state,
        'prompt':        'select_account',
        'access_type':   'online',
    }
    from urllib.parse import urlencode
    return redirect(f"{GOOGLE_AUTH_URL}?{urlencode(params)}")

@app.route('/login/google/callback')
def login_google_callback():
    db = get_db()
    settings = get_settings(db)
    client_id     = settings.get('google_client_id')     or GOOGLE_CLIENT_ID
    client_secret = settings.get('google_client_secret') or GOOGLE_CLIENT_SECRET

    error = request.args.get('error')
    if error:
        flash(f'Login Google dibatalkan: {error}', 'warning')
        return redirect(url_for('login'))

    state = request.args.get('state', '')
    if state != session.pop('oauth_state', None):
        flash('State tidak valid. Coba lagi.', 'danger')
        return redirect(url_for('login'))

    code = request.args.get('code')
    if not code:
        flash('Tidak mendapat kode otorisasi dari Google.', 'danger')
        return redirect(url_for('login'))

    callback_url = _google_callback_url(settings)
    try:
        token_resp = req_lib.post(GOOGLE_TOKEN_URL, data={
            'code':          code,
            'client_id':     client_id,
            'client_secret': client_secret,
            'redirect_uri':  callback_url,
            'grant_type':    'authorization_code',
        }, timeout=10)
        token_data = token_resp.json()
        access_token = token_data.get('access_token')
        if not access_token:
            raise ValueError(token_data.get('error_description', 'Tidak ada access token'))

        user_resp = req_lib.get(GOOGLE_USERINFO_URL,
                                headers={'Authorization': f'Bearer {access_token}'}, timeout=10)
        ginfo = user_resp.json()
    except Exception as e:
        flash(f'Gagal menghubungi Google: {e}', 'danger')
        return redirect(url_for('login'))

    google_email = ginfo.get('email', '').lower().strip()
    google_id    = ginfo.get('sub', '')
    google_name  = ginfo.get('name', '')
    verified     = ginfo.get('email_verified', False)

    if not google_email or not verified:
        flash('Email Google tidak terverifikasi.', 'danger')
        return redirect(url_for('login'))

    # Allowed domain check (optional — jika ada setting google_workspace_domain)
    allowed_domain = (settings.get('google_workspace_domain') or '').strip().lower()
    if allowed_domain:
        domain = google_email.split('@')[-1]
        if domain != allowed_domain:
            flash(f'Hanya email @{allowed_domain} yang diizinkan masuk.', 'danger')
            audit_log('login_google_rejected', 'users', 0,
                      f'Email ditolak (domain): {google_email}', app_slug='portal')
            return redirect(url_for('login'))

    # Cari user berdasarkan google_id dulu, lalu email
    user = db.execute("SELECT * FROM users WHERE google_id=? AND is_active=1",
                      (google_id,)).fetchone()
    if not user:
        user = db.execute("SELECT * FROM users WHERE LOWER(email)=? AND is_active=1",
                          (google_email,)).fetchone()

    emp = None
    if not user:
        # Cek apakah ada karyawan dengan email yang sama dan sudah punya user manual
        emp = db.execute("SELECT * FROM employees WHERE LOWER(email)=? AND is_active=1",
                         (google_email,)).fetchone()
        if emp and emp['user_id']:
            # Merge: hubungkan google_id ke user manual yang sudah ada
            existing = db.execute('SELECT * FROM users WHERE id=? AND is_active=1',
                                  (emp['user_id'],)).fetchone()
            if existing:
                db.execute("UPDATE users SET google_id=?, email=? WHERE id=?",
                           (google_id, google_email, existing['id']))
                db.commit()
                user = db.execute('SELECT * FROM users WHERE id=?', (existing['id'],)).fetchone()
                audit_log('merge_google', 'users', user['id'],
                          f'Akun manual digabung dengan Google ({google_email})', app_slug='portal')

    is_new_user = False
    if not user:
        # Auto-create user baru tanpa akses app — admin harus grant via Akses Aplikasi
        username_base = google_email.split('@')[0].lower().replace('.', '_')
        username = username_base
        suffix = 1
        while db.execute('SELECT id FROM users WHERE username=?', (username,)).fetchone():
            username = f'{username_base}{suffix}'
            suffix += 1
        cur = db.execute(
            'INSERT INTO users(username,password_hash,full_name,role,email,google_id,is_active) VALUES(?,?,?,?,?,?,1)',
            (username, generate_password_hash(secrets.token_hex(32), method='pbkdf2:sha256'),
             google_name, 'viewer', google_email, google_id))
        db.commit()
        new_uid = cur.lastrowid
        # Auto-link ke karyawan jika email cocok dan belum punya user
        if emp and not emp['user_id']:
            db.execute('UPDATE employees SET user_id=? WHERE id=?', (new_uid, emp['id']))
            db.commit()
        # Berikan default akses ke AttendanceCore dengan role user
        db.execute(
            'INSERT INTO user_app_access(user_id, app_slug, app_role, is_active) VALUES(?,?,?,1)',
            (new_uid, 'attendance', 'user')
        )
        db.commit()
        user = db.execute('SELECT * FROM users WHERE id=?', (new_uid,)).fetchone()
        is_new_user = True
        audit_log('register_google', 'users', user['id'],
                  f'Akun baru via Google ({google_email}) — diberikan akses default AttendanceCore (user)', app_slug='portal')

    # Pastikan tidak ada sisa akses aktif dari akun lama (seharusnya tidak ada, tapi defensive)
    # Tidak lakukan apa-apa — akses dikontrol sepenuhnya via user_app_access

    # Update google_id dan last_login
    db.execute("UPDATE users SET google_id=?, last_login=? WHERE id=?",
               (google_id, datetime.now().strftime('%Y-%m-%d %H:%M:%S'), user['id']))
    db.commit()

    session['user_id']   = user['id']
    session['username']  = user['username']
    session['user_name'] = user['full_name'] or google_name or user['username']
    session['user_role'] = user['role']
    session['show_mfa_prompt'] = not bool(user['mfa_enabled'])
    audit_log('login_google', 'users', user['id'],
              f'Login via Google ({google_email})', app_slug='portal')
    if is_new_user:
        flash(
            f'Selamat datang, {session["user_name"]}! '
            'Akun Anda telah terdaftar dengan akses default ke AttendanceCore. Hubungi administrator jika Anda memerlukan akses ke aplikasi lainnya.',
            'info'
        )
    else:
        flash(f'Selamat datang, {session["user_name"]}!', 'success')
    return redirect(url_for('portal'))

@app.route('/mfa/challenge', methods=['GET', 'POST'])
@login_required
def mfa_challenge():
    if request.method == 'POST':
        code = request.form.get('code', '').strip()
        db   = get_db()
        user = db.execute('SELECT * FROM users WHERE id=?', (session['user_id'],)).fetchone()
        if user and verify_totp(user['totp_secret'], code):
            sk = session.pop('mfa_session_key', 'mfa_verified')
            session[sk] = datetime.now().timestamp()
            return_app = session.pop('mfa_return_app_slug', None)
            if return_app:
                session['active_app'] = return_app
            return redirect(session.pop('mfa_return_to', url_for('index')))
        flash('Kode MFA salah', 'danger')
    return render_template('mfa_challenge.html')

@app.route('/profile/mfa/setup', methods=['GET', 'POST'])
@login_required
def mfa_setup():
    db   = get_db()
    user = db.execute('SELECT * FROM users WHERE id=?', (session['user_id'],)).fetchone()
    if request.method == 'POST':
        action = request.form.get('action')
        if action == 'generate':
            secret = generate_totp_secret()
            session['pending_totp_secret'] = secret
            uri    = get_totp_uri(secret, user['username'])
            qr     = qr_png_base64(uri)
            return render_template('mfa_setup.html', user=user, secret=secret, qr=qr, step='verify')
        elif action == 'verify':
            secret = session.get('pending_totp_secret', '')
            code   = request.form.get('code', '').strip()
            if secret and verify_totp(secret, code):
                db.execute('UPDATE users SET totp_secret=?, mfa_enabled=1 WHERE id=?',
                           (secret, user['id']))
                db.commit()
                session.pop('pending_totp_secret', None)
                flash('Google Authenticator MFA berhasil diaktifkan! Silakan pilih aplikasi.', 'success')
                return redirect(url_for('portal'))
            flash('Kode verifikasi salah. Scan ulang QR dan coba lagi.', 'danger')
            uri = get_totp_uri(secret, user['username'])
            qr  = qr_png_base64(uri)
            return render_template('mfa_setup.html', user=user, secret=secret, qr=qr, step='verify')
        elif action == 'disable':
            code = request.form.get('code', '').strip()
            if user['totp_secret'] and verify_totp(user['totp_secret'], code):
                db.execute("UPDATE users SET totp_secret='', mfa_enabled=0 WHERE id=?", (user['id'],))
                db.commit()
                flash('MFA berhasil dinonaktifkan.', 'warning')
                return redirect(url_for('mfa_setup'))
            flash('Kode Authenticator salah. MFA tidak dinonaktifkan.', 'danger')
    mfa_on = bool(user['mfa_enabled'] and user['totp_secret'])
    if mfa_on:
        session['show_mfa_prompt'] = False
    return render_template('mfa_setup.html', user=user, step='intro', mfa_on=mfa_on)

@app.route('/mfa/dismiss-prompt', methods=['POST'])
@login_required
def mfa_dismiss_prompt():
    """User pilih 'Aktifkan Nanti' — sembunyikan prompt untuk sesi ini."""
    session['show_mfa_prompt'] = False
    return ('', 204)

@app.route('/admin/users/<int:uid>/send-reset', methods=['POST'])
@superadmin_required
def admin_send_reset(uid):
    """Superadmin kirim link reset password ke user tertentu."""
    db   = get_db()
    user = db.execute('SELECT * FROM users WHERE id=? AND is_active=1', (uid,)).fetchone()
    if not user:
        flash('User tidak ditemukan atau tidak aktif.', 'danger')
        return redirect(url_for('users_list'))
    # Hapus token lama
    db.execute('DELETE FROM password_reset_tokens WHERE user_id=?', (uid,))
    token   = secrets.token_urlsafe(48)
    expires = (datetime.now() + timedelta(hours=1)).strftime('%Y-%m-%d %H:%M:%S')
    db.execute('INSERT INTO password_reset_tokens(user_id, token, expires_at) VALUES(?,?,?)',
               (uid, token, expires))
    db.commit()
    settings   = get_settings(db)
    reset_link = f"{get_base_url(settings)}/reset-password/{token}"
    sent = _send_reset_notifications(dict(user), reset_link, settings, db)
    if sent:
        flash(f'Link reset password dikirim ke {user["username"]} via: {", ".join(sent)}.', 'success')
    else:
        flash(f'Tidak ada kontak terdaftar untuk {user["username"]}. '
              f'Salin link ini dan kirim manual: {reset_link}', 'warning')
    return redirect(url_for('users_list'))

@app.route('/admin/users/<int:uid>/mfa-reset', methods=['POST'])
@superadmin_required
def admin_mfa_reset(uid):
    db = get_db()
    db.execute("UPDATE users SET totp_secret='', mfa_enabled=0 WHERE id=?", (uid,))
    db.commit()
    flash('MFA user berhasil direset. User harus setup ulang saat login berikutnya.', 'warning')
    return redirect(url_for('users_list'))

# ─── Forgot / Reset Password ───────────────────────────────────────────────────

def get_base_url(settings=None):
    """Kembalikan URL publik aplikasi. Prioritas: setting app_url > ProxyFix request.host_url."""
    if settings:
        configured = (settings.get('app_url') or '').strip().rstrip('/')
        if configured:
            return configured
    return request.host_url.rstrip('/')

RESET_TOKEN_TTL = 3600  # 1 jam

def _send_reset_notifications(user, reset_link, settings, db=None):
    """Kirim link reset password via email, Telegram, dan WhatsApp.
    Mengirim ke semua kontak unik yang terdaftar di user fields maupun data karyawan."""
    emails = set()
    phones = set()
    telegram_ids = set()

    # Ambil dari user
    if user.get('email'):
        emails.add(user['email'].strip())
    if user.get('phone'):
        phones.add(user['phone'].strip())
    if user.get('telegram_id'):
        telegram_ids.add(user['telegram_id'].strip())

    # Ambil dari employee
    if db is not None:
        uid = user['id'] if hasattr(user, 'keys') else user.get('id')
        emp = db.execute(
            'SELECT email, phone, telegram_id FROM employees WHERE user_id=? AND is_active=1',
            (uid,)
        ).fetchone()
        if emp:
            if emp['email']:
                emails.add(emp['email'].strip())
            if emp['phone']:
                phones.add(emp['phone'].strip())
            if emp['telegram_id']:
                telegram_ids.add(emp['telegram_id'].strip())

    sent = []
    display_name = user.get('full_name') or user.get('username')
    subject  = 'Reset Password — Aplikasi TalentCore'
    body_html = f'''
<p>Halo <b>{display_name}</b>,</p>
<p>Anda (atau seseorang) meminta reset password untuk akun <b>{user.get('username')}</b>.</p>
<p><a href="{reset_link}" style="background:#1a7a3a;color:#fff;padding:10px 20px;border-radius:6px;
   text-decoration:none;font-weight:bold">Klik di sini untuk reset password</a></p>
<p>Link ini hanya berlaku <b>1 jam</b> dan <b>langsung kadaluarsa setelah dibuka satu kali</b>.</p>
<p>Jika bukan Anda yang meminta, abaikan email ini.</p>
<hr><p style="color:#888;font-size:12px">Aplikasi TalentCore</p>
'''

    # Email
    for email in sorted(emails):
        if email:
            ok, _ = send_email(settings, email, subject, body_html)
            if ok and 'email' not in sent:
                sent.append('email')

    # Telegram
    bot_token = settings.get('telegram_bot_token','').strip()
    if bot_token:
        for tg_id in sorted(telegram_ids):
            if tg_id:
                tg_msg = (f'🔐 *Reset Password*\n\nHalo {display_name},\n'
                          f'Klik link berikut untuk reset password akun `{user.get("username")}`:\n\n'
                          f'{reset_link}\n\n_Link berlaku 1 jam dan sekali pakai._')
                ok, _ = send_telegram(bot_token, tg_id, tg_msg)
                if ok and 'telegram' not in sent:
                    sent.append('telegram')

    # WhatsApp
    wa_url     = settings.get('openwa_url','').strip()
    wa_key     = settings.get('openwa_api_key','').strip()
    wa_session = get_openwa_session(settings)
    if wa_url:
        for phone in sorted(phones):
            if phone:
                wa_msg = (f'🔐 Reset Password\n\nHalo {display_name},\n'
                          f'Link reset password akun *{user.get("username")}*:\n\n{reset_link}\n\n'
                          f'Link berlaku 1 jam dan langsung kadaluarsa setelah dibuka.')
                ok, _ = send_whatsapp(wa_url, wa_key, wa_session, phone, wa_msg)
                if ok and 'whatsapp' not in sent:
                    sent.append('whatsapp')

    return sent


@app.route('/forgot-password', methods=['GET', 'POST'])
def forgot_password():
    if 'user_id' in session:
        return redirect(url_for('index'))
    if request.method == 'POST':
        identifier = request.form.get('identifier', '').strip()
        db = get_db()
        user = db.execute(
            "SELECT * FROM users WHERE (username=? OR email=?) AND is_active=1",
            (identifier, identifier)
        ).fetchone()
        if user:
            # Invalidate old tokens
            db.execute("DELETE FROM password_reset_tokens WHERE user_id=?", (user['id'],))
            token    = secrets.token_urlsafe(48)
            expires  = (datetime.now() + timedelta(hours=1)).strftime('%Y-%m-%d %H:%M:%S')
            db.execute(
                "INSERT INTO password_reset_tokens(user_id, token, expires_at) VALUES(?,?,?)",
                (user['id'], token, expires)
            )
            db.commit()
            settings   = get_settings(db)
            reset_link = f"{get_base_url(settings)}/reset-password/{token}"
            sent = _send_reset_notifications(dict(user), reset_link, settings, db)
            contacts = get_user_contacts(db, dict(user))
            if sent:
                flash(f'Link reset password dikirim via: {", ".join(sent)}. '
                      f'Cek email/Telegram/WhatsApp Anda.', 'success')
            elif not any([contacts.get('email'), contacts.get('phone'), contacts.get('telegram_id')]):
                flash('Tidak ada kontak (email/WA/Telegram) yang terdaftar pada akun ini. '
                      'Hubungi admin untuk reset password.', 'danger')
            else:
                flash('Gagal mengirim notifikasi — periksa konfigurasi email/WA/Telegram di pengaturan. '
                      'Hubungi admin untuk mendapatkan link reset.', 'warning')
        else:
            # Pesan generik agar tidak bisa enumerate user
            flash('Jika username/email terdaftar, link reset akan dikirim ke kontak yang tersimpan.', 'info')
        return redirect(url_for('forgot_password'))
    return render_template('forgot_password.html')


@app.route('/reset-password/<token>', methods=['GET', 'POST'])
def reset_password(token):
    if 'user_id' in session:
        return redirect(url_for('index'))
    db  = get_db()
    row = db.execute(
        "SELECT * FROM password_reset_tokens WHERE token=?", (token,)
    ).fetchone()

    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    if not row or row['used'] or row['expires_at'] < now:
        flash('Link reset tidak valid atau sudah kadaluarsa. Silakan minta link baru.', 'danger')
        return redirect(url_for('forgot_password'))

    user = db.execute("SELECT * FROM users WHERE id=?", (row['user_id'],)).fetchone()
    if not user:
        flash('Akun tidak ditemukan.', 'danger')
        return redirect(url_for('forgot_password'))

    if request.method == 'POST':
        new_pass  = request.form.get('password', '').strip()
        new_pass2 = request.form.get('password2', '').strip()
        if len(new_pass) < 6:
            flash('Password minimal 6 karakter.', 'danger')
        elif new_pass != new_pass2:
            flash('Konfirmasi password tidak cocok.', 'danger')
        else:
            db.execute("UPDATE users SET password_hash=? WHERE id=?",
                       (generate_password_hash(new_pass, method='pbkdf2:sha256'), user['id']))
            # Mark token as used — link langsung kadaluarsa setelah dipakai
            db.execute("UPDATE password_reset_tokens SET used=1 WHERE token=?", (token,))
            db.commit()
            flash('Password berhasil diubah. Silakan login.', 'success')
            return redirect(url_for('login'))
        return render_template('reset_password.html', token=token, user=user)

    # GET: mark token as used agar link tidak bisa dibuka ulang
    # Tapi jangan invalidate dulu — baru invalidate setelah POST sukses
    # (agar form bisa disubmit dari halaman yang sama)
    return render_template('reset_password.html', token=token, user=user)

@app.route('/logout')
def logout():
    audit_log('logout', 'users', session.get('user_id',''), 'User logout', app_slug='portal')
    session.clear()
    flash('Anda telah logout', 'info')
    return redirect(url_for('login'))

# ─── User Management (Superadmin) ─────────────────────────────────────────────

@app.route('/users')
@superadmin_required
def users_list():
    db    = get_db()
    users = db.execute('SELECT * FROM users ORDER BY role, username').fetchall()
    # Map user_id → linked employee contact (untuk ditampilkan di tabel)
    emps  = db.execute(
        'SELECT user_id, email, phone, telegram_id, name FROM employees WHERE user_id IS NOT NULL AND is_active=1'
    ).fetchall()
    linked_emps = {e['user_id']: e for e in emps}
    return render_template('users.html', users=users, linked_emps=linked_emps)

@app.route('/users/add', methods=['GET', 'POST'])
@superadmin_required
def user_add():
    if request.method == 'POST':
        username  = request.form['username'].strip()
        full_name = request.form.get('full_name','').strip()
        password  = request.form.get('password','')
        role      = request.form.get('role','admin')
        email     = request.form.get('email','').strip()
        phone     = request.form.get('phone','').strip()
        telegram  = request.form.get('telegram_id','').strip()
        if not username or not password:
            flash('Username dan password wajib diisi', 'danger')
        else:
            db = get_db()
            try:
                db.execute(
                    'INSERT INTO users(username,password_hash,full_name,role,email,phone,telegram_id) VALUES(?,?,?,?,?,?,?)',
                    (username, generate_password_hash(password, method='pbkdf2:sha256'),
                     full_name, role, email, phone, telegram))
                db.commit()
                flash(f'User {username} berhasil dibuat', 'success')
                return redirect(url_for('users_list'))
            except Exception as e:
                db.rollback()
                if _is_db_integrity_error(e):
                    flash('Username sudah digunakan', 'danger')
                else:
                    raise
    return render_template('user_form.html', user=None)

@app.route('/users/<int:uid>/edit', methods=['GET', 'POST'])
@superadmin_required
def user_edit(uid):
    db = get_db()
    user = db.execute('SELECT * FROM users WHERE id=?', (uid,)).fetchone()
    if not user:
        flash('User tidak ditemukan', 'danger')
        return redirect(url_for('users_list'))
    if request.method == 'POST':
        full_name  = request.form.get('full_name','').strip()
        role       = request.form.get('role', user['role'])
        is_active  = 1 if request.form.get('is_active') else 0
        email      = request.form.get('email','').strip()
        phone      = request.form.get('phone','').strip()
        telegram   = request.form.get('telegram_id','').strip()
        new_pass   = request.form.get('password','').strip()
        if new_pass:
            db.execute(
                'UPDATE users SET full_name=?,role=?,is_active=?,email=?,phone=?,telegram_id=?,password_hash=? WHERE id=?',
                (full_name, role, is_active, email, phone, telegram,
                 generate_password_hash(new_pass, method='pbkdf2:sha256'), uid))
        else:
            db.execute(
                'UPDATE users SET full_name=?,role=?,is_active=?,email=?,phone=?,telegram_id=? WHERE id=?',
                (full_name, role, is_active, email, phone, telegram, uid))
        db.commit()
        flash('User diperbarui', 'success')
        if uid == session['user_id']:
            session['user_name'] = full_name or session['username']
            session['user_role'] = role
        return redirect(url_for('users_list'))
    linked_emp = db.execute(
        'SELECT name, email, phone, telegram_id FROM employees WHERE user_id=? AND is_active=1',
        (uid,)
    ).fetchone()
    return render_template('user_form.html', user=user, linked_emp=linked_emp)

@app.route('/users/<int:uid>/delete', methods=['POST'])
@superadmin_required
def user_delete(uid):
    if uid == session['user_id']:
        flash('Tidak bisa menghapus akun sendiri', 'danger')
        return redirect(url_for('users_list'))
    db = get_db()
    db.execute('DELETE FROM users WHERE id=?', (uid,))
    db.commit()
    flash('User dihapus', 'warning')
    return redirect(url_for('users_list'))

@app.route('/profile', methods=['GET', 'POST'])
@login_required
def profile():
    db = get_db()
    user = db.execute('SELECT * FROM users WHERE id=?', (session['user_id'],)).fetchone()
    if not user:
        session.clear()
        flash('Sesi Anda tidak valid atau pengguna tidak ditemukan', 'danger')
        return redirect(url_for('login'))
    if request.method == 'POST':
        full_name = request.form.get('full_name','').strip()
        old_pass  = request.form.get('old_password','')
        new_pass  = request.form.get('new_password','').strip()
        if new_pass:
            if not check_password_hash(user['password_hash'], old_pass):
                flash('Password lama salah', 'danger')
                return render_template('profile.html', user=user)
            db.execute('UPDATE users SET full_name=?,password_hash=? WHERE id=?',
                       (full_name, generate_password_hash(new_pass), session['user_id']))
        else:
            db.execute('UPDATE users SET full_name=? WHERE id=?',
                       (full_name, session['user_id']))
        db.commit()
        session['user_name'] = full_name or session['username']
        flash('Profil diperbarui', 'success')
        return redirect(url_for('profile'))
    return render_template('profile.html', user=user)

# ─── Dashboard ─────────────────────────────────────────────────────────────────

@app.route('/portal')
@login_required
def portal():
    session.pop('active_app', None)   # clear app saat kembali ke portal
    db   = get_db()
    uid  = session.get('user_id')
    role = session.get('user_role', '')
    all_apps = db.execute('SELECT * FROM superapp_apps WHERE is_active=1 ORDER BY sort_order, name').fetchall()
    if role == 'superadmin':
        accessible = {a['slug'] for a in all_apps}
    else:
        rows = db.execute(
            'SELECT app_slug FROM user_app_access WHERE user_id=? AND is_active=1', (uid,)
        ).fetchall()
        accessible = {r['app_slug'] for r in rows}
    return render_template('portal.html', apps=all_apps, accessible=accessible)

@app.route('/portal/open/<slug>')
@login_required
def portal_open(slug):
    db  = get_db()
    app_row = db.execute('SELECT * FROM superapp_apps WHERE slug=? AND is_active=1', (slug,)).fetchone()
    if not app_row:
        flash('Aplikasi tidak ditemukan.', 'danger')
        return redirect(url_for('portal'))
    # Cek akses (superadmin bypass)
    if session.get('user_role') != 'superadmin':
        acc = db.execute('SELECT 1 FROM user_app_access WHERE user_id=? AND app_slug=? AND is_active=1',
                         (session['user_id'], slug)).fetchone()
        if not acc:
            flash('Anda tidak memiliki akses ke aplikasi ini.', 'danger')
            return redirect(url_for('portal'))
    if app_row['is_coming_soon']:
        flash('Aplikasi ini belum tersedia.', 'info')
        return redirect(url_for('portal'))
    session['active_app'] = slug
    audit_log('open_app', 'superapp_apps', slug, f"Buka aplikasi: {app_row['name']}", app_slug='portal')
    return redirect(app_row['url'])

@app.route('/portal/settings', methods=['GET', 'POST'])
@login_required
def portal_settings():
    if not is_portal_admin():
        flash('Akses ditolak.', 'danger')
        return redirect(url_for('portal'))
    db    = get_db()
    users = db.execute('SELECT id, username, full_name, role, email, google_id FROM users WHERE is_active=1 ORDER BY role DESC, username').fetchall()
    apps  = db.execute('SELECT * FROM superapp_apps WHERE is_active=1 ORDER BY sort_order').fetchall()

    if request.method == 'POST':
        # Hapus semua akses non-superadmin lalu rebuild dari form
        db.execute("DELETE FROM user_app_access WHERE user_id IN (SELECT id FROM users WHERE role != 'superadmin')")
        for u in users:
            if u['role'] == 'superadmin':
                continue
            for a in apps:
                key    = f"access_{u['id']}_{a['slug']}"
                role_k = f"role_{u['id']}_{a['slug']}"
                if request.form.get(key):
                    app_role = request.form.get(role_k, 'user')
                    db.execute('''INSERT INTO user_app_access(user_id,app_slug,app_role,is_active)
                                  VALUES(?,?,?,1)
                                  ON CONFLICT(user_id,app_slug) DO UPDATE SET app_role=excluded.app_role,is_active=1''',
                               (u['id'], a['slug'], app_role))
        db.commit()
        flash('Pengaturan akses berhasil disimpan.', 'success')
        return redirect(url_for('portal_settings'))

    # Baca akses saat ini
    access_map = {}
    rows = db.execute('SELECT user_id, app_slug, app_role, is_active FROM user_app_access').fetchall()
    for r in rows:
        access_map[(r['user_id'], r['app_slug'])] = {'role': r['app_role'], 'active': r['is_active']}

    # Roles per app — include global roles (app_slug='') untuk setiap app
    all_roles    = db.execute("SELECT name, description, app_slug FROM roles ORDER BY is_system DESC, name").fetchall()
    global_roles = [r for r in all_roles if r['app_slug'] == '']
    roles_by_app = {}
    for a in apps:
        app_specific = [r for r in all_roles if r['app_slug'] == a['slug']]
        roles_by_app[a['slug']] = global_roles + app_specific

    employees = db.execute('SELECT id, name, email, divisi, user_id FROM employees WHERE is_active=1 ORDER BY name').fetchall()

    return render_template('portal_settings.html', users=users, apps=apps,
                           access_map=access_map, roles_by_app=roles_by_app,
                           employees=employees)

@app.route('/portal/settings/mass', methods=['POST'])
@login_required
def portal_settings_mass():
    if not is_portal_admin():
        flash('Akses ditolak.', 'danger')
        return redirect(url_for('portal'))
    db = get_db()
    employee_ids = request.form.getlist('employee_ids')
    app_slug = request.form.get('app_slug')
    app_role = request.form.get('app_role', 'user')

    if not employee_ids or not app_slug:
        flash('Karyawan dan Aplikasi wajib dipilih.', 'danger')
        return redirect(url_for('portal_settings'))

    success_count = 0
    try:
        for eid in employee_ids:
            emp = db.execute('SELECT id, name, email, user_id FROM employees WHERE id=?', (eid,)).fetchone()
            if not emp or not emp['email']:
                continue

            email_lower = emp['email'].lower().strip()

            user = None
            if emp['user_id']:
                user = db.execute('SELECT id, role FROM users WHERE id=?', (emp['user_id'],)).fetchone()
            if not user:
                user = db.execute('SELECT id, role FROM users WHERE LOWER(email)=? OR LOWER(username)=?', 
                                  (email_lower, email_lower)).fetchone()

            if user:
                uid = user['id']
                if not emp['user_id']:
                    db.execute('UPDATE employees SET user_id=? WHERE id=?', (uid, eid))
            else:
                p_hash = generate_password_hash('hive2026', method='pbkdf2:sha256')
                cur = db.execute('''
                    INSERT INTO users (username, password_hash, full_name, role, email, is_active)
                    VALUES (?, ?, ?, 'viewer', ?, 1)
                ''', (emp['email'].strip(), p_hash, emp['name'], emp['email'].strip()))
                uid = cur.lastrowid

                db.execute('UPDATE employees SET user_id=? WHERE id=?', (uid, eid))

            check_user = db.execute('SELECT role FROM users WHERE id=?', (uid,)).fetchone()
            if check_user and check_user['role'] == 'superadmin':
                continue

            db.execute('''INSERT INTO user_app_access(user_id,app_slug,app_role,is_active)
                          VALUES(?, ?, ?, 1)
                          ON CONFLICT(user_id,app_slug) DO UPDATE SET app_role=excluded.app_role, is_active=1''',
                       (uid, app_slug, app_role))
            success_count += 1

        db.commit()
        flash(f'Akses aplikasi "{app_slug}" dengan role "{app_role}" berhasil diberikan ke {success_count} karyawan (akun baru dibuat jika belum ada).', 'success')
    except Exception as e:
        db.rollback()
        flash(f'Gagal menerapkan akses massal: {str(e)}', 'danger')

    return redirect(url_for('portal_settings'))



def is_portal_admin():
    """True jika user adalah superadmin atau admin (dapat akses portal management)."""
    return session.get('user_role') in ('superadmin', 'admin')

def is_superadmin():
    return session.get('user_role') == 'superadmin'

# ─── Portal: Kelola User ───────────────────────────────────────────────────────

@app.route('/portal/api/employees-search')
@login_required
def portal_emp_search():
    q   = request.args.get('q', '').strip()
    db  = get_db()
    if len(q) < 2:
        return jsonify([])
    rows = db.execute('''
        SELECT id, name, jabatan, divisi, email, phone, telegram_id
        FROM employees WHERE is_active=1 AND user_id IS NULL
        AND (name LIKE ? OR jabatan LIKE ? OR divisi LIKE ?)
        ORDER BY name LIMIT 20
    ''', (f'%{q}%', f'%{q}%', f'%{q}%')).fetchall()
    return jsonify([dict(r) for r in rows])

@app.route('/portal/users')
@login_required
def portal_users():
    if not is_portal_admin():
        flash('Akses ditolak.', 'danger')
        return redirect(url_for('portal'))
    db    = get_db()
    users = db.execute('SELECT * FROM users ORDER BY role, username').fetchall()
    emps  = db.execute(
        'SELECT user_id, email, phone, telegram_id, name FROM employees WHERE user_id IS NOT NULL AND is_active=1'
    ).fetchall()
    linked_emps = {e['user_id']: e for e in emps}
    apps = db.execute('SELECT slug, name FROM superapp_apps WHERE is_active=1 ORDER BY sort_order').fetchall()
    # Akses per user
    access_rows = db.execute('SELECT user_id, app_slug, app_role FROM user_app_access WHERE is_active=1').fetchall()
    user_access = {}
    for r in access_rows:
        user_access.setdefault(r['user_id'], {})[r['app_slug']] = r['app_role']
    # Deteksi semua duplikat email — termasuk email dari tabel employees
    from collections import defaultdict
    # effective email: users.email jika ada, fallback ke linked employee email
    def eff_email(u):
        e = (u['email'] or '').strip().lower()
        if not e:
            emp = linked_emps.get(u['id'])
            if emp:
                e = (emp['email'] or '').strip().lower()
        return e

    email_groups = defaultdict(list)
    for u in users:
        if not u['is_active']:
            continue
        em = eff_email(u)
        if em:
            email_groups[em].append(u)
    duplicate_emails = {email: grp for email, grp in email_groups.items() if len(grp) > 1}
    # merge_candidates: google_uid -> manual_uid
    merge_candidates = {}
    for grp in duplicate_emails.values():
        google_u = next((u for u in grp if u['google_id']), None)
        manual_u = next((u for u in grp if not u['google_id']), None)
        if google_u and manual_u:
            merge_candidates[google_u['id']] = manual_u['id']
    return render_template('portal_users.html', users=users, linked_emps=linked_emps,
                           apps=apps, user_access=user_access,
                           merge_candidates=merge_candidates,
                           duplicate_count=len(duplicate_emails))

@app.route('/portal/users/merge-all-duplicates', methods=['POST'])
@login_required
def portal_user_merge_all():
    if not is_portal_admin():
        flash('Akses ditolak.', 'danger')
        return redirect(url_for('portal_users'))
    db = get_db()
    users = db.execute('SELECT * FROM users WHERE is_active=1').fetchall()
    # Bangun map user_id -> employee email untuk fallback
    emp_emails = {r['user_id']: (r['email'] or '').strip().lower()
                  for r in db.execute('SELECT user_id, email FROM employees WHERE user_id IS NOT NULL AND is_active=1').fetchall()}
    from collections import defaultdict
    email_groups = defaultdict(list)
    for u in users:
        em = (u['email'] or '').strip().lower() or emp_emails.get(u['id'], '')
        if em:
            email_groups[em].append(dict(u))
    merged = 0
    for email, grp in email_groups.items():
        if len(grp) < 2:
            continue
        # Prioritas: user dengan google_id jadi "donor", user tanpa google_id jadi "penerima"
        # Jika keduanya manual/google, pilih yang lebih lama (id lebih kecil) sebagai penerima
        google_users = [u for u in grp if u['google_id']]
        manual_users = [u for u in grp if not u['google_id']]
        if not google_users:
            # Dua user manual: nonaktifkan yang lebih baru
            grp_sorted = sorted(grp, key=lambda u: u['id'])
            keep, drop = grp_sorted[0], grp_sorted[1]
        else:
            keep  = manual_users[0] if manual_users else sorted(google_users, key=lambda u: u['id'])[0]
            donor = google_users[0]
            # Pastikan email tersimpan di users.email untuk keep
            keep_email = (keep.get('email') or '').strip() or emp_emails.get(keep['id'], '')
            if keep_email and not keep.get('email'):
                db.execute("UPDATE users SET email=? WHERE id=?", (keep_email, keep['id']))
            # Pindahkan google_id ke keep
            db.execute("UPDATE users SET google_id=? WHERE id=?", (donor['google_id'], keep['id']))
            # Pindahkan akses app (query sekali per donor, bukan dalam outer loop)
            donor_access = db.execute('SELECT * FROM user_app_access WHERE user_id=?', (donor['id'],)).fetchall()
            for a in donor_access:
                db.execute('INSERT OR IGNORE INTO user_app_access(user_id,app_slug,app_role,is_active) VALUES(?,?,?,?)',
                           (keep['id'], a['app_slug'], a['app_role'], a['is_active']))
            # Update link karyawan
            db.execute("UPDATE employees SET user_id=? WHERE user_id=?", (keep['id'], donor['id']))
            drop = donor
        db.execute("UPDATE users SET is_active=0, email='' WHERE id=?", (drop['id'],))
        audit_log('merge_user_bulk', 'users', keep['id'],
                  f'Duplikat email {email}: akun {drop["username"]} digabung ke {keep["username"]}',
                  app_slug='portal')
        merged += 1
    db.commit()
    flash(f'{merged} duplikat email berhasil digabungkan.', 'success')
    return redirect(url_for('portal_users'))

@app.route('/portal/users/<int:google_uid>/merge/<int:manual_uid>', methods=['POST'])
@login_required
def portal_user_merge(google_uid, manual_uid):
    if not is_portal_admin():
        flash('Akses ditolak.', 'danger')
        return redirect(url_for('portal_users'))
    db = get_db()
    g_user = db.execute('SELECT * FROM users WHERE id=?', (google_uid,)).fetchone()
    m_user = db.execute('SELECT * FROM users WHERE id=?', (manual_uid,)).fetchone()
    if not g_user or not m_user:
        flash('User tidak ditemukan.', 'danger')
        return redirect(url_for('portal_users'))
    # Pindahkan google_id ke user manual
    db.execute("UPDATE users SET google_id=?, email=COALESCE(NULLIF(email,''),?) WHERE id=?",
               (g_user['google_id'], g_user['email'], manual_uid))
    # Pindahkan akses app dari Google user ke user manual (INSERT OR IGNORE)
    access = db.execute('SELECT * FROM user_app_access WHERE user_id=?', (google_uid,)).fetchall()
    for a in access:
        db.execute('''INSERT OR IGNORE INTO user_app_access(user_id,app_slug,app_role,is_active)
                      VALUES(?,?,?,?)''', (manual_uid, a['app_slug'], a['app_role'], a['is_active']))
    # Nonaktifkan akun Google duplikat
    db.execute("UPDATE users SET is_active=0 WHERE id=?", (google_uid,))
    # Update link karyawan jika ada
    db.execute("UPDATE employees SET user_id=? WHERE user_id=?", (manual_uid, google_uid))
    db.commit()
    audit_log('merge_user', 'users', manual_uid,
              f'Akun Google {g_user["username"]} digabung ke {m_user["username"]}', app_slug='portal')
    flash(f'Akun Google {g_user["username"]} berhasil digabung ke {m_user["username"]}.', 'success')
    return redirect(url_for('portal_users'))

@app.route('/portal/users/add', methods=['GET', 'POST'])
@login_required
def portal_user_add():
    if not is_portal_admin():
        flash('Akses ditolak.', 'danger')
        return redirect(url_for('portal'))
    db   = get_db()
    apps = db.execute('SELECT slug, name, icon FROM superapp_apps WHERE is_active=1 ORDER BY sort_order').fetchall()
    all_roles    = db.execute("SELECT name, description, app_slug FROM roles ORDER BY is_system DESC, name").fetchall()
    global_roles = [r for r in all_roles if r['app_slug'] == '']
    roles_by_app = {}
    for a in apps:
        specific = [r for r in all_roles if r['app_slug'] == a['slug']]
        roles_by_app[a['slug']] = global_roles + specific

    if request.method == 'POST':
        username  = request.form.get('username', '').strip()
        full_name = request.form.get('full_name', '').strip()
        password  = request.form.get('password', '')
        # role portal: hanya set jika access_portal dicentang, default viewer
        role      = request.form.get('role', 'viewer') if request.form.get('access_portal') else 'viewer'
        email     = request.form.get('email', '').strip()
        phone     = request.form.get('phone', '').strip()
        telegram  = request.form.get('telegram_id', '').strip()
        emp_id    = request.form.get('emp_id', type=int)
        if not username or not password:
            flash('Username dan password wajib diisi', 'danger')
        elif email and db.execute("SELECT id FROM users WHERE LOWER(email)=? AND is_active=1", (email.lower(),)).fetchone():
            flash(f'Email {email} sudah digunakan oleh user lain.', 'danger')
        else:
            try:
                cur = db.execute(
                    'INSERT INTO users(username,password_hash,full_name,role,email,phone,telegram_id) VALUES(?,?,?,?,?,?,?)',
                    (username, generate_password_hash(password, method='pbkdf2:sha256'),
                     full_name, role, email, phone, telegram))
                new_uid = cur.lastrowid
                if emp_id:
                    db.execute('UPDATE employees SET user_id=? WHERE id=? AND user_id IS NULL', (new_uid, emp_id))
                for a in apps:
                    if request.form.get(f'access_{a["slug"]}'):
                        app_role = request.form.get(f'role_{a["slug"]}', 'admin')
                        db.execute('''INSERT INTO user_app_access(user_id,app_slug,app_role,is_active) VALUES(?,?,?,1)
                                      ON CONFLICT(user_id,app_slug) DO UPDATE SET app_role=excluded.app_role,is_active=1''',
                                   (new_uid, a['slug'], app_role))
                db.commit()
                flash(f'User {username} berhasil dibuat', 'success')
                return redirect(url_for('portal_users'))
            except Exception as e:
                db.rollback()
                if _is_db_integrity_error(e):
                    flash('Username sudah digunakan', 'danger')
                else:
                    raise
    free_emps = db.execute(
        'SELECT id,name,jabatan,divisi FROM employees WHERE is_active=1 AND user_id IS NULL ORDER BY name'
    ).fetchall()
    return render_template('portal_user_form.html', user=None, apps=apps, free_emps=free_emps,
                           global_roles=global_roles, roles_by_app=roles_by_app, user_access={})

@app.route('/portal/users/<int:uid>/edit', methods=['GET', 'POST'])
@login_required
def portal_user_edit(uid):
    if not is_portal_admin():
        flash('Akses ditolak.', 'danger')
        return redirect(url_for('portal'))
    db   = get_db()
    user = db.execute('SELECT * FROM users WHERE id=?', (uid,)).fetchone()
    if not user:
        flash('User tidak ditemukan', 'danger')
        return redirect(url_for('portal_users'))
    apps = db.execute('SELECT slug, name, icon FROM superapp_apps WHERE is_active=1 ORDER BY sort_order').fetchall()
    all_roles    = db.execute("SELECT name, description, app_slug FROM roles ORDER BY is_system DESC, name").fetchall()
    global_roles = [r for r in all_roles if r['app_slug'] == '']
    roles_by_app = {}
    for a in apps:
        specific = [r for r in all_roles if r['app_slug'] == a['slug']]
        roles_by_app[a['slug']] = global_roles + specific

    if request.method == 'POST':
        full_name = request.form.get('full_name', '').strip()
        role      = request.form.get('role', 'viewer') if request.form.get('access_portal') else 'viewer'
        is_active = 1 if request.form.get('is_active') else 0
        email     = request.form.get('email', '').strip()
        phone     = request.form.get('phone', '').strip()
        telegram  = request.form.get('telegram_id', '').strip()
        new_pass  = request.form.get('password', '').strip()
        if email:
            dup = db.execute("SELECT id FROM users WHERE LOWER(email)=? AND is_active=1 AND id!=?",
                             (email.lower(), uid)).fetchone()
            if dup:
                flash(f'Email {email} sudah digunakan oleh user lain.', 'danger')
                linked_emp = db.execute('SELECT id,name,email,phone,telegram_id FROM employees WHERE user_id=? AND is_active=1', (uid,)).fetchone()
                user_access = {r['app_slug']: r['app_role'] for r in
                               db.execute('SELECT app_slug,app_role FROM user_app_access WHERE user_id=? AND is_active=1',(uid,)).fetchall()}
                return render_template('portal_user_form.html', user=user, apps=apps,
                                       linked_emp=linked_emp, free_emps=[],
                                       global_roles=global_roles, roles_by_app=roles_by_app,
                                       user_access=user_access)
        if new_pass:
            db.execute('UPDATE users SET full_name=?,role=?,is_active=?,email=?,phone=?,telegram_id=?,password_hash=? WHERE id=?',
                       (full_name, role, is_active, email, phone, telegram,
                        generate_password_hash(new_pass, method='pbkdf2:sha256'), uid))
        else:
            db.execute('UPDATE users SET full_name=?,role=?,is_active=?,email=?,phone=?,telegram_id=? WHERE id=?',
                       (full_name, role, is_active, email, phone, telegram, uid))
        # Update per-app access: hapus lama, insert baru
        db.execute('DELETE FROM user_app_access WHERE user_id=?', (uid,))
        for a in apps:
            if request.form.get(f'access_{a["slug"]}'):
                app_role = request.form.get(f'role_{a["slug"]}', 'admin')
                db.execute('INSERT INTO user_app_access(user_id,app_slug,app_role,is_active) VALUES(?,?,?,1)',
                           (uid, a['slug'], app_role))
        db.commit()
        if uid == session['user_id']:
            session['user_name'] = full_name or session['username']
            session['user_role'] = role
        flash('User diperbarui', 'success')
        return redirect(url_for('portal_users'))
    linked_emp  = db.execute(
        'SELECT id,name,email,phone,telegram_id FROM employees WHERE user_id=? AND is_active=1', (uid,)
    ).fetchone()
    user_access = {r['app_slug']: r['app_role'] for r in
                   db.execute('SELECT app_slug,app_role FROM user_app_access WHERE user_id=? AND is_active=1',(uid,)).fetchall()}
    return render_template('portal_user_form.html', user=user, apps=apps,
                           linked_emp=linked_emp, free_emps=[],
                           global_roles=global_roles, roles_by_app=roles_by_app,
                           user_access=user_access)

@app.route('/portal/users/<int:uid>/delete', methods=['POST'])
@login_required
def portal_user_delete(uid):
    if not is_portal_admin():
        return jsonify({'error': 'forbidden'}), 403
    if uid == session['user_id']:
        flash('Tidak bisa menonaktifkan akun sendiri', 'danger')
        return redirect(url_for('portal_users'))
    db = get_db()
    u = db.execute('SELECT username FROM users WHERE id=?', (uid,)).fetchone()
    db.execute('UPDATE users SET is_active=0 WHERE id=?', (uid,))
    db.commit()
    flash(f'User {u["username"] if u else uid} dinonaktifkan', 'warning')
    return redirect(url_for('portal_users'))

@app.route('/portal/users/<int:uid>/remove', methods=['POST'])
@login_required
def portal_user_remove(uid):
    """Hapus permanen — hanya superadmin."""
    if session.get('user_role') != 'superadmin':
        flash('Hanya Superadmin yang bisa menghapus user secara permanen.', 'danger')
        return redirect(url_for('portal_users'))
    if uid == session['user_id']:
        flash('Tidak bisa menghapus akun sendiri.', 'danger')
        return redirect(url_for('portal_users'))
    db = get_db()
    u = db.execute('SELECT * FROM users WHERE id=?', (uid,)).fetchone()
    if not u:
        flash('User tidak ditemukan.', 'danger')
        return redirect(url_for('portal_users'))
    if u['role'] == 'superadmin':
        # Pastikan masih ada superadmin lain
        cnt = db.execute("SELECT COUNT(*) FROM users WHERE role='superadmin' AND is_active=1 AND id!=?",
                         (uid,)).fetchone()[0]
        if cnt == 0:
            flash('Tidak bisa menghapus satu-satunya Superadmin.', 'danger')
            return redirect(url_for('portal_users'))
    # Cleanup data terkait
    db.execute('DELETE FROM user_app_access WHERE user_id=?', (uid,))
    db.execute('UPDATE employees SET user_id=NULL WHERE user_id=?', (uid,))
    db.execute('DELETE FROM users WHERE id=?', (uid,))
    db.commit()
    audit_log('remove_user', 'users', uid,
              f'User {u["username"]} ({u["email"] or "-"}) dihapus permanen', app_slug='portal')
    flash(f'User {u["username"]} berhasil dihapus secara permanen.', 'success')
    return redirect(url_for('portal_users'))

@app.route('/portal/users/<int:uid>/send-reset', methods=['POST'])
@login_required
def portal_user_send_reset(uid):
    if not is_portal_admin():
        return redirect(url_for('portal'))
    # Delegasi ke fungsi existing
    from flask import current_app
    db   = get_db()
    user = db.execute('SELECT * FROM users WHERE id=?', (uid,)).fetchone()
    if user:
        settings = get_settings(db)
        token    = secrets.token_urlsafe(48)
        expires  = (datetime.now() + timedelta(hours=1)).strftime('%Y-%m-%d %H:%M:%S')
        db.execute('DELETE FROM password_reset_tokens WHERE user_id=?', (user['id'],))
        db.execute('INSERT INTO password_reset_tokens(user_id, token, expires_at) VALUES(?,?,?)',
                   (user['id'], token, expires))
        db.commit()
        reset_link = f"{get_base_url(settings)}/reset-password/{token}"
        sent = _send_reset_notifications(dict(user), reset_link, settings, db)
        if sent:
            flash(f'Link reset dikirim via: {", ".join(sent)}.', 'success')
        else:
            flash(f'Tidak ada kontak — salin link manual: {reset_link}', 'warning')
    return redirect(url_for('portal_users'))

@app.route('/portal/users/<int:uid>/mfa-reset', methods=['POST'])
@login_required
def portal_user_mfa_reset(uid):
    if not is_portal_admin():
        return redirect(url_for('portal'))
    db = get_db()
    db.execute('UPDATE users SET totp_secret="", mfa_enabled=0 WHERE id=?', (uid,))
    db.commit()
    flash('MFA user berhasil direset', 'warning')
    return redirect(url_for('portal_users'))

# ─── Portal: Role & Permission ─────────────────────────────────────────────────

@app.route('/portal/roles', methods=['GET', 'POST'])
@login_required
def portal_roles():
    if not is_portal_admin():
        flash('Akses ditolak.', 'danger')
        return redirect(url_for('portal'))
    db         = get_db()
    active_app = request.args.get('app', 'evaluasi')
    superadmin = is_superadmin()

    if request.method == 'POST':
        action   = request.form.get('action')
        app_slug = request.form.get('app_slug', 'evaluasi')
        if action == 'add':
            if not superadmin:
                flash('Hanya superadmin yang dapat menambah role.', 'danger')
            else:
                name = request.form.get('name', '').strip().lower().replace(' ', '_')
                desc = request.form.get('description', '').strip()
                if name:
                    try:
                        db.execute('INSERT INTO roles(name,description,is_system,app_slug) VALUES(?,?,0,?)',
                                   (name, desc, app_slug))
                        db.commit()
                        flash(f'Role "{name}" ditambahkan', 'success')
                    except Exception:
                        flash('Nama role sudah ada', 'danger')
        elif action == 'delete':
            if not superadmin:
                flash('Hanya superadmin yang dapat menghapus role.', 'danger')
            else:
                rname  = request.form.get('role_name', '')
                is_sys = db.execute('SELECT is_system FROM roles WHERE name=?', (rname,)).fetchone()
                if is_sys and is_sys['is_system']:
                    flash('Role sistem tidak bisa dihapus', 'danger')
                else:
                    db.execute('DELETE FROM roles WHERE name=?', (rname,))
                    db.execute('DELETE FROM role_permissions WHERE role_name=?', (rname,))
                    db.commit()
                    flash(f'Role "{rname}" dihapus', 'warning')
        elif action == 'save_perms':
            rname     = request.form.get('role_name', '')
            
            # 1. Save menu assignments (app-scoped)
            db.execute('''
                DELETE FROM role_menus 
                WHERE role_name=? AND menu_id IN (SELECT id FROM app_menus WHERE app_slug=?)
            ''', (rname, app_slug))
            
            selected_menus = request.form.getlist('menus')
            for mid in selected_menus:
                if mid.isdigit():
                    db.execute('INSERT OR IGNORE INTO role_menus(role_name,menu_id) VALUES(?,?)', (rname, int(mid)))
            
            # 2. Derive permissions automatically from assigned menus (app-scoped)
            app_perms = APP_PERMISSIONS.get(app_slug, {})
            for perm in app_perms:
                db.execute('DELETE FROM role_permissions WHERE role_name=? AND permission=?', (rname, perm))
            
            if app_slug == 'portal':
                db.execute('DELETE FROM role_permissions WHERE role_name=? AND permission IN (?,?)', 
                           (rname, 'manage_users', 'manage_roles'))

            if rname == 'superadmin':
                for perm in ALL_PERMISSIONS:
                    db.execute('INSERT OR IGNORE INTO role_permissions(role_name,permission) VALUES(?,?)', (rname, perm))
            else:
                rows = db.execute('''
                    SELECT DISTINCT m.required_permission FROM app_menus m
                    JOIN role_menus rm ON m.id = rm.menu_id
                    WHERE rm.role_name=? AND m.required_permission != '' AND m.app_slug=?
                ''', (rname, app_slug)).fetchall()
                for row in rows:
                    perm = row['required_permission']
                    if superadmin or perm not in CRITICAL_PERMISSIONS:
                        db.execute('INSERT OR IGNORE INTO role_permissions(role_name,permission) VALUES(?,?)',
                                   (rname, perm))
            
            db.commit()
            flash(f'Menu dan Hak Akses role "{rname}" diperbarui', 'success')
        return redirect(url_for('portal_roles', app=active_app))

    apps_list     = db.execute('SELECT slug, name FROM superapp_apps WHERE is_active=1 ORDER BY sort_order').fetchall()
    roles         = db.execute("SELECT * FROM roles WHERE app_slug=? OR app_slug='' ORDER BY is_system DESC, name",
                               (active_app,)).fetchall()
    perms_by_role = {r['name']: get_role_permissions(db, r['name']) for r in roles}
    app_perms     = APP_PERMISSIONS.get(active_app, {})
    
    # Load menus of active app hierarchically
    all_app_menus = db.execute('SELECT * FROM app_menus WHERE app_slug=? ORDER BY sort_order, id', (active_app,)).fetchall()
    app_menu_tree = []
    children_by_parent = {}
    for row in all_app_menus:
        m = dict(row)
        if m['parent_id'] is None:
            app_menu_tree.append(m)
        else:
            children_by_parent.setdefault(m['parent_id'], []).append(m)
    for p in app_menu_tree:
        p['children'] = children_by_parent.get(p['id'], [])

    assigned_menus_by_role = {}
    for role in roles:
        rows = db.execute('SELECT menu_id FROM role_menus WHERE role_name=?', (role['name'],)).fetchall()
        assigned_menus_by_role[role['name']] = {r['menu_id'] for r in rows}

    return render_template('portal_roles.html', roles=roles, perms_by_role=perms_by_role,
                           app_perms=app_perms, apps_list=apps_list, active_app=active_app,
                           critical_permissions=CRITICAL_PERMISSIONS, is_superadmin=superadmin,
                           app_menu_tree=app_menu_tree, assigned_menus_by_role=assigned_menus_by_role)

PORTAL_SYSTEM_KEYS = [
    'app_url',
    'smtp_host', 'smtp_port', 'smtp_user', 'smtp_password', 'smtp_from', 'smtp_ssl',
    'telegram_bot_token', 'telegram_default_chat_id',
    'openwa_url', 'openwa_api_key', 'openwa_session_id', 'openwa_enabled',
    'openwa_session_evaluasi', 'openwa_session_support', 'openwa_session_booking', 'openwa_session_aset',
    'google_client_id', 'google_client_secret', 'google_workspace_domain', 'google_oauth_enabled',
    'recaptcha_site_key', 'recaptcha_secret_key', 'recaptcha_enabled',
    'chatbot_enabled',
    'ai_provider', 'ai_api_key', 'ai_model', 'ai_base_url',
]

@app.route('/portal/system-settings', methods=['GET', 'POST'])
@login_required
def portal_system_settings():
    if not is_portal_admin():
        flash('Akses ditolak.', 'danger')
        return redirect(url_for('portal'))
    db = get_db()
    if request.method == 'POST':
        for k in PORTAL_SYSTEM_KEYS:
            if k in ('smtp_ssl', 'openwa_enabled', 'google_oauth_enabled', 'recaptcha_enabled', 'chatbot_enabled'):
                v = '1' if request.form.get(k) else '0'
            else:
                v = request.form.get(k, '').strip()
            save_setting(db, k, v)
        db.commit()
        flash('Pengaturan sistem disimpan', 'success')
        return redirect(url_for('portal_system_settings'))
    cfg = get_settings(db)
    return render_template('portal_system_settings.html', cfg=cfg)

@app.route('/portal/system-settings/test-email', methods=['POST'])
@login_required
def portal_test_email():
    if not is_portal_admin():
        return jsonify({'ok': False, 'msg': 'Akses ditolak'})
    db = get_db()
    cfg = get_settings(db)
    to_email = request.form.get('test_email', '').strip()
    if not to_email:
        return jsonify({'ok': False, 'msg': 'Masukkan alamat email tujuan test'})
    ok, err = send_email(cfg, to_email, 'Test Email — Hive',
                         '<h3>Test berhasil!</h3><p>Konfigurasi email sudah benar.</p>')
    return jsonify({'ok': ok, 'msg': 'Email berhasil dikirim' if ok else str(err)})

@app.route('/portal/system-settings/test-telegram', methods=['POST'])
@login_required
def portal_test_telegram():
    if not is_portal_admin():
        return jsonify({'ok': False, 'msg': 'Akses ditolak'})
    db = get_db()
    cfg = get_settings(db)
    bot_token = cfg.get('telegram_bot_token', '').strip()
    chat_id   = request.form.get('test_chat_id', '').strip() or cfg.get('telegram_default_chat_id', '').strip()
    if not bot_token or not chat_id:
        return jsonify({'ok': False, 'msg': 'Bot token dan chat ID harus diisi'})
    ok, err = send_telegram(bot_token, chat_id,
                            '✅ <b>Test berhasil!</b>\n\nKonfigurasi Telegram sudah benar.')
    return jsonify({'ok': ok, 'msg': 'Pesan Telegram berhasil dikirim' if ok else str(err)})

@app.route('/portal/system-settings/test-whatsapp', methods=['POST'])
@login_required
def portal_test_whatsapp():
    if not is_portal_admin():
        return jsonify({'ok': False, 'msg': 'Akses ditolak'})
    db  = get_db()
    cfg = get_settings(db)
    wa_url     = cfg.get('openwa_url', '').strip()
    wa_key     = cfg.get('openwa_api_key', '').strip()
    test_app   = request.form.get('app_slug', '').strip()
    wa_session = get_openwa_session(cfg, test_app) if test_app else get_openwa_session(cfg)
    phone      = request.form.get('test_wa_phone', '').strip()
    if not wa_url or not phone:
        return jsonify({'ok': False, 'msg': 'URL OpenWA dan nomor HP harus diisi'})
    app_label  = {'evaluasi': 'TalentCore', 'support': 'SupportCore',
                  'booking': 'BookingCore', 'aset': 'AssetCore'}.get(test_app, 'Hive')
    ok, err = send_whatsapp(wa_url, wa_key, wa_session, phone,
                            f'✅ *Test berhasil!*\n\nSesi: `{wa_session}`\nAplikasi: {app_label}')
    chat_id = normalize_phone_wa(phone)
    return jsonify({'ok': ok, 'chat_id': chat_id,
                    'msg': f'Pesan terkirim ke {chat_id}' if ok else str(err)})

@app.route('/portal/system-settings/save-ai', methods=['POST'])
@login_required
def portal_save_ai_settings():
    if not is_portal_admin():
        return jsonify({'ok': False, 'msg': 'Akses ditolak'})
    try:
        db = get_db()
        AI_KEYS = ['chatbot_enabled', 'ai_provider', 'ai_api_key', 'ai_model', 'ai_base_url']
        saved = {}
        for k in AI_KEYS:
            if k == 'chatbot_enabled':
                v = '1' if request.form.get(k) else '0'
            else:
                v = request.form.get(k, '').strip()
            save_setting(db, k, v)
            saved[k] = v
        db.commit()
        # Verify: baca kembali dari DB
        verify = get_settings(db)
        return jsonify({
            'ok': True,
            'msg': 'Konfigurasi AI Assistant berhasil disimpan.',
            'saved': saved,
            'verified_enabled': verify.get('chatbot_enabled','?'),
            'verified_key_len': len(verify.get('ai_api_key','')),
            'verified_model': verify.get('ai_model',''),
        })
    except Exception as ex:
        return jsonify({'ok': False, 'msg': f'Error saat menyimpan: {str(ex)}'}), 500

@app.route('/portal/system-settings/diag-ai', methods=['GET'])
@login_required
def portal_diag_ai():
    """Diagnostik AI settings — superadmin only."""
    if not is_portal_admin():
        return jsonify({'error': 'Akses ditolak'}), 403
    try:
        db = get_db()
        cfg = get_settings(db)
        api_key = cfg.get('ai_api_key', '')
        provider = cfg.get('ai_provider', 'gemini')
        # Cek openai package
        try:
            import openai as _oai
            oai_version = getattr(_oai, '__version__', 'ok')
        except ImportError:
            oai_version = 'NOT INSTALLED'
        # Cek anthropic package
        try:
            import anthropic as _ant
            ant_version = getattr(_ant, '__version__', 'ok')
        except ImportError:
            ant_version = 'NOT INSTALLED'
        # Default model per provider
        _default_models = {
            'gemini': 'gemini-2.0-flash', 'openai': 'gpt-4o',
            'claude': 'claude-sonnet-4-6', 'ollama': 'llama3', 'openwebui': 'llama3',
        }
        default_model = _default_models.get(provider, 'gemini-2.0-flash')
        return jsonify({
            'chatbot_enabled':    cfg.get('chatbot_enabled', '0'),
            'ai_provider':        provider,
            'ai_model':           cfg.get('ai_model', f'(kosong → default {default_model})'),
            'ai_api_key_length':  len(api_key),
            'ai_api_key_prefix':  api_key[:7] + '...' if len(api_key) > 7 else '(kosong)',
            'openai_sdk':         oai_version,
            'anthropic_sdk':      ant_version,
            'app_version':        VERSION,
        })
    except Exception as ex:
        return jsonify({'error': str(ex)}), 500

@app.route('/portal/system-settings/reload', methods=['POST'])
@login_required
def portal_reload_app():
    if not is_portal_admin():
        return jsonify({'ok': False, 'msg': 'Akses ditolak'})
    try:
        # Cari PID master gunicorn (parent dari proses ini)
        ppid = os.getppid()
        os.kill(ppid, signal.SIGHUP)
        return jsonify({'ok': True, 'msg': f'Sinyal reload dikirim ke gunicorn master (PID {ppid}). Worker akan restart dalam beberapa detik.'})
    except Exception as ex:
        return jsonify({'ok': False, 'msg': str(ex)})

@app.route('/portal/system-settings/version-info', methods=['GET'])
@login_required
def portal_version_info():
    if not is_portal_admin():
        return jsonify({'ok': False})
    info = {'deployed': VERSION, 'release_date': RELEASE_DATE}
    # Coba baca dari .git_info (ditulis oleh deploy script — lebih andal dari git commands di server)
    git_info_path = os.path.join(os.path.dirname(__file__), '.git_info')
    try:
        lines = open(git_info_path).read().splitlines()
        info.update({'git_hash': lines[0] if len(lines)>0 else '-',
                     'git_msg':  lines[1] if len(lines)>1 else '-',
                     'git_date': lines[2] if len(lines)>2 else '-'})
    except Exception:
        # Fallback: coba git command langsung (untuk dev environment)
        try:
            git_hash = subprocess.check_output(['git', 'rev-parse', '--short', 'HEAD'],
                                               cwd=os.path.dirname(__file__), stderr=subprocess.DEVNULL,
                                               timeout=3).decode().strip()
            git_msg  = subprocess.check_output(['git', 'log', '-1', '--pretty=%s'],
                                               cwd=os.path.dirname(__file__), stderr=subprocess.DEVNULL,
                                               timeout=3).decode().strip()
            git_date = subprocess.check_output(['git', 'log', '-1', '--pretty=%ci'],
                                               cwd=os.path.dirname(__file__), stderr=subprocess.DEVNULL,
                                               timeout=3).decode().strip()
            info.update({'git_hash': git_hash, 'git_msg': git_msg, 'git_date': git_date})
        except Exception:
            info.update({'git_hash': '-', 'git_msg': '-', 'git_date': '-'})
    return jsonify(info)

@app.route('/')
@login_required
def index():
    db = get_db()
    employees = db.execute('''
        SELECT emp.*,
               (SELECT COUNT(*) FROM evaluations WHERE employee_id=emp.id) AS eval_count,
               (SELECT MAX(periode) FROM evaluations WHERE employee_id=emp.id) AS latest_periode,
               (SELECT final_total FROM evaluations WHERE employee_id=emp.id ORDER BY id DESC LIMIT 1) AS latest_score,
               (SELECT status FROM evaluations WHERE employee_id=emp.id ORDER BY id DESC LIMIT 1) AS latest_status,
               (SELECT id FROM evaluations WHERE employee_id=emp.id ORDER BY id DESC LIMIT 1) AS latest_eval_id
        FROM employees emp WHERE emp.is_active=1
        ORDER BY emp.divisi, emp.name
    ''').fetchall()

    today = date.today()
    contracts_alert = db.execute('''
        SELECT *, julianday(contract_end) - julianday('now') AS days_left
        FROM employees WHERE employment_type='kontrak' AND contract_end != ''
        AND is_active=1 ORDER BY contract_end
    ''').fetchall()
    expiring_soon = [e for e in contracts_alert if e['days_left'] is not None and e['days_left'] <= 30]

    divisi_list  = get_divisi_list(db)
    divisi_stats = {}
    for d in divisi_list:
        row = db.execute('''
            SELECT COUNT(DISTINCT emp.id) AS total,
                   COUNT(DISTINCT CASE WHEN ev.status='final' THEN ev.employee_id END) AS done
            FROM employees emp LEFT JOIN evaluations ev ON ev.employee_id=emp.id AND ev.status='final'
            WHERE emp.divisi=? AND emp.is_active=1
        ''', (d,)).fetchone()
        divisi_stats[d] = row
    return render_template('index.html', employees=employees, divisi_stats=divisi_stats,
                           expiring_soon=expiring_soon)

# ─── Employee CRUD ──────────────────────────────────────────────────────────────

@app.route('/emp/add', methods=['GET', 'POST'])
@login_required
def emp_add():
    if request.method == 'POST':
        name  = request.form['name'].strip()
        divisi = request.form['divisi']
        if not name:
            flash('Nama tidak boleh kosong', 'danger')
            db = get_db()
            employees_all = db.execute('SELECT id,name,jabatan,divisi,level FROM employees WHERE is_active=1 ORDER BY name').fetchall()
            return render_template('employee_form.html', emp=None, employees_all=employees_all)
        db = get_db()
        def _int_or_none(v):
            return int(v) if v and str(v).isdigit() else None
        emp_type = request.form.get('employment_type','tetap')
        rate_md_raw = request.form.get('rate_mandays','').strip()
        rate_md = float(rate_md_raw) if rate_md_raw and emp_type == 'staff_worker' else None
        db.execute('''INSERT INTO employees(name,jabatan,divisi,level,employment_type,
                      contract_start,contract_end,rate_mandays,email,phone,telegram_id,notes,
                      supervisor_id,leader_id,manager_id)
                      VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)''', (
            name,
            request.form.get('jabatan','').strip(),
            divisi,
            request.form.get('level','Staff'),
            emp_type,
            request.form.get('contract_start','') if emp_type in ('kontrak','staff_worker') else '',
            request.form.get('contract_end','')   if emp_type in ('kontrak','staff_worker') else '',
            rate_md,
            request.form.get('email','').strip(),
            request.form.get('phone','').strip(),
            normalize_telegram_id(request.form.get('telegram_id','')),
            request.form.get('notes','').strip(),
            _int_or_none(request.form.get('supervisor_id','')),
            _int_or_none(request.form.get('leader_id','')),
            _int_or_none(request.form.get('manager_id','')),
        ))
        db.commit()
        flash(f'Karyawan {name} berhasil ditambahkan', 'success')
        return redirect(url_for('index'))
    db = get_db()
    employees_all = db.execute('SELECT id,name,jabatan,divisi,level FROM employees WHERE is_active=1 ORDER BY name').fetchall()
    return render_template('employee_form.html', emp=None, employees_all=employees_all)

@app.route('/emp/<int:emp_id>/edit', methods=['GET', 'POST'])
@login_required
def emp_edit(emp_id):
    db = get_db()
    emp = db.execute('SELECT * FROM employees WHERE id=?', (emp_id,)).fetchone()
    if not emp:
        flash('Karyawan tidak ditemukan', 'danger')
        return redirect(url_for('index'))
    if request.method == 'POST':
        action = request.form.get('action', 'save')
        if action == 'promote':
            # Role promotion only for superadmin
            if session.get('user_role') != 'superadmin':
                flash('Hanya Superadmin yang bisa promote role', 'danger')
                return redirect(url_for('emp_edit', emp_id=emp_id))
            new_level = request.form.get('promote_level', emp['level'])
            new_jabatan = request.form.get('promote_jabatan', emp['jabatan'])
            db.execute('UPDATE employees SET level=?,jabatan=? WHERE id=?',
                       (new_level, new_jabatan, emp_id))
            db.commit()
            flash(f'{emp["name"]} dipromosikan ke {new_level} — {new_jabatan}', 'success')
            return redirect(url_for('emp_edit', emp_id=emp_id))
        # Normal save
        def _int_or_none(v):
            return int(v) if v and str(v).isdigit() else None
        emp_type = request.form.get('employment_type', 'tetap')
        rate_md_raw = request.form.get('rate_mandays','').strip()
        rate_md = float(rate_md_raw) if rate_md_raw and emp_type == 'staff_worker' else None
        db.execute('''UPDATE employees SET name=?,jabatan=?,divisi=?,level=?,
                      employment_type=?,contract_start=?,contract_end=?,rate_mandays=?,
                      email=?,phone=?,telegram_id=?,notes=?,
                      supervisor_id=?,leader_id=?,manager_id=? WHERE id=?''', (
            request.form['name'].strip(),
            request.form.get('jabatan','').strip(),
            request.form['divisi'],
            request.form.get('level','Staff'),
            emp_type,
            request.form.get('contract_start','') if emp_type in ('kontrak','staff_worker') else '',
            request.form.get('contract_end','')   if emp_type in ('kontrak','staff_worker') else '',
            rate_md,
            request.form.get('email','').strip(),
            request.form.get('phone','').strip(),
            normalize_telegram_id(request.form.get('telegram_id','')),
            request.form.get('notes','').strip(),
            _int_or_none(request.form.get('supervisor_id','')),
            _int_or_none(request.form.get('leader_id','')),
            _int_or_none(request.form.get('manager_id','')),
            emp_id,
        ))
        db.commit()
        flash('Data karyawan diperbarui', 'success')
        return redirect(url_for('emp_edit', emp_id=emp_id))
    # Reminder logs for this employee
    logs = db.execute('''SELECT * FROM reminder_logs WHERE employee_id=?
                         ORDER BY created_at DESC LIMIT 10''', (emp_id,)).fetchall()
    employees_all = db.execute('SELECT id,name,jabatan,divisi,level FROM employees WHERE is_active=1 AND id!=? ORDER BY name', (emp_id,)).fetchall()
    linked_user = db.execute('SELECT id,username,role,is_active FROM users WHERE id=?', (emp['user_id'],)).fetchone() if emp['user_id'] else None
    apps = db.execute('SELECT slug, name, icon FROM superapp_apps WHERE is_active=1 ORDER BY sort_order').fetchall()
    all_roles    = db.execute("SELECT name, description, app_slug FROM roles ORDER BY is_system DESC, name").fetchall()
    global_roles = [r for r in all_roles if r['app_slug'] == '']
    roles_by_app = {}
    for a in apps:
        specific = [r for r in all_roles if r['app_slug'] == a['slug']]
        roles_by_app[a['slug']] = global_roles + specific
    user_access = {}
    if emp['user_id']:
        user_access = {r['app_slug']: r['app_role'] for r in
                       db.execute('SELECT app_slug,app_role FROM user_app_access WHERE user_id=? AND is_active=1', (emp['user_id'],)).fetchall()}
    return render_template('employee_form.html', emp=emp, logs=logs,
                           employees_all=employees_all, linked_user=linked_user,
                           apps=apps, global_roles=global_roles,
                           roles_by_app=roles_by_app, user_access=user_access)

@app.route('/emp/<int:emp_id>/delete', methods=['POST'])
@superadmin_required
def emp_delete(emp_id):
    db = get_db()
    emp = db.execute('SELECT name, user_id FROM employees WHERE id=?', (emp_id,)).fetchone()
    if emp:
        db.execute('UPDATE employees SET is_active=0 WHERE id=?', (emp_id,))
        if emp['user_id']:
            db.execute('UPDATE users SET is_active=0 WHERE id=?', (emp['user_id'],))
        db.commit()
        flash(f'Karyawan {emp["name"]} dinonaktifkan', 'warning')
    return redirect(url_for('index'))

@app.route('/emp/bulk-deactivate', methods=['POST'])
@superadmin_required
def emp_bulk_deactivate():
    db = get_db()
    ids_str = request.form.get('ids', '')
    if not ids_str:
        flash('Tidak ada karyawan yang dipilih.', 'warning')
        return redirect(url_for('karyawan'))
        
    try:
        emp_ids = [int(x.strip()) for x in ids_str.split(',') if x.strip()]
    except ValueError:
        flash('ID karyawan tidak valid.', 'danger')
        return redirect(url_for('karyawan'))
        
    if not emp_ids:
        flash('Tidak ada karyawan yang dipilih.', 'warning')
        return redirect(url_for('karyawan'))
        
    # Get names of deactivated employees and user IDs
    placeholders = ','.join('?' for _ in emp_ids)
    rows = db.execute(f'SELECT name, user_id FROM employees WHERE id IN ({placeholders})', emp_ids).fetchall()
    
    names = [r['name'] for r in rows]
    user_ids = [r['user_id'] for r in rows if r['user_id']]
    
    # Update employees
    db.execute(f'UPDATE employees SET is_active=0 WHERE id IN ({placeholders})', emp_ids)
    
    # Update linked users
    if user_ids:
        user_placeholders = ','.join('?' for _ in user_ids)
        db.execute(f'UPDATE users SET is_active=0 WHERE id IN ({user_placeholders})', user_ids)
        
    db.commit()
    
    flash(f'Berhasil menonaktifkan {len(names)} karyawan: {", ".join(names)}', 'warning')
    return redirect(url_for('karyawan'))

# ─── Import Excel: Karyawan & Gaji ────────────────────────────────────────────

@app.route('/emp/import/template')
@permission_required('manage_employees')
def emp_import_template():
    """Download template Excel untuk import karyawan."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.worksheet.datavalidation import DataValidation

    wb = Workbook()
    ws = wb.active
    ws.title = 'Data Karyawan'

    headers = [
        ('name',            'Nama Lengkap *',         30, 'Wajib diisi'),
        ('jabatan',         'Jabatan',                 25, 'Contoh: Staff IT, Manager HRD'),
        ('divisi',          'Divisi *',                20, 'Wajib. Harus sesuai divisi di sistem'),
        ('level',           'Level',                   15, 'Staff / Leader / Manager / Director'),
        ('employment_type', 'Tipe *',                  12, 'tetap / kontrak'),
        ('contract_start',  'Mulai Kontrak',           15, 'Format: YYYY-MM-DD, kosongkan jika tetap'),
        ('contract_end',    'Akhir Kontrak',           15, 'Format: YYYY-MM-DD, kosongkan jika tetap'),
        ('email',           'Email',                   28, 'Alamat email karyawan'),
        ('phone',           'No. HP/WA',               18, 'Contoh: 628123456789'),
        ('telegram_id',     'Telegram ID',             18, 'Contoh: @username atau chat_id'),
        ('notes',           'Catatan',                 30, 'Catatan tambahan (opsional)'),
    ]

    hdr_fill = PatternFill('solid', fgColor='1F4E79')
    req_fill = PatternFill('solid', fgColor='2E7D32')
    tip_fill = PatternFill('solid', fgColor='F5F5F5')
    hdr_font = Font(color='FFFFFF', bold=True, size=10)
    tip_font = Font(color='555555', italic=True, size=9)
    thin = Side(style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Row 1: header; Row 2: tips; Row 3+: data
    for col_idx, (field, label, width, tip) in enumerate(headers, 1):
        fill = req_fill if '*' in label else hdr_fill
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = hdr_font; cell.fill = fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border

        tip_cell = ws.cell(row=2, column=col_idx, value=tip)
        tip_cell.font = tip_font; tip_cell.fill = tip_fill
        tip_cell.alignment = Alignment(wrap_text=True, vertical='top')
        tip_cell.border = border

        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width

    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 36
    ws.freeze_panes = 'A3'

    # Data validation: tipe
    dv_type = DataValidation(type='list', formula1='"tetap,kontrak"', showDropDown=False)
    ws.add_data_validation(dv_type)
    dv_type.sqref = 'E3:E1000'

    # Contoh baris data
    sample = ['Budi Santoso', 'Staff IT', 'IT', 'Staff', 'tetap', '', '', 'budi@company.com', '6281234567890', '', '']
    for col_idx, val in enumerate(sample, 1):
        c = ws.cell(row=3, column=col_idx, value=val)
        c.border = border

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return Response(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    headers={'Content-Disposition': 'attachment; filename=template_import_karyawan.xlsx'})


@app.route('/emp/import', methods=['POST'])
@permission_required('manage_employees')
def emp_import():
    """Import karyawan dari file Excel."""
    from openpyxl import load_workbook
    f = request.files.get('file')
    if not f or not f.filename.endswith('.xlsx'):
        flash('Upload file .xlsx yang valid', 'danger')
        return redirect(url_for('karyawan'))

    try:
        wb = load_workbook(f, read_only=True, data_only=True)
        ws = wb.active
        db = get_db()
        now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

        # Header row = 1, tip row = 2, data starts row 3
        FIELD_MAP = {
            'nama lengkap': 'name', 'jabatan': 'jabatan', 'divisi': 'divisi',
            'level': 'level', 'tipe': 'employment_type',
            'mulai kontrak': 'contract_start', 'akhir kontrak': 'contract_end',
            'email': 'email', 'no. hp/wa': 'phone', 'telegram id': 'telegram_id',
            'catatan': 'notes',
        }
        # Read header row to get column mapping
        hdr_row = [str(c.value or '').lower().strip().rstrip(' *') for c in ws[1]]
        col_map = {}
        for idx, hdr in enumerate(hdr_row):
            field = FIELD_MAP.get(hdr)
            if field:
                col_map[field] = idx

        inserted = updated = skipped = 0
        errors = []
        mode = request.form.get('mode', 'insert')  # insert | upsert

        for row_num, row in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
            if all(v is None or str(v).strip() == '' for v in row):
                continue
            def get(field):
                idx = col_map.get(field)
                v = row[idx] if idx is not None and idx < len(row) else None
                return str(v).strip() if v is not None else ''

            name   = get('name')
            divisi = get('divisi')
            if not name or not divisi:
                errors.append(f'Baris {row_num}: name/divisi wajib diisi')
                skipped += 1
                continue

            emp_type = get('employment_type') or 'tetap'
            if emp_type not in ('tetap', 'kontrak'):
                emp_type = 'tetap'

            fields = {
                'name': name, 'jabatan': get('jabatan'), 'divisi': divisi,
                'level': get('level') or 'Staff', 'employment_type': emp_type,
                'contract_start': get('contract_start'), 'contract_end': get('contract_end'),
                'email': get('email'), 'phone': get('phone'),
                'telegram_id': get('telegram_id'), 'notes': get('notes'),
                'is_active': 1,
            }

            existing = db.execute('SELECT id FROM employees WHERE name=? AND divisi=?',
                                  (name, divisi)).fetchone()
            if existing:
                if mode == 'upsert':
                    sets = ', '.join(f'{k}=?' for k in fields if k != 'name')
                    vals = [fields[k] for k in fields if k != 'name'] + [existing['id']]
                    db.execute(f'UPDATE employees SET {sets} WHERE id=?', vals)
                    updated += 1
                else:
                    skipped += 1
            else:
                cols = ', '.join(fields.keys())
                phs  = ', '.join('?' for _ in fields)
                db.execute(f'INSERT INTO employees ({cols}) VALUES ({phs})', list(fields.values()))
                inserted += 1

        db.commit()
        msg = f'Import selesai: {inserted} ditambah, {updated} diperbarui, {skipped} dilewati'
        if errors:
            msg += f'. {len(errors)} baris error (lihat log)'
        flash(msg, 'success' if not errors else 'warning')
    except Exception as e:
        flash(f'Gagal import: {e}', 'danger')

    return redirect(url_for('karyawan'))


@app.route('/salary/import/template')
@permission_required('manage_salary')
@mfa_challenge_required('mfa_salary_verified')
def salary_import_template():
    """Download template Excel untuk import data gaji."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = 'Data Gaji'

    db = get_db()
    year = date.today().year
    emps = db.execute(
        "SELECT id, name, jabatan, divisi, employment_type FROM employees WHERE is_active=1 ORDER BY employment_type DESC, divisi, name"
    ).fetchall()

    headers = [
        ('employee_id',   'ID Karyawan *', 12),
        ('name',          'Nama Karyawan', 30),
        ('jabatan',       'Jabatan',       22),
        ('divisi',        'Divisi',        15),
        ('year',          'Tahun *',       10),
        ('base_salary',   'SALARY (Gaji Pokok)', 20),
        ('al_001',        'AL_001 (Tj. Jabatan)', 20),
        ('al_002',        'AL_002 (Tj. Komunikasi)', 22),
        ('al_003',        'AL_003 (Tj. Performance)', 22),
        ('al_004',        'AL_004 (Tj. Kehadiran)', 20),
        ('increase_pct',  '% Kenaikan', 14),
        ('increase_date', 'Bulan Kenaikan', 16),
        ('notes',         'Catatan', 25),
    ]

    hdr_fill = PatternFill('solid', fgColor='1F4E79')
    req_fill = PatternFill('solid', fgColor='2E7D32')
    emp_fill = PatternFill('solid', fgColor='EFF7EF')
    num_fill = PatternFill('solid', fgColor='FAFEFF')
    hdr_font = Font(color='FFFFFF', bold=True, size=10)
    thin = Side(style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    from openpyxl.styles import numbers as xl_numbers

    for col_idx, (field, label, width) in enumerate(headers, 1):
        fill = req_fill if '*' in label else hdr_fill
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = hdr_font; cell.fill = fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width

    ws.row_dimensions[1].height = 32
    ws.freeze_panes = 'F2'

    from openpyxl.utils import get_column_letter
    for row_idx, emp in enumerate(emps, start=2):
        row_fill = PatternFill('solid', fgColor='FFF8E1' if emp['employment_type'] == 'kontrak' else 'F0F9FF')
        vals = [emp['id'], emp['name'], emp['jabatan'] or '', emp['divisi'], year,
                0, 0, 0, 0, 0, 0, '', '']
        for col_idx, val in enumerate(vals, 1):
            c = ws.cell(row=row_idx, column=col_idx, value=val)
            c.border = border
            # Read-only info columns
            if col_idx <= 4:
                c.fill = emp_fill
                if col_idx == 1:
                    c.font = Font(bold=True, color='1F4E79')
            elif col_idx == 5:
                c.fill = emp_fill
            elif 6 <= col_idx <= 10:
                c.fill = num_fill
                c.number_format = '#,##0'
            elif col_idx == 11:
                c.fill = num_fill
                c.number_format = '0.00'

    ws.row_dimensions[1].height = 30

    # Second sheet: instructions
    ws2 = wb.create_sheet('Petunjuk')
    instructions = [
        ['PETUNJUK IMPORT DATA GAJI'],
        [''],
        ['1. Jangan ubah kolom ID Karyawan — digunakan sebagai kunci pencarian'],
        ['2. Kolom Tahun wajib diisi dengan angka 4 digit (mis. 2024)'],
        ['3. Kolom gaji diisi angka tanpa titik/koma (mis. 5000000)'],
        ['4. % Kenaikan: angka desimal mis. 7.5 untuk 7.5%'],
        ['5. Bulan Kenaikan: format bebas mis. Apr-2025'],
        ['6. Data yang sudah ada di tahun tersebut akan DITIMPA (update)'],
        ['7. Baris dengan ID kosong akan dilewati'],
    ]
    for r, row in enumerate(instructions, 1):
        c = ws2.cell(row=r, column=1, value=row[0])
        if r == 1:
            c.font = Font(bold=True, size=13, color='1F4E79')
        ws2.column_dimensions['A'].width = 65

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return Response(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    headers={'Content-Disposition': 'attachment; filename=template_import_gaji.xlsx'})


@app.route('/salary/import', methods=['POST'])
@permission_required('manage_salary')
@mfa_challenge_required('mfa_salary_verified')
def salary_import():
    """Import data gaji dari file Excel."""
    from openpyxl import load_workbook
    f = request.files.get('file')
    if not f or not f.filename.endswith('.xlsx'):
        flash('Upload file .xlsx yang valid', 'danger')
        return redirect(url_for('salary_table'))

    try:
        wb   = load_workbook(f, read_only=True, data_only=True)
        ws   = wb.active
        db   = get_db()
        now  = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

        # Header row 1 → field map
        FIELD_MAP = {
            'id karyawan': 'employee_id', 'tahun': 'year',
            'salary (gaji pokok)': 'base_salary', 'al_001 (tj. jabatan)': 'al_001',
            'al_002 (tj. komunikasi)': 'al_002', 'al_003 (tj. performance)': 'al_003',
            'al_004 (tj. kehadiran)': 'al_004',
            '% kenaikan': 'increase_pct', 'bulan kenaikan': 'increase_date',
            'catatan': 'notes',
        }
        hdr_row = [str(c.value or '').lower().strip().rstrip(' *') for c in ws[1]]
        col_map = {field: idx for idx, hdr in enumerate(hdr_row)
                   for key, field in FIELD_MAP.items() if hdr == key}

        upserted = skipped = 0
        errors   = []

        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if all(v is None or str(v).strip() == '' for v in row):
                continue
            def get(field, default=''):
                idx = col_map.get(field)
                v = row[idx] if idx is not None and idx < len(row) else None
                return v if v is not None else default

            emp_id = get('employee_id')
            year   = get('year')
            try:
                emp_id = int(emp_id)
                year   = int(year)
            except (TypeError, ValueError):
                errors.append(f'Baris {row_num}: ID/tahun tidak valid')
                skipped += 1
                continue

            emp = db.execute('SELECT id FROM employees WHERE id=?', (emp_id,)).fetchone()
            if not emp:
                errors.append(f'Baris {row_num}: ID {emp_id} tidak ditemukan')
                skipped += 1
                continue

            def num(field):
                v = get(field, 0)
                try: return float(v) if v != '' else 0.0
                except (TypeError, ValueError): return 0.0

            db.execute('''
                INSERT INTO employee_salary
                    (employee_id, year, base_salary, al_001, al_002, al_003, al_004,
                     increase_pct, increase_date, notes, updated_at)
                VALUES (?,?,?,?,?,?,?,?,?,?,?)
                ON CONFLICT(employee_id, year) DO UPDATE SET
                    base_salary=excluded.base_salary, al_001=excluded.al_001,
                    al_002=excluded.al_002, al_003=excluded.al_003,
                    al_004=excluded.al_004, increase_pct=excluded.increase_pct,
                    increase_date=excluded.increase_date, notes=excluded.notes,
                    updated_at=excluded.updated_at
            ''', (emp_id, year,
                  _fenc(num('base_salary')), _fenc(num('al_001')), _fenc(num('al_002')),
                  _fenc(num('al_003')), _fenc(num('al_004')),
                  num('increase_pct'), str(get('increase_date', '')), str(get('notes', '')), now))
            upserted += 1

        db.commit()
        msg = f'Import gaji selesai: {upserted} baris disimpan, {skipped} dilewati'
        if errors:
            msg += f'. Error: ' + '; '.join(errors[:5])
            if len(errors) > 5:
                msg += f' (+{len(errors)-5} lainnya)'
        flash(msg, 'success' if not errors else 'warning')
    except Exception as e:
        flash(f'Gagal import gaji: {e}', 'danger')

    year_param = request.form.get('year_redirect', date.today().year)
    return redirect(url_for('salary_table', year=year_param))

# ─── Contract Management ───────────────────────────────────────────────────────

@app.route('/karyawan')
@login_required
def karyawan():
    db = get_db()
    today = date.today()
    _kontrak_raw = db.execute('''
        SELECT *, julianday(contract_end) - julianday('now') AS days_left
        FROM employees
        WHERE employment_type IN ('kontrak','staff_worker') AND is_active = 1 AND divisi != 'Telegram Core'
        ORDER BY CASE WHEN contract_end IS NULL OR contract_end='' THEN 1 ELSE 0 END,
                 contract_end ASC
    ''').fetchall()
    # days_left NULL (contract_end kosong) → 9999 agar selectattr di template tidak error
    kontrak = [{**dict(r), 'days_left': r['days_left'] if r['days_left'] is not None else 9999}
               for r in _kontrak_raw]
    tetap = db.execute('''
        SELECT * FROM employees
        WHERE employment_type = 'tetap' AND is_active = 1 AND divisi != 'Telegram Core'
        ORDER BY divisi, name
    ''').fetchall()
    telegram_core = db.execute('''
        SELECT * FROM employees
        WHERE divisi = 'Telegram Core' AND is_active = 1
        ORDER BY name
    ''').fetchall()
    emp_user_map = {r['id']: r for r in db.execute('''
        SELECT e.id, u.username, u.role, u.is_active AS u_active
        FROM employees e JOIN users u ON u.id = e.user_id
        WHERE e.is_active=1
    ''').fetchall()}
    apps = db.execute('SELECT slug, name, icon FROM superapp_apps WHERE is_active=1 ORDER BY sort_order').fetchall()
    all_roles    = db.execute("SELECT name, description, app_slug FROM roles ORDER BY is_system DESC, name").fetchall()
    global_roles = [r for r in all_roles if r['app_slug'] == '']
    roles_by_app = {}
    for a in apps:
        specific = [r for r in all_roles if r['app_slug'] == a['slug']]
        roles_by_app[a['slug']] = global_roles + specific
    # Evaluasi terbaru per karyawan (untuk tombol trigger review)
    current_year = str(date.today().year)
    eval_rows = db.execute(
        """SELECT e.employee_id, e.id AS eval_id, e.periode, e.review_status, e.status
           FROM evaluations e
           WHERE e.status NOT IN ('approved','closed')
             AND e.periode=?
           ORDER BY e.created_at DESC""", (current_year,)
    ).fetchall()
    # Map: emp_id -> evaluasi terbaru tahun ini
    eval_map = {}
    for row in eval_rows:
        if row['employee_id'] not in eval_map:
            eval_map[row['employee_id']] = dict(row)
    return render_template('karyawan.html', kontrak=kontrak, tetap=tetap, telegram_core=telegram_core, today=today,
                           emp_user_map=emp_user_map, global_roles=global_roles,
                           apps=apps, roles_by_app=roles_by_app,
                           eval_map=eval_map, current_year=current_year)

@app.route('/contracts')
@login_required
def contracts():
    return redirect(url_for('karyawan'))

@app.route('/contracts/remind/<int:emp_id>', methods=['POST'])
@login_required
def contract_remind_one(emp_id):
    db = get_db()
    emp = db.execute('SELECT * FROM employees WHERE id=?', (emp_id,)).fetchone()
    if not emp or not emp['contract_end']:
        flash('Data karyawan tidak valid', 'danger')
        return redirect(url_for('contracts'))
    settings = get_settings(db)
    try:
        end_date  = date.fromisoformat(emp['contract_end'])
        days_left = (end_date - date.today()).days
    except Exception:
        flash('Format tanggal kontrak tidak valid', 'danger')
        return redirect(url_for('contracts'))

    who = session.get('username', 'manual')
    s, f = _send_contract_notification(db, emp, days_left, settings, who)
    db.commit()
    if s == 0 and f == 0:
        flash('Tidak ada channel notifikasi yang dikonfigurasi (email / telegram / whatsapp)', 'warning')
    else:
        flash(f'Reminder dikirim — {s} berhasil, {f} gagal (termasuk ke Manajerial/Leader/Manager)', 'success' if f == 0 else 'warning')
    return redirect(url_for('karyawan'))

@app.route('/contracts/remind-all', methods=['POST'])
@login_required
def contract_remind_all():
    sent, failed = run_contract_reminders(triggered_by=session.get('username','manual'))
    flash(f'Selesai — {sent} berhasil, {failed} gagal', 'success' if failed == 0 else 'warning')
    return redirect(url_for('karyawan'))

@app.route('/reminders')
@login_required
def reminder_log():
    db   = get_db()
    logs = db.execute('''
        SELECT rl.*, emp.name AS emp_name, emp.divisi
        FROM reminder_logs rl
        LEFT JOIN employees emp ON emp.id = rl.employee_id
        ORDER BY rl.created_at DESC LIMIT 200
    ''').fetchall()
    stats = db.execute('''
        SELECT channel, status, COUNT(*) AS cnt
        FROM reminder_logs GROUP BY channel, status
    ''').fetchall()
    return render_template('reminder_log.html', logs=logs, stats=stats)

# ─── SupportCore ───────────────────────────────────────────────────────────────

def sc_require(perm):
    """Check SupportCore permission, redirect to /support if denied."""
    db = get_db()
    if not has_permission(session.get('user_role', ''), perm, db):
        flash(f'Akses ditolak — permission "{perm}" diperlukan', 'danger')
        return False
    return True

@app.route('/support/')
@login_required
def sc_index():
    db = get_db()
    counts = {
        'customers':    db.execute('SELECT COUNT(*) FROM sc_customers WHERE is_active=1').fetchone()[0],
        'services':     db.execute('SELECT COUNT(*) FROM sc_services WHERE is_active=1').fetchone()[0],
        'support_types':db.execute('SELECT COUNT(*) FROM sc_support_types WHERE is_active=1').fetchone()[0],
        'sla_categories':db.execute('SELECT COUNT(*) FROM sc_sla_categories WHERE is_active=1').fetchone()[0],
    }
    return render_template('sc_index.html', counts=counts)

# ── Master Customer ────────────────────────────────────────────────────────────
@app.route('/support/customers')
@login_required
def sc_customers():
    if not sc_require('sc_view'): return redirect(url_for('sc_index'))
    from datetime import date as _date, timedelta
    db    = get_db()
    today = _date.today().isoformat()
    soon  = (_date.today() + timedelta(days=90)).isoformat()
    rows  = db.execute('SELECT * FROM sc_customers ORDER BY name').fetchall()
    # Contract status per customer: green / orange / red
    contract_status = {}
    for r in rows:
        ctr = db.execute('''SELECT end_date FROM sc_contracts
                            WHERE customer_id=? AND status='active'
                            ORDER BY end_date DESC LIMIT 1''', (r['id'],)).fetchone()
        if not ctr:
            contract_status[r['id']] = 'none'       # merah
        elif ctr['end_date'] <= soon:
            contract_status[r['id']] = 'expiring'   # orange
        else:
            contract_status[r['id']] = 'active'     # hijau
    return render_template('sc_customers.html', rows=rows, contract_status=contract_status)

@app.route('/support/api/customer-pics/<int:cid>')
@login_required
def sc_api_customer_pics(cid):
    db = get_db()
    return jsonify(_sc_customer_pic_info(db, cid))

@app.route('/support/customers/add', methods=['GET', 'POST'])
@login_required
def sc_customer_add():
    if not sc_require('sc_manage_customers'): return redirect(url_for('sc_customers'))
    db = get_db()
    helpdesk, implementor, coleader, sales = _sc_pic_employees(db)
    if request.method == 'POST':
        code = request.form.get('code', '').strip().upper()
        name = request.form.get('name', '').strip()
        if not code or not name:
            flash('Kode dan nama wajib diisi', 'danger')
        else:
            try:
                db.execute(
                    '''INSERT INTO sc_customers
                       (code,name,address,contact_person,phone,email,notes,customer_type,
                        pic_helpdesk_id,pic_helpdesk_backup_id,pic_implementor_id,pic_coleader_id,
                        pic_sales_id,telegram_group_id)
                       VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?)''',
                    (code, name,
                     request.form.get('address','').strip(),
                     request.form.get('contact_person','').strip(),
                     request.form.get('phone','').strip(),
                     request.form.get('email','').strip(),
                     request.form.get('notes','').strip(),
                     request.form.get('customer_type','aktif'),
                     request.form.get('pic_helpdesk_id') or None,
                     request.form.get('pic_helpdesk_backup_id') or None,
                     request.form.get('pic_implementor_id') or None,
                     request.form.get('pic_coleader_id') or None,
                     request.form.get('pic_sales_id') or None,
                     request.form.get('telegram_group_id','').strip()))
                db.commit()
                flash(f'Customer "{name}" ditambahkan', 'success')
                return redirect(url_for('sc_customers'))
            except Exception:
                flash('Kode customer sudah ada', 'danger')
    return render_template('sc_customer_form.html', row=None,
                           helpdesk=helpdesk, implementor=implementor, coleader=coleader, sales=sales)

@app.route('/support/customers/<int:cid>/edit', methods=['GET', 'POST'])
@login_required
def sc_customer_edit(cid):
    if not sc_require('sc_manage_customers'): return redirect(url_for('sc_customers'))
    db  = get_db()
    row = db.execute('SELECT * FROM sc_customers WHERE id=?', (cid,)).fetchone()
    if not row:
        flash('Customer tidak ditemukan', 'danger')
        return redirect(url_for('sc_customers'))
    helpdesk, implementor, coleader, sales = _sc_pic_employees(db)
    if request.method == 'POST':
        name      = request.form.get('name', '').strip()
        is_active = 1 if request.form.get('is_active') else 0
        db.execute('''UPDATE sc_customers
                      SET name=?,address=?,contact_person=?,phone=?,email=?,notes=?,is_active=?,
                          customer_type=?,
                          pic_helpdesk_id=?,pic_helpdesk_backup_id=?,pic_implementor_id=?,pic_coleader_id=?,
                          pic_sales_id=?,telegram_group_id=?
                      WHERE id=?''',
                   (name,
                    request.form.get('address','').strip(),
                    request.form.get('contact_person','').strip(),
                    request.form.get('phone','').strip(),
                    request.form.get('email','').strip(),
                    request.form.get('notes','').strip(),
                    is_active,
                    request.form.get('customer_type','aktif'),
                    request.form.get('pic_helpdesk_id') or None,
                    request.form.get('pic_helpdesk_backup_id') or None,
                    request.form.get('pic_implementor_id') or None,
                    request.form.get('pic_coleader_id') or None,
                    request.form.get('pic_sales_id') or None,
                    request.form.get('telegram_group_id','').strip(),
                    cid))
        db.commit()
        flash('Customer diperbarui', 'success')
        return redirect(url_for('sc_customers'))
    return render_template('sc_customer_form.html', row=row,
                           helpdesk=helpdesk, implementor=implementor, coleader=coleader, sales=sales)

@app.route('/support/customers/<int:cid>/delete', methods=['POST'])
@login_required
def sc_customer_delete(cid):
    if not sc_require('sc_manage_customers'): return redirect(url_for('sc_customers'))
    db = get_db()
    db.execute('UPDATE sc_customers SET is_active=0 WHERE id=?', (cid,))
    db.commit()
    flash('Customer dinonaktifkan', 'warning')
    return redirect(url_for('sc_customers'))

# ── Notification helper ────────────────────────────────────────────────────────
def _save_ticket_attachments(db, ticket_id, section):
    """Simpan file upload dari request.files (multiple) ke tabel sc_ticket_attachments."""
    files = request.files.getlist(f'attach_{section}')
    for f in files:
        if not f or not f.filename:
            continue
        url = _save_upload_file(f, f'tickets/{ticket_id}/{section}', ALLOWED_ATTACHMENT_EXT)
        if url:
            db.execute('''INSERT INTO sc_ticket_attachments(ticket_id,section,filename,original_name,uploaded_by,uploaded_by_name)
                          VALUES(?,?,?,?,?,?)''',
                       (ticket_id, section, url, f.filename,
                        session.get('user_id'), session.get('user_name','')))

def _sync_external_assignees(db, table, fk_col, fk_val, names):
    """Ganti seluruh external assignees dengan daftar nama baru."""
    db.execute(f'DELETE FROM {table} WHERE {fk_col}=?', (fk_val,))
    for name in names:
        name = name.strip()
        if name:
            db.execute(f'INSERT INTO {table}({fk_col},name,added_by) VALUES(?,?,?)',
                       (fk_val, name, session.get('user_id')))

def _sc_ticket_history(db, ticket_id, action, field_name='', old_value='', new_value='', notes=''):
    """Catat history perubahan tiket (tidak menimpa, selalu tambah baris baru)."""
    try:
        db.execute('''INSERT INTO sc_ticket_history
                      (ticket_id,changed_by,changed_by_name,action,field_name,old_value,new_value,notes)
                      VALUES(?,?,?,?,?,?,?,?)''',
                   (ticket_id, session.get('user_id'), session.get('user_name',''),
                    action, field_name, str(old_value or ''), str(new_value or ''), notes))
        db.commit()
    except Exception:
        pass

def _sc_sync_assignees(db, ticket_id, employee_ids):
    """Sync multiple assignee ke sc_ticket_assignees."""
    old_ids = {r['employee_id'] for r in
               db.execute('SELECT employee_id FROM sc_ticket_assignees WHERE ticket_id=?', (ticket_id,)).fetchall()}
    new_ids = set(int(i) for i in employee_ids if i)
    # tambah yang baru
    for eid in new_ids - old_ids:
        emp = db.execute('SELECT name, divisi FROM employees WHERE id=?', (eid,)).fetchone()
        db.execute('INSERT OR IGNORE INTO sc_ticket_assignees(ticket_id,employee_id,divisi) VALUES(?,?,?)',
                   (ticket_id, eid, emp['divisi'] if emp else ''))
        if emp:
            _sc_ticket_history(db, ticket_id, 'assignee_added', 'assignee', '', emp['name'],
                               f"Assignee ditambahkan: {emp['name']}")
    # hapus yang dicopot
    for eid in old_ids - new_ids:
        emp = db.execute('SELECT name FROM employees WHERE id=?', (eid,)).fetchone()
        db.execute('DELETE FROM sc_ticket_assignees WHERE ticket_id=? AND employee_id=?', (ticket_id, eid))
        if emp:
            _sc_ticket_history(db, ticket_id, 'assignee_removed', 'assignee', emp['name'], '',
                               f"Assignee dicopot: {emp['name']}")
    db.commit()

def _sc_notify_ticket(db, ticket_id, event='created'):
    """Send Telegram notification to customer group when ticket event occurs."""
    try:
        bot_token = db.execute("SELECT value FROM app_settings WHERE key='telegram_bot_token'").fetchone()
        if not bot_token or not bot_token['value']:
            return
        t = db.execute('''SELECT t.*, cu.name as customer_name, cu.telegram_group_id,
                          st.name as type_name, e.name as assignee_name
                          FROM sc_tickets t
                          JOIN sc_customers cu ON cu.id=t.customer_id
                          JOIN sc_support_types st ON st.id=t.support_type_id
                          LEFT JOIN employees e ON e.id=t.assignee_id
                          WHERE t.id=?''', (ticket_id,)).fetchone()
        if not t or not t['telegram_group_id']:
            return
        labels = {'created': '🎫 Tiket Baru', 'assigned': '👤 Tiket Ditugaskan', 'status': '🔄 Status Tiket Diperbarui'}
        header = labels.get(event, '📋 Update Tiket')
        msg = f"{header}\n\n*{t['ticket_no']}* — {t['subject']}\nCustomer: {t['customer_name']}\nTipe: {t['type_name']}\nStatus: {t['status']}"
        if t['assignee_name']:
            msg += f"\nAssignee: {t['assignee_name']}"
        send_telegram_bg(bot_token['value'], t['telegram_group_id'], msg)
    except Exception:
        pass

# ── Auto-log tiket ke Project Performance evaluasi karyawan ───────────────────
def _sc_log_ticket_to_eval(db, ticket_id, event='assigned'):
    """Catat aktivitas tiket ke project_entries evaluasi karyawan (assignee)."""
    try:
        from datetime import date as _date
        t = db.execute('''SELECT t.*, cu.name as customer_name, st.name as type_name
                          FROM sc_tickets t
                          JOIN sc_customers cu ON cu.id=t.customer_id
                          JOIN sc_support_types st ON st.id=t.support_type_id
                          WHERE t.id=?''', (ticket_id,)).fetchone()
        if not t or not t['assignee_id']:
            return
        periode = str(_date.today().year)
        # Cari atau buat evaluasi periode ini untuk assignee
        ev = db.execute('SELECT id FROM evaluations WHERE employee_id=? AND periode=?',
                        (t['assignee_id'], periode)).fetchone()
        if not ev:
            cur = db.execute('INSERT INTO evaluations(employee_id, periode, status) VALUES(?,?,?)',
                             (t['assignee_id'], periode, 'draft'))
            eval_id = cur.lastrowid
        else:
            eval_id = ev['id']
        # Cek apakah tiket ini sudah ada di project_entries
        existing = db.execute(
            "SELECT id FROM project_entries WHERE eval_id=? AND project_name=?",
            (eval_id, t['ticket_no'])).fetchone()
        status_map = {
            'assigned':    'ON PROGRESS',
            'in_progress': 'ON PROGRESS',
            'resolved':    'DONE',
            'closed':      'DONE',
            'hold':        'HOLD',
            'feedback':    'ON PROGRESS',
            'rejected':    'CANCELLED',
        }
        entry_status = status_map.get(event, status_map.get(t['status'], 'ON PROGRESS'))
        detail = f"[{t['type_name']}] {t['subject']} — {t['customer_name']}"
        if existing:
            db.execute('UPDATE project_entries SET status=?, detail_task=? WHERE id=?',
                       (entry_status, detail, existing['id']))
        else:
            max_order = db.execute('SELECT COALESCE(MAX(sort_order),0) FROM project_entries WHERE eval_id=? AND entry_type=?',
                                   (eval_id, 'history')).fetchone()[0]
            db.execute('''INSERT INTO project_entries(eval_id,entry_type,project_name,detail_task,status,notes,sort_order)
                          VALUES(?,?,?,?,?,?,?)''',
                       (eval_id, 'history', t['ticket_no'], detail, entry_status,
                        f"Auto dari SupportCore #{t['ticket_no']}", max_order + 1))
        db.commit()
    except Exception:
        pass

# ── Master Apps / Modules ──────────────────────────────────────────────────────
@app.route('/support/apps')
@login_required
def sc_apps():
    if not sc_require('sc_view'): return redirect(url_for('sc_index'))
    db   = get_db()
    apps = db.execute('SELECT * FROM sc_apps ORDER BY name').fetchall()
    mods = db.execute('''SELECT m.*, a.name as app_name FROM sc_modules m
                         JOIN sc_apps a ON a.id=m.app_id ORDER BY a.name, m.name''').fetchall()
    return render_template('sc_apps.html', apps=apps, mods=mods)

@app.route('/support/apps/add', methods=['GET','POST'])
@login_required
def sc_app_add():
    if not sc_require('sc_manage_apps'): return redirect(url_for('sc_apps'))
    db = get_db()
    if request.method == 'POST':
        name = request.form.get('name','').strip()
        if not name:
            flash('Nama aplikasi wajib diisi', 'danger')
        else:
            try:
                db.execute('INSERT INTO sc_apps(name,description) VALUES(?,?)',
                           (name, request.form.get('description','').strip()))
                db.commit()
                flash(f'Aplikasi "{name}" ditambahkan', 'success')
                return redirect(url_for('sc_apps'))
            except Exception:
                flash('Nama aplikasi sudah ada', 'danger')
    return render_template('sc_apps.html', apps=db.execute('SELECT * FROM sc_apps ORDER BY name').fetchall(),
                           mods=db.execute('SELECT m.*, a.name as app_name FROM sc_modules m JOIN sc_apps a ON a.id=m.app_id ORDER BY a.name, m.name').fetchall(),
                           add_app=True)

@app.route('/support/apps/<int:aid>/delete', methods=['POST'])
@login_required
def sc_app_delete(aid):
    if not sc_require('sc_manage_apps'): return redirect(url_for('sc_apps'))
    db = get_db()
    db.execute('DELETE FROM sc_apps WHERE id=?', (aid,))
    db.commit()
    flash('Aplikasi dihapus', 'warning')
    return redirect(url_for('sc_apps'))

@app.route('/support/modules/add', methods=['POST'])
@login_required
def sc_module_add():
    if not sc_require('sc_manage_apps'): return redirect(url_for('sc_apps'))
    db = get_db()
    app_id = request.form.get('app_id')
    name   = request.form.get('name','').strip()
    if not app_id or not name:
        flash('Aplikasi dan nama modul wajib diisi', 'danger')
    else:
        try:
            db.execute('INSERT INTO sc_modules(app_id,name,description) VALUES(?,?,?)',
                       (app_id, name, request.form.get('description','').strip()))
            db.commit()
            flash(f'Modul "{name}" ditambahkan', 'success')
        except Exception:
            flash('Modul sudah ada di aplikasi tersebut', 'danger')
    return redirect(url_for('sc_apps'))

@app.route('/support/modules/<int:mid>/delete', methods=['POST'])
@login_required
def sc_module_delete(mid):
    if not sc_require('sc_manage_apps'): return redirect(url_for('sc_apps'))
    db = get_db()
    db.execute('DELETE FROM sc_modules WHERE id=?', (mid,))
    db.commit()
    flash('Modul dihapus', 'warning')
    return redirect(url_for('sc_apps'))

@app.route('/support/api/modules')
@login_required
def sc_api_modules():
    db = get_db()
    mods = db.execute('''SELECT m.id, m.name, a.id as app_id, a.name as app_name
                         FROM sc_modules m JOIN sc_apps a ON a.id=m.app_id
                         WHERE m.is_active=1 ORDER BY a.name, m.name''').fetchall()
    return jsonify([dict(r) for r in mods])

# ── Master Services ────────────────────────────────────────────────────────────
@app.route('/support/services')
@login_required
def sc_services():
    if not sc_require('sc_view'): return redirect(url_for('sc_index'))
    db = get_db()
    rows = db.execute('SELECT * FROM sc_services ORDER BY name').fetchall()
    return render_template('sc_services.html', rows=rows)

@app.route('/support/services/add', methods=['GET', 'POST'])
@login_required
def sc_service_add():
    if not sc_require('sc_manage_services'): return redirect(url_for('sc_services'))
    db = get_db()
    if request.method == 'POST':
        code = request.form.get('code', '').strip().upper()
        name = request.form.get('name', '').strip()
        if not code or not name:
            flash('Kode dan nama wajib diisi', 'danger')
        else:
            try:
                db.execute('INSERT INTO sc_services(code,name,description) VALUES(?,?,?)',
                           (code, name, request.form.get('description','').strip()))
                db.commit()
                flash(f'Layanan "{name}" ditambahkan', 'success')
                return redirect(url_for('sc_services'))
            except Exception:
                flash('Kode layanan sudah ada', 'danger')
    return render_template('sc_service_form.html', row=None)

@app.route('/support/services/<int:sid>/edit', methods=['GET', 'POST'])
@login_required
def sc_service_edit(sid):
    if not sc_require('sc_manage_services'): return redirect(url_for('sc_services'))
    db  = get_db()
    row = db.execute('SELECT * FROM sc_services WHERE id=?', (sid,)).fetchone()
    if not row:
        flash('Layanan tidak ditemukan', 'danger')
        return redirect(url_for('sc_services'))
    if request.method == 'POST':
        is_active = 1 if request.form.get('is_active') else 0
        db.execute('UPDATE sc_services SET name=?,description=?,is_active=? WHERE id=?',
                   (request.form.get('name','').strip(),
                    request.form.get('description','').strip(),
                    is_active, sid))
        db.commit()
        flash('Layanan diperbarui', 'success')
        return redirect(url_for('sc_services'))
    return render_template('sc_service_form.html', row=row)

@app.route('/support/services/<int:sid>/delete', methods=['POST'])
@login_required
def sc_service_delete(sid):
    if not sc_require('sc_manage_services'): return redirect(url_for('sc_services'))
    db = get_db()
    db.execute('UPDATE sc_services SET is_active=0 WHERE id=?', (sid,))
    db.commit()
    flash('Layanan dinonaktifkan', 'warning')
    return redirect(url_for('sc_services'))

# ── Master Support Types ───────────────────────────────────────────────────────
@app.route('/support/support-types')
@login_required
def sc_support_types():
    if not sc_require('sc_view'): return redirect(url_for('sc_index'))
    db = get_db()
    rows = db.execute('SELECT * FROM sc_support_types ORDER BY code').fetchall()
    return render_template('sc_support_types.html', rows=rows)

@app.route('/support/support-types/add', methods=['GET', 'POST'])
@login_required
def sc_support_type_add():
    if not sc_require('sc_manage_types'): return redirect(url_for('sc_support_types'))
    db = get_db()
    if request.method == 'POST':
        code = request.form.get('code', '').strip().upper()
        name = request.form.get('name', '').strip()
        if not code or not name:
            flash('Kode dan nama wajib diisi', 'danger')
        else:
            try:
                db.execute('INSERT INTO sc_support_types(code,name,description) VALUES(?,?,?)',
                           (code, name, request.form.get('description','').strip()))
                db.commit()
                flash(f'Tipe support "{name}" ditambahkan', 'success')
                return redirect(url_for('sc_support_types'))
            except Exception:
                flash('Kode tipe sudah ada', 'danger')
    return render_template('sc_support_type_form.html', row=None)

@app.route('/support/support-types/<int:tid>/edit', methods=['GET', 'POST'])
@login_required
def sc_support_type_edit(tid):
    if not sc_require('sc_manage_types'): return redirect(url_for('sc_support_types'))
    db  = get_db()
    row = db.execute('SELECT * FROM sc_support_types WHERE id=?', (tid,)).fetchone()
    if not row:
        flash('Tipe support tidak ditemukan', 'danger')
        return redirect(url_for('sc_support_types'))
    if request.method == 'POST':
        is_active = 1 if request.form.get('is_active') else 0
        db.execute('UPDATE sc_support_types SET name=?,description=?,is_active=? WHERE id=?',
                   (request.form.get('name','').strip(),
                    request.form.get('description','').strip(),
                    is_active, tid))
        db.commit()
        flash('Tipe support diperbarui', 'success')
        return redirect(url_for('sc_support_types'))
    return render_template('sc_support_type_form.html', row=row)

@app.route('/support/support-types/<int:tid>/delete', methods=['POST'])
@login_required
def sc_support_type_delete(tid):
    if not sc_require('sc_manage_types'): return redirect(url_for('sc_support_types'))
    db = get_db()
    db.execute('UPDATE sc_support_types SET is_active=0 WHERE id=?', (tid,))
    db.commit()
    flash('Tipe support dinonaktifkan', 'warning')
    return redirect(url_for('sc_support_types'))

# ── Master SLA Categories ──────────────────────────────────────────────────────
@app.route('/support/sla-categories')
@login_required
def sc_sla_categories():
    if not sc_require('sc_view'): return redirect(url_for('sc_index'))
    db = get_db()
    rows = db.execute('SELECT * FROM sc_sla_categories ORDER BY priority, name').fetchall()
    return render_template('sc_sla_categories.html', rows=rows)

@app.route('/support/sla-categories/add', methods=['GET', 'POST'])
@login_required
def sc_sla_category_add():
    if not sc_require('sc_manage_sla'): return redirect(url_for('sc_sla_categories'))
    db = get_db()
    if request.method == 'POST':
        code = request.form.get('code', '').strip().upper()
        name = request.form.get('name', '').strip()
        if not code or not name:
            flash('Kode dan nama wajib diisi', 'danger')
        else:
            try:
                wta_raw = request.form.get('workaround_time_hours','').strip()
                db.execute('''INSERT INTO sc_sla_categories
                              (code,name,priority,response_time_hours,workaround_time_hours,
                               resolution_time_hours,maintenance_type,description)
                              VALUES(?,?,?,?,?,?,?,?)''',
                           (code, name,
                            request.form.get('priority','Medium'),
                            float(request.form.get('response_time_hours', 4) or 4),
                            float(wta_raw) if wta_raw else None,
                            float(request.form.get('resolution_time_hours', 24) or 24),
                            request.form.get('maintenance_type','corrective'),
                            request.form.get('description','').strip()))
                db.commit()
                flash(f'Kategori SLA "{name}" ditambahkan', 'success')
                return redirect(url_for('sc_sla_categories'))
            except Exception:
                flash('Kode kategori sudah ada', 'danger')
    return render_template('sc_sla_category_form.html', row=None)

@app.route('/support/sla-categories/<int:kid>/edit', methods=['GET', 'POST'])
@login_required
def sc_sla_category_edit(kid):
    if not sc_require('sc_manage_sla'): return redirect(url_for('sc_sla_categories'))
    db  = get_db()
    row = db.execute('SELECT * FROM sc_sla_categories WHERE id=?', (kid,)).fetchone()
    if not row:
        flash('Kategori SLA tidak ditemukan', 'danger')
        return redirect(url_for('sc_sla_categories'))
    if request.method == 'POST':
        is_active = 1 if request.form.get('is_active') else 0
        wta_raw = request.form.get('workaround_time_hours','').strip()
        db.execute('''UPDATE sc_sla_categories
                      SET name=?,priority=?,response_time_hours=?,workaround_time_hours=?,
                          resolution_time_hours=?,maintenance_type=?,description=?,is_active=?
                      WHERE id=?''',
                   (request.form.get('name','').strip(),
                    request.form.get('priority','Medium'),
                    float(request.form.get('response_time_hours', 4) or 4),
                    float(wta_raw) if wta_raw else None,
                    float(request.form.get('resolution_time_hours', 24) or 24),
                    request.form.get('maintenance_type','corrective'),
                    request.form.get('description','').strip(),
                    is_active, kid))
        db.commit()
        flash('Kategori SLA diperbarui', 'success')
        return redirect(url_for('sc_sla_categories'))
    return render_template('sc_sla_category_form.html', row=row)

@app.route('/support/sla-categories/<int:kid>/delete', methods=['POST'])
@login_required
def sc_sla_category_delete(kid):
    if not sc_require('sc_manage_sla'): return redirect(url_for('sc_sla_categories'))
    db = get_db()
    db.execute('UPDATE sc_sla_categories SET is_active=0 WHERE id=?', (kid,))
    db.commit()
    flash('Kategori SLA dinonaktifkan', 'warning')
    return redirect(url_for('sc_sla_categories'))

# ─── SupportCore Helpers ───────────────────────────────────────────────────────

def generate_ticket_no(db):
    from datetime import datetime as _dt
    year = datetime.now().year
    last = db.execute(
        "SELECT ticket_no FROM sc_tickets WHERE ticket_no LIKE ? ORDER BY id DESC LIMIT 1",
        (f'TKT-{year}-%',)
    ).fetchone()
    seq = (int(last['ticket_no'].rsplit('-', 1)[-1]) + 1) if last else 1
    return f'TKT-{year}-{seq:04d}'

def calc_sla(reported_at_str, done_at_str, limit_hours):
    """Returns ('met'|'violated'|'pending', hours_taken_or_None)"""
    if not reported_at_str:
        return 'pending', None
    from datetime import datetime as _dt
    fmt = '%Y-%m-%d %H:%M:%S'
    try:
        t0 = datetime.strptime(reported_at_str[:19], fmt)
        if done_at_str:
            t1 = datetime.strptime(done_at_str[:19], fmt)
            hours = (t1 - t0).total_seconds() / 3600
            return ('met' if hours <= limit_hours else 'violated'), round(hours, 1)
        return 'pending', None
    except Exception:
        return 'pending', None

# ─── SupportCore: Contracts ────────────────────────────────────────────────────

@app.route('/support/contracts')
@login_required
def sc_contracts():
    if not sc_require('sc_view'): return redirect(url_for('sc_index'))
    db = get_db()
    from datetime import date as _date
    today = _date.today().isoformat()
    customer_id = request.args.get('customer_id', '')
    status_f    = request.args.get('status', '')
    q = '''SELECT c.*, cu.name as customer_name
           FROM sc_contracts c
           JOIN sc_customers cu ON cu.id = c.customer_id
           WHERE 1=1'''
    params = []
    if customer_id:
        q += ' AND c.customer_id=?'; params.append(customer_id)
    if status_f:
        q += ' AND c.status=?'; params.append(status_f)
    q += ' ORDER BY c.start_date DESC'
    rows = db.execute(q, params).fetchall()
    customers = db.execute('SELECT id, name FROM sc_customers WHERE is_active=1 ORDER BY name').fetchall()
    # Compute display status
    contracts = []
    for r in rows:
        d = dict(r)
        if d['status'] not in ('terminated', 'draft') and d['end_date'] < today:
            d['display_status'] = 'expired'
        else:
            d['display_status'] = d['status']
        contracts.append(d)
    return render_template('sc_contracts.html', contracts=contracts, customers=customers,
                           filter_customer=customer_id, filter_status=status_f, today=today)

def _sc_pic_employees(db):
    """Return (helpdesk, implementor, coleader, sales) employee lists for PIC dropdowns."""
    helpdesk   = db.execute("SELECT id,name,jabatan FROM employees WHERE is_active=1 AND LOWER(divisi) LIKE '%helpdesk%' ORDER BY name").fetchall()
    implementor= db.execute("SELECT id,name,jabatan FROM employees WHERE is_active=1 AND LOWER(divisi) LIKE '%implementor%' ORDER BY name").fetchall()
    coleader   = db.execute("SELECT id,name,jabatan FROM employees WHERE is_active=1 AND LOWER(level) LIKE '%co-leader%' ORDER BY name").fetchall()
    sales      = db.execute("SELECT id,name,jabatan FROM employees WHERE is_active=1 AND (LOWER(divisi) LIKE '%sales%' OR LOWER(jabatan) LIKE '%sales%') ORDER BY name").fetchall()
    return helpdesk, implementor, coleader, sales

def _sc_customer_pic_info(db, customer_id):
    """Return dict of PIC names for a customer (for AJAX & template display)."""
    c = db.execute('SELECT pic_helpdesk_id,pic_helpdesk_backup_id,pic_implementor_id,pic_coleader_id FROM sc_customers WHERE id=?', (customer_id,)).fetchone()
    if not c:
        return {}
    def emp_name(eid):
        if not eid: return None
        r = db.execute('SELECT name, jabatan FROM employees WHERE id=?', (eid,)).fetchone()
        return f"{r['name']} ({r['jabatan']})" if r else None
    return {
        'pic_helpdesk':        emp_name(c['pic_helpdesk_id']),
        'pic_helpdesk_backup': emp_name(c['pic_helpdesk_backup_id']),
        'pic_implementor':     emp_name(c['pic_implementor_id']),
        'pic_coleader':        emp_name(c['pic_coleader_id']),
    }

def _sc_contract_lookups(db):
    """Data lookup yang dibutuhkan form kontrak."""
    customers     = db.execute('SELECT id, name FROM sc_customers WHERE is_active=1 ORDER BY name').fetchall()
    services      = db.execute('SELECT id, name FROM sc_services WHERE is_active=1 ORDER BY name').fetchall()
    support_types = db.execute('SELECT id, code, name FROM sc_support_types WHERE is_active=1 ORDER BY code').fetchall()
    helpdesk_emps = db.execute(
        "SELECT id, name, jabatan FROM employees WHERE is_active=1 AND LOWER(divisi) LIKE '%helpdesk%' ORDER BY name"
    ).fetchall()
    return customers, services, support_types, helpdesk_emps

@app.route('/support/contracts/add', methods=['GET','POST'])
@login_required
def sc_contract_add():
    if not sc_require('sc_manage_contracts'): return redirect(url_for('sc_contracts'))
    db = get_db()
    customers, services, support_types, helpdesk_emps = _sc_contract_lookups(db)
    if request.method == 'POST':
        code      = request.form['code'].strip()
        cid       = request.form['customer_id']
        title     = request.form['title'].strip()
        desc      = request.form.get('description','').strip()
        sd        = request.form['start_date']
        ed        = request.form['end_date']
        val       = float(request.form.get('contract_value') or 0)
        status    = request.form.get('status','active')
        notes     = request.form.get('notes','').strip()
        pic_id    = request.form.get('pic_helpdesk_id') or None
        svc_ids   = request.form.getlist('service_ids')
        type_ids  = request.form.getlist('support_type_ids')
        try:
            db.execute('''INSERT INTO sc_contracts(code,customer_id,title,description,start_date,end_date,
                          contract_value,status,notes,pic_helpdesk_id,created_by)
                          VALUES(?,?,?,?,?,?,?,?,?,?,?)''',
                       (code, cid, title, desc, sd, ed, val, status, notes, pic_id, session.get('user_id')))
            ctr_id = db.execute('SELECT last_insert_rowid()').fetchone()[0]
            for sid in svc_ids:
                db.execute('INSERT OR IGNORE INTO sc_contract_services(contract_id,service_id) VALUES(?,?)', (ctr_id, sid))
            for tid in type_ids:
                db.execute('INSERT OR IGNORE INTO sc_contract_support_types(contract_id,support_type_id) VALUES(?,?)', (ctr_id, tid))
            db.commit()
            flash('Kontrak berhasil ditambahkan', 'success')
            return redirect(url_for('sc_contracts'))
        except Exception as e:
            flash(f'Error: {e}', 'danger')
    return render_template('sc_contract_form.html', row=None,
                           customers=customers, services=services,
                           support_types=support_types, helpdesk_emps=helpdesk_emps,
                           sel_services=[], sel_types=[], sel_pic=None)

@app.route('/support/contracts/<int:cid>/edit', methods=['GET','POST'])
@login_required
def sc_contract_edit(cid):
    if not sc_require('sc_manage_contracts'): return redirect(url_for('sc_contracts'))
    db  = get_db()
    row = db.execute('SELECT * FROM sc_contracts WHERE id=?', (cid,)).fetchone()
    if not row: abort(404)
    customers, services, support_types, helpdesk_emps = _sc_contract_lookups(db)
    sel_services = [r['service_id']      for r in db.execute('SELECT service_id FROM sc_contract_services WHERE contract_id=?',      (cid,)).fetchall()]
    sel_types    = [r['support_type_id'] for r in db.execute('SELECT support_type_id FROM sc_contract_support_types WHERE contract_id=?', (cid,)).fetchall()]
    if request.method == 'POST':
        code     = request.form['code'].strip()
        custid   = request.form['customer_id']
        title    = request.form['title'].strip()
        desc     = request.form.get('description','').strip()
        sd       = request.form['start_date']
        ed       = request.form['end_date']
        val      = float(request.form.get('contract_value') or 0)
        status   = request.form.get('status','active')
        notes    = request.form.get('notes','').strip()
        pic_id   = request.form.get('pic_helpdesk_id') or None
        svc_ids  = request.form.getlist('service_ids')
        type_ids = request.form.getlist('support_type_ids')
        try:
            db.execute('''UPDATE sc_contracts SET code=?,customer_id=?,title=?,description=?,start_date=?,
                          end_date=?,contract_value=?,status=?,notes=?,pic_helpdesk_id=? WHERE id=?''',
                       (code, custid, title, desc, sd, ed, val, status, notes, pic_id, cid))
            db.execute('DELETE FROM sc_contract_services WHERE contract_id=?', (cid,))
            db.execute('DELETE FROM sc_contract_support_types WHERE contract_id=?', (cid,))
            for sid in svc_ids:
                db.execute('INSERT OR IGNORE INTO sc_contract_services(contract_id,service_id) VALUES(?,?)', (cid, sid))
            for tid in type_ids:
                db.execute('INSERT OR IGNORE INTO sc_contract_support_types(contract_id,support_type_id) VALUES(?,?)', (cid, tid))
            db.commit()
            flash('Kontrak diperbarui', 'success')
            return redirect(url_for('sc_contracts'))
        except Exception as e:
            flash(f'Error: {e}', 'danger')
    return render_template('sc_contract_form.html', row=row,
                           customers=customers, services=services,
                           support_types=support_types, helpdesk_emps=helpdesk_emps,
                           sel_services=sel_services, sel_types=sel_types,
                           sel_pic=row['pic_helpdesk_id'])

@app.route('/support/contracts/<int:cid>')
@login_required
def sc_contract_detail(cid):
    if not sc_require('sc_view'): return redirect(url_for('sc_index'))
    db = get_db()
    from datetime import date as _date
    today = _date.today().isoformat()
    row = db.execute('''SELECT c.*, cu.name as customer_name
                        FROM sc_contracts c JOIN sc_customers cu ON cu.id=c.customer_id
                        WHERE c.id=?''', (cid,)).fetchone()
    if not row: abort(404)
    ctr = dict(row)
    if ctr['status'] not in ('terminated','draft') and ctr['end_date'] < today:
        ctr['display_status'] = 'expired'
    else:
        ctr['display_status'] = ctr['status']
    svcs = db.execute('''SELECT s.name FROM sc_contract_services cs
                         JOIN sc_services s ON s.id=cs.service_id
                         WHERE cs.contract_id=?''', (cid,)).fetchall()
    tickets = db.execute('''SELECT t.*, st.name as type_name, sc.name as sla_name,
                             sc.response_time_hours, sc.resolution_time_hours
                             FROM sc_tickets t
                             JOIN sc_support_types st ON st.id=t.support_type_id
                             LEFT JOIN sc_sla_categories sc ON sc.id=t.sla_category_id
                             WHERE t.contract_id=?
                             ORDER BY t.reported_at DESC''', (cid,)).fetchall()
    sla_stats = {'met': 0, 'violated': 0, 'pending': 0}
    ticket_list = []
    for t in tickets:
        d = dict(t)
        d['sla_status'], _ = calc_sla(d['reported_at'], d['responded_at'], d['response_time_hours'] or 999)
        sla_stats[d['sla_status']] += 1
        ticket_list.append(d)
    return render_template('sc_contract_detail.html', ctr=ctr, svcs=svcs, tickets=ticket_list, sla_stats=sla_stats)

@app.route('/support/contracts/<int:cid>/delete', methods=['POST'])
@login_required
def sc_contract_delete(cid):
    if not sc_require('sc_manage_contracts'): return redirect(url_for('sc_contracts'))
    db = get_db()
    db.execute("UPDATE sc_contracts SET status='terminated' WHERE id=?", (cid,))
    db.commit()
    flash('Kontrak diterminasi', 'warning')
    return redirect(url_for('sc_contracts'))

# ─── SupportCore: Tickets ──────────────────────────────────────────────────────

@app.route('/support/tickets')
@login_required
def sc_tickets():
    if not sc_require('sc_view'): return redirect(url_for('sc_index'))
    db = get_db()
    status_f  = request.args.get('status','')
    cust_f    = request.args.get('customer_id','')
    type_f    = request.args.get('support_type_id','')
    date_from = request.args.get('date_from','')
    date_to   = request.args.get('date_to','')
    q = '''SELECT t.*, cu.name as customer_name, st.name as type_name,
                  sc.name as sla_name, sc.response_time_hours
           FROM sc_tickets t
           JOIN sc_customers cu ON cu.id=t.customer_id
           JOIN sc_support_types st ON st.id=t.support_type_id
           LEFT JOIN sc_sla_categories sc ON sc.id=t.sla_category_id
           WHERE 1=1'''
    params = []
    if status_f:
        q += ' AND t.status=?'; params.append(status_f)
    if cust_f:
        q += ' AND t.customer_id=?'; params.append(cust_f)
    if type_f:
        q += ' AND t.support_type_id=?'; params.append(type_f)
    if date_from:
        q += ' AND t.reported_at >= ?'; params.append(date_from)
    if date_to:
        q += ' AND t.reported_at <= ?'; params.append(date_to + ' 23:59:59')
    q += ' ORDER BY t.reported_at DESC'
    rows = db.execute(q, params).fetchall()
    tickets = []
    for r in rows:
        d = dict(r)
        d['sla_status'], d['sla_hours'] = calc_sla(d['reported_at'], d['responded_at'], d['response_time_hours'] or 999)
        tickets.append(d)
    customers    = db.execute('SELECT id, name FROM sc_customers WHERE is_active=1 ORDER BY name').fetchall()
    support_types = db.execute('SELECT id, name FROM sc_support_types WHERE is_active=1 ORDER BY name').fetchall()
    return render_template('sc_tickets.html', tickets=tickets, customers=customers,
                           support_types=support_types, filter_status=status_f,
                           filter_customer=cust_f, filter_type=type_f,
                           filter_date_from=date_from, filter_date_to=date_to)

def _sc_ticket_lookups(db):
    customers     = db.execute('SELECT id, name FROM sc_customers WHERE is_active=1 ORDER BY name').fetchall()
    support_types = db.execute('SELECT id, name FROM sc_support_types WHERE is_active=1 ORDER BY name').fetchall()
    sla_cats      = db.execute('SELECT id, name FROM sc_sla_categories WHERE is_active=1 ORDER BY name').fetchall()
    contracts     = db.execute('SELECT id, code, title, customer_id FROM sc_contracts ORDER BY start_date DESC').fetchall()
    modules       = db.execute('''SELECT m.id, m.name, a.id as app_id, a.name as app_name
                                  FROM sc_modules m JOIN sc_apps a ON a.id=m.app_id
                                  WHERE m.is_active=1 ORDER BY a.name, m.name''').fetchall()
    employees     = db.execute("SELECT id, name, jabatan, divisi FROM employees WHERE is_active=1 ORDER BY divisi, name").fetchall()
    return customers, support_types, sla_cats, contracts, modules, employees

@app.route('/support/tickets/add', methods=['GET','POST'])
@login_required
def sc_ticket_add():
    if not sc_require('sc_manage_tickets'): return redirect(url_for('sc_tickets'))
    db = get_db()
    customers, support_types, sla_cats, contracts, modules, employees = _sc_ticket_lookups(db)
    if request.method == 'POST':
        ticket_no     = generate_ticket_no(db)
        contract_id   = request.form.get('contract_id') or None
        customer_id   = request.form['customer_id']
        support_type_id = request.form['support_type_id']
        sla_cat_id    = request.form.get('sla_category_id') or None
        subject       = request.form['subject'].strip()
        description   = request.form.get('description','').strip()
        reported_by   = request.form.get('reported_by','').strip()
        reported_at   = request.form.get('reported_at','').strip() or None
        notes         = request.form.get('notes','').strip()
        module_id     = request.form.get('module_id') or None
        assignee_ids  = request.form.getlist('assignee_ids')
        status_note   = request.form.get('status_note','').strip()
        mandays       = request.form.get('mandays','').strip() or None
        pct_done      = request.form.get('pct_done',0)
        solution_type = request.form.get('solution_type','').strip()
        solution_note = request.form.get('solution_note','').strip()
        due_date      = request.form.get('due_date','').strip() or None
        work_start_date = request.form.get('work_start_date','').strip() or None
        media_lapor   = request.form.get('media_lapor','').strip()
        priority      = request.form.get('priority','Medium').strip()
        from datetime import datetime as _dt
        if not reported_at:
            reported_at = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        try:
            cur = db.execute('''INSERT INTO sc_tickets(ticket_no,contract_id,customer_id,support_type_id,
                          sla_category_id,subject,description,reported_by,reported_at,notes,created_by,
                          module_id,status_note,mandays,pct_done,solution_type,solution_note,
                          due_date,work_start_date,media_lapor,priority)
                          VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)''',
                       (ticket_no, contract_id, customer_id, support_type_id, sla_cat_id,
                        subject, description, reported_by, reported_at, notes, session.get('user_id'),
                        module_id, status_note, mandays, pct_done,
                        solution_type, solution_note, due_date, work_start_date, media_lapor, priority))
            new_id = cur.lastrowid
            db.commit()
            _sc_ticket_history(db, new_id, 'created', notes=f'Tiket dibuat oleh {session.get("user_name","")}')
            for sec in ('description', 'status_note', 'solution_note'):
                _save_ticket_attachments(db, new_id, sec)
            ext_names = [n for n in request.form.get('external_assignees','').split('\n') if n.strip()]
            _sync_external_assignees(db, 'sc_ticket_external_assignees', 'ticket_id', new_id, ext_names)
            db.commit()
            _sc_sync_assignees(db, new_id, assignee_ids)
            _sc_notify_ticket(db, new_id, 'created')
            if assignee_ids:
                _sc_notify_ticket(db, new_id, 'assigned')
                for aid in assignee_ids:
                    if aid: _sc_log_ticket_to_eval(db, new_id, 'assigned')
            flash(f'Tiket {ticket_no} berhasil dibuat', 'success')
            return redirect(url_for('sc_tickets'))
        except Exception as e:
            flash(f'Error: {e}', 'danger')
    return render_template('sc_ticket_form.html', row=None, sel_assignees=[], ext_assignees=[],
                           customers=customers,
                           support_types=support_types, sla_cats=sla_cats, contracts=contracts,
                           modules=modules, employees=employees,
                           sc_ticket_statuses=SC_TICKET_STATUSES,
                           sc_solution_types=SC_SOLUTION_TYPES,
                           sc_media_lapor=SC_MEDIA_LAPOR,
                           sc_ticket_priorities=SC_TICKET_PRIORITIES)

@app.route('/support/tickets/<int:tid>/edit', methods=['GET','POST'])
@login_required
def sc_ticket_edit(tid):
    if not sc_require('sc_manage_tickets'): return redirect(url_for('sc_tickets'))
    db = get_db()
    row = db.execute('SELECT * FROM sc_tickets WHERE id=?', (tid,)).fetchone()
    if not row: abort(404)
    customers, support_types, sla_cats, contracts, modules, employees = _sc_ticket_lookups(db)
    sel_assignees = [r['employee_id'] for r in
                     db.execute('SELECT employee_id FROM sc_ticket_assignees WHERE ticket_id=?', (tid,)).fetchall()]
    if request.method == 'POST':
        contract_id     = request.form.get('contract_id') or None
        customer_id     = request.form['customer_id']
        support_type_id = request.form['support_type_id']
        sla_cat_id      = request.form.get('sla_category_id') or None
        subject         = request.form['subject'].strip()
        description     = request.form.get('description','').strip()
        reported_by     = request.form.get('reported_by','').strip()
        reported_at     = request.form.get('reported_at','').strip() or None
        notes           = request.form.get('notes','').strip()
        module_id       = request.form.get('module_id') or None
        assignee_ids    = request.form.getlist('assignee_ids')
        status_note     = request.form.get('status_note','').strip()
        mandays         = request.form.get('mandays','').strip() or None
        pct_done        = request.form.get('pct_done',0)
        solution_type   = request.form.get('solution_type','').strip()
        solution_note   = request.form.get('solution_note','').strip()
        due_date        = request.form.get('due_date','').strip() or None
        work_start_date = request.form.get('work_start_date','').strip() or None
        media_lapor     = request.form.get('media_lapor','').strip()
        priority        = request.form.get('priority','Medium').strip()
        # Track changed fields for history
        changed = []
        if row['status_note'] != status_note: changed.append(('status_note','Keterangan',row['status_note'],status_note))
        if str(row['pct_done'] or 0) != str(pct_done): changed.append(('pct_done','% Done',row['pct_done'],pct_done))
        if (row['solution_type'] or '') != solution_type: changed.append(('solution_type','Tipe Solusi',row['solution_type'],solution_type))
        if (row['priority'] or 'Medium') != priority: changed.append(('priority','Prioritas',row['priority'],priority))
        try:
            db.execute('''UPDATE sc_tickets SET contract_id=?,customer_id=?,support_type_id=?,
                          sla_category_id=?,subject=?,description=?,reported_by=?,reported_at=?,notes=?,
                          module_id=?,status_note=?,mandays=?,pct_done=?,
                          solution_type=?,solution_note=?,due_date=?,work_start_date=?,media_lapor=?,priority=?
                          WHERE id=?''',
                       (contract_id, customer_id, support_type_id, sla_cat_id,
                        subject, description, reported_by, reported_at, notes,
                        module_id, status_note, mandays, pct_done,
                        solution_type, solution_note, due_date, work_start_date, media_lapor, priority, tid))
            db.commit()
            for sec in ('description', 'status_note', 'solution_note'):
                _save_ticket_attachments(db, tid, sec)
            ext_names = [n for n in request.form.get('external_assignees','').split('\n') if n.strip()]
            _sync_external_assignees(db, 'sc_ticket_external_assignees', 'ticket_id', tid, ext_names)
            db.commit()
            for field, label, old, new in changed:
                _sc_ticket_history(db, tid, 'update', field, old, new, f'{label} diubah')
            _sc_sync_assignees(db, tid, assignee_ids)
            for aid in assignee_ids:
                if aid: _sc_log_ticket_to_eval(db, tid, 'assigned')
            flash('Tiket diperbarui', 'success')
            return redirect(url_for('sc_ticket_detail', tid=tid))
        except Exception as e:
            flash(f'Error: {e}', 'danger')
    sel_ext = db.execute(
        'SELECT name FROM sc_ticket_external_assignees WHERE ticket_id=? ORDER BY id', (tid,)).fetchall()
    return render_template('sc_ticket_form.html', row=row, sel_assignees=sel_assignees,
                           ext_assignees=[r['name'] for r in sel_ext],
                           customers=customers,
                           support_types=support_types, sla_cats=sla_cats, contracts=contracts,
                           modules=modules, employees=employees,
                           sc_ticket_statuses=SC_TICKET_STATUSES,
                           sc_solution_types=SC_SOLUTION_TYPES,
                           sc_media_lapor=SC_MEDIA_LAPOR,
                           sc_ticket_priorities=SC_TICKET_PRIORITIES)

@app.route('/support/tickets/<int:tid>')
@login_required
def sc_ticket_detail(tid):
    if not sc_require('sc_view'): return redirect(url_for('sc_index'))
    db = get_db()
    row = db.execute('''SELECT t.*, cu.name as customer_name, st.name as type_name,
                        sc.name as sla_name, sc.response_time_hours, sc.resolution_time_hours,
                        co.code as contract_code, co.title as contract_title
                        FROM sc_tickets t
                        JOIN sc_customers cu ON cu.id=t.customer_id
                        JOIN sc_support_types st ON st.id=t.support_type_id
                        LEFT JOIN sc_sla_categories sc ON sc.id=t.sla_category_id
                        LEFT JOIN sc_contracts co ON co.id=t.contract_id
                        WHERE t.id=?''', (tid,)).fetchone()
    if not row: abort(404)
    t = dict(row)
    t['resp_sla'], t['resp_hours'] = calc_sla(t['reported_at'], t['responded_at'], t['response_time_hours'] or 999)
    t['res_sla'],  t['res_hours']  = calc_sla(t['reported_at'], t['resolved_at'],  t['resolution_time_hours'] or 999)
    history  = db.execute('SELECT * FROM sc_ticket_history WHERE ticket_id=? ORDER BY id DESC', (tid,)).fetchall()
    assignees = db.execute('''SELECT e.name, e.jabatan, e.divisi, ta.role_note
                              FROM sc_ticket_assignees ta
                              JOIN employees e ON e.id=ta.employee_id
                              WHERE ta.ticket_id=? ORDER BY e.divisi, e.name''', (tid,)).fetchall()
    attachments = db.execute('SELECT * FROM sc_ticket_attachments WHERE ticket_id=? ORDER BY section, id',
                             (tid,)).fetchall()
    att_by_section = {}
    for a in attachments:
        att_by_section.setdefault(a['section'], []).append(a)
    ext_assignees = db.execute(
        'SELECT * FROM sc_ticket_external_assignees WHERE ticket_id=? ORDER BY id', (tid,)).fetchall()
    can_manage = has_permission(session.get('user_role',''), 'sc_manage_tickets', db)
    return render_template('sc_ticket_detail.html', t=t, history=history, assignees=assignees,
                           ext_assignees=ext_assignees,
                           att_by_section=att_by_section, can_manage=can_manage,
                           sc_ticket_statuses=SC_TICKET_STATUSES)

@app.route('/support/tickets/<int:tid>/attachments/<int:att_id>/delete', methods=['POST'])
@login_required
def sc_ticket_attachment_delete(tid, att_id):
    if not sc_require('sc_manage_tickets'): return redirect(url_for('sc_ticket_detail', tid=tid))
    db = get_db()
    att = db.execute('SELECT * FROM sc_ticket_attachments WHERE id=? AND ticket_id=?', (att_id, tid)).fetchone()
    if att:
        try:
            import os as _os
            fpath = _os.path.join(_os.path.dirname(__file__), att['filename'].lstrip('/'))
            if _os.path.exists(fpath):
                _os.remove(fpath)
        except Exception:
            pass
        db.execute('DELETE FROM sc_ticket_attachments WHERE id=?', (att_id,))
        db.commit()
    return redirect(url_for('sc_ticket_detail', tid=tid))

@app.route('/support/tickets/<int:tid>/status', methods=['POST'])
@login_required
def sc_ticket_status(tid):
    db  = get_db()
    row = db.execute('SELECT * FROM sc_tickets WHERE id=?', (tid,)).fetchone()
    if not row: abort(404)
    # Allow any assignee OR users with sc_manage_tickets permission
    user_id    = session.get('user_id')
    emp        = db.execute('SELECT id FROM employees WHERE user_id=?', (user_id,)).fetchone()
    is_assignee = emp and db.execute('SELECT 1 FROM sc_ticket_assignees WHERE ticket_id=? AND employee_id=?',
                                     (tid, emp['id'])).fetchone()
    if not is_assignee and not sc_require('sc_manage_tickets'):
        return redirect(url_for('sc_ticket_detail', tid=tid))
    _allowed_statuses = tuple(s for s, *_ in SC_TICKET_STATUSES)
    new_status  = request.form.get('new_status','')
    if new_status not in _allowed_statuses:
        flash('Status tidak valid.', 'danger')
        return redirect(url_for('sc_ticket_detail', tid=tid))
    status_note = request.form.get('status_note','').strip()
    mandays     = request.form.get('mandays','').strip() or None
    pct_done    = request.form.get('pct_done', row['pct_done'])
    solution_type = request.form.get('solution_type', row['solution_type'] or '')
    solution_note = request.form.get('solution_note', row['solution_note'] or '')
    due_date    = request.form.get('due_date','').strip() or row['due_date']
    work_start  = request.form.get('work_start_date','').strip() or row['work_start_date']
    from datetime import datetime as _dt
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    updates = {'status': new_status, 'status_note': status_note,
               'mandays': mandays, 'pct_done': pct_done,
               'solution_type': solution_type, 'solution_note': solution_note,
               'due_date': due_date, 'work_start_date': work_start}
    if new_status == 'in_progress' and not row['responded_at']:
        updates['responded_at'] = now
    if new_status in ('resolved','feedback') and not row['resolved_at']:
        updates['resolved_at'] = now
    if new_status == 'closed':
        updates['closed_at'] = now
    sets = ', '.join(f'{k}=?' for k in updates)
    db.execute(f'UPDATE sc_tickets SET {sets} WHERE id=?', list(updates.values()) + [tid])
    if str(row['pct_done'] or 0) != str(pct_done):
        _sc_ticket_history(db, tid, 'update', 'pct_done', str(row['pct_done'] or 0), str(pct_done), '% Done diubah')
    db.commit()
    _sc_ticket_history(db, tid, 'status_change', 'status', row['status'], new_status,
                       status_note or f'Status diubah ke {new_status}')
    _sc_notify_ticket(db, tid, 'status')
    _sc_log_ticket_to_eval(db, tid, new_status)
    flash(f'Status tiket diubah ke {new_status}', 'success')
    return redirect(url_for('sc_ticket_detail', tid=tid))

# ─── SupportCore: SLA Monitor ──────────────────────────────────────────────────

@app.route('/support/sla-monitor')
@login_required
def sc_sla_monitor():
    if not sc_require('sc_view_reports'): return redirect(url_for('sc_index'))
    db = get_db()
    from datetime import date as _date
    month = request.args.get('month', _date.today().strftime('%Y-%m'))
    date_from = month + '-01'
    # last day
    import calendar
    try:
        y, m = int(month[:4]), int(month[5:7])
        if not (1 <= m <= 12): raise ValueError
    except (ValueError, IndexError):
        from datetime import date as _d2
        month = _d2.today().strftime('%Y-%m')
        y, m = int(month[:4]), int(month[5:7])
    last_day = calendar.monthrange(y, m)[1]
    date_to = f'{month}-{last_day:02d} 23:59:59'
    tickets = db.execute('''SELECT t.*, cu.name as customer_name, st.name as type_name,
                             sc.response_time_hours, sc.resolution_time_hours
                             FROM sc_tickets t
                             JOIN sc_customers cu ON cu.id=t.customer_id
                             JOIN sc_support_types st ON st.id=t.support_type_id
                             LEFT JOIN sc_sla_categories sc ON sc.id=t.sla_category_id
                             WHERE t.reported_at >= ? AND t.reported_at <= ?
                             ORDER BY t.reported_at DESC''', (date_from, date_to)).fetchall()
    total = len(tickets)
    open_cnt = sum(1 for t in tickets if t['status'] == 'open')
    inprog   = sum(1 for t in tickets if t['status'] == 'in_progress')
    sla_met  = 0; sla_vio = 0
    by_customer = {}
    by_type     = {}
    for t in tickets:
        st, _ = calc_sla(t['reported_at'], t['responded_at'], t['response_time_hours'] or 999)
        if st == 'met':   sla_met += 1
        elif st == 'violated': sla_vio += 1
        cn = t['customer_name']
        if cn not in by_customer:
            by_customer[cn] = {'total':0,'met':0,'violated':0,'pending':0}
        by_customer[cn]['total'] += 1
        by_customer[cn][st] += 1
        tn = t['type_name']
        if tn not in by_type:
            by_type[tn] = {'total':0,'met':0,'violated':0,'pending':0}
        by_type[tn]['total'] += 1
        by_type[tn][st] += 1
    # Monthly trend last 6 months
    import json
    trend_labels = []
    trend_met    = []
    trend_vio    = []
    for i in range(5, -1, -1):
        mm = m - i
        yy = y
        while mm <= 0: mm += 12; yy -= 1
        label = f'{yy}-{mm:02d}'
        ld = calendar.monthrange(yy, mm)[1]
        df = f'{label}-01'
        dt = f'{label}-{ld:02d} 23:59:59'
        ts = db.execute('''SELECT t.reported_at, t.responded_at, sc.response_time_hours
                           FROM sc_tickets t
                           LEFT JOIN sc_sla_categories sc ON sc.id=t.sla_category_id
                           WHERE t.reported_at >= ? AND t.reported_at <= ?''', (df, dt)).fetchall()
        met_c = sum(1 for x in ts if calc_sla(x['reported_at'], x['responded_at'], x['response_time_hours'] or 999)[0] == 'met')
        vio_c = sum(1 for x in ts if calc_sla(x['reported_at'], x['responded_at'], x['response_time_hours'] or 999)[0] == 'violated')
        trend_labels.append(label)
        trend_met.append(met_c)
        trend_vio.append(vio_c)
    sla_pct = round(sla_met / total * 100, 1) if total > 0 else 0
    return render_template('sc_sla_monitor.html',
                           total=total, open_cnt=open_cnt, inprog=inprog,
                           sla_met=sla_met, sla_vio=sla_vio, sla_pct=sla_pct,
                           by_customer=by_customer, by_type=by_type,
                           trend_labels=json.dumps(trend_labels),
                           trend_met=json.dumps(trend_met),
                           trend_vio=json.dumps(trend_vio),
                           month=month)

# ─── SupportCore: Presales & POC ──────────────────────────────────────────────

def _sc_presales_no(db):
    from datetime import date as _d
    prefix = f"PSL-{_d.today().year}-"
    last = db.execute("SELECT req_no FROM sc_presales_requests WHERE req_no LIKE ? ORDER BY id DESC LIMIT 1",
                      (prefix + '%',)).fetchone()
    seq = int(last['req_no'].split('-')[-1]) + 1 if last else 1
    return f"{prefix}{seq:04d}"

SC_PRESALES_STATUSES = [
    ('new',         'Baru',       'secondary'),
    ('in_progress', 'In Progress','primary'),
    ('done',        'Selesai',    'success'),
    ('cancelled',   'Dibatalkan', 'danger'),
]

@app.route('/support/presales')
@login_required
def sc_presales():
    if not sc_require('sc_view'): return redirect(url_for('sc_index'))
    db   = get_db()
    type_filter   = request.args.get('type', '')
    status_filter = request.args.get('status', '')
    q = '''SELECT r.*, cu.name as customer_name, cu.customer_type
           FROM sc_presales_requests r
           JOIN sc_customers cu ON cu.id=r.customer_id'''
    conds, params = [], []
    if type_filter:   conds.append('r.request_type=?'); params.append(type_filter)
    if status_filter: conds.append('r.status=?');       params.append(status_filter)
    if conds: q += ' WHERE ' + ' AND '.join(conds)
    q += ' ORDER BY r.id DESC'
    rows = db.execute(q, params).fetchall()
    # Ambil assignees per request
    assignees_map = {}
    for r in rows:
        assignees_map[r['id']] = db.execute(
            '''SELECT e.name, e.divisi FROM sc_presales_assignees pa
               JOIN employees e ON e.id=pa.employee_id WHERE pa.request_id=?''', (r['id'],)).fetchall()
    return render_template('sc_presales.html', rows=rows, assignees_map=assignees_map,
                           sc_presales_statuses=SC_PRESALES_STATUSES,
                           type_filter=type_filter, status_filter=status_filter)

@app.route('/support/presales/add', methods=['GET','POST'])
@login_required
def sc_presales_add():
    if not sc_require('sc_manage_presales'): return redirect(url_for('sc_presales'))
    db = get_db()
    calon_customers = db.execute(
        "SELECT id, name, code FROM sc_customers WHERE is_active=1 ORDER BY name").fetchall()
    employees = db.execute("SELECT id, name, jabatan, divisi FROM employees WHERE is_active=1 ORDER BY divisi, name").fetchall()
    if request.method == 'POST':
        customer_id  = request.form.get('customer_id')
        req_type     = request.form.get('request_type', 'presales')
        subject      = request.form.get('subject','').strip()
        description  = request.form.get('description','').strip()
        assignee_ids = request.form.getlist('assignee_ids')
        if not customer_id or not subject:
            flash('Customer dan subjek wajib diisi', 'danger')
        else:
            req_no = _sc_presales_no(db)
            try:
                cur = db.execute('''INSERT INTO sc_presales_requests
                                    (req_no,customer_id,request_type,subject,description,created_by)
                                    VALUES(?,?,?,?,?,?)''',
                                 (req_no, customer_id, req_type, subject, description, session.get('user_id')))
                req_id = cur.lastrowid
                db.commit()
                for eid in assignee_ids:
                    if eid:
                        emp = db.execute('SELECT divisi FROM employees WHERE id=?', (eid,)).fetchone()
                        db.execute('INSERT OR IGNORE INTO sc_presales_assignees(request_id,employee_id,divisi) VALUES(?,?,?)',
                                   (req_id, eid, emp['divisi'] if emp else ''))
                db.commit()
                db.execute('''INSERT INTO sc_presales_history(request_id,changed_by,changed_by_name,action,notes)
                              VALUES(?,?,?,?,?)''',
                           (req_id, session.get('user_id'), session.get('user_name',''),
                            'created', f'Request {req_type} dibuat: {subject}'))
                db.commit()
                flash(f'Request {req_no} berhasil dibuat', 'success')
                return redirect(url_for('sc_presales'))
            except Exception as e:
                flash(f'Error: {e}', 'danger')
    return render_template('sc_presales_form.html', row=None, sel_assignees=[],
                           calon_customers=calon_customers, employees=employees,
                           sc_presales_statuses=SC_PRESALES_STATUSES)

@app.route('/support/presales/<int:rid>/edit', methods=['GET','POST'])
@login_required
def sc_presales_edit(rid):
    if not sc_require('sc_manage_presales'): return redirect(url_for('sc_presales'))
    db  = get_db()
    row = db.execute('SELECT * FROM sc_presales_requests WHERE id=?', (rid,)).fetchone()
    if not row: abort(404)
    calon_customers = db.execute("SELECT id, name, code FROM sc_customers WHERE is_active=1 ORDER BY name").fetchall()
    employees = db.execute("SELECT id, name, jabatan, divisi FROM employees WHERE is_active=1 ORDER BY divisi, name").fetchall()
    sel_assignees = [r['employee_id'] for r in
                     db.execute('SELECT employee_id FROM sc_presales_assignees WHERE request_id=?', (rid,)).fetchall()]
    if request.method == 'POST':
        new_status   = request.form.get('status', row['status'])
        status_note  = request.form.get('status_note','').strip()
        subject      = request.form.get('subject','').strip()
        description  = request.form.get('description','').strip()
        assignee_ids = request.form.getlist('assignee_ids')
        old_status   = row['status']
        db.execute('UPDATE sc_presales_requests SET status=?,status_note=?,subject=?,description=? WHERE id=?',
                   (new_status, status_note, subject, description, rid))
        db.commit()
        # Sync assignees
        old_ids = set(sel_assignees)
        new_ids = set(int(i) for i in assignee_ids if i)
        for eid in new_ids - old_ids:
            emp = db.execute('SELECT divisi FROM employees WHERE id=?', (eid,)).fetchone()
            db.execute('INSERT OR IGNORE INTO sc_presales_assignees(request_id,employee_id,divisi) VALUES(?,?,?)',
                       (rid, eid, emp['divisi'] if emp else ''))
        for eid in old_ids - new_ids:
            db.execute('DELETE FROM sc_presales_assignees WHERE request_id=? AND employee_id=?', (rid, eid))
        db.commit()
        note_parts = []
        if old_status != new_status: note_parts.append(f'Status: {old_status} → {new_status}')
        if status_note: note_parts.append(status_note)
        db.execute('''INSERT INTO sc_presales_history(request_id,changed_by,changed_by_name,action,notes)
                      VALUES(?,?,?,?,?)''',
                   (rid, session.get('user_id'), session.get('user_name',''),
                    'status_change' if old_status != new_status else 'update',
                    '; '.join(note_parts) or 'Diperbarui'))
        db.commit()
        flash('Request diperbarui', 'success')
        return redirect(url_for('sc_presales_detail', rid=rid))
    return render_template('sc_presales_form.html', row=row, sel_assignees=sel_assignees,
                           calon_customers=calon_customers, employees=employees,
                           sc_presales_statuses=SC_PRESALES_STATUSES)

@app.route('/support/presales/<int:rid>')
@login_required
def sc_presales_detail(rid):
    if not sc_require('sc_view'): return redirect(url_for('sc_index'))
    db  = get_db()
    row = db.execute('''SELECT r.*, cu.name as customer_name, cu.customer_type
                        FROM sc_presales_requests r
                        JOIN sc_customers cu ON cu.id=r.customer_id
                        WHERE r.id=?''', (rid,)).fetchone()
    if not row: abort(404)
    history   = db.execute('SELECT * FROM sc_presales_history WHERE request_id=? ORDER BY id DESC', (rid,)).fetchall()
    assignees = db.execute('''SELECT e.name, e.jabatan, e.divisi, pa.role_note
                              FROM sc_presales_assignees pa
                              JOIN employees e ON e.id=pa.employee_id
                              WHERE pa.request_id=? ORDER BY e.divisi, e.name''', (rid,)).fetchall()
    return render_template('sc_presales_detail.html', row=row, history=history, assignees=assignees,
                           sc_presales_statuses=SC_PRESALES_STATUSES)

@app.route('/support/presales/<int:rid>/delete', methods=['POST'])
@login_required
def sc_presales_delete(rid):
    if not sc_require('sc_manage_presales'): return redirect(url_for('sc_presales'))
    db = get_db()
    db.execute('DELETE FROM sc_presales_requests WHERE id=?', (rid,))
    db.commit()
    flash('Request dihapus', 'warning')
    return redirect(url_for('sc_presales'))

# ─── SupportCore: Reports ──────────────────────────────────────────────────────

@app.route('/support/reports')
@login_required
def sc_reports():
    if not sc_require('sc_view_reports'): return redirect(url_for('sc_index'))
    db = get_db()
    from datetime import date as _date
    today       = _date.today()
    date_from   = request.args.get('date_from', today.replace(day=1).strftime('%Y-%m-%d'))
    date_to     = request.args.get('date_to',   today.strftime('%Y-%m-%d'))
    cust_f      = request.args.get('customer_id', '')
    creator_f   = request.args.get('created_by',  '')
    pic_f       = request.args.get('pic_id',      '')
    assignee_f  = request.args.get('assignee_id', '')

    q_base = '''FROM sc_tickets t
                JOIN sc_customers cu ON cu.id=t.customer_id
                JOIN sc_support_types st ON st.id=t.support_type_id
                LEFT JOIN sc_sla_categories sc ON sc.id=t.sla_category_id
                LEFT JOIN employees cr ON cr.id=t.created_by
                LEFT JOIN employees as_e ON as_e.id=t.assigned_to
                WHERE date(t.reported_at) BETWEEN ? AND ?'''
    params = [date_from, date_to]
    if cust_f:
        q_base += ' AND t.customer_id=?'; params.append(cust_f)
    if creator_f:
        q_base += ' AND t.created_by=?'; params.append(creator_f)
    if pic_f:
        # filter tiket yang customernya memiliki PIC = pic_f
        q_base += (' AND cu.id IN ('
                   'SELECT id FROM sc_customers WHERE '
                   'pic_helpdesk_id=? OR pic_helpdesk_backup_id=? OR '
                   'pic_implementor_id=? OR pic_coleader_id=? OR pic_sales_id=?)')
        params.extend([pic_f]*5)
    if assignee_f:
        q_base += (' AND (t.assigned_to=? OR t.id IN ('
                   'SELECT ticket_id FROM sc_ticket_assignees WHERE employee_id=?))')
        params.extend([assignee_f, assignee_f])

    tickets = db.execute(
        'SELECT t.*, cu.name as customer_name, st.name as type_name, '
        'sc.response_time_hours, sc.resolution_time_hours, '
        'cr.name as creator_name, as_e.name as assignee_name ' + q_base
        + ' ORDER BY t.reported_at', params).fetchall()

    # Monthly summary
    monthly = {}
    for t in tickets:
        mo = t['reported_at'][:7]
        if mo not in monthly:
            monthly[mo] = {'total':0,'open':0,'closed':0,'corrective':0,'preventive':0,'onsite':0}
        monthly[mo]['total'] += 1
        if t['status'] in ('open','in_progress'): monthly[mo]['open'] += 1
        else: monthly[mo]['closed'] += 1
        tn = (t['type_name'] or '').lower()
        if 'corrective' in tn: monthly[mo]['corrective'] += 1
        elif 'preventive' in tn: monthly[mo]['preventive'] += 1
        elif 'onsite' in tn: monthly[mo]['onsite'] += 1
    filter_year = date_from[:4]  # kept for template compat in card title

    # Avg response & resolution time
    resp_times, res_times = [], []
    for t in tickets:
        _, rh = calc_sla(t['reported_at'], t['responded_at'], 999)
        if rh is not None: resp_times.append(rh)
        _, reh = calc_sla(t['reported_at'], t['resolved_at'], 999)
        if reh is not None: res_times.append(reh)
    avg_resp = round(sum(resp_times)/len(resp_times), 1) if resp_times else None
    avg_res  = round(sum(res_times)/len(res_times), 1)  if res_times  else None

    # Top customers
    cust_count = {}
    for t in tickets:
        cn = t['customer_name']
        cust_count[cn] = cust_count.get(cn, 0) + 1
    top_customers = sorted(cust_count.items(), key=lambda x: -x[1])[:10]

    customers = db.execute('SELECT id, name FROM sc_customers WHERE is_active=1 ORDER BY name').fetchall()
    employees = db.execute('SELECT id, name, divisi FROM employees WHERE is_active=1 ORDER BY name').fetchall()

    return render_template('sc_reports.html', monthly=monthly, avg_resp=avg_resp,
                           avg_res=avg_res, top_customers=top_customers,
                           customers=customers, employees=employees,
                           filter_date_from=date_from, filter_date_to=date_to,
                           filter_year=filter_year,
                           filter_customer=cust_f, filter_creator=creator_f,
                           filter_pic=pic_f, filter_assignee=assignee_f,
                           total_tickets=len(tickets))

# ─── Settings ─────────────────────────────────────────────────────────────────

TALENTCORE_SETTINGS_KEYS = [
    'app_name', 'reminder_days', 'reminder_enabled',
    'notification_emails', 'notification_telegram_ids', 'openwa_extra_phones',
]

@app.route('/settings', methods=['GET', 'POST'])
@login_required
def settings():
    db = get_db()
    if not has_permission(session.get('user_role', ''), 'manage_settings', db):
        flash('Akses ditolak.', 'danger')
        return redirect(url_for('index'))
    if request.method == 'POST':
        for k in TALENTCORE_SETTINGS_KEYS:
            v = request.form.get(k, '').strip()
            if k == 'reminder_enabled':
                v = '1' if request.form.get('reminder_enabled') else '0'
            save_setting(db, k, v)
        db.commit()
        flash('Pengaturan TalentCore disimpan', 'success')
        return redirect(url_for('settings'))
    cfg = get_settings(db)
    return render_template('settings.html', cfg=cfg)

@app.route('/settings/test-email', methods=['POST'])
@superadmin_required
def test_email():
    db = get_db()
    settings_data = get_settings(db)
    to_email = request.form.get('test_email','').strip()
    if not to_email:
        return jsonify({'ok': False, 'msg': 'Masukkan alamat email tujuan test'})
    ok, err = send_email(settings_data, to_email,
                         'Test Email — TalentCore',
                         '<h3>Test berhasil!</h3><p>Konfigurasi email sudah benar.</p>')
    return jsonify({'ok': ok, 'msg': 'Email berhasil dikirim' if ok else str(err)})

@app.route('/settings/test-telegram', methods=['POST'])
@superadmin_required
def test_telegram():
    db = get_db()
    cfg = get_settings(db)
    bot_token = cfg.get('telegram_bot_token','').strip()
    chat_id   = request.form.get('test_chat_id','').strip() or cfg.get('telegram_default_chat_id','').strip()
    if not bot_token or not chat_id:
        return jsonify({'ok': False, 'msg': 'Bot token dan chat ID harus diisi'})
    ok, err = send_telegram(bot_token, chat_id,
                            '✅ <b>Test berhasil!</b>\n\nKonfigurasi Telegram sudah benar.')
    return jsonify({'ok': ok, 'msg': 'Pesan Telegram berhasil dikirim' if ok else str(err)})

@app.route('/settings/test-whatsapp', methods=['POST'])
@superadmin_required
def test_whatsapp():
    db  = get_db()
    cfg = get_settings(db)
    wa_url     = request.form.get('test_wa_url', '').strip() or cfg.get('openwa_url', '').strip()
    wa_key     = cfg.get('openwa_api_key', '').strip()
    wa_session = get_openwa_session(cfg, 'evaluasi')
    phone      = request.form.get('test_wa_phone', '').strip()
    if not wa_url or not phone:
        return jsonify({'ok': False, 'msg': 'URL OpenWA dan nomor HP harus diisi'})
    ok, err = send_whatsapp(wa_url, wa_key, wa_session, phone,
                            '✅ *Test berhasil!*\n\nKonfigurasi OpenWA WhatsApp sudah terhubung dengan TalentCore.')
    chat_id = normalize_phone_wa(phone)
    return jsonify({'ok': ok, 'chat_id': chat_id,
                    'msg': f'Pesan terkirim ke {chat_id}' if ok else str(err)})

@app.route('/settings/run-reminders', methods=['POST'])
@login_required
def run_reminders_now():
    db = get_db()
    if not has_permission(session.get('user_role', ''), 'manage_settings', db):
        return jsonify({'ok': False, 'msg': 'Akses ditolak'})
    sent, failed = run_contract_reminders(triggered_by=session.get('username','manual'))
    return jsonify({'sent': sent, 'failed': failed,
                    'msg': f'{sent} reminder dikirim, {failed} gagal'})

# ─── Evaluation Routes (unchanged logic, login_required added) ─────────────────

@app.route('/eval/new/<int:emp_id>', methods=['POST'])
@login_required
def eval_new(emp_id):
    db = get_db()
    periode = request.form.get('periode', str(date.today().year)).strip()
    from datetime import date as _d, timedelta as _td
    today = _d.today()
    date_to = today.strftime('%Y-%m-%d')
    date_from = (today - _td(days=90)).strftime('%Y-%m-%d')
    cur = db.execute('INSERT INTO evaluations(employee_id, periode, task_date_from, task_date_to) VALUES(?,?,?,?)',
                     (emp_id, periode, date_from, date_to))
    eval_id = cur.lastrowid
    for slot in range(1, 6):
        db.execute('INSERT INTO peer_reviews(eval_id, slot) VALUES(?,?)', (eval_id, slot))
    db.commit()

    emp = db.execute('SELECT * FROM employees WHERE id=?', (emp_id,)).fetchone()
    if emp and (emp['email'] or emp['telegram_id']):
        results, any_ok, link = _send_self_assessment(
            db, eval_id, emp, periode,
            request.host_url.rstrip('/'),
            triggered_by=session.get('username', 'auto-new')
        )
        db.commit()
        if any_ok:
            flash(f'Evaluasi dibuat & link self-assessment dikirim — ' + ' | '.join(results), 'success')
        else:
            flash(f'Evaluasi dibuat. Gagal kirim notifikasi ({"; ".join(results)}). '
                  f'Kirim manual dari halaman Summary.', 'warning')
    else:
        flash('Evaluasi berhasil dibuat. Email/Telegram karyawan belum diisi — '
              'kirim link self-assessment manual dari halaman Summary.', 'info')

    return redirect(url_for('eval_project', eval_id=eval_id))

@app.route('/emp/<int:emp_id>/trigger-review', methods=['POST'])
@login_required
def emp_trigger_review(emp_id):
    """Buat atau temukan evaluasi untuk periode yang dipilih, kirim link self-assessment
    ke karyawan, lalu kembali ke halaman Manajemen Karyawan.
    """
    db  = get_db()
    emp = db.execute('SELECT * FROM employees WHERE id=? AND is_active=1', (emp_id,)).fetchone()
    if not emp:
        flash('Karyawan tidak ditemukan.', 'danger')
        return redirect(url_for('karyawan'))

    periode = request.form.get('periode', str(date.today().year)).strip()
    if not periode:
        flash('Periode tidak boleh kosong.', 'danger')
        return redirect(url_for('karyawan'))

    # Cari evaluasi yang sudah ada untuk periode ini (status bukan approved/closed)
    existing = db.execute(
        "SELECT * FROM evaluations WHERE employee_id=? AND periode=? AND status NOT IN ('approved','closed')",
        (emp_id, periode)
    ).fetchone()

    if existing:
        eval_id      = existing['id']
        created_new  = False
    else:
        from datetime import date as _d, timedelta as _td
        today = _d.today()
        date_to = today.strftime('%Y-%m-%d')
        date_from = (today - _td(days=90)).strftime('%Y-%m-%d')
        cur      = db.execute('INSERT INTO evaluations(employee_id, periode, task_date_from, task_date_to) VALUES(?,?,?,?)',
                              (emp_id, periode, date_from, date_to))
        eval_id  = cur.lastrowid
        for slot in range(1, 6):
            db.execute('INSERT INTO peer_reviews(eval_id, slot) VALUES(?,?)', (eval_id, slot))
        db.commit()
        created_new = True

    if not (emp['email'] or emp['telegram_id']):
        label = 'dibuat baru' if created_new else 'sudah ada'
        flash(
            f'Evaluasi periode {periode} untuk {emp["name"]} {label} (ID #{eval_id}). '
            'Email & Telegram karyawan belum diisi — kirim link self-assessment manual dari halaman Summary.',
            'info' if created_new else 'warning'
        )
        return redirect(url_for('karyawan'))

    results, any_ok, link = _send_self_assessment(
        db, eval_id, emp, periode,
        request.host_url.rstrip('/'),
        triggered_by=session.get('username', 'manual')
    )
    db.commit()
    prefix = f'Evaluasi baru (#{eval_id}) dibuat & ' if created_new else f'Evaluasi #{eval_id} ada & '
    if any_ok:
        flash(prefix + 'link self-assessment dikirim — ' + ' | '.join(results), 'success')
    else:
        flash(
            prefix + 'gagal mengirim notifikasi (' + '; '.join(results) + '). '
            f'Link manual: {link}',
            'warning'
        )
    return redirect(url_for('karyawan'))

@app.route('/eval/<int:eval_id>/project', methods=['GET', 'POST'])
@login_required
def eval_project(eval_id):
    db = get_db()
    ev = get_eval_or_404(db, eval_id)
    if not ev:
        flash('Evaluasi tidak ditemukan', 'danger')
        return redirect(url_for('index'))
    if request.method == 'POST':
        try:
            db.execute('DELETE FROM project_entries WHERE eval_id=?', (eval_id,))
            for t in ['history','top_task','improvement']:
                names    = request.form.getlist(f'{t}_project_name')
                tasks    = request.form.getlist(f'{t}_detail_task')
                statuses = request.form.getlist(f'{t}_status')
                notes_l  = request.form.getlist(f'{t}_notes')
                for i, name in enumerate(names):
                    if name.strip() or (i < len(tasks) and tasks[i].strip()):
                        db.execute('''INSERT INTO project_entries
                                      (eval_id,entry_type,project_name,detail_task,status,notes,sort_order)
                                      VALUES(?,?,?,?,?,?,?)''',
                                   (eval_id, t, name.strip(),
                                    tasks[i] if i < len(tasks) else '',
                                    statuses[i] if i < len(statuses) else 'DONE',
                                    notes_l[i] if i < len(notes_l) else '', i))
            pp_score = int(request.form.get('pp_score', 0))
            db.execute('UPDATE evaluations SET pp_score=? WHERE id=?', (pp_score, eval_id))
            db.commit()
        except Exception:
            db.execute('ROLLBACK')
            flash('Terjadi kesalahan saat menyimpan data project.', 'danger')
            return redirect(url_for('eval_project', eval_id=eval_id))
        recalc(db, eval_id)
        if request.form.get('action') == 'next':
            return redirect(url_for('eval_hardskill', eval_id=eval_id))
        flash('Project Performance disimpan', 'success')
        return redirect(url_for('eval_project', eval_id=eval_id))
    entries = {t: db.execute(
        'SELECT * FROM project_entries WHERE eval_id=? AND entry_type=? ORDER BY sort_order',
        (eval_id, t)).fetchall() for t in ['history','top_task','improvement']}
    # ── Auto-suggest dari sistem: task PC + ticket SC + POC ──
    sys_projects = _get_sys_project_data(db, ev)
    return render_template('eval_project.html', ev=ev, entries=entries,
                           sys_projects=sys_projects)


def _get_sys_project_data(db, ev):
    """Ambil data project/task/tiket/POC dari sistem untuk karyawan & periode evaluasi."""
    emp_id    = ev['employee_id']
    date_from = ev['task_date_from'] or ''
    date_to   = ev['task_date_to']   or ''

    # Filter tanggal (jika ada)
    def _date_filter(col_start, col_end=''):
        if date_from and date_to:
            return f" AND ({col_start} IS NULL OR {col_start} <= '{date_to}')"
        return ''

    # 1. ProjectCore — proyek yang diikuti (sebagai member / task assignee)
    pc_projs = db.execute('''
        SELECT DISTINCT p.name AS proj_name, p.status, p.start_date, p.end_date,
               pm.role AS member_role,
               COUNT(t.id) FILTER (WHERE ta2.employee_id IS NOT NULL) AS task_count,
               COUNT(t.id) FILTER (WHERE ta2.employee_id IS NOT NULL AND t.status='done') AS done_count,
               COUNT(t.id) FILTER (WHERE ta2.employee_id IS NOT NULL AND t.status NOT IN ('done','cancelled')
                   AND t.due_date IS NOT NULL AND t.due_date <> '' AND NULLIF(t.due_date,'')::date < CURRENT_DATE) AS overtime_count
        FROM pc_projects p
        LEFT JOIN pc_members pm ON pm.project_id=p.id AND pm.employee_id=?
        LEFT JOIN pc_tasks t ON t.project_id=p.id
        LEFT JOIN pc_task_assignees ta2 ON ta2.task_id=t.id AND ta2.employee_id=?
        WHERE (pm.employee_id=? OR ta2.employee_id=?)
        GROUP BY p.id, p.name, p.status, p.start_date, p.end_date, pm.role
        ORDER BY p.start_date DESC NULLS LAST
    ''', (emp_id, emp_id, emp_id, emp_id)).fetchall()

    # 2. SupportCore — tiket yang ditangani
    sc_tickets = db.execute('''
        SELECT COUNT(*) AS total,
               COUNT(*) FILTER (WHERE t.status='resolved') AS resolved,
               COUNT(*) FILTER (WHERE t.status='open') AS open_count,
               COUNT(DISTINCT t.customer_id) AS customer_count,
               STRING_AGG(DISTINCT c.name, ', ') AS customer_names
        FROM sc_ticket_assignees ta
        JOIN sc_tickets t ON t.id=ta.ticket_id
        LEFT JOIN sc_customers c ON c.id=t.customer_id
        WHERE ta.employee_id=?
    ''', (emp_id,)).fetchone()

    # 3. SupportCore — contract/customer list
    sc_customers = db.execute('''
        SELECT DISTINCT c.name AS cust_name, COUNT(DISTINCT t.id) AS ticket_count
        FROM sc_ticket_assignees ta
        JOIN sc_tickets t ON t.id=ta.ticket_id
        JOIN sc_customers c ON c.id=t.customer_id
        WHERE ta.employee_id=?
        GROUP BY c.id, c.name ORDER BY ticket_count DESC LIMIT 10
    ''', (emp_id,)).fetchall()

    # 4. POC / Presales
    poc_rows = db.execute('''
        SELECT r.subject AS title, r.status, c.name AS customer_name, r.created_at AS start_date
        FROM sc_presales_assignees pa
        JOIN sc_presales_requests r ON r.id=pa.request_id
        LEFT JOIN sc_customers c ON c.id=r.customer_id
        WHERE pa.employee_id=?
        ORDER BY r.created_at DESC NULLS LAST
    ''', (emp_id,)).fetchall()

    # 5. Task summary per difficulty
    task_diff = db.execute('''
        SELECT t.difficulty, t.status, COUNT(*) AS cnt
        FROM pc_task_assignees ta
        JOIN pc_tasks t ON t.id=ta.task_id
        WHERE ta.employee_id=?
        GROUP BY t.difficulty, t.status
    ''', (emp_id,)).fetchall()

    return {
        'pc_projs':    [dict(r) for r in pc_projs],
        'sc_tickets':  dict(sc_tickets) if sc_tickets else {},
        'sc_customers':[dict(r) for r in sc_customers],
        'poc_rows':    [dict(r) for r in poc_rows],
        'task_diff':   [dict(r) for r in task_diff],
    }


@app.route('/eval/<int:eval_id>/project/autofill', methods=['POST'])
@login_required
def eval_project_autofill(eval_id):
    """Auto-isi project_entries dari data sistem (PC + SC + POC)."""
    db    = get_db()
    ev    = get_eval_or_404(db, eval_id)
    emp_id = ev['employee_id']
    if not ev:
        return jsonify({'ok': False, 'msg': 'Evaluasi tidak ditemukan'})

    def _proj_status(s):
        s = (s or '').lower()
        if s in ('completed', 'done', 'selesai'):   return 'DONE'
        if s in ('cancelled', 'canceled', 'batal'): return 'CANCELLED'
        return 'ONPROGRESS'

    db.execute('DELETE FROM project_entries WHERE eval_id=?', (eval_id,))
    order = 0

    # ── HISTORY: semua proyek PC yang diikuti, bawa status proyek ──────────────
    pc_projs = db.execute('''
        SELECT DISTINCT p.id, p.name AS proj_name, p.status AS proj_status,
               pm.role AS member_role,
               COUNT(ta2.id) AS task_count,
               COUNT(ta2.id) FILTER (WHERE t.status=\'done\') AS done_count,
               COUNT(ta2.id) FILTER (WHERE t.status NOT IN (\'done\',\'cancelled\')
                   AND t.due_date IS NOT NULL AND t.due_date <> \'\'
                   AND NULLIF(t.due_date,\'\')::date < CURRENT_DATE) AS overtime_count
        FROM pc_projects p
        LEFT JOIN pc_members pm ON pm.project_id=p.id AND pm.employee_id=?
        LEFT JOIN pc_tasks t ON t.project_id=p.id
        LEFT JOIN pc_task_assignees ta2 ON ta2.task_id=t.id AND ta2.employee_id=?
        WHERE (pm.employee_id=? OR ta2.employee_id=?) AND p.deleted_at IS NULL
        GROUP BY p.id, p.name, p.status, pm.role
        ORDER BY task_count DESC
    ''', (emp_id, emp_id, emp_id, emp_id)).fetchall()

    for p in pc_projs:
        role   = p['member_role'] or 'Anggota Tim'
        detail = (f"Role: {role}. "
                  f"Task: {p['task_count']} total, {p['done_count']} selesai"
                  + (f", {p['overtime_count']} melewati deadline" if p['overtime_count'] else '') + ".")
        db.execute('''INSERT INTO project_entries
                      (eval_id,entry_type,project_name,detail_task,status,sort_order)
                      VALUES(?,?,?,?,?,?)''',
                   (eval_id, 'history', p['proj_name'], detail,
                    _proj_status(p['proj_status']), order))
        order += 1

    # SC tiket per customer — masuk history juga sebagai support project
    sc_per_cust = db.execute('''
        SELECT c.name AS cust_name,
               COUNT(t.id) AS total,
               COUNT(t.id) FILTER (WHERE t.status=\'resolved\') AS resolved,
               COUNT(t.id) FILTER (WHERE t.status=\'open\') AS open_count
        FROM sc_ticket_assignees ta
        JOIN sc_tickets t ON t.id=ta.ticket_id
        JOIN sc_customers c ON c.id=t.customer_id
        WHERE ta.employee_id=?
        GROUP BY c.id, c.name ORDER BY total DESC
    ''', (emp_id,)).fetchall()

    for sc in sc_per_cust:
        detail = (f"Handle {sc['total']} tiket support "
                  f"({sc['resolved']} resolved, {sc['open_count']} open).")
        st = 'DONE' if sc['open_count'] == 0 and sc['resolved'] > 0 else 'ONPROGRESS'
        db.execute('''INSERT INTO project_entries
                      (eval_id,entry_type,project_name,detail_task,status,sort_order)
                      VALUES(?,?,?,?,?,?)''',
                   (eval_id, 'history', f"Support — {sc['cust_name']}", detail, st, order))
        order += 1

    # ── TOP TASK: project + SC customer dengan task/tiket terbanyak ────────────
    # Gabung pc_projs dan sc_per_cust, urutkan by total, ambil top entry
    top_candidates = []
    for p in pc_projs:
        top_candidates.append({
            'name':   p['proj_name'],
            'total':  p['task_count'],
            'done':   p['done_count'],
            'src':    'project',
            'detail': (f"Proyek dengan {p['task_count']} task ({p['done_count']} selesai). "
                       f"Kontribusi terbesar sebagai {p['member_role'] or 'Anggota'}."),
            'status': _proj_status(p['proj_status']),
        })
    for sc in sc_per_cust:
        top_candidates.append({
            'name':   f"Support — {sc['cust_name']}",
            'total':  sc['total'],
            'done':   sc['resolved'],
            'src':    'sc',
            'detail': (f"Customer dengan {sc['total']} tiket support ditangani "
                       f"({sc['resolved']} resolved). Pencapaian handling terbanyak."),
            'status': 'DONE' if sc['open_count'] == 0 else 'ONPROGRESS',
        })

    # Sort by total DESC, ambil top 3
    top_candidates.sort(key=lambda x: x['total'], reverse=True)
    for tc in top_candidates[:3]:
        db.execute('''INSERT INTO project_entries
                      (eval_id,entry_type,project_name,detail_task,status,sort_order)
                      VALUES(?,?,?,?,?,?)''',
                   (eval_id, 'top_task', tc['name'], tc['detail'], tc['status'], order))
        order += 1

    # ── IMPROVEMENT: POC/Presales (1 baris per POC) + 1 baris kosong manual ──
    poc_rows = db.execute('''
        SELECT r.subject, r.status, r.request_type, c.name AS cust_name, r.created_at
        FROM sc_presales_assignees pa
        JOIN sc_presales_requests r ON r.id=pa.request_id
        LEFT JOIN sc_customers c ON c.id=r.customer_id
        WHERE pa.employee_id=?
        ORDER BY r.created_at DESC NULLS LAST
    ''', (emp_id,)).fetchall()

    for poc in poc_rows:
        rtype  = (poc['request_type'] or 'POC').upper()
        detail = f"[{rtype}] Customer: {poc['cust_name'] or '-'}. Status: {poc['status'] or '-'}."
        st     = 'DONE' if (poc['status'] or '').lower() in ('completed','done','selesai') else 'ONPROGRESS'
        db.execute('''INSERT INTO project_entries
                      (eval_id,entry_type,project_name,detail_task,status,sort_order)
                      VALUES(?,?,?,?,?,?)''',
                   (eval_id, 'improvement', poc['subject'] or 'POC/Presales', detail, st, order))
        order += 1

    # Baris kosong untuk input manual (training, dll)
    for _ in range(2):
        db.execute('''INSERT INTO project_entries
                      (eval_id,entry_type,project_name,detail_task,status,sort_order)
                      VALUES(?,?,?,?,?,?)''',
                   (eval_id, 'improvement', '', '', 'ONPROGRESS', order))
        order += 1

    db.commit()
    return jsonify({'ok': True, 'inserted': order})


@app.route('/eval/<int:eval_id>/hardskill', methods=['GET', 'POST'])
@login_required
def eval_hardskill(eval_id):
    db = get_db()
    ev = get_eval_or_404(db, eval_id)
    if not ev: return redirect(url_for('index'))
    if request.method == 'POST':
        for key, val in request.form.items():
            if key.startswith('score_'):
                item_id = int(key.split('_')[1])
                score   = max(0, min(5, int(val or 0)))
                db.execute('''INSERT INTO skill_scores(eval_id,skill_item_id,score) VALUES(?,?,?)
                              ON CONFLICT(eval_id,skill_item_id) DO UPDATE SET score=excluded.score''',
                           (eval_id, item_id, score))
        db.commit(); recalc(db, eval_id)
        if request.form.get('action') == 'next':
            return redirect(url_for('eval_ability', eval_id=eval_id))
        flash('Hard Skill disimpan', 'success')
        return redirect(url_for('eval_hardskill', eval_id=eval_id))
    categories = db.execute('''SELECT sc.id, sc.name FROM skill_categories sc
                               WHERE sc.divisi=? ORDER BY sc.sort_order''', (ev['divisi'],)).fetchall()
    items_by_cat = {cat['id']: db.execute('''
        SELECT si.*, COALESCE(ss.score,0) AS current_score,
               ROUND(((COALESCE(ss.score,0)/4.0)*si.bobot)::numeric, 2) AS final_value
        FROM skill_items si LEFT JOIN skill_scores ss ON ss.skill_item_id=si.id AND ss.eval_id=?
        WHERE si.category_id=? ORDER BY si.sort_order
    ''', (eval_id, cat['id'])).fetchall() for cat in categories}
    hs_total = db.execute('SELECT hs_total FROM evaluations WHERE id=?', (eval_id,)).fetchone()['hs_total']
    return render_template('eval_hardskill.html', ev=ev, categories=categories,
                           items_by_cat=items_by_cat, hs_total=hs_total)

@app.route('/eval/<int:eval_id>/ability', methods=['GET', 'POST'])
@login_required
def eval_ability(eval_id):
    db = get_db()
    ev = get_eval_or_404(db, eval_id)
    if not ev: return redirect(url_for('index'))
    if request.method == 'POST':
        for key, val in request.form.items():
            if key.startswith('level_'):
                item_id = int(key.split('_')[1])
                level   = val if val in ('A','B','C','D') else ''
                db.execute('''INSERT INTO ability_scores(eval_id,ability_item_id,level) VALUES(?,?,?)
                              ON CONFLICT(eval_id,ability_item_id) DO UPDATE SET level=excluded.level''',
                           (eval_id, item_id, level))
        db.commit(); recalc(db, eval_id)
        if request.form.get('action') == 'next':
            return redirect(url_for('eval_competency', eval_id=eval_id))
        flash('Ability disimpan', 'success')
        return redirect(url_for('eval_ability', eval_id=eval_id))
    items    = db.execute('''SELECT ai.*, COALESCE(abs.level,'') AS current_level
                             FROM ability_items ai
                             LEFT JOIN ability_scores abs ON abs.ability_item_id=ai.id AND abs.eval_id=?
                             WHERE ai.divisi=? ORDER BY ai.sort_order''',
                          (eval_id, ev['divisi'])).fetchall()
    ev_row   = db.execute('SELECT ability_score FROM evaluations WHERE id=?', (eval_id,)).fetchone()
    # Data kontekstual untuk referensi evaluator
    sys_ctx = _ability_sys_context(db, ev)
    return render_template('eval_ability.html', ev=ev, items=items,
                           ability_score=ev_row['ability_score'], sys_ctx=sys_ctx)


def _ability_sys_context(db, ev):
    """Kumpulkan data kontekstual dari sistem untuk bantu penilaian ability."""
    emp_id    = ev['employee_id']
    date_from = ev['task_date_from'] or ''
    date_to   = ev['task_date_to']   or ''

    perf = calc_task_perf(db, emp_id, date_from, date_to)
    ana  = calc_task_analytics(db, emp_id, date_from, date_to)

    # Tiket SC
    sc = db.execute('''
        SELECT COUNT(*) AS total,
               COUNT(*) FILTER (WHERE t.status='resolved') AS resolved,
               COUNT(DISTINCT t.customer_id) AS customer_count
        FROM sc_ticket_assignees ta
        JOIN sc_tickets t ON t.id=ta.ticket_id
        WHERE ta.employee_id=?
    ''', (emp_id,)).fetchone()

    # POC
    poc_count = db.execute('''
        SELECT COUNT(*) AS c FROM sc_presales_assignees WHERE employee_id=?
    ''', (emp_id,)).fetchone()['c']

    # Difficulty breakdown
    diff_rows = db.execute('''
        SELECT COALESCE(t.difficulty,'Normal') AS difficulty, COUNT(*) AS cnt
        FROM pc_task_assignees ta JOIN pc_tasks t ON t.id=ta.task_id
        WHERE ta.employee_id=?
        GROUP BY t.difficulty
    ''', (emp_id,)).fetchall()
    diff_map = {r['difficulty']: r['cnt'] for r in diff_rows}

    # Self-assigned (inisiatif)
    self_assigned = db.execute('''
        SELECT COUNT(*) AS c FROM pc_task_assignees WHERE employee_id=? AND self_assigned=1
    ''', (emp_id,)).fetchone()['c']

    return {
        'task_score':    perf.get('task_score', 0),
        'ontime_rate':   ana.get('ontime_rate'),
        'total_done':    ana.get('total_done', 0),
        'open_overtime': ana.get('open_overtime', 0),
        'concurrent_max':ana.get('concurrent_max', 0),
        'sc_total':      dict(sc).get('total', 0) if sc else 0,
        'sc_resolved':   dict(sc).get('resolved', 0) if sc else 0,
        'sc_customers':  dict(sc).get('customer_count', 0) if sc else 0,
        'poc_count':     poc_count,
        'diff_hard':     diff_map.get('Hard', diff_map.get('hard', diff_map.get('Sulit', 0))),
        'diff_normal':   diff_map.get('Normal', diff_map.get('normal', 0)),
        'diff_easy':     diff_map.get('Easy', diff_map.get('easy', diff_map.get('Mudah', 0))),
        'self_assigned': self_assigned,
    }


@app.route('/eval/<int:eval_id>/ability/ai-suggest', methods=['POST'])
@login_required
def eval_ability_ai_suggest(eval_id):
    """Analisa ability per item menggunakan AI berdasarkan data sistem."""
    db  = get_db()
    ev  = get_eval_or_404(db, eval_id)
    emp = db.execute('SELECT * FROM employees WHERE id=?', (ev['employee_id'],)).fetchone()
    if not ev:
        return jsonify({'ok': False, 'msg': 'Evaluasi tidak ditemukan'})

    items = db.execute('''SELECT ai.* FROM ability_items ai
                          WHERE ai.divisi=? ORDER BY ai.sort_order''',
                       (ev['divisi'],)).fetchall()
    ctx   = _ability_sys_context(db, ev)

    prompt = f"""Kamu adalah sistem evaluasi kinerja AI yang objektif.
Karyawan: {emp['name']} | Jabatan: {emp['jabatan'] or '-'} | Divisi: {ev['divisi']}
Periode Evaluasi: {ev['periode']}

DATA KINERJA SISTEM:
- Task Score: {ctx['task_score']} pts
- Ontime Rate: {ctx['ontime_rate']}%
- Task Selesai: {ctx['total_done']} | Open Overtime: {ctx['open_overtime']}
- Max Concurrent Task: {ctx['concurrent_max']}
- Task Sulit (Hard): {ctx['diff_hard']} | Normal: {ctx['diff_normal']} | Mudah: {ctx['diff_easy']}
- Inisiatif (self-assigned): {ctx['self_assigned']} task
- Tiket Support: {ctx['sc_total']} ({ctx['sc_resolved']} resolved) | Customer: {ctx['sc_customers']}
- POC/Presales: {ctx['poc_count']}

DEFINISI LEVEL:
A = Sangat Baik / Outstanding (melampaui ekspektasi, konsisten top performer)
B = Baik / Good (memenuhi dan kadang melampaui ekspektasi)
C = Cukup / Average (memenuhi ekspektasi dasar)
D = Perlu Peningkatan / Needs Improvement (di bawah ekspektasi)

Untuk setiap item ability berikut, tentukan level (A/B/C/D) dan penjelasan singkat (1 kalimat) berdasarkan data di atas.
Format response HANYA JSON array:
[{{"id": <id>, "level": "<A|B|C|D>", "reason": "<1 kalimat>"}}]

DAFTAR ABILITY ITEMS:
"""
    for item in items:
        prompt += f'\n- ID {item["id"]}: {item["name"]}'
        prompt += f'\n  A: {item["desc_a"]} | B: {item["desc_b"]} | C: {item["desc_c"]} | D: {item["desc_d"]}'

    # Coba Ollama dulu, fallback ke rule-based
    result = _call_ollama_or_fallback(prompt, items, ctx)
    return jsonify({'ok': True, 'suggestions': result})


def _call_ollama_or_fallback(prompt, items, ctx):
    """Panggil Ollama untuk AI suggest, fallback ke rule-based jika gagal."""
    # Coba Ollama
    try:
        import requests as _req
        resp = _req.post('http://localhost:11434/api/generate',
                         json={'model': 'llama3', 'prompt': prompt, 'stream': False},
                         timeout=30)
        if resp.ok:
            import json as _json, re as _re
            text = resp.json().get('response', '')
            match = _re.search(r'\[.*?\]', text, _re.DOTALL)
            if match:
                return _json.loads(match.group(0))
    except Exception:
        pass

    # Rule-based fallback berdasarkan data sistem
    score     = ctx.get('task_score', 0)
    ontime    = ctx.get('ontime_rate') or 0
    overtime  = ctx.get('open_overtime', 0)
    diff_hard = ctx.get('diff_hard', 0)
    self_init = ctx.get('self_assigned', 0)

    if score >= 80 and ontime >= 80 and overtime == 0:
        default_level = 'A'
    elif score >= 60 and ontime >= 65:
        default_level = 'B'
    elif score >= 40 and ontime >= 40:
        default_level = 'C'
    else:
        default_level = 'D'

    suggestions = []
    for item in items:
        name = (item['name'] or '').lower()
        # Penyesuaian per tipe ability
        if 'mandiri' in name or 'inisiatif' in name:
            lvl = 'A' if self_init >= 3 else ('B' if self_init >= 1 else default_level)
            reason = f"Self-assigned {self_init} task menunjukkan tingkat inisiatif {'tinggi' if self_init >= 3 else 'cukup'}."
        elif 'tekno' in name or 'tools' in name or 'teknis' in name:
            lvl = 'A' if diff_hard >= 3 else ('B' if diff_hard >= 1 else default_level)
            reason = f"Menangani {diff_hard} task Hard menunjukkan penguasaan teknis {'solid' if diff_hard >= 3 else 'cukup'}."
        elif 'customer' in name or 'pelanggan' in name:
            n_cust = ctx.get('sc_customers', 0)
            lvl = 'A' if n_cust >= 3 else ('B' if n_cust >= 1 else 'C')
            reason = f"Menangani {n_cust} customer dalam periode ini."
        elif 'proses' in name or 'improve' in name or 'kontribusi' in name:
            lvl = 'A' if score >= 85 else ('B' if score >= 65 else default_level)
            reason = f"Task score {score} mencerminkan kontribusi pada peningkatan kualitas proses."
        else:
            lvl    = default_level
            reason = f"Berdasarkan task score {score} dan ontime rate {ontime}%."

        suggestions.append({'id': item['id'], 'level': lvl, 'reason': reason})
    return suggestions


@app.route('/eval/<int:eval_id>/competency', methods=['GET', 'POST'])
@login_required
def eval_competency(eval_id):
    db = get_db()
    ev = get_eval_or_404(db, eval_id)
    if not ev: return redirect(url_for('index'))
    if request.method == 'POST':
        for key, val in request.form.items():
            if key.startswith('rating_'):
                item_id = int(key.split('_')[1])
                rating  = max(0, min(5, int(val or 0)))
                db.execute('''INSERT INTO competency_scores(eval_id,competency_item_id,rating) VALUES(?,?,?)
                              ON CONFLICT(eval_id,competency_item_id) DO UPDATE SET rating=excluded.rating''',
                           (eval_id, item_id, rating))
        db.commit(); recalc(db, eval_id)
        if request.form.get('action') == 'next':
            return redirect(url_for('eval_summary', eval_id=eval_id))
        flash('Competency disimpan', 'success')
        return redirect(url_for('eval_competency', eval_id=eval_id))
    ev_row   = db.execute('SELECT hs_total, competency_total FROM evaluations WHERE id=?', (eval_id,)).fetchone()
    hs_total = ev_row['hs_total']
    items    = db.execute('''SELECT ci.*, COALESCE(cs.rating,0) AS current_rating
                             FROM competency_items ci
                             LEFT JOIN competency_scores cs ON cs.competency_item_id=ci.id AND cs.eval_id=?
                             WHERE ci.divisi=? ORDER BY ci.sort_order''',
                          (eval_id, ev['divisi'])).fetchall()
    items_with_val = [{'item': i, 'val': round(hs_total*i['bobot'],2) if i['is_hardskill']
                       else round(i['current_rating']*20.0*i['bobot'],2)} for i in items]
    return render_template('eval_competency.html', ev=ev, items_with_val=items_with_val,
                           hs_total=hs_total, competency_total=ev_row['competency_total'])

@app.route('/eval/<int:eval_id>/summary', methods=['GET', 'POST'])
@login_required
def eval_summary(eval_id):
    db = get_db()
    ev = get_eval_or_404(db, eval_id)
    if not ev: return redirect(url_for('index'))
    if request.method == 'POST':
        evaluator = request.form.get('evaluator','').strip()
        overall   = request.form.get('overall_assessment','').strip()
        plan      = request.form.get('development_plan','').strip()
        action    = request.form.get('action','save')
        status    = 'final' if action == 'finalize' else 'draft'
        db.execute('''UPDATE evaluations SET evaluator=?,overall_assessment=?,
                      development_plan=?,status=? WHERE id=?''',
                   (evaluator, overall, plan, status, eval_id))
        for slot in range(1, 6):
            def _int_or_none(v):
                try: return max(1, min(5, int(v))) if v and str(v).strip() else None
                except: return None
            db.execute('''UPDATE peer_reviews SET
                reviewer_name=?, feedback=?,
                dim_kerjasama=?, dim_komunikasi=?, dim_keandalan=?, dim_inisiatif=?, dim_kualitas=?
                WHERE eval_id=? AND slot=?''',
                (request.form.get(f'peer_name_{slot}','').strip(),
                 request.form.get(f'peer_feedback_{slot}','').strip(),
                 _int_or_none(request.form.get(f'peer_dim_kerjasama_{slot}')),
                 _int_or_none(request.form.get(f'peer_dim_komunikasi_{slot}')),
                 _int_or_none(request.form.get(f'peer_dim_keandalan_{slot}')),
                 _int_or_none(request.form.get(f'peer_dim_inisiatif_{slot}')),
                 _int_or_none(request.form.get(f'peer_dim_kualitas_{slot}')),
                 eval_id, slot))
        db.commit()
        if action == 'finalize':
            flash('Evaluasi berhasil difinalkan!', 'success')
            return redirect(url_for('index'))
        flash('Summary disimpan', 'success')
        return redirect(url_for('eval_summary', eval_id=eval_id))
    recalc(db, eval_id)
    ev    = get_eval_or_404(db, eval_id)
    peers = db.execute('SELECT * FROM peer_reviews WHERE eval_id=? ORDER BY slot', (eval_id,)).fetchall()
    entries = {t: db.execute(
        'SELECT * FROM project_entries WHERE eval_id=? AND entry_type=? ORDER BY sort_order',
        (eval_id, t)).fetchall() for t in ['history','top_task','improvement']}
    competency_items = db.execute('''SELECT ci.*, COALESCE(cs.rating,0) AS current_rating
                                     FROM competency_items ci
                                     LEFT JOIN competency_scores cs ON cs.competency_item_id=ci.id AND cs.eval_id=?
                                     WHERE ci.divisi=? ORDER BY ci.sort_order''',
                                  (eval_id, ev['divisi'])).fetchall()
    emp = db.execute('''SELECT e.*,
        e1.name AS sup_name, e2.name AS lead_name, e3.name AS mgr_name
        FROM employees e
        LEFT JOIN employees e1 ON e1.id = e.supervisor_id
        LEFT JOIN employees e2 ON e2.id = e.leader_id
        LEFT JOIN employees e3 ON e3.id = e.manager_id
        WHERE e.id=?''', (ev['employee_id'],)).fetchone()
    eval_token = db.execute('SELECT * FROM eval_tokens WHERE eval_id=?', (eval_id,)).fetchone()
    all_reviews = db.execute('''SELECT er.*, u.full_name, u.username
        FROM eval_reviews er JOIN users u ON u.id = er.reviewer_user_id
        WHERE er.eval_id=? ORDER BY er.submitted_at''', (eval_id,)).fetchall()
    base_url = request.host_url.rstrip('/')
    self_link = f"{base_url}/assess/{eval_token['token']}" if eval_token else None
    accumulated = calc_accumulated_metrics(db, ev, emp)
    import json as _json
    self_data = {}
    if ev.get('self_assessment_json'):
        try:
            self_data = _json.loads(ev['self_assessment_json'])
        except:
            pass
    return render_template('eval_summary.html', ev=ev, peers=peers, entries=entries,
                           competency_items=competency_items, emp=emp,
                           eval_token=eval_token, self_link=self_link,
                           all_reviews=all_reviews, accumulated=accumulated,
                           self_data=self_data)

@app.route('/eval/<int:eval_id>/delete', methods=['POST'])
@superadmin_required
def eval_delete(eval_id):
    db = get_db()
    db.execute('DELETE FROM evaluations WHERE id=?', (eval_id,))
    db.commit()
    flash('Evaluasi dihapus', 'warning')
    return redirect(url_for('index'))

# ─── AI Evaluation Summary (Ollama / OpenAI-compat) ──────────────────────────

def _build_eval_ai_prompt(db, eval_id):
    """Kumpulkan semua data evaluasi sebagai teks terstruktur untuk prompt AI."""
    import json as _json
    ev  = db.execute('SELECT * FROM evaluations WHERE id=?', (eval_id,)).fetchone()
    if not ev:
        return None, None
    emp = db.execute('SELECT * FROM employees WHERE id=?', (ev['employee_id'],)).fetchone()
    if not emp:
        return None, None

    # Task analytics
    perf = calc_task_perf(db, emp['id'], ev['task_date_from'] or '', ev['task_date_to'] or '')
    ana  = calc_task_analytics(db, emp['id'], ev['task_date_from'] or '', ev['task_date_to'] or '')

    # Skill scores
    skill_rows = db.execute('''
        SELECT sc.name AS cat, si.name AS item, ss.score
        FROM skill_scores ss
        JOIN skill_items si ON si.id=ss.skill_item_id
        JOIN skill_categories sc ON sc.id=si.category_id
        WHERE ss.eval_id=? ORDER BY sc.name, si.name
    ''', (eval_id,)).fetchall()

    # Ability scores
    ability_rows = db.execute('''
        SELECT ai.name, s.level
        FROM ability_scores s JOIN ability_items ai ON ai.id=s.ability_item_id
        WHERE s.eval_id=?
    ''', (eval_id,)).fetchall()

    # Competency scores
    comp_rows = db.execute('''
        SELECT ci.point_measurement, cs.rating
        FROM competency_scores cs JOIN competency_items ci ON ci.id=cs.competency_item_id
        WHERE cs.eval_id=?
    ''', (eval_id,)).fetchall()

    # Peer reviews (structured dims jika ada, fallback ke feedback text)
    peer_rows = db.execute(
        'SELECT * FROM peer_reviews WHERE eval_id=? ORDER BY slot', (eval_id,)
    ).fetchall()

    # Benchmark untuk divisi+level
    bm = get_benchmark_for_emp(db, emp)
    acc = calc_accumulated_metrics(db, ev, emp)

    # Format prompt
    lines = [
        f"=== DATA EVALUASI KARYAWAN ===",
        f"Nama      : {emp['name']}",
        f"Jabatan   : {emp['jabatan']} — {emp['divisi']}",
        f"Level     : {emp['level']}",
        f"Periode   : {ev['periode']}",
        f"Status Karyawan: {'Kontrak' if emp['employment_type']=='kontrak' else 'Tetap'}",
        "",
        f"=== AKUMULASI EVALUASI KINERJA (5 PILAR) ===",
        f"1. Task Load (Data Task)  : {acc['task_score']:.1f} / 100 (periode: {acc['date_from']} s/d {acc['date_to']})",
        f"2. Project Performance     : {acc['pp_total']:.1f} / 100",
        f"3. Self-Improvement/Train  : {acc['training_score']:.1f} / 100",
        f"4. Soft Skill (Kompetensi) : {acc['soft_skill_score']:.1f} / 100",
        f"5. Hard Skill (Keahlian)   : {acc['hs_total']:.1f} / 100",
        f"NILAI INDEKS AKUMULASI     : {acc['overall']:.1f} / 100",
        "",
        f"=== DETAIL KINERJA TASK (otomatis dari sistem) ===",
        f"Task Score     : {perf['task_score']:.1f} / 100 (benchmark {bm} poin/bulan)",
        f"Total Poin Raw : {perf['total_raw']:.1f}",
        f"Durasi Periode : {perf['months']:.1f} bulan",
        f"Task Selesai   : {ana['total_done']} task",
        f"Task Open      : {ana['total_open']} task (overtime: {ana['open_overtime']})",
        f"Ontime Rate    : {ana['ontime_rate']}%" if ana['ontime_rate'] is not None else "Ontime Rate: N/A",
        f"Max Concurrent : {ana['concurrent_max']} task bersamaan",
        "",
        "Breakdown sumber poin:",
    ]
    for b in perf['breakdown']:
        lines.append(f"  - {b['label']}: {b['count']}x = {b['raw_pts']:.1f} poin")

    if skill_rows:
        lines += ["", "=== SKILL SCORES (1-5) ==="]
        cat_cur = None
        for r in skill_rows:
            if r['cat'] != cat_cur:
                lines.append(f"  [{r['cat']}]")
                cat_cur = r['cat']
            lines.append(f"    - {r['item']}: {r['score']}/5")

    if ability_rows:
        lines += ["", "=== ABILITY LEVEL (A=Terbaik, D=Perlu Bimbingan) ==="]
        for r in ability_rows:
            lines.append(f"  - {r['name']}: Level {r['level']}")

    if comp_rows:
        lines += ["", "=== KOMPETENSI (1-5) ==="]
        for r in comp_rows:
            lines.append(f"  - {r['point_measurement']}: {r['rating']}/5")

    if peer_rows:
        lines += ["", "=== PEER REVIEW ==="]
        for pr in peer_rows:
            lines.append(f"  Reviewer: {pr['reviewer_name'] or 'Anonim'}")
            # Structured dimensions
            dims = {
                'Kerjasama': pr['dim_kerjasama'],
                'Komunikasi': pr['dim_komunikasi'],
                'Keandalan': pr['dim_keandalan'],
                'Inisiatif': pr['dim_inisiatif'],
                'Kualitas Kerja': pr['dim_kualitas'],
            }
            dim_text = ', '.join(f"{k}:{v}/5" for k, v in dims.items() if v is not None)
            if dim_text:
                lines.append(f"    Dimensi: {dim_text}")
            if pr['feedback']:
                lines.append(f"    Feedback: {pr['feedback']}")

    if ev['self_notes']:
        lines += ["", "=== SELF ASSESSMENT ===", f"Catatan Diri : {ev['self_notes']}"]
    if ev['self_achievements']:
        lines.append(f"Pencapaian   : {ev['self_achievements']}")
    if ev['self_improvements']:
        lines.append(f"Rencana Improve: {ev['self_improvements']}")

    prompt_data = '\n'.join(lines)
    return prompt_data, emp


@app.route('/eval/<int:eval_id>/ai-summary', methods=['POST'])
@login_required
def eval_ai_summary(eval_id):
    """Generate ringkasan evaluasi + rekomendasi menggunakan Ollama/AI."""
    db       = get_db()
    ev       = get_eval_or_404(db, eval_id)
    settings = get_settings(db)

    ai_provider = settings.get('ai_provider', '')
    ai_key      = settings.get('ai_api_key', '')
    ai_model    = settings.get('ai_model', '')
    ai_base_url = settings.get('ai_base_url', '')

    if not ai_provider:
        return {'ok': False, 'error': 'Konfigurasi AI belum diset. Hubungi administrator.'}, 400

    prompt_data, emp = _build_eval_ai_prompt(db, eval_id)
    if not prompt_data:
        return {'ok': False, 'error': 'Data evaluasi tidak ditemukan.'}, 404

    system_prompt = (
        "Kamu adalah sistem analisa kinerja SDM yang objektif dan profesional. "
        "Tugasmu adalah menganalisa data kinerja karyawan secara komprehensif dan menghasilkan:\n"
        "1. RINGKASAN KINERJA: narasi singkat 2-3 paragraf yang menjelaskan kekuatan, area lemah, "
        "   dan pola kinerja berdasarkan data kuantitatif (bukan opini).\n"
        "2. REKOMENDASI: poin-poin konkret (format: - Item: Rekomendasi) mencakup:\n"
        "   - Kenaikan Gaji: berikan estimasi % yang wajar berdasarkan performa\n"
        "   - Jenjang Karir: pertahankan/naikan level/PIP/tidak diperpanjang kontrak\n"
        "   - Distribusi Task: apakah beban kerja sudah optimal atau perlu diseimbangkan\n"
        "   - Skill Development: area yang perlu ditingkatkan\n"
        "   - Risiko: jika ada indikator burnout (concurrent terlalu tinggi) atau underperform\n"
        "Gunakan bahasa Indonesia yang formal dan profesional. "
        "Basis semua analisa pada data numerik yang diberikan, bukan asumsi. "
        "Hindari penilaian yang subyektif tanpa dasar data."
    )

    user_message = (
        f"{prompt_data}\n\n"
        "Berdasarkan seluruh data di atas, berikan:\n"
        "1. RINGKASAN KINERJA\n"
        "2. REKOMENDASI (dalam poin-poin)\n\n"
        "Pisahkan dua bagian tersebut dengan garis '---'."
    )

    try:
        if ai_provider == 'anthropic':
            import anthropic as _ant
            client = _ant.Anthropic(api_key=ai_key)
            resp = client.messages.create(
                model=ai_model or 'claude-sonnet-4-6',
                max_tokens=1500,
                system=system_prompt,
                messages=[{'role': 'user', 'content': user_message}],
            )
            result_text = resp.content[0].text if resp.content else ''
        else:
            # Ollama / OpenAI-compat / openai
            import openai as _oai
            kwargs = {'api_key': ai_key or 'ollama'}
            if ai_base_url:
                kwargs['base_url'] = ai_base_url
            client = _oai.OpenAI(**kwargs)
            resp = client.chat.completions.create(
                model=ai_model or 'llama3',
                max_tokens=1500,
                messages=[
                    {'role': 'system', 'content': system_prompt},
                    {'role': 'user',   'content': user_message},
                ],
            )
            result_text = resp.choices[0].message.content or ''

        # Split ringkasan vs rekomendasi
        if '---' in result_text:
            parts = result_text.split('---', 1)
            summary_text = parts[0].strip()
            recom_text   = parts[1].strip()
        else:
            summary_text = result_text.strip()
            recom_text   = ''

        now = datetime.now().strftime('%Y-%m-%d %H:%M')
        db.execute(
            "UPDATE evaluations SET ai_summary=?, ai_recommendation=?, ai_generated_at=? WHERE id=?",
            (summary_text, recom_text, now, eval_id)
        )
        db.commit()
        return {'ok': True, 'summary': summary_text, 'recommendation': recom_text, 'generated_at': now}

    except Exception as e:
        return {'ok': False, 'error': f'Gagal menghubungi AI: {e}'}, 500


# ─── Self-Assessment & Review Routes ──────────────────────────────────────────

def _send_self_assessment(db, eval_id, emp, periode, base_url, triggered_by='auto'):
    """Kirim link self-assessment via email dan/atau Telegram. Return (results, any_ok, link)."""
    token    = get_or_create_eval_token(db, eval_id)
    link     = f'{base_url}/assess/{token}'
    settings = get_settings(db)
    now_str  = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    results  = []

    if emp['email']:
        subject   = f'TalentCore Self-Assessment {emp["name"]} — Periode {periode}'
        html_body = f"""
<h3 style="color:#1e2a3a">TalentCore Self-Assessment</h3>
<p>Yth. <strong>{emp['name']}</strong>,</p>
<p>Anda diminta mengisi <b>penilaian diri (self-assessment)</b> untuk TalentCore
   periode <strong>{periode}</strong>.</p>
<p>Silakan klik tombol di bawah untuk mengisi form:</p>
<p style="margin:20px 0">
  <a href="{link}" style="background:#4da8da;color:#fff;padding:12px 24px;border-radius:6px;
     text-decoration:none;font-weight:bold">&#128221; Isi Self-Assessment</a>
</p>
<p style="color:#666;font-size:.9em">Atau buka link: <a href="{link}">{link}</a></p>
<hr>
<p style="color:#999;font-size:.85em">
  Mohon segera mengisi sebelum batas waktu. Hubungi tim HR/evaluator jika ada pertanyaan.
</p>"""
        ok, err = send_email(settings, emp['email'], subject, html_body)
        log_reminder(db, emp['id'], 'email', subject, html_body, ok, err, triggered_by)
        if ok:
            db.execute('''INSERT INTO eval_tokens(eval_id, token, email_sent_to, sent_at)
                          VALUES(?,?,?,?)
                          ON CONFLICT(eval_id) DO UPDATE SET token=excluded.token,email_sent_to=excluded.email_sent_to,sent_at=excluded.sent_at''',
                       (eval_id, token, emp['email'], now_str))
            results.append(f'Email {emp["email"]} ✓')
        else:
            results.append(f'Email gagal: {err}')

    bot_token = settings.get('telegram_bot_token', '').strip()
    tg_id     = normalize_telegram_id(emp['telegram_id'] or '')
    if bot_token and tg_id:
        tg_msg = (
            f"📝 <b>TalentCore Self-Assessment</b>\n\n"
            f"Yth. <b>{emp['name']}</b>,\n\n"
            f"Anda diminta mengisi penilaian diri (self-assessment) untuk TalentCore "
            f"periode <b>{periode}</b>.\n\n"
            f"Silakan buka link berikut:\n<a href=\"{link}\">{link}</a>\n\n"
            f"<i>Mohon segera mengisi sebelum batas waktu.</i>"
        )
        ok, err = send_telegram(bot_token, tg_id, tg_msg)
        log_reminder(db, emp['id'], 'telegram', f'Self-Assessment {emp["name"]}', tg_msg, ok, err, triggered_by)
        if ok:
            results.append(f'Telegram {tg_id} ✓')
        else:
            results.append(f'Telegram {tg_id} gagal: {err}')
    elif emp['telegram_id'] and not bot_token:
        results.append('Telegram: bot token belum dikonfigurasi')

    # ── WhatsApp (OpenWA) ──
    wa_url     = settings.get('openwa_url', '').strip()
    wa_key     = settings.get('openwa_api_key', '').strip()
    wa_session = get_openwa_session(settings, 'evaluasi')
    wa_enabled = settings.get('openwa_enabled', '0') == '1'
    emp_phone  = emp['phone'] or ''
    if wa_enabled and wa_url and emp_phone:
        wa_chat_id = normalize_phone_wa(emp_phone)
        wa_msg = (
            f"📝 *TalentCore Self-Assessment*\n\n"
            f"Yth. *{emp['name']}*,\n\n"
            f"Anda diminta mengisi penilaian diri (self-assessment) untuk TalentCore "
            f"periode *{periode}*.\n\n"
            f"Silakan buka link berikut:\n{link}\n\n"
            f"_Mohon segera mengisi sebelum batas waktu._"
        )
        ok, err = send_whatsapp(wa_url, wa_key, wa_session, emp_phone, wa_msg)
        log_reminder(db, emp['id'], 'whatsapp', f'Self-Assessment {emp["name"]}', wa_msg, ok, err, triggered_by)
        if ok:
            results.append(f'WhatsApp {wa_chat_id} ✓')
        else:
            results.append(f'WhatsApp gagal: {err}')
    elif wa_enabled and wa_url and not emp_phone:
        results.append('WhatsApp: nomor HP karyawan belum diisi')

    any_ok = any('✓' in r for r in results)
    if any_ok:
        db.execute("UPDATE evaluations SET review_status='pending_self' WHERE id=? AND review_status='draft'",
                   (eval_id,))
    return results, any_ok, link


@app.route('/eval/<int:eval_id>/send-self-link', methods=['POST'])
@login_required
def eval_send_self_link(eval_id):
    db = get_db()
    ev = get_eval_or_404(db, eval_id)
    if not ev:
        flash('Evaluasi tidak ditemukan', 'danger')
        return redirect(url_for('index'))
    emp = db.execute('SELECT * FROM employees WHERE id=?', (ev['employee_id'],)).fetchone()
    if not emp or (not emp['email'] and not emp['telegram_id']):
        flash('Email dan Telegram karyawan belum diisi. Lengkapi data karyawan terlebih dahulu.', 'warning')
        return redirect(url_for('eval_summary', eval_id=eval_id))

    results, any_ok, link = _send_self_assessment(
        db, eval_id, emp, ev['periode'],
        request.host_url.rstrip('/'),
        triggered_by=session.get('username', 'manual')
    )
    db.commit()
    if any_ok:
        flash('Self-assessment dikirim — ' + ' | '.join(results), 'success')
    else:
        flash('Gagal mengirim — ' + ' | '.join(results) + f'. Link manual: {link}', 'warning')
    return redirect(url_for('eval_summary', eval_id=eval_id))


@app.route('/assess/<token>', methods=['GET', 'POST'])
def self_assess(token):
    """Public — no login required. Employee fills self-assessment via token link."""
    db = _get_raw_db()
    try:
        tok = db.execute("SELECT * FROM eval_tokens WHERE token=?", (token,)).fetchone()
        if not tok:
            return render_template('eval_self.html', error='Link tidak valid atau sudah kadaluarsa.')

        ev_row = db.execute('''SELECT e.*, emp.name AS emp_name, emp.jabatan, emp.divisi, emp.id AS emp_id
                               FROM evaluations e JOIN employees emp ON emp.id=e.employee_id
                               WHERE e.id=?''', (tok['eval_id'],)).fetchone()
        if not ev_row:
            return render_template('eval_self.html', error='Data evaluasi tidak ditemukan.')

        import json as _json
        self_data = {}
        if ev_row.get('self_assessment_json'):
            try:
                self_data = _json.loads(ev_row['self_assessment_json'])
            except:
                self_data = {}

        if tok['status'] == 'submitted':
            return render_template('eval_self.html', ev=ev_row, already_submitted=True, self_data=self_data)

        if request.method == 'POST':
            now_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            ans_dict = {}
            for key in request.form.keys():
                if key.startswith(('s1_', 's2_', 's3_', 's4_', 's5_', 's6_')):
                    ans_dict[key] = request.form.get(key, '').strip()
            
            # Map to standard fields for backward compatibility
            self_ach = f"Pencapaian OKR:\n{request.form.get('s1_q2', '')}\n\nInovasi & Efisiensi:\n{request.form.get('s2_inovasi', '')}"
            self_nt  = request.form.get('s1_q3', '')
            self_imp = request.form.get('s6_komitmen', '')
            json_str = _json.dumps(ans_dict)

            db.execute('''UPDATE evaluations SET self_achievements=?, self_notes=?,
                          self_improvements=?, self_assessment_json=?,
                          review_status=CASE WHEN review_status='pending_self' THEN 'self_filled'
                                             ELSE review_status END
                          WHERE id=?''', (
                self_ach,
                self_nt,
                self_imp,
                json_str,
                tok['eval_id']
            ))
            db.execute("UPDATE eval_tokens SET status='submitted', submitted_at=? WHERE id=?",
                       (now_str, tok['id']))
            # Create pending review rows for assigned supervisors (employee hierarchy → user)
            emp = db.execute('SELECT * FROM employees WHERE id=?', (ev_row['emp_id'],)).fetchone()
            for emp_ref_id, role in [
                (emp['supervisor_id'], 'Atasan Langsung'),
                (emp['leader_id'],     'Leader'),
                (emp['manager_id'],    'Manager'),
            ]:
                if emp_ref_id:
                    rev_emp = db.execute('SELECT user_id FROM employees WHERE id=?', (emp_ref_id,)).fetchone()
                    if rev_emp and rev_emp['user_id']:
                        db.execute('''INSERT OR IGNORE INTO eval_reviews(eval_id,reviewer_user_id,reviewer_role)
                                      VALUES(?,?,?)''', (tok['eval_id'], rev_emp['user_id'], role))
            db.commit()
            return render_template('eval_self.html', ev=ev_row, submitted=True, self_data=ans_dict)

        if not tok['accessed_at']:
            db.execute('UPDATE eval_tokens SET accessed_at=? WHERE id=?',
                       (datetime.now().strftime('%Y-%m-%d %H:%M:%S'), tok['id']))
            db.commit()
        return render_template('eval_self.html', ev=ev_row, token=token, self_data=self_data)
    finally:
        db.close()


@app.route('/reviews')
@login_required
def reviews():
    db  = get_db()
    uid = session['user_id']
    role = session.get('user_role', '')
    if role == 'superadmin':
        evals = db.execute('''
            SELECT e.*, emp.name AS emp_name, emp.jabatan, emp.divisi,
                   emp.supervisor_id, emp.leader_id, emp.manager_id,
                   u.full_name AS reviewed_by_name
            FROM evaluations e
            JOIN employees emp ON emp.id = e.employee_id
            LEFT JOIN users u ON u.id = e.reviewed_by
            WHERE e.review_status != 'draft'
            ORDER BY CASE e.review_status WHEN 'self_filled' THEN 0 WHEN 'pending_self' THEN 1 ELSE 2 END,
                     e.id DESC
        ''').fetchall()
    else:
        my_emp_id = get_user_emp_id(db, uid) or -1
        evals = db.execute('''
            SELECT e.*, emp.name AS emp_name, emp.jabatan, emp.divisi,
                   emp.supervisor_id, emp.leader_id, emp.manager_id,
                   u.full_name AS reviewed_by_name
            FROM evaluations e
            JOIN employees emp ON emp.id = e.employee_id
            LEFT JOIN users u ON u.id = e.reviewed_by
            WHERE e.review_status != 'draft'
            AND (emp.supervisor_id=? OR emp.leader_id=? OR emp.manager_id=?)
            ORDER BY CASE e.review_status WHEN 'self_filled' THEN 0 WHEN 'pending_self' THEN 1 ELSE 2 END,
                     e.id DESC
        ''', (my_emp_id, my_emp_id, my_emp_id)).fetchall()
    my_emp_id = get_user_emp_id(db, uid) if role != 'superadmin' else None
    # mark own pending_review rows for each eval
    my_reviews = {r['eval_id']: r for r in db.execute(
        'SELECT * FROM eval_reviews WHERE reviewer_user_id=?', (uid,)).fetchall()}
    return render_template('reviews.html', evals=evals, my_reviews=my_reviews, my_emp_id=my_emp_id)


@app.route('/eval/<int:eval_id>/review', methods=['GET', 'POST'])
@login_required
def eval_review(eval_id):
    db  = get_db()
    uid = session['user_id']
    if not can_review_eval(db, uid, session.get('user_role',''), eval_id):
        flash('Anda tidak memiliki akses mereview evaluasi ini', 'danger')
        return redirect(url_for('index'))

    ev  = get_eval_or_404(db, eval_id)
    emp = db.execute('''SELECT e.*,
        e1.name AS sup_name, e2.name AS lead_name, e3.name AS mgr_name
        FROM employees e
        LEFT JOIN employees e1 ON e1.id=e.supervisor_id
        LEFT JOIN employees e2 ON e2.id=e.leader_id
        LEFT JOIN employees e3 ON e3.id=e.manager_id
        WHERE e.id=?''', (ev['employee_id'],)).fetchone()

    my_emp_id = get_user_emp_id(db, uid)
    def _reviewer_role():
        if my_emp_id and my_emp_id == emp['supervisor_id']: return 'Atasan Langsung'
        if my_emp_id and my_emp_id == emp['leader_id']:     return 'Leader'
        if my_emp_id and my_emp_id == emp['manager_id']:    return 'Manager'
        if session.get('user_role') == 'superadmin':        return 'Superadmin'
        return ''

    if request.method == 'POST':
        action = request.form.get('action', 'save_review')
        notes  = request.form.get('review_notes','').strip()
        now_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        reviewer_role = _reviewer_role()

        db.execute('''INSERT INTO eval_reviews
                      (eval_id, reviewer_user_id, reviewer_role, notes, status, submitted_at)
                      VALUES(?,?,?,?,'submitted',?)
                      ON CONFLICT(eval_id,reviewer_user_id) DO UPDATE SET reviewer_role=excluded.reviewer_role,notes=excluded.notes,status=excluded.status,submitted_at=excluded.submitted_at''',
                   (eval_id, uid, reviewer_role, notes, now_str))

        if action == 'approve':
            db.execute('''UPDATE evaluations SET review_status='approved',
                          reviewed_by=?, reviewed_at=?, review_notes=? WHERE id=?''',
                       (uid, now_str, notes, eval_id))
            db.commit()
            flash('Evaluasi disetujui', 'success')
            return redirect(url_for('reviews'))
        elif action == 'reject':
            db.execute('''UPDATE evaluations SET review_status='rejected',
                          reviewed_by=?, reviewed_at=?, review_notes=? WHERE id=?''',
                       (uid, now_str, notes, eval_id))
            db.commit()
            flash('Evaluasi dikembalikan untuk perbaikan', 'warning')
            return redirect(url_for('reviews'))
        else:
            db.commit()
            flash('Catatan review disimpan', 'success')
            return redirect(url_for('eval_review', eval_id=eval_id))

    my_review = db.execute('SELECT * FROM eval_reviews WHERE eval_id=? AND reviewer_user_id=?',
                            (eval_id, uid)).fetchone()
    all_reviews = db.execute('''SELECT er.*, u.full_name, u.username
        FROM eval_reviews er JOIN users u ON u.id=er.reviewer_user_id
        WHERE er.eval_id=? ORDER BY er.submitted_at''', (eval_id,)).fetchall()
    competency_items = db.execute('''SELECT ci.*, COALESCE(cs.rating,0) AS current_rating
        FROM competency_items ci
        LEFT JOIN competency_scores cs ON cs.competency_item_id=ci.id AND cs.eval_id=?
        WHERE ci.divisi=? ORDER BY ci.sort_order''', (eval_id, ev['divisi'])).fetchall()

    reviewer_role = _reviewer_role()

    # ── Data analisa sistem untuk approver ──
    date_from = ev['task_date_from'] or ''
    date_to   = ev['task_date_to']   or ''
    perf = calc_task_perf(db, ev['employee_id'], date_from, date_to,
                          benchmark_per_month=get_benchmark_for_emp(db, emp))
    ana  = calc_task_analytics(db, ev['employee_id'], date_from, date_to)
    bm   = get_benchmark_for_emp(db, emp)

    # Peer review averages per dimensi
    peer_rows = db.execute('''SELECT dim_kerjasama, dim_komunikasi, dim_keandalan,
                                      dim_inisiatif, dim_kualitas
                               FROM peer_reviews WHERE eval_id=? AND dim_kerjasama IS NOT NULL''',
                           (eval_id,)).fetchall()
    peer_count = len(peer_rows)
    peer_avgs  = {}
    if peer_rows:
        dims = ['Kerjasama','Komunikasi','Keandalan','Inisiatif','Kualitas']
        cols = ['dim_kerjasama','dim_komunikasi','dim_keandalan','dim_inisiatif','dim_kualitas']
        for label, col in zip(dims, cols):
            vals = [r[col] for r in peer_rows if r[col] is not None]
            if vals:
                peer_avgs[label] = round(sum(vals)/len(vals), 1)

    return render_template('eval_review.html', ev=ev, emp=emp,
                           my_review=my_review, all_reviews=all_reviews,
                           competency_items=competency_items,
                           reviewer_role=reviewer_role,
                           perf=perf, ana=ana, bm=bm,
                           peer_avgs=peer_avgs, peer_count=peer_count)


@app.route('/api/eval/<int:eval_id>/score')
@login_required
def api_score(eval_id):
    db = get_db()
    recalc(db, eval_id)
    row = db.execute('SELECT pp_total,hs_total,ability_score,competency_total,final_total FROM evaluations WHERE id=?',
                     (eval_id,)).fetchone()
    return (dict(row) if row else {'error':'not found'})


# ─── Export Evaluasi ───────────────────────────────────────────────────────────

def _eval_export_data(db, eval_id):
    """Kumpulkan semua data evaluasi untuk export (Excel/PDF)."""
    ev  = db.execute('SELECT * FROM evaluations WHERE id=?', (eval_id,)).fetchone()
    if not ev:
        return None
    emp = db.execute('''SELECT e.*, e1.name AS sup_name, e2.name AS lead_name, e3.name AS mgr_name
        FROM employees e
        LEFT JOIN employees e1 ON e1.id=e.supervisor_id
        LEFT JOIN employees e2 ON e2.id=e.leader_id
        LEFT JOIN employees e3 ON e3.id=e.manager_id
        WHERE e.id=?''', (ev['employee_id'],)).fetchone()

    recalc(db, eval_id)
    ev = db.execute('SELECT * FROM evaluations WHERE id=?', (eval_id,)).fetchone()

    perf = calc_task_perf(db, emp['id'], ev['task_date_from'] or '', ev['task_date_to'] or '')
    ana  = calc_task_analytics(db, emp['id'], ev['task_date_from'] or '', ev['task_date_to'] or '')
    bm   = get_benchmark_for_emp(db, emp)

    skill_rows = db.execute('''
        SELECT sc.name AS cat, si.name AS item, ss.score
        FROM skill_scores ss
        JOIN skill_items si ON si.id=ss.skill_item_id
        JOIN skill_categories sc ON sc.id=si.category_id
        WHERE ss.eval_id=? ORDER BY sc.name, si.name''', (eval_id,)).fetchall()

    ability_rows = db.execute('''
        SELECT ai.name, ai.desc_a, ai.desc_b, ai.desc_c, ai.desc_d, s.level
        FROM ability_scores s JOIN ability_items ai ON ai.id=s.ability_item_id
        WHERE s.eval_id=?''', (eval_id,)).fetchall()

    comp_rows = db.execute('''
        SELECT ci.point_measurement, ci.bobot, ci.is_hardskill, cs.rating
        FROM competency_scores cs JOIN competency_items ci ON ci.id=cs.competency_item_id
        WHERE cs.eval_id=? ORDER BY ci.sort_order''', (eval_id,)).fetchall()

    peer_rows = db.execute(
        'SELECT * FROM peer_reviews WHERE eval_id=? ORDER BY slot', (eval_id,)
    ).fetchall()

    proj_rows = db.execute(
        'SELECT * FROM project_entries WHERE eval_id=? ORDER BY entry_type, sort_order', (eval_id,)
    ).fetchall()

    all_reviews = db.execute('''SELECT er.*, u.full_name
        FROM eval_reviews er JOIN users u ON u.id=er.reviewer_user_id
        WHERE er.eval_id=? ORDER BY er.submitted_at''', (eval_id,)).fetchall()

    # Gaji tahun berjalan + tahun sebelumnya (jika ada izin tidak perlu disini — hanya untuk context)
    sal_year = date.today().year
    sal = db.execute(
        'SELECT * FROM employee_salary WHERE employee_id=? AND year=?',
        (emp['id'], sal_year)
    ).fetchone()
    sal_prev = db.execute(
        'SELECT * FROM employee_salary WHERE employee_id=? AND year=?',
        (emp['id'], sal_year - 1)
    ).fetchone()
    sal_dec  = _dec_sal_row(sal)      if sal      else {}
    sal_prev_dec = _dec_sal_row(sal_prev) if sal_prev else {}

    accumulated = calc_accumulated_metrics(db, ev, emp)
    import json as _json
    self_data = {}
    if ev.get('self_assessment_json'):
        try:
            self_data = _json.loads(ev['self_assessment_json'])
        except:
            pass
    return {
        'ev': dict(ev), 'emp': dict(emp),
        'perf': perf, 'ana': ana, 'bm': bm,
        'skill_rows': [dict(r) for r in skill_rows],
        'ability_rows': [dict(r) for r in ability_rows],
        'comp_rows': [dict(r) for r in comp_rows],
        'peer_rows': [dict(r) for r in peer_rows],
        'proj_rows': [dict(r) for r in proj_rows],
        'all_reviews': [dict(r) for r in all_reviews],
        'sal': sal_dec, 'sal_prev': sal_prev_dec, 'sal_year': sal_year,
        'accumulated': accumulated,
        'self_data': self_data,
    }


@app.route('/eval/<int:eval_id>/export/excel')
@login_required
def eval_export_excel(eval_id):
    """Export evaluasi ke Excel (.xlsx) — untuk lampiran pengajuan gaji."""
    try:
        import openpyxl
        from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                                      numbers as _onum)
        from openpyxl.utils import get_column_letter
        import io
    except ImportError:
        flash('Library openpyxl tidak terinstall. Jalankan: pip install openpyxl', 'danger')
        return redirect(url_for('eval_review', eval_id=eval_id))

    db   = get_db()
    data = _eval_export_data(db, eval_id)
    if not data:
        abort(404)
    ev, emp = data['ev'], data['emp']

    wb = openpyxl.Workbook()

    # ── Styles ────────────────────────────────────────────────────────────────
    BLUE   = 'FF1E3A5F'
    LBLUE  = 'FFD9E8F7'
    GREEN  = 'FF198754'
    LGREEN = 'FFD1E8DB'
    GRAY   = 'FFF4F6F9'
    GOLD   = 'FFAD8B2A'
    WHITE  = 'FFFFFFFF'
    RED    = 'FFDC3545'

    def _hdr(ws, row, col, text, bold=True, bg=BLUE, fg=WHITE, size=11, wrap=False, merge_to=None):
        c = ws.cell(row=row, column=col, value=text)
        c.font = Font(bold=bold, color=fg, size=size)
        c.fill = PatternFill('solid', fgColor=bg)
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=wrap)
        if merge_to:
            ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=merge_to)
        return c

    def _cell(ws, row, col, val, bold=False, bg=None, fg='FF000000', align='left',
              num_fmt=None, wrap=False):
        c = ws.cell(row=row, column=col, value=val)
        c.font = Font(bold=bold, color=fg, size=10)
        if bg:
            c.fill = PatternFill('solid', fgColor=bg)
        c.alignment = Alignment(horizontal=align, vertical='center', wrap_text=wrap)
        if num_fmt:
            c.number_format = num_fmt
        return c

    def _border_range(ws, min_row, max_row, min_col, max_col):
        thin = Side(style='thin', color='FFCCCCCC')
        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                ws.cell(r, c).border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ═════════════════════════════════════════════════════════════════════════
    # SHEET 1: Halaman Utama / Ringkasan Evaluasi
    # ═════════════════════════════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = 'Evaluasi Kinerja'
    ws1.column_dimensions['A'].width = 28
    ws1.column_dimensions['B'].width = 30
    ws1.column_dimensions['C'].width = 18
    ws1.column_dimensions['D'].width = 18
    ws1.column_dimensions['E'].width = 18
    ws1.column_dimensions['F'].width = 18
    ws1.row_dimensions[1].height = 36
    ws1.row_dimensions[2].height = 22

    # Header judul
    _hdr(ws1, 1, 1, 'FORM EVALUASI KINERJA KARYAWAN', size=14, merge_to=6)
    _hdr(ws1, 2, 1, 'Hive — TalentCore', size=10, bg=LBLUE, fg=BLUE, merge_to=6)

    # Info Karyawan
    r = 4
    _hdr(ws1, r, 1, 'INFORMASI KARYAWAN', bg=BLUE, merge_to=6)
    r += 1
    info = [
        ('Nama Karyawan', emp['name']),
        ('Jabatan', emp.get('jabatan') or '—'),
        ('Divisi', emp.get('divisi') or '—'),
        ('Level / Grade', emp.get('level') or '—'),
        ('Status', 'Tetap' if emp.get('employment_type') == 'tetap' else 'Kontrak'),
        ('Periode Evaluasi', ev['periode']),
        ('Evaluator', ev.get('evaluator') or '—'),
        ('Atasan Langsung', emp.get('sup_name') or '—'),
        ('Manager', emp.get('mgr_name') or '—'),
        ('Status Evaluasi', ev.get('review_status', '').upper()),
    ]
    for label, val in info:
        _cell(ws1, r, 1, label, bold=True, bg=GRAY)
        _cell(ws1, r, 2, val, merge_to=None)
        ws1.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
        r += 1

    r += 1
    # Rekapitulasi Nilai
    _hdr(ws1, r, 1, 'REKAPITULASI NILAI EVALUASI', bg=BLUE, merge_to=6)
    r += 1
    _hdr(ws1, r, 1, 'Komponen', bg=LBLUE, fg=BLUE, bold=True)
    _hdr(ws1, r, 2, 'Bobot', bg=LBLUE, fg=BLUE, bold=True)
    _hdr(ws1, r, 3, 'Nilai Mentah', bg=LBLUE, fg=BLUE, bold=True)
    _hdr(ws1, r, 4, 'Nilai Tertimbang', bg=LBLUE, fg=BLUE, bold=True)
    _hdr(ws1, r, 5, 'Task Score', bg=LBLUE, fg=BLUE, bold=True)
    _hdr(ws1, r, 6, 'Benchmark', bg=LBLUE, fg=BLUE, bold=True)
    r += 1
    _cell(ws1, r, 1, 'Project Performance (Task)', bold=True)
    _cell(ws1, r, 2, '30%', align='center')
    _cell(ws1, r, 3, round(ev.get('pp_total') or 0, 1), align='center')
    _cell(ws1, r, 4, round((ev.get('pp_total') or 0) * 0.3, 2), align='center')
    _cell(ws1, r, 5, data['perf']['task_score'], align='center', bold=True,
          bg=LGREEN if data['perf']['task_score'] >= 70 else None)
    _cell(ws1, r, 6, data['bm'], align='center')
    r += 1
    _cell(ws1, r, 1, 'Competency (+ Hard Skill)', bold=True)
    _cell(ws1, r, 2, '70%', align='center')
    _cell(ws1, r, 3, round(ev.get('competency_total') or 0, 2), align='center')
    _cell(ws1, r, 4, round((ev.get('competency_total') or 0) * 0.7, 2), align='center')
    r += 1
    _cell(ws1, r, 1, 'NILAI AKHIR', bold=True, bg=LGREEN)
    _cell(ws1, r, 2, '100%', align='center', bold=True, bg=LGREEN)
    ws1.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
    c_final = ws1.cell(r, 3, round(ev.get('final_total') or 0, 2))
    c_final.font = Font(bold=True, size=14,
                        color=GREEN if (ev.get('final_total') or 0) >= 70 else RED)
    c_final.alignment = Alignment(horizontal='center', vertical='center')
    c_final.fill = PatternFill('solid', fgColor=LGREEN)
    r += 1

    # Keterangan predikat
    ft = ev.get('final_total') or 0
    predikat = ('Luar Biasa' if ft >= 80 else 'Sangat Baik' if ft >= 70 else
                'Baik' if ft >= 60 else 'Cukup' if ft >= 50 else 'Perlu Perbaikan')
    _cell(ws1, r, 1, 'Predikat', bold=True, bg=GRAY)
    _cell(ws1, r, 2, predikat, bold=True,
          fg=GREEN if ft >= 70 else (GOLD if ft >= 50 else RED))
    ws1.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
    r += 2

    # Task Analytics
    _hdr(ws1, r, 1, 'ANALITIK KINERJA TASK (OTOMATIS SISTEM)', bg=BLUE, merge_to=6)
    r += 1
    ana_info = [
        ('Task Score (vs Benchmark)',
         f"{data['perf']['task_score']} / 100 (benchmark {data['bm']} pts/bln)"),
        ('Total Task Selesai', data['ana']['total_done']),
        ('Task Masih Berjalan', data['ana']['total_open']),
        ('Task Melewati Deadline', data['ana']['open_overtime']),
        ('Ontime Rate', f"{data['ana']['ontime_rate']}%" if data['ana']['ontime_rate'] is not None else 'N/A'),
        ('Maks. Concurrent Task', data['ana']['concurrent_max']),
        ('Durasi Periode', f"{data['perf']['months']:.1f} bulan"),
    ]
    for lbl, val in ana_info:
        _cell(ws1, r, 1, lbl, bold=True, bg=GRAY)
        _cell(ws1, r, 2, str(val))
        ws1.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
        r += 1

    r += 1
    # AI Rekomendasi (jika ada)
    if ev.get('ai_recommendation'):
        _hdr(ws1, r, 1, 'REKOMENDASI SISTEM (AI)', bg=GREEN, merge_to=6)
        r += 1
        c = ws1.cell(r, 1, ev['ai_recommendation'])
        c.alignment = Alignment(wrap_text=True, vertical='top')
        c.font = Font(size=10)
        ws1.merge_cells(start_row=r, start_column=1, end_row=r+10, end_column=6)
        ws1.row_dimensions[r].height = 120
        r += 12

    # Overall assessment
    if ev.get('overall_assessment'):
        _hdr(ws1, r, 1, 'PENILAIAN EVALUATOR', bg=BLUE, merge_to=6)
        r += 1
        c = ws1.cell(r, 1, ev['overall_assessment'])
        c.alignment = Alignment(wrap_text=True, vertical='top')
        c.font = Font(size=10)
        ws1.merge_cells(start_row=r, start_column=1, end_row=r+4, end_column=6)
        ws1.row_dimensions[r].height = 60
        r += 6

    if ev.get('development_plan'):
        _hdr(ws1, r, 1, 'RENCANA PENGEMBANGAN', bg=BLUE, merge_to=6)
        r += 1
        c = ws1.cell(r, 1, ev['development_plan'])
        c.alignment = Alignment(wrap_text=True, vertical='top')
        c.font = Font(size=10)
        ws1.merge_cells(start_row=r, start_column=1, end_row=r+4, end_column=6)
        ws1.row_dimensions[r].height = 60
        r += 6

    # Tanda tangan
    r += 1
    _hdr(ws1, r, 1, 'PERSETUJUAN', bg=BLUE, merge_to=6)
    r += 1
    _cell(ws1, r, 1, 'Karyawan', bold=True, align='center', bg=GRAY)
    _cell(ws1, r, 3, 'Atasan Langsung', bold=True, align='center', bg=GRAY)
    _cell(ws1, r, 5, 'Manager / HR', bold=True, align='center', bg=GRAY)
    ws1.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    ws1.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
    ws1.merge_cells(start_row=r, start_column=5, end_row=r, end_column=6)
    r += 1
    for col in [1, 3, 5]:
        ws1.row_dimensions[r].height = 50
        ws1.merge_cells(start_row=r, start_column=col, end_row=r, end_column=col+1)
    r += 1
    _cell(ws1, r, 1, emp['name'], bold=True, align='center')
    _cell(ws1, r, 3, emp.get('sup_name') or '____________________', bold=True, align='center')
    _cell(ws1, r, 5, emp.get('mgr_name') or '____________________', bold=True, align='center')
    ws1.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    ws1.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
    ws1.merge_cells(start_row=r, start_column=5, end_row=r, end_column=6)
    r += 1
    tgl = date.today().strftime('%d %B %Y')
    for col in [1, 3, 5]:
        _cell(ws1, r, col, f'Tanggal: {tgl}', align='center', fg='FF666666')
        ws1.merge_cells(start_row=r, start_column=col, end_row=r, end_column=col+1)

    _border_range(ws1, 4, r, 1, 6)

    # ═════════════════════════════════════════════════════════════════════════
    # SHEET 2: Detail Skill & Kompetensi
    # ═════════════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet('Skill & Kompetensi')
    ws2.column_dimensions['A'].width = 24
    ws2.column_dimensions['B'].width = 30
    ws2.column_dimensions['C'].width = 14
    ws2.column_dimensions['D'].width = 14

    _hdr(ws2, 1, 1, 'DETAIL SKILL SCORE', bg=BLUE, merge_to=4, size=12)
    r2 = 3
    if data['skill_rows']:
        _hdr(ws2, r2, 1, 'Kategori', bg=LBLUE, fg=BLUE)
        _hdr(ws2, r2, 2, 'Item Skill', bg=LBLUE, fg=BLUE)
        _hdr(ws2, r2, 3, 'Score', bg=LBLUE, fg=BLUE)
        _hdr(ws2, r2, 4, 'Maks', bg=LBLUE, fg=BLUE)
        r2 += 1
        cat_cur = None
        for row in data['skill_rows']:
            bg = GRAY if row['cat'] != cat_cur else None
            cat_cur = row['cat']
            _cell(ws2, r2, 1, row['cat'] if bg else '', bold=bool(bg), bg=bg)
            _cell(ws2, r2, 2, row['item'])
            sc = row['score'] or 0
            _cell(ws2, r2, 3, sc, align='center', bold=True,
                  fg=GREEN if sc >= 4 else (GOLD if sc >= 3 else RED))
            _cell(ws2, r2, 4, 5, align='center', fg='FF888888')
            r2 += 1
        _border_range(ws2, 3, r2-1, 1, 4)

    r2 += 2
    _hdr(ws2, r2, 1, 'ABILITY LEVEL', bg=BLUE, merge_to=4, size=12)
    r2 += 1
    if data['ability_rows']:
        _hdr(ws2, r2, 1, 'Aspek Kemampuan', bg=LBLUE, fg=BLUE, merge_to=2)
        _hdr(ws2, r2, 3, 'Level', bg=LBLUE, fg=BLUE)
        _hdr(ws2, r2, 4, 'Deskripsi', bg=LBLUE, fg=BLUE)
        r2 += 1
        level_desc = {'A': 'Terbaik — jadi referensi tim', 'B': 'Baik — mandiri',
                      'C': 'Cukup — perlu supervisi', 'D': 'Perlu bimbingan intensif'}
        level_color = {'A': GREEN, 'B': '0d6efd', 'C': GOLD, 'D': RED}
        for row in data['ability_rows']:
            lvl = (row['level'] or '').upper()
            ws2.merge_cells(start_row=r2, start_column=1, end_row=r2, end_column=2)
            _cell(ws2, r2, 1, row['name'])
            _cell(ws2, r2, 3, lvl, align='center', bold=True,
                  fg='FF' + level_color.get(lvl, '000000').lstrip('#').lstrip('FF'))
            _cell(ws2, r2, 4, level_desc.get(lvl, ''), fg='FF555555', wrap=True)
            r2 += 1
        _border_range(ws2, r2 - len(data['ability_rows']) - 1, r2-1, 1, 4)

    r2 += 2
    _hdr(ws2, r2, 1, 'KOMPETENSI', bg=BLUE, merge_to=4, size=12)
    r2 += 1
    if data['comp_rows']:
        _hdr(ws2, r2, 1, 'Poin Pengukuran', bg=LBLUE, fg=BLUE, merge_to=2)
        _hdr(ws2, r2, 3, 'Bobot', bg=LBLUE, fg=BLUE)
        _hdr(ws2, r2, 4, 'Rating (1-5)', bg=LBLUE, fg=BLUE)
        r2 += 1
        for row in data['comp_rows']:
            ws2.merge_cells(start_row=r2, start_column=1, end_row=r2, end_column=2)
            lbl = row['point_measurement']
            if row['is_hardskill']:
                lbl += ' [HS Auto]'
            _cell(ws2, r2, 1, lbl)
            _cell(ws2, r2, 3, f"{int(row['bobot']*100)}%", align='center')
            rtg = row['rating'] or 0
            _cell(ws2, r2, 4, rtg, align='center', bold=True,
                  fg=GREEN if rtg >= 4 else (GOLD if rtg >= 3 else RED))
            r2 += 1
        _border_range(ws2, r2 - len(data['comp_rows']) - 1, r2-1, 1, 4)

    # ═════════════════════════════════════════════════════════════════════════
    # SHEET 3: Peer Review
    # ═════════════════════════════════════════════════════════════════════════
    ws3 = wb.create_sheet('Peer Review')
    ws3.column_dimensions['A'].width = 20
    ws3.column_dimensions['B'].width = 14
    ws3.column_dimensions['C'].width = 14
    ws3.column_dimensions['D'].width = 14
    ws3.column_dimensions['E'].width = 14
    ws3.column_dimensions['F'].width = 14
    ws3.column_dimensions['G'].width = 40
    _hdr(ws3, 1, 1, 'PEER REVIEW TERSTRUKTUR', bg=BLUE, merge_to=7, size=12)
    r3 = 3
    _hdr(ws3, r3, 1, 'Reviewer', bg=LBLUE, fg=BLUE)
    for ci, lbl in enumerate(['Kerjasama', 'Komunikasi', 'Keandalan', 'Inisiatif', 'Kualitas'], 2):
        _hdr(ws3, r3, ci, lbl, bg=LBLUE, fg=BLUE)
    _hdr(ws3, r3, 7, 'Narasi / Saran', bg=LBLUE, fg=BLUE)
    r3 += 1
    dim_keys = ['dim_kerjasama','dim_komunikasi','dim_keandalan','dim_inisiatif','dim_kualitas']
    all_avgs = []
    for pr in data['peer_rows']:
        if not pr.get('reviewer_name') and not pr.get('feedback'):
            continue
        _cell(ws3, r3, 1, pr.get('reviewer_name') or '—', bold=True)
        dims = [pr.get(k) for k in dim_keys]
        for ci, v in enumerate(dims, 2):
            _cell(ws3, r3, ci, v if v is not None else '—', align='center',
                  fg=GREEN if (v and v >= 4) else (GOLD if (v and v >= 3) else RED) if v else 'FF888888')
        _cell(ws3, r3, 7, pr.get('feedback') or '', wrap=True)
        ws3.row_dimensions[r3].height = 40
        valid = [v for v in dims if v is not None]
        if valid:
            all_avgs.append(sum(valid) / len(valid))
        r3 += 1
    if all_avgs:
        overall_avg = round(sum(all_avgs) / len(all_avgs), 2)
        _cell(ws3, r3, 1, 'Rata-rata Peer', bold=True, bg=LGREEN)
        _cell(ws3, r3, 2, overall_avg, align='center', bold=True, bg=LGREEN,
              fg=GREEN if overall_avg >= 4 else (GOLD if overall_avg >= 3 else RED))
        ws3.merge_cells(start_row=r3, start_column=2, end_row=r3, end_column=6)
        r3 += 1
    _border_range(ws3, 3, r3-1, 1, 7)

    # ═════════════════════════════════════════════════════════════════════════
    # SHEET 4: Task Performance + Pengajuan Gaji
    # ═════════════════════════════════════════════════════════════════════════
    ws4 = wb.create_sheet('Task & Gaji')
    ws4.column_dimensions['A'].width = 30
    ws4.column_dimensions['B'].width = 22
    ws4.column_dimensions['C'].width = 18
    ws4.column_dimensions['D'].width = 18
    ws4.column_dimensions['E'].width = 18
    _hdr(ws4, 1, 1, 'DETAIL TASK PERFORMANCE', bg=BLUE, merge_to=5, size=12)
    r4 = 3
    _hdr(ws4, r4, 1, 'Sumber', bg=LBLUE, fg=BLUE)
    _hdr(ws4, r4, 2, 'Jumlah', bg=LBLUE, fg=BLUE)
    _hdr(ws4, r4, 3, 'Poin Raw', bg=LBLUE, fg=BLUE)
    ws4.merge_cells(start_row=r4, start_column=4, end_row=r4, end_column=5)
    _hdr(ws4, r4, 4, 'Keterangan', bg=LBLUE, fg=BLUE)
    r4 += 1
    for b in data['perf']['breakdown']:
        _cell(ws4, r4, 1, b['label'])
        _cell(ws4, r4, 2, b['count'], align='center')
        _cell(ws4, r4, 3, b['raw_pts'], align='center')
        r4 += 1
    _cell(ws4, r4, 1, 'Total Poin Raw', bold=True, bg=LGREEN)
    _cell(ws4, r4, 2, '', bg=LGREEN)
    _cell(ws4, r4, 3, data['perf']['total_raw'], bold=True, align='center', bg=LGREEN)
    r4 += 1
    _cell(ws4, r4, 1, 'Task Score (normalized)', bold=True, bg=LGREEN)
    _cell(ws4, r4, 2, f"Benchmark: {data['bm']} pts/bln", fg='FF555555', bg=LGREEN)
    _cell(ws4, r4, 3, f"{data['perf']['task_score']} / 100", bold=True, align='center', bg=LGREEN,
          fg=GREEN if data['perf']['task_score'] >= 70 else RED)
    _border_range(ws4, 3, r4, 1, 5)

    # Pengajuan Kenaikan Gaji
    r4 += 3
    _hdr(ws4, r4, 1, 'DATA GAJI & PENGAJUAN KENAIKAN', bg=BLUE, merge_to=5, size=12)
    r4 += 1
    sal = data['sal']
    sal_prev = data['sal_prev']
    sal_year = data['sal_year']
    gaji_labels = [
        ('Gaji Pokok', 'base_salary'),
        ('Tunjangan Jabatan', 'al_001'),
        ('Tunjangan Komunikasi', 'al_002'),
        ('Tunjangan Performance', 'al_003'),
        ('Tunjangan Kehadiran', 'al_004'),
    ]
    _hdr(ws4, r4, 1, 'Komponen', bg=LBLUE, fg=BLUE)
    _hdr(ws4, r4, 2, f'Tahun {sal_year-1}', bg=LBLUE, fg=BLUE)
    _hdr(ws4, r4, 3, f'Tahun {sal_year}', bg=LBLUE, fg=BLUE)
    _hdr(ws4, r4, 4, 'Selisih', bg=LBLUE, fg=BLUE)
    _hdr(ws4, r4, 5, '% Naik', bg=LBLUE, fg=BLUE)
    r4 += 1
    total_cur = total_prv = 0
    rp_fmt = '#,##0'
    for lbl, key in gaji_labels:
        cur = sal.get(key, 0) or 0
        prv = sal_prev.get(f'p_{key}', sal_prev.get(key, 0)) or 0
        diff = cur - prv
        pct  = round(diff / prv * 100, 1) if prv else 0
        _cell(ws4, r4, 1, lbl, bold=True)
        _cell(ws4, r4, 2, prv, align='right', num_fmt=rp_fmt)
        _cell(ws4, r4, 3, cur, align='right', num_fmt=rp_fmt)
        _cell(ws4, r4, 4, diff, align='right', num_fmt=rp_fmt,
              fg=GREEN if diff > 0 else (RED if diff < 0 else 'FF000000'))
        _cell(ws4, r4, 5, f'{pct}%', align='center',
              fg=GREEN if pct > 0 else (RED if pct < 0 else 'FF000000'))
        total_cur += cur; total_prv += prv
        r4 += 1
    # Total
    diff_tot = total_cur - total_prv
    pct_tot  = round(diff_tot / total_prv * 100, 1) if total_prv else 0
    _cell(ws4, r4, 1, 'TOTAL TAKE HOME', bold=True, bg=LGREEN)
    _cell(ws4, r4, 2, total_prv, bold=True, align='right', num_fmt=rp_fmt, bg=LGREEN)
    _cell(ws4, r4, 3, total_cur, bold=True, align='right', num_fmt=rp_fmt, bg=LGREEN,
          fg=GREEN)
    _cell(ws4, r4, 4, diff_tot, bold=True, align='right', num_fmt=rp_fmt, bg=LGREEN,
          fg=GREEN if diff_tot > 0 else RED)
    _cell(ws4, r4, 5, f'{pct_tot}%', bold=True, align='center', bg=LGREEN,
          fg=GREEN if pct_tot > 0 else RED)
    r4 += 1
    if sal.get('increase_pct'):
        _cell(ws4, r4, 1, 'Kenaikan Gaji Tercatat', bold=True)
        _cell(ws4, r4, 2, f"{sal.get('increase_pct')}%", bold=True, fg=GREEN)
    if sal.get('notes'):
        r4 += 1
        _cell(ws4, r4, 1, 'Catatan', bold=True)
        _cell(ws4, r4, 2, sal.get('notes', ''), wrap=True)
        ws4.merge_cells(start_row=r4, start_column=2, end_row=r4, end_column=5)
    _border_range(ws4, r4 - len(gaji_labels) - 4, r4, 1, 5)

    # Approval reviews
    if data['all_reviews']:
        r4 += 3
        _hdr(ws4, r4, 1, 'PERSETUJUAN REVIEWER', bg=BLUE, merge_to=5, size=12)
        r4 += 1
        _hdr(ws4, r4, 1, 'Reviewer', bg=LBLUE, fg=BLUE)
        _hdr(ws4, r4, 2, 'Role', bg=LBLUE, fg=BLUE)
        _hdr(ws4, r4, 3, 'Status', bg=LBLUE, fg=BLUE)
        _hdr(ws4, r4, 4, 'Tanggal', bg=LBLUE, fg=BLUE)
        _hdr(ws4, r4, 5, 'Catatan', bg=LBLUE, fg=BLUE)
        r4 += 1
        for rv in data['all_reviews']:
            _cell(ws4, r4, 1, rv.get('full_name') or '—', bold=True)
            _cell(ws4, r4, 2, rv.get('reviewer_role') or '—')
            st = rv.get('status','')
            _cell(ws4, r4, 3, st.upper(), align='center',
                  fg=GREEN if st == 'submitted' else GOLD)
            _cell(ws4, r4, 4, (rv.get('submitted_at') or '')[:16])
            _cell(ws4, r4, 5, rv.get('notes') or '—', wrap=True)
            ws4.row_dimensions[r4].height = 30
            r4 += 1
        _border_range(ws4, r4 - len(data['all_reviews']) - 1, r4-1, 1, 5)

    # Output
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    fname = f"Evaluasi_{emp['name'].replace(' ','_')}_{ev['periode']}_{date.today()}.xlsx"
    return send_file(
        out,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=fname
    )


@app.route('/eval/<int:eval_id>/export/pdf')
@login_required
def eval_export_pdf(eval_id):
    """Halaman cetak evaluasi (HTML) — user print to PDF dari browser."""
    db   = get_db()
    data = _eval_export_data(db, eval_id)
    if not data:
        abort(404)
    from datetime import datetime as _dtnow
    return render_template('eval_print.html', **data,
                           salary_cols=SALARY_COLS, salary_total=_salary_total,
                           now=_dtnow.now().strftime('%d %b %Y %H:%M'))


# ─── Admin Routes ──────────────────────────────────────────────────────────────




@app.route('/admin')
@login_required
def admin():
    db = get_db()
    data = {}
    divisi_list = get_divisi_list(db)
    for d in divisi_list:
        cats = db.execute('''SELECT sc.id, sc.name, COUNT(si.id) AS item_count
                             FROM skill_categories sc LEFT JOIN skill_items si ON si.category_id=sc.id
                             WHERE sc.divisi=? GROUP BY sc.id ORDER BY sc.sort_order''', (d,)).fetchall()
        comp = db.execute('SELECT COUNT(*) AS c FROM competency_items WHERE divisi=?', (d,)).fetchone()['c']
        abl  = db.execute('SELECT COUNT(*) AS c FROM ability_items WHERE divisi=?', (d,)).fetchone()['c']
        data[d] = {'cats': cats, 'comp_count': comp, 'ability_count': abl}
    return render_template('admin.html', data=data)

@app.route('/admin/divisi/<path:divisi>')
@login_required
def admin_divisi(divisi):
    db = get_db()
    cats = db.execute('SELECT * FROM skill_categories WHERE divisi=? ORDER BY sort_order', (divisi,)).fetchall()
    _all_items = db.execute(
        'SELECT * FROM skill_items WHERE category_id IN (SELECT id FROM skill_categories WHERE divisi=?) ORDER BY sort_order',
        (divisi,)
    ).fetchall()
    items_by_cat: dict = {}
    for _it in _all_items:
        items_by_cat.setdefault(_it['category_id'], []).append(_it)
    comp_items    = db.execute('SELECT * FROM competency_items WHERE divisi=? ORDER BY sort_order', (divisi,)).fetchall()
    ability_items = db.execute('SELECT * FROM ability_items WHERE divisi=? ORDER BY sort_order', (divisi,)).fetchall()
    return render_template('admin_divisi.html', divisi=divisi, cats=cats,
                           items_by_cat=items_by_cat, comp_items=comp_items,
                           ability_items=ability_items)

# ─── Division Management ──────────────────────────────────────────────────────

@app.route('/admin/divisions', methods=['GET', 'POST'])
@superadmin_required
def admin_divisions():
    db = get_db()
    if request.method == 'POST':
        action = request.form.get('action')
        if action == 'add':
            name = request.form.get('name', '').strip()
            desc = request.form.get('description', '').strip()
            if name:
                try:
                    max_order = db.execute('SELECT COALESCE(MAX(sort_order),0) FROM divisions').fetchone()[0]
                    db.execute('INSERT INTO divisions(name, description, sort_order) VALUES(?,?,?)',
                               (name, desc, max_order + 1))
                    db.commit()
                    flash(f'Divisi "{name}" ditambahkan', 'success')
                except Exception:
                    flash('Nama divisi sudah ada', 'danger')
        elif action == 'edit':
            div_id = request.form.get('id')
            name   = request.form.get('name', '').strip()
            desc   = request.form.get('description', '').strip()
            sort_o = request.form.get('sort_order', '0')
            is_act = 1 if request.form.get('is_active') else 0
            if div_id and name:
                db.execute('UPDATE divisions SET name=?,description=?,sort_order=?,is_active=? WHERE id=?',
                           (name, desc, int(sort_o), is_act, int(div_id)))
                db.commit()
                flash('Divisi diperbarui', 'success')
        elif action == 'delete':
            div_id = request.form.get('id')
            if div_id:
                emp_count = db.execute('SELECT COUNT(*) FROM employees WHERE divisi=(SELECT name FROM divisions WHERE id=?) AND is_active=1',
                                       (int(div_id),)).fetchone()[0]
                if emp_count > 0:
                    flash(f'Tidak bisa hapus divisi — masih ada {emp_count} karyawan aktif', 'danger')
                else:
                    db.execute('DELETE FROM divisions WHERE id=?', (int(div_id),))
                    db.commit()
                    flash('Divisi dihapus', 'warning')
        return redirect(url_for('admin_divisions'))
    divisions = db.execute('SELECT * FROM divisions ORDER BY sort_order, name').fetchall()
    emp_counts = {r['divisi']: r['c'] for r in db.execute(
        'SELECT divisi, COUNT(*) AS c FROM employees WHERE is_active=1 GROUP BY divisi').fetchall()}
    return render_template('admin_divisions.html', divisions=divisions, emp_counts=emp_counts)


# ─── Promote Employee to App User ─────────────────────────────────────────────

@app.route('/emp/<int:emp_id>/promote-role', methods=['POST'])
@superadmin_required
def emp_promote_role(emp_id):
    db  = get_db()
    emp = db.execute('SELECT * FROM employees WHERE id=?', (emp_id,)).fetchone()
    if not emp:
        flash('Karyawan tidak ditemukan', 'danger')
        return redirect(url_for('karyawan'))

    # role portal: hanya set jika access_portal dicentang, default viewer
    role = request.form.get('role', 'viewer') if request.form.get('access_portal') else 'viewer'
    apps = db.execute('SELECT slug, name FROM superapp_apps WHERE is_active=1').fetchall()

    def _save_app_access(uid):
        db.execute('DELETE FROM user_app_access WHERE user_id=?', (uid,))
        for a in apps:
            if request.form.get(f'access_{a["slug"]}'):
                app_role = request.form.get(f'role_{a["slug"]}', 'admin')
                db.execute('INSERT INTO user_app_access(user_id,app_slug,app_role,is_active) VALUES(?,?,?,1)',
                           (uid, a['slug'], app_role))

    if emp['user_id']:
        db.execute('UPDATE users SET role=? WHERE id=?', (role, emp['user_id']))
        _save_app_access(emp['user_id'])
        db.commit()
        flash(f'Role & akses {emp["name"]} diperbarui', 'success')
    else:
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '').strip()
        if not username:
            flash('Username diperlukan', 'danger')
            return redirect(url_for('karyawan'))
        existing = db.execute('SELECT id FROM users WHERE username=?', (username,)).fetchone()
        if existing:
            db.execute('UPDATE employees SET user_id=? WHERE id=?', (existing['id'], emp_id))
            db.execute('UPDATE users SET role=?, full_name=?, is_active=1 WHERE id=?',
                       (role, emp['name'], existing['id']))
            _save_app_access(existing['id'])
            db.commit()
            flash(f'{emp["name"]} dihubungkan ke akun "{username}"', 'success')
        else:
            if not password:
                flash('Password diperlukan untuk akun baru', 'danger')
                return redirect(url_for('karyawan'))
            new_uid = db.execute(
                'INSERT INTO users(username,password_hash,full_name,role) VALUES(?,?,?,?)',
                (username, generate_password_hash(password), emp['name'], role)
            ).lastrowid
            db.execute('UPDATE employees SET user_id=? WHERE id=?', (new_uid, emp_id))
            _save_app_access(new_uid)
            db.commit()
            flash(f'Akun "{username}" dibuat untuk {emp["name"]}', 'success')
    return redirect(url_for('karyawan'))


# ─── Role Management ──────────────────────────────────────────────────────────

@app.route('/admin/roles', methods=['GET', 'POST'])
@permission_required('manage_roles')
def admin_roles():
    db = get_db()
    if request.method == 'POST':
        action = request.form.get('action')
        if action == 'add':
            name = request.form.get('name', '').strip().lower().replace(' ', '_')
            desc = request.form.get('description', '').strip()
            if name:
                try:
                    db.execute('INSERT INTO roles(name,description,is_system,app_slug) VALUES(?,?,0,?)', (name, desc, 'evaluasi'))
                    db.commit()
                    flash(f'Role "{name}" ditambahkan', 'success')
                except Exception:
                    flash('Nama role sudah ada', 'danger')
        elif action == 'delete':
            rname = request.form.get('role_name', '')
            is_sys = db.execute('SELECT is_system FROM roles WHERE name=?', (rname,)).fetchone()
            if is_sys and is_sys['is_system']:
                flash('Role sistem tidak bisa dihapus', 'danger')
            else:
                db.execute('DELETE FROM roles WHERE name=?', (rname,))
                db.execute('DELETE FROM role_permissions WHERE role_name=?', (rname,))
                db.commit()
                flash(f'Role "{rname}" dihapus', 'warning')
        return redirect(url_for('admin_roles'))
    roles = db.execute("SELECT * FROM roles WHERE app_slug='evaluasi' OR app_slug='' ORDER BY is_system DESC, name").fetchall()
    perms_by_role = {}
    for r in roles:
        perms_by_role[r['name']] = get_role_permissions(db, r['name'])
    return render_template('admin_roles.html', roles=roles, perms_by_role=perms_by_role,
                           all_permissions=ALL_PERMISSIONS)

@app.route('/admin/roles/<role_name>/permissions', methods=['POST'])
@permission_required('manage_roles')
def admin_role_perms(role_name):
    db = get_db()
    selected = set(request.form.getlist('permissions'))
    db.execute('DELETE FROM role_permissions WHERE role_name=?', (role_name,))
    for perm in selected:
        if perm in ALL_PERMISSIONS:
            db.execute('INSERT OR IGNORE INTO role_permissions(role_name,permission) VALUES(?,?)',
                       (role_name, perm))
    db.commit()
    flash(f'Permission role "{role_name}" diperbarui', 'success')
    return redirect(url_for('admin_roles'))

# ─── Template Editor ──────────────────────────────────────────────────────────

@app.route('/admin/template/<path:divisi>', methods=['POST'])
@permission_required('manage_template')
def admin_template_edit(divisi):
    db     = get_db()
    action = request.form.get('action')

    if action == 'add_category':
        name  = request.form.get('name', '').strip()
        if name:
            max_o = db.execute('SELECT COALESCE(MAX(sort_order),0) FROM skill_categories WHERE divisi=?', (divisi,)).fetchone()[0]
            db.execute('INSERT INTO skill_categories(divisi,name,sort_order) VALUES(?,?,?)', (divisi, name, max_o+1))
            db.commit()
            flash(f'Kategori "{name}" ditambahkan', 'success')

    elif action == 'edit_category':
        cid  = request.form.get('id')
        name = request.form.get('name', '').strip()
        if cid and name:
            db.execute('UPDATE skill_categories SET name=? WHERE id=? AND divisi=?', (name, cid, divisi))
            db.commit()
            flash('Kategori diperbarui', 'success')

    elif action == 'delete_category':
        cid = request.form.get('id')
        if cid:
            db.execute('DELETE FROM skill_categories WHERE id=? AND divisi=?', (cid, divisi))
            db.commit()
            flash('Kategori dihapus', 'warning')

    elif action == 'add_skill':
        cat_id = request.form.get('category_id')
        name   = request.form.get('name', '').strip()
        desc   = request.form.get('description', '').strip()
        bobot  = request.form.get('bobot', '1')
        if cat_id and name:
            max_o = db.execute('SELECT COALESCE(MAX(sort_order),0) FROM skill_items WHERE category_id=?', (cat_id,)).fetchone()[0]
            db.execute('INSERT INTO skill_items(category_id,name,description,bobot,sort_order) VALUES(?,?,?,?,?)',
                       (cat_id, name, desc, float(bobot), max_o+1))
            db.commit()
            flash(f'Skill "{name}" ditambahkan', 'success')

    elif action == 'edit_skill':
        sid   = request.form.get('id')
        name  = request.form.get('name', '').strip()
        desc  = request.form.get('description', '').strip()
        bobot = request.form.get('bobot', '1')
        if sid and name:
            db.execute('UPDATE skill_items SET name=?,description=?,bobot=? WHERE id=?',
                       (name, desc, float(bobot), sid))
            db.commit()
            flash('Skill diperbarui', 'success')

    elif action == 'delete_skill':
        sid = request.form.get('id')
        if sid:
            db.execute('DELETE FROM skill_items WHERE id=?', (sid,))
            db.commit()
            flash('Skill dihapus', 'warning')

    elif action == 'add_competency':
        measurement = request.form.get('point_measurement', '').strip()
        bobot       = request.form.get('bobot', '0')
        is_hs       = 1 if request.form.get('is_hardskill') else 0
        if measurement:
            max_o = db.execute('SELECT COALESCE(MAX(sort_order),0) FROM competency_items WHERE divisi=?', (divisi,)).fetchone()[0]
            db.execute('INSERT INTO competency_items(divisi,point_measurement,bobot,sort_order,is_hardskill) VALUES(?,?,?,?,?)',
                       (divisi, measurement, float(bobot), max_o+1, is_hs))
            db.commit()
            flash('Kompetensi ditambahkan', 'success')

    elif action == 'edit_competency':
        cid  = request.form.get('id')
        meas = request.form.get('point_measurement', '').strip()
        bob  = request.form.get('bobot', '0')
        is_hs = 1 if request.form.get('is_hardskill') else 0
        if cid and meas:
            db.execute('UPDATE competency_items SET point_measurement=?,bobot=?,is_hardskill=? WHERE id=?',
                       (meas, float(bob), is_hs, cid))
            db.commit()
            flash('Kompetensi diperbarui', 'success')

    elif action == 'delete_competency':
        cid = request.form.get('id')
        if cid:
            db.execute('DELETE FROM competency_items WHERE id=?', (cid,))
            db.commit()
            flash('Kompetensi dihapus', 'warning')

    elif action == 'add_ability':
        name = request.form.get('name','').strip()
        if name:
            max_o = db.execute('SELECT COALESCE(MAX(sort_order),0) FROM ability_items WHERE divisi=?', (divisi,)).fetchone()[0]
            db.execute('INSERT INTO ability_items(divisi,name,desc_a,desc_b,desc_c,desc_d,sort_order) VALUES(?,?,?,?,?,?,?)',
                       (divisi, name, request.form.get('desc_a',''), request.form.get('desc_b',''),
                        request.form.get('desc_c',''), request.form.get('desc_d',''), max_o+1))
            db.commit()
            flash('Ability ditambahkan', 'success')

    elif action == 'edit_ability':
        aid = request.form.get('id')
        if aid:
            db.execute('UPDATE ability_items SET name=?,desc_a=?,desc_b=?,desc_c=?,desc_d=? WHERE id=?',
                       (request.form.get('name',''), request.form.get('desc_a',''), request.form.get('desc_b',''),
                        request.form.get('desc_c',''), request.form.get('desc_d',''), aid))
            db.commit()
            flash('Ability diperbarui', 'success')

    elif action == 'delete_ability':
        aid = request.form.get('id')
        if aid:
            db.execute('DELETE FROM ability_items WHERE id=?', (aid,))
            db.commit()
            flash('Ability dihapus', 'warning')

    return redirect(url_for('admin_divisi', divisi=divisi) + '#' + request.form.get('tab','tab-skill'))

# ─── Salary (MFA protected) ───────────────────────────────────────────────────

# Kolom yang dienkripsi di tabel employee_salary
SALARY_ENC_FIELDS = frozenset(['base_salary', 'al_001', 'al_002', 'al_003', 'al_004'])

def _get_fernet():
    if not _CRYPTO_OK:
        return None
    key = os.environ.get('FIELD_ENCRYPT_KEY', '')
    if not key:
        return None
    try:
        return Fernet(key.encode() if isinstance(key, str) else key)
    except Exception:
        return None

def _fenc(val):
    """Enkripsi nilai numerik gaji. Return string Fernet token, atau float jika key tidak ada."""
    if val is None:
        return None
    f = _get_fernet()
    if f is None:
        return float(val) if not isinstance(val, str) else val
    return f.encrypt(str(float(val)).encode()).decode()

def _fdec(val):
    """Dekripsi nilai gaji. Handle: None, float legacy, string token terenkripsi."""
    if val is None:
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)  # data lama belum terenkripsi
    f = _get_fernet()
    if f is None:
        try:
            return float(val)
        except (ValueError, TypeError):
            return 0.0
    try:
        return float(f.decrypt(val.encode() if isinstance(val, str) else val).decode())
    except Exception:
        try:
            return float(val)  # fallback: coba parse langsung (data lama)
        except (ValueError, TypeError):
            return 0.0

def _dec_sal_row(row):
    """Konversi sqlite3.Row salary ke dict dengan nilai terdekripsi."""
    d = dict(row)
    for field in SALARY_ENC_FIELDS:
        if field in d:
            d[field] = _fdec(d[field])
        prev = f'p_{field}'
        if prev in d:
            d[prev] = _fdec(d[prev])
    return d

SALARY_COLS = [
    ('base_salary', 'SALARY',    'Gaji Pokok'),
    ('al_001',      'AL_001',    'Tunjangan Jabatan'),
    ('al_002',      'AL_002',    'Tunjangan Komunikasi'),
    ('al_003',      'AL_003',    'Tunjangan Performance'),
    ('al_004',      'AL_004',    'Tunjangan Kehadiran'),
]

def _salary_total(row):
    if not row:
        return 0
    return sum(_fdec(row[col]) for col, *_ in SALARY_COLS)

@app.route('/salary')
@permission_required('view_salary')
@mfa_challenge_required('mfa_salary_verified')
def salary_table():
    db       = get_db()
    year     = request.args.get('year', date.today().year, type=int)
    prev_year = year - 1
    emps = db.execute('''
        SELECT e.*, s.id AS sal_id,
               s.base_salary, s.al_001, s.al_002, s.al_003, s.al_004,
               s.increase_pct, s.increase_date, s.notes,
               p.base_salary AS p_base_salary, p.al_001 AS p_al_001,
               p.al_002 AS p_al_002, p.al_003 AS p_al_003, p.al_004 AS p_al_004
        FROM employees e
        LEFT JOIN employee_salary s ON s.employee_id = e.id AND s.year = ?
        LEFT JOIN employee_salary p ON p.employee_id = e.id AND p.year = ?
        WHERE e.is_active = 1
        ORDER BY 
            CASE WHEN e.divisi = 'Telegram Core' THEN 3
                 WHEN e.employment_type = 'tetap' THEN 1
                 ELSE 2 END ASC,
            e.divisi, e.name
    ''', (year, prev_year)).fetchall()
    emps = [_dec_sal_row(r) for r in emps]
    
    today = date.today()
    def get_active_fraction(emp, Y):
        if emp.get('employment_type') == 'tetap' or emp.get('divisi') == 'Telegram Core':
            return 1.0
        start_str = emp.get('contract_start')
        end_str = emp.get('contract_end')
        if not start_str and not end_str:
            return 1.0
        s_y = date(Y, 1, 1)
        e_y = date(Y, 12, 31)
        try:
            c_start = datetime.strptime(start_str, '%Y-%m-%d').date() if start_str else s_y
        except Exception:
            c_start = s_y
        try:
            c_end = datetime.strptime(end_str, '%Y-%m-%d').date() if end_str else e_y
        except Exception:
            c_end = e_y
        o_start = max(s_y, c_start)
        o_end = min(e_y, c_end)
        if o_start > o_end:
            return 0.0
        active_days = (o_end - o_start).days + 1
        total_days = 366.0 if (Y % 4 == 0 and (Y % 100 != 0 or Y % 400 == 0)) else 365.0
        return active_days / total_days

    for emp in emps:
        days_left = None
        is_expired = False
        if emp.get('contract_end'):
            try:
                end_date = datetime.strptime(emp['contract_end'], '%Y-%m-%d').date()
                days_left = (end_date - today).days
                is_expired = end_date < today
            except Exception:
                pass
        emp['days_left'] = days_left if days_left is not None else 9999
        emp['is_expired'] = is_expired
        emp['active_fraction'] = get_active_fraction(emp, year)
        emp['active_fraction_prev'] = get_active_fraction(emp, prev_year)
        emp['active_fraction_next'] = get_active_fraction(emp, year + 1)
    years = [r['year'] for r in db.execute(
        'SELECT DISTINCT year FROM employee_salary ORDER BY year DESC').fetchall()]
    if year not in years:
        years = sorted(set(years + [year]), reverse=True)
    return render_template('salary.html', emps=emps, year=year, prev_year=prev_year, years=years,
                           salary_cols=SALARY_COLS, salary_total=_salary_total,
                           can_edit='manage_salary' in get_role_permissions(db, session.get('user_role','')))

@app.route('/salary/save', methods=['POST'])
@permission_required('manage_salary')
@mfa_challenge_required('mfa_salary_verified')
def salary_save():
    db      = get_db()
    emp_id  = request.form.get('employee_id', type=int)
    year    = request.form.get('year', type=int)
    field   = request.form.get('field', '').strip()
    value   = request.form.get('value', '').strip()
    valid_fields = {c for c, *_ in SALARY_COLS} | {'increase_pct', 'increase_date', 'notes'}
    if not emp_id or not year or field not in valid_fields:
        return jsonify({'ok': False, 'msg': 'Parameter tidak valid'})
    try:
        if field in ('notes', 'increase_date'):
            val = value
        elif field in SALARY_ENC_FIELDS:
            val = _fenc(float(value))
        else:
            val = float(value)
    except ValueError:
        val = _fenc(0.0) if field in SALARY_ENC_FIELDS else 0.0
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    db.execute(f'''
        INSERT INTO employee_salary(employee_id, year, {field}, updated_at)
        VALUES(?,?,?,?)
        ON CONFLICT(employee_id, year) DO UPDATE SET {field}=excluded.{field}, updated_at=excluded.updated_at
    ''', (emp_id, year, val, now))
    db.commit()
    # Return updated row totals (decrypt before computing)
    row = db.execute('SELECT * FROM employee_salary WHERE employee_id=? AND year=?',
                     (emp_id, year)).fetchone()
    row = _dec_sal_row(row) if row else row
    total = _salary_total(row)
    next_pct  = (row['increase_pct'] or 0) if row else 0
    next_total = round(total * (1 + next_pct / 100))
    return jsonify({'ok': True, 'total': total, 'next_total': next_total})

@app.route('/salary/add-year', methods=['POST'])
@permission_required('manage_salary')
@mfa_challenge_required('mfa_salary_verified')
def salary_add_year():
    """Buat data gaji untuk tahun baru (salin dari tahun sumber jika ada)."""
    db       = get_db()
    year     = request.form.get('year', type=int)
    src_year = request.form.get('src_year', type=int)
    now      = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    if not year or year < 2000 or year > 2100:
        return jsonify({'ok': False, 'msg': 'Tahun tidak valid'})

    # Copy from source year if provided and data exists
    copied = 0
    if src_year:
        rows = db.execute('SELECT * FROM employee_salary WHERE year=?', (src_year,)).fetchall()
        pct_apply = request.form.get('apply_increase', '0') == '1'
        for r in rows:
            pct    = (r['increase_pct'] or 0) if pct_apply else 0
            factor = 1 + pct / 100
            try:
                db.execute('''
                    INSERT OR IGNORE INTO employee_salary
                    (employee_id, year, base_salary, al_001, al_002, al_003, al_004,
                     increase_pct, notes, updated_at)
                    VALUES(?,?,?,?,?,?,?,0,'',?)
                ''', (r['employee_id'], year,
                      _fenc(round(_fdec(r['base_salary']) * factor)),
                      _fenc(round(_fdec(r['al_001']) * factor)),
                      _fenc(round(_fdec(r['al_002']) * factor)),
                      _fenc(round(_fdec(r['al_003']) * factor)),
                      _fenc(round(_fdec(r['al_004']) * factor)),
                      now))
                copied += 1
            except Exception:
                pass
    else:
        # Buat baris kosong untuk semua karyawan aktif
        emps = db.execute("SELECT id FROM employees WHERE is_active=1").fetchall()
        for e in emps:
            try:
                db.execute('''
                    INSERT OR IGNORE INTO employee_salary(employee_id, year, updated_at)
                    VALUES(?,?,?)
                ''', (e['id'], year, now))
                copied += 1
            except Exception:
                pass
    db.commit()

    years_available = [r['year'] for r in db.execute(
        'SELECT DISTINCT year FROM employee_salary ORDER BY year DESC').fetchall()]
    msg = (f'{copied} data gaji disalin dari tahun {src_year} ke {year}'
           if src_year else f'Tahun {year} dibuat dengan {copied} baris kosong')
    return jsonify({'ok': True, 'copied': copied, 'msg': msg, 'years': years_available})

@app.route('/emp/<int:emp_id>/salary', methods=['GET'])
@permission_required('view_salary')
@mfa_challenge_required('mfa_salary_verified')
def emp_salary_view(emp_id):
    db  = get_db()
    emp = db.execute('SELECT id, name FROM employees WHERE id=?', (emp_id,)).fetchone()
    if not emp:
        return jsonify({'ok': False, 'msg': 'Karyawan tidak ditemukan'})
    return jsonify({'ok': True, 'name': emp['name']})

@app.route('/emp/<int:emp_id>/salary', methods=['POST'])
@permission_required('manage_salary')
@mfa_challenge_required('mfa_salary_verified')
def emp_salary_update(emp_id):
    return jsonify({'ok': True, 'msg': 'Gunakan /salary untuk edit komponen gaji'})

# ─── Error handlers (audit trail) ─────────────────────────────────────────────

@app.errorhandler(404)
def err_404(e):
    try:
        db = get_db()
        db.execute('''INSERT INTO audit_errors(app_slug,user_id,username,url,method,error_code,error_type,error_msg,ip)
                      VALUES(?,?,?,?,?,?,?,?,?)''',
                   (session.get('active_app','portal'), session.get('user_id'), session.get('user_name',''),
                    request.path, request.method, 404, 'NotFound', str(e),
                    _real_ip()))
        db.commit()
    except Exception:
        pass
    return render_template('error.html', code=404, msg='Halaman tidak ditemukan'), 404

@app.errorhandler(500)
def err_500(e):
    import traceback as _tb
    tb = _tb.format_exc()
    # Gunakan koneksi BARU — koneksi lama (g.db) mungkin dalam state aborted transaction
    # (misal setelah psycopg2 error), sehingga tidak bisa digunakan untuk INSERT.
    _logged = False
    try:
        _econn = _pg_connect()
        _econn.autocommit = False
        _edb = _DBWrapper(_econn, is_pg=True)
        _edb.execute('''INSERT INTO audit_errors(app_slug,user_id,username,url,method,error_code,error_type,error_msg,traceback,ip)
                        VALUES(?,?,?,?,?,?,?,?,?,?)''',
                     (session.get('active_app','portal'), session.get('user_id'), session.get('user_name',''),
                      request.path, request.method, 500, type(e).__name__, str(e), tb[:3000],
                      _real_ip()))
        _edb.commit()
        _edb.close()
        _logged = True
    except Exception:
        pass
    if not _logged:
        # Fallback: tulis ke file log jika DB tidak bisa dihubungi
        try:
            import datetime as _dt
            _log_path = os.path.join(os.path.dirname(__file__), 'error_fallback.log')
            with open(_log_path, 'a', encoding='utf-8') as _lf:
                _lf.write(f"[{_dt.datetime.now().isoformat()}] 500 {request.method} {request.path} "
                          f"ip={_real_ip()} user={session.get('user_name','')} "
                          f"err={type(e).__name__}: {e}\n{tb[:2000]}\n---\n")
        except Exception:
            pass
    return render_template('error.html', code=500, msg='Terjadi kesalahan server'), 500

# ─── Portal: Update Center ─────────────────────────────────────────────────────

@app.route('/portal/update')
@login_required
def portal_update():
    role = session.get('user_role', '')
    db   = get_db()
    settings = get_settings(db)
    notify_roles  = [r.strip() for r in settings.get('update_notify_roles',  'superadmin,admin').split(',')]
    trigger_roles = [r.strip() for r in settings.get('update_trigger_roles', 'superadmin').split(',')]
    if role not in notify_roles:
        abort(403)

    # Semua tag tersedia dari GitHub (cache di app_settings sebagai JSON)
    import json
    all_tags_raw = settings.get('update_all_tags', '[]')
    try:
        _raw = json.loads(all_tags_raw)
        if _raw and isinstance(_raw[0], str):
            _raw = [{'tag': t, 'notes': ''} for t in _raw]
        def _vp(v):
            try: return [int(x) for x in v.strip().lstrip('v').split('.')]
            except: return [0]
        _cur = _vp(VERSION)
        for item in _raw:
            _v = _vp(item.get('tag', ''))
            if _v == _cur:
                item['status'] = 'installed'
            elif _v > _cur:
                item['status'] = 'newer'   # lebih baru dari terpasang
            else:
                item['status'] = 'old'
        all_tags = _raw[:5]
        # Hitung update_available langsung dari all_tags (bukan dari DB)
        # agar selalu sinkron dengan daftar versi yang ditampilkan
        _newer = [i for i in all_tags if i.get('status') == 'newer']
        _latest_newer = max(_newer, key=lambda i: _vp(i['tag']), default=None)
        update_available_live = bool(_newer)
        latest_version_live   = _latest_newer['tag'].lstrip('v') if _latest_newer else settings.get('update_latest_version', '')
    except Exception:
        all_tags             = []
        update_available_live = settings.get('update_available', '0') == '1'
        latest_version_live   = settings.get('update_latest_version', '')

    # Status deploy sedang berjalan
    log_tail  = ''
    log_size  = 0
    deploy_done = False
    if os.path.exists(UPDATE_LOG_FILE):
        with open(UPDATE_LOG_FILE, 'r', errors='replace') as f:
            log_tail = f.read()
        log_size    = os.path.getsize(UPDATE_LOG_FILE)
        deploy_done = UPDATE_DONE_MARKER in log_tail or UPDATE_FAIL_MARKER in log_tail

    deploy_running = os.path.exists(UPDATE_TRIGGER_FILE) and not deploy_done
    # Cleanup trigger file jika log sudah menandai selesai/gagal
    if deploy_done and os.path.exists(UPDATE_TRIGGER_FILE):
        try: os.remove(UPDATE_TRIGGER_FILE)
        except Exception: pass

    return render_template('update_center.html',
        settings       = settings,
        current_version= VERSION,
        latest_version = latest_version_live,
        latest_tag     = settings.get('update_latest_tag', ''),
        update_available = update_available_live,
        release_notes  = settings.get('update_release_notes', ''),
        check_last     = settings.get('update_check_last', ''),
        all_tags       = all_tags,
        can_trigger    = role in trigger_roles,
        deploy_running = deploy_running,
        deploy_done    = deploy_done,
        log_tail       = log_tail[-8000:] if log_tail else '',
        log_size       = log_size,
    )


@app.route('/portal/update/check', methods=['POST'])
@login_required
def portal_update_check():
    role = session.get('user_role', '')
    settings = get_settings(get_db())
    notify_roles = [r.strip() for r in settings.get('update_notify_roles', 'superadmin,admin').split(',')]
    if role not in notify_roles:
        abort(403)
    check_for_updates()
    # Baca hasil cek untuk flash message yang informatif
    _db2 = get_db()
    _s2  = get_settings(_db2)
    _lv  = _s2.get('update_latest_version', '')
    if _s2.get('update_available', '0') == '1':
        flash(f'Update tersedia: v{_lv}', 'info')
    elif _lv:
        flash(f'Hive sudah versi terbaru (v{_lv}).', 'success')
    else:
        flash('Pengecekan selesai.', 'secondary')
    return redirect(url_for('portal_update'))


@app.route('/portal/update/trigger', methods=['POST'])
@login_required
def portal_update_trigger():
    role = session.get('user_role', '')
    db   = get_db()
    settings = get_settings(db)
    trigger_roles = [r.strip() for r in settings.get('update_trigger_roles', 'superadmin').split(',')]
    if role not in trigger_roles:
        abort(403)

    target_version = request.form.get('version', '').strip()
    if not target_version:
        flash('Versi tidak valid.', 'danger')
        return redirect(url_for('portal_update'))

    # Tulis trigger file — systemd path unit akan mendeteksi dan menjalankan deploy
    try:
        # Bersihkan log lama
        if os.path.exists(UPDATE_LOG_FILE):
            os.remove(UPDATE_LOG_FILE)
        with open(UPDATE_TRIGGER_FILE, 'w') as f:
            f.write(f'{target_version}\n')
        flash(f'Update ke v{target_version} dimulai! Pantau progress di bawah.', 'success')
    except Exception as e:
        flash(f'Gagal memulai update: {e}', 'danger')

    return redirect(url_for('portal_update'))


@app.route('/portal/update/log-poll')
@login_required
def portal_update_log_poll():
    """Non-blocking poll endpoint — kembalikan JSON chunk log baru dari posisi terakhir."""
    from flask import jsonify
    role = session.get('user_role', '')
    settings = get_settings(get_db())
    notify_roles = [r.strip() for r in settings.get('update_notify_roles', 'superadmin,admin').split(',')]
    if role not in notify_roles:
        abort(403)

    pos     = request.args.get('pos', 0, type=int)
    lines   = []
    new_pos = pos
    done    = False
    failed  = False

    if os.path.exists(UPDATE_LOG_FILE):
        try:
            with open(UPDATE_LOG_FILE, 'r', errors='replace') as f:
                f.seek(pos)
                chunk = f.read(8192)
                new_pos = f.tell()
            if chunk:
                lines = chunk.splitlines()
                done   = UPDATE_DONE_MARKER in chunk
                failed = UPDATE_FAIL_MARKER in chunk
        except Exception:
            pass

    running = os.path.exists(UPDATE_TRIGGER_FILE) and not done and not failed
    # Cleanup trigger file jika deploy selesai
    if (done or failed) and os.path.exists(UPDATE_TRIGGER_FILE):
        try: os.remove(UPDATE_TRIGGER_FILE)
        except Exception: pass

    return jsonify({'lines': lines, 'pos': new_pos, 'done': done, 'failed': failed, 'running': running})


@app.route('/portal/update/cancel', methods=['POST'])
@login_required
def portal_update_cancel():
    role = session.get('user_role', '')
    settings = get_settings(get_db())
    trigger_roles = [r.strip() for r in settings.get('update_trigger_roles', 'superadmin').split(',')]
    if role not in trigger_roles:
        abort(403)
    if os.path.exists(UPDATE_TRIGGER_FILE):
        os.remove(UPDATE_TRIGGER_FILE)
        flash('Trigger dibatalkan (jika deploy sudah berjalan, tidak bisa dihentikan di tengah jalan).', 'warning')
    return redirect(url_for('portal_update'))


# ─── Kinerja Task: Individual & Tim ──────────────────────────────────────────

@app.route('/kinerja/individu/<int:emp_id>')
@login_required
def kinerja_individu(emp_id):
    db   = get_db()
    emp  = db.execute('SELECT * FROM employees WHERE id=?', (emp_id,)).fetchone()
    if not emp:
        abort(404)

    date_from = request.args.get('from', '')
    date_to   = request.args.get('to', '')
    benchmark = float(request.args.get('benchmark', 100))

    perf = calc_task_perf(db, emp_id, date_from, date_to, benchmark)

    # Riwayat evaluasi terakhir
    evals = db.execute('''
        SELECT id, periode, final_total, task_score, status, task_date_from, task_date_to
        FROM evaluations WHERE employee_id=? ORDER BY id DESC LIMIT 6
    ''', (emp_id,)).fetchall()

    # Skor rata-rata tim satu divisi (untuk perbandingan)
    divisi_avg = None
    if emp['divisi']:
        row = db.execute('''
            SELECT AVG(e.final_total) AS avg_score
            FROM evaluations e JOIN employees em ON em.id=e.employee_id
            WHERE em.divisi=? AND e.status='final'
        ''', (emp['divisi'],)).fetchone()
        divisi_avg = round(row['avg_score'], 1) if row and row['avg_score'] else None

    return render_template('kinerja_individu.html',
        emp=emp, perf=perf, evals=evals,
        divisi_avg=divisi_avg,
        date_from=date_from, date_to=date_to, benchmark=benchmark,
    )


@app.route('/kinerja/tim')
@login_required
def kinerja_tim():
    db      = get_db()
    divisi  = request.args.get('divisi', '')
    date_from = request.args.get('from', '')
    date_to   = request.args.get('to', '')
    benchmark = float(request.args.get('benchmark', 100))

    # Daftar semua divisi
    divisi_list = [r['divisi'] for r in db.execute(
        "SELECT DISTINCT divisi FROM employees WHERE is_active=1 AND divisi!='' ORDER BY divisi"
    ).fetchall()]

    # Karyawan di divisi terpilih (atau semua jika tidak filter)
    if divisi:
        emps = db.execute(
            "SELECT * FROM employees WHERE is_active=1 AND divisi=? ORDER BY name", (divisi,)
        ).fetchall()
    else:
        emps = db.execute(
            "SELECT * FROM employees WHERE is_active=1 ORDER BY divisi, name"
        ).fetchall()

    members = []
    for emp in emps:
        bm   = benchmark if benchmark != 100 else get_benchmark_for_emp(db, emp)
        perf = calc_task_perf(db, emp['id'], date_from, date_to, bm)
        last_eval = db.execute('''
            SELECT final_total, task_score, status, periode FROM evaluations
            WHERE employee_id=? ORDER BY id DESC LIMIT 1
        ''', (emp['id'],)).fetchone()
        members.append({
            'emp':        emp,
            'perf':       perf,
            'bm':         bm,
            'last_eval':  last_eval,
        })

    # Urutkan: task_score tertinggi dulu
    members.sort(key=lambda x: x['perf']['task_score'], reverse=True)
    # Tambahkan rank
    for i, m in enumerate(members, 1):
        m['rank'] = i

    # Statistik tim
    scores = [m['perf']['task_score'] for m in members if m['perf']['task_score'] > 0]
    tim_avg  = round(sum(scores) / len(scores), 1) if scores else 0
    tim_max  = max(scores) if scores else 0
    tim_min  = min(scores) if scores else 0

    return render_template('kinerja_tim.html',
        members=members, divisi=divisi, divisi_list=divisi_list,
        date_from=date_from, date_to=date_to, benchmark=benchmark,
        tim_avg=tim_avg, tim_max=tim_max, tim_min=tim_min,
    )


@app.route('/kinerja/analitik')
@login_required
def kinerja_analitik():
    db        = get_db()
    divisi    = request.args.get('divisi', '')
    date_from = request.args.get('from', '')
    date_to   = request.args.get('to', '')
    benchmark = float(request.args.get('benchmark', 100))
    level     = request.args.get('level', '')

    divisi_list = [r['divisi'] for r in db.execute(
        "SELECT DISTINCT divisi FROM employees WHERE is_active=1 AND divisi!='' ORDER BY divisi"
    ).fetchall()]

    # Filter karyawan
    q = "SELECT * FROM employees WHERE is_active=1"
    params = []
    if divisi:
        q += " AND divisi=?"; params.append(divisi)
    if level:
        q += " AND level=?"; params.append(level)
    q += " ORDER BY divisi, name"
    emps = db.execute(q, params).fetchall()

    members = []
    for emp in emps:
        analytics = calc_task_analytics(db, emp['id'], date_from, date_to)
        perf      = calc_task_perf(db, emp['id'], date_from, date_to, benchmark)
        members.append({
            'emp':       emp,
            'a':         analytics,
            'perf':      perf,
        })

    # Urutkan: total task selesai tertinggi
    members.sort(key=lambda x: x['a']['total_done'], reverse=True)
    for i, m in enumerate(members, 1):
        m['rank'] = i

    # Aggregate per divisi
    div_stats = {}
    for m in members:
        d = m['emp']['divisi'] or '—'
        if d not in div_stats:
            div_stats[d] = {'count': 0, 'total_done': 0, 'ontime': 0,
                            'delay': 0, 'overtime': 0, 'score_sum': 0}
        s = div_stats[d]
        s['count']      += 1
        s['total_done'] += m['a']['total_done']
        s['ontime']     += m['a']['done_ontime']
        s['delay']      += m['a']['done_delay']
        s['overtime']   += m['a']['open_overtime']
        s['score_sum']  += m['perf']['task_score']
    for d in div_stats:
        n = div_stats[d]['count']
        div_stats[d]['avg_score'] = round(div_stats[d]['score_sum'] / n, 1) if n else 0

    return render_template('kinerja_analitik.html',
        members=members, divisi=divisi, divisi_list=divisi_list,
        date_from=date_from, date_to=date_to, benchmark=benchmark,
        level=level, div_stats=div_stats,
    )


@app.route('/api/kinerja/task-score/<int:emp_id>')
@login_required
def api_task_score(emp_id):
    import json as _json
    db = get_db()
    date_from = request.args.get('from', '')
    date_to   = request.args.get('to', '')
    benchmark = float(request.args.get('benchmark', 100))
    perf = calc_task_perf(db, emp_id, date_from, date_to, benchmark)
    return app.response_class(
        response=_json.dumps(perf, ensure_ascii=False),
        mimetype='application/json'
    )


@app.route('/kinerja/divisi')
@login_required
def kinerja_divisi():
    """Dashboard kinerja per divisi: siapa yang perlu diratakan, siapa yang peak."""
    db        = get_db()
    divisi    = request.args.get('divisi', '')
    date_from = request.args.get('from', '')
    date_to   = request.args.get('to', '')

    divisi_list = [r['divisi'] for r in db.execute(
        "SELECT DISTINCT divisi FROM employees WHERE is_active=1 AND divisi!='' ORDER BY divisi"
    ).fetchall()]

    if divisi:
        emps = db.execute(
            "SELECT * FROM employees WHERE is_active=1 AND divisi=? ORDER BY name", (divisi,)
        ).fetchall()
    else:
        emps = db.execute(
            "SELECT * FROM employees WHERE is_active=1 ORDER BY divisi, name"
        ).fetchall()

    members = []
    for emp in emps:
        bm   = get_benchmark_for_emp(db, emp)
        perf = calc_task_perf(db, emp['id'], date_from, date_to, bm)
        ana  = calc_task_analytics(db, emp['id'], date_from, date_to)

        # Ambil skill rata-rata dari evaluasi terbaru
        last_eval = db.execute(
            "SELECT id, final_total, task_score, periode FROM evaluations "
            "WHERE employee_id=? ORDER BY id DESC LIMIT 1", (emp['id'],)
        ).fetchone()
        avg_skill = None
        if last_eval:
            row = db.execute(
                "SELECT AVG(score) AS avg FROM skill_scores WHERE eval_id=?",
                (last_eval['id'],)
            ).fetchone()
            avg_skill = round(row['avg'], 2) if row and row['avg'] else None

        # Distribusi task berdasarkan difficulty
        diff_dist = {}
        for row in db.execute('''
            SELECT COALESCE(t.difficulty,'Normal') AS diff, COUNT(*) AS cnt
            FROM pc_task_assignees ta
            JOIN pc_tasks t ON t.id=ta.task_id
            WHERE ta.employee_id=?
            GROUP BY diff
        ''', (emp['id'],)).fetchall():
            diff_dist[row['diff']] = row['cnt']

        # Hitung peer review rata-rata (structured dimensions)
        peer_avg = db.execute('''
            SELECT AVG(dim_kerjasama) AS kj, AVG(dim_komunikasi) AS km,
                   AVG(dim_keandalan) AS kd, AVG(dim_inisiatif) AS ini,
                   AVG(dim_kualitas) AS kl
            FROM peer_reviews
            WHERE eval_id IN (SELECT id FROM evaluations WHERE employee_id=?)
              AND dim_kerjasama IS NOT NULL
        ''', (emp['id'],)).fetchone()

        members.append({
            'emp':       emp,
            'perf':      perf,
            'ana':       ana,
            'bm':        bm,
            'last_eval': last_eval,
            'avg_skill': avg_skill,
            'diff_dist': diff_dist,
            'peer_avg':  dict(peer_avg) if peer_avg else {},
        })

    # Kelompokkan per divisi untuk perbandingan
    div_groups = {}
    for m in members:
        d = m['emp']['divisi'] or '—'
        div_groups.setdefault(d, []).append(m)

    div_summary = {}
    for d, mlist in div_groups.items():
        scores = [m['perf']['task_score'] for m in mlist]
        workloads = [m['ana']['total_done'] + m['ana']['total_open'] for m in mlist]
        avg_s  = round(sum(scores) / len(scores), 1) if scores else 0
        max_s  = max(scores) if scores else 0
        min_s  = min(scores) if scores else 0
        std_s  = 0.0
        if len(scores) > 1:
            import math
            mean = sum(scores) / len(scores)
            std_s = round(math.sqrt(sum((x - mean) ** 2 for x in scores) / len(scores)), 1)

        # Identifikasi siapa yang overloaded vs underloaded (task count)
        avg_wl = sum(workloads) / len(workloads) if workloads else 0
        for m in mlist:
            wl = m['ana']['total_done'] + m['ana']['total_open']
            if avg_wl > 0:
                m['wl_ratio'] = round(wl / avg_wl, 2)  # >1.3 = overloaded, <0.7 = underloaded
            else:
                m['wl_ratio'] = 1.0
            m['wl_flag'] = (
                'overloaded'  if m['wl_ratio'] > 1.3 else
                'underloaded' if m['wl_ratio'] < 0.7 else
                'balanced'
            )
            m['score_flag'] = (
                'peak'   if m['perf']['task_score'] >= avg_s * 1.2 else
                'low'    if m['perf']['task_score'] <= avg_s * 0.8 else
                'normal'
            )

        div_summary[d] = {
            'avg': avg_s, 'max': max_s, 'min': min_s, 'std': std_s,
            'count': len(mlist),
        }

    # Urutkan members dalam setiap divisi: skor tertinggi dulu
    for d in div_groups:
        div_groups[d].sort(key=lambda x: x['perf']['task_score'], reverse=True)

    # Rekomendasi pemerataan: pasangkan yang overloaded + peak dengan yang underloaded + low
    rebalance_hints = []
    for d, mlist in div_groups.items():
        overloaded = [m for m in mlist if m['wl_flag'] == 'overloaded']
        underloaded = [m for m in mlist if m['wl_flag'] == 'underloaded']
        if overloaded and underloaded:
            for ov in overloaded[:2]:
                for un in underloaded[:2]:
                    rebalance_hints.append({
                        'divisi': d,
                        'from_emp': ov['emp']['name'],
                        'from_wl': ov['wl_ratio'],
                        'from_score': ov['perf']['task_score'],
                        'to_emp': un['emp']['name'],
                        'to_wl': un['wl_ratio'],
                        'to_score': un['perf']['task_score'],
                    })

    return render_template('kinerja_divisi.html',
        div_groups=div_groups, div_summary=div_summary,
        divisi=divisi, divisi_list=divisi_list,
        date_from=date_from, date_to=date_to,
        rebalance_hints=rebalance_hints,
    )


@app.route('/api/kinerja/divisi-data')
@login_required
def api_kinerja_divisi_data():
    """JSON endpoint untuk chart performa per divisi."""
    import json as _json
    db        = get_db()
    date_from = request.args.get('from', '')
    date_to   = request.args.get('to', '')

    divisi_list = [r['divisi'] for r in db.execute(
        "SELECT DISTINCT divisi FROM employees WHERE is_active=1 AND divisi!='' ORDER BY divisi"
    ).fetchall()]

    result = []
    for div in divisi_list:
        emps = db.execute(
            "SELECT * FROM employees WHERE is_active=1 AND divisi=? ORDER BY name", (div,)
        ).fetchall()
        members_data = []
        for emp in emps:
            bm   = get_benchmark_for_emp(db, emp)
            perf = calc_task_perf(db, emp['id'], date_from, date_to, bm)
            ana  = calc_task_analytics(db, emp['id'], date_from, date_to)
            members_data.append({
                'name':         emp['name'],
                'level':        emp['level'],
                'task_score':   perf['task_score'],
                'total_done':   ana['total_done'],
                'total_open':   ana['total_open'],
                'open_overtime':ana['open_overtime'],
                'ontime_rate':  ana['ontime_rate'],
                'concurrent_max': ana['concurrent_max'],
                'benchmark':    bm,
            })
        result.append({'divisi': div, 'members': members_data})

    return app.response_class(
        response=_json.dumps(result, ensure_ascii=False),
        mimetype='application/json'
    )


# ─── AI Chatbot ────────────────────────────────────────────────────────────────

CHATBOT_TOOLS = [
    {
        "name": "execute_sql_query",
        "description": "Execute a read-only SQL SELECT query on the PostgreSQL database to retrieve application data. Use this as the primary tool to answer any questions about tickets, projects, employees, assets, bookings, or statistics.",
        "input_schema": {
            "type": "object",
            "properties": {
                "query": {"type": "string", "description": "The SQL SELECT query to execute. Must be read-only and return no more than 50 rows. Use table joins if necessary."}
            },
            "required": ["query"]
        }
    },
    {
        "name": "cari_tiket_support",
        "description": "Cari tiket support berdasarkan kata kunci subjek, nomor tiket, customer, atau status. Gunakan untuk pertanyaan tentang masalah/issue yang pernah dilaporkan.",
        "input_schema": {
            "type": "object",
            "properties": {
                "keyword": {"type": "string", "description": "Kata kunci pencarian (subjek, nomor tiket, atau nama customer)"},
                "status":  {"type": "string", "description": "Filter status: open, in_progress, resolved, closed (opsional)"},
                "limit":   {"type": "integer", "description": "Jumlah hasil maksimum (default 5)"}
            },
            "required": ["keyword"]
        }
    },
    {
        "name": "detail_tiket",
        "description": "Ambil detail lengkap satu tiket support termasuk deskripsi, riwayat, dan solusi.",
        "input_schema": {
            "type": "object",
            "properties": {
                "ticket_no": {"type": "string", "description": "Nomor tiket (misal TKT-001)"}
            },
            "required": ["ticket_no"]
        }
    },
    {
        "name": "cari_project",
        "description": "Cari project dan task berdasarkan nama atau status.",
        "input_schema": {
            "type": "object",
            "properties": {
                "keyword": {"type": "string", "description": "Nama project atau kata kunci"},
                "status":  {"type": "string", "description": "Filter status project (opsional): active, completed, on_hold"}
            },
            "required": ["keyword"]
        }
    },
    {
        "name": "cari_karyawan",
        "description": "Cari data karyawan berdasarkan nama, jabatan, atau divisi. TIDAK menampilkan data gaji.",
        "input_schema": {
            "type": "object",
            "properties": {
                "keyword": {"type": "string", "description": "Nama, jabatan, atau divisi karyawan"}
            },
            "required": ["keyword"]
        }
    },
    {
        "name": "cari_aset",
        "description": "Cari data aset, lisensi software, atau infrastruktur IT.",
        "input_schema": {
            "type": "object",
            "properties": {
                "keyword": {"type": "string", "description": "Nama aset, vendor, atau kategori"},
                "category": {"type": "string", "description": "Kategori aset (opsional)"}
            },
            "required": ["keyword"]
        }
    },
    {
        "name": "statistik_aplikasi",
        "description": "Ambil ringkasan statistik aplikasi: jumlah tiket, project aktif, karyawan, aset, dll.",
        "input_schema": {
            "type": "object",
            "properties": {
                "module": {"type": "string", "description": "Modul: support, project, talent, asset, atau all"}
            },
            "required": ["module"]
        }
    },
    {
        "name": "detail_karyawan",
        "description": "Ambil detail lengkap satu karyawan: jabatan, divisi, tipe kontrak, masa kerja, dan status probasi. Gunakan untuk pertanyaan seperti 'apakah X sudah selesai probasi?' atau 'berapa lama masa kerja si Y?'.",
        "input_schema": {
            "type": "object",
            "properties": {
                "nama": {"type": "string", "description": "Nama karyawan yang dicari"}
            },
            "required": ["nama"]
        }
    },
]

def _chatbot_exec_tool(db, name, inp):
    """Eksekusi tool chatbot dan return hasil sebagai string."""
    try:
        if name == 'execute_sql_query':
            query = inp.get('query', '').strip()
            if not query:
                return "Query kosong."
            
            clean_query = query.strip()
            import re
            # Remove multi-line comments
            clean_query = re.sub(r'/\*.*?\*/', '', clean_query, flags=re.DOTALL)
            # Remove single-line comments
            clean_query = re.sub(r'--.*$', '', clean_query, flags=re.MULTILINE)
            clean_query = clean_query.strip()
            
            clean_query_upper = clean_query.upper()
            if not clean_query_upper.startswith('SELECT'):
                return "Error: Hanya query SELECT (read-only) yang diizinkan."
                
            if any(w in clean_query_upper for w in ('SALARY', 'PASSWORD_HASH', 'RATE_MANDAYS')):
                return "Error: Akses ditolak. Informasi sensitif (gaji, rate mandays, password hash) bersifat rahasia dan tidak dapat diakses."
                
            if 'LIMIT' not in clean_query_upper:
                clean_query = clean_query.rstrip(';')
                clean_query += " LIMIT 50"
                
            rows = db.execute(clean_query).fetchall()
            if not rows:
                return "Query berhasil dieksekusi, tetapi tidak mengembalikan baris data."
                
            out = []
            keys = rows[0].keys()
            out.append(" | ".join(keys))
            out.append("-" * (len(keys) * 12))
            for r in rows:
                row_str = " | ".join(str(r[k]) if r[k] is not None else 'NULL' for k in keys)
                out.append(row_str)
            return "\n".join(out)

        elif name == 'cari_tiket_support':
            kw = f"%{inp.get('keyword','')}%"
            st = inp.get('status','')
            lm = min(int(inp.get('limit', 5)), 20)
            where = "AND t.status=?" if st else ""
            params = [kw, kw, kw] + ([st] if st else []) + [lm]
            rows = db.execute(f'''
                SELECT t.ticket_no, t.subject, t.status, t.priority,
                       cu.name as customer, t.reported_at, t.resolved_at
                FROM sc_tickets t
                JOIN sc_customers cu ON cu.id=t.customer_id
                WHERE (t.ticket_no LIKE ? OR t.subject LIKE ? OR cu.name LIKE ?) {where}
                ORDER BY t.id DESC LIMIT ?
            ''', params).fetchall()
            if not rows: return "Tidak ditemukan tiket yang cocok."
            out = []
            for r in rows:
                out.append(f"[{r['ticket_no']}] {r['subject']} | Status: {r['status']} | Prioritas: {r['priority'] or 'Medium'} | Customer: {r['customer']} | Lapor: {r['reported_at']}")
            return "\n".join(out)

        elif name == 'detail_tiket':
            tno = inp.get('ticket_no','').strip()
            row = db.execute('''
                SELECT t.*, cu.name as customer, st.name as type_name
                FROM sc_tickets t
                JOIN sc_customers cu ON cu.id=t.customer_id
                JOIN sc_support_types st ON st.id=t.support_type_id
                WHERE t.ticket_no=?
            ''', (tno,)).fetchone()
            if not row: return f"Tiket {tno} tidak ditemukan."
            hist = db.execute('''
                SELECT action, notes, new_value, created_at FROM sc_ticket_history
                WHERE ticket_id=? ORDER BY id DESC LIMIT 5
            ''', (row['id'],)).fetchall()
            out = [
                f"Tiket: {row['ticket_no']}",
                f"Subjek: {row['subject']}",
                f"Customer: {row['customer']}",
                f"Tipe: {row['type_name']}",
                f"Status: {row['status']} | Prioritas: {row['priority'] or 'Medium'}",
                f"Deskripsi: {row['description'] or '—'}",
                f"Solusi: {row['solution_note'] or '—'}",
            ]
            if hist:
                out.append("Riwayat terakhir:")
                for h in hist:
                    out.append(f"  [{h['created_at']}] {h['action']} — {h['notes'] or h['new_value'] or ''}")
            return "\n".join(out)

        elif name == 'cari_project':
            kw = f"%{inp.get('keyword','')}%"
            st = inp.get('status','')
            where = "AND p.status=?" if st else ""
            params = [kw, kw] + ([st] if st else [])
            rows = db.execute(f'''
                SELECT p.name, p.status, p.start_date, p.end_date, p.description,
                       COUNT(DISTINCT t.id) as task_count
                FROM pc_projects p
                LEFT JOIN pc_tasks t ON t.project_id=p.id
                WHERE (p.name LIKE ? OR p.description LIKE ?) {where}
                GROUP BY p.id ORDER BY p.id DESC LIMIT 10
            ''', params).fetchall()
            if not rows: return "Tidak ditemukan project yang cocok."
            out = []
            for r in rows:
                out.append(f"[{r['status']}] {r['name']} | Tasks: {r['task_count']} | {r['start_date'] or '?'} s/d {r['end_date'] or '?'}")
            return "\n".join(out)

        elif name == 'cari_karyawan':
            kw = f"%{inp.get('keyword','')}%"
            rows = db.execute('''
                SELECT name, jabatan, divisi, email, employment_type, contract_start FROM employees
                WHERE (name LIKE ? OR jabatan LIKE ? OR divisi LIKE ?) AND is_active=1
                ORDER BY divisi, name LIMIT 15
            ''', (kw, kw, kw)).fetchall()
            if not rows: return "Tidak ditemukan karyawan yang cocok."
            out = []
            for r in rows:
                tipe = r['employment_type'] or 'tetap'
                out.append(f"{r['name']} | {r['jabatan'] or '—'} | {r['divisi'] or '—'} | {tipe} | {r['email'] or '—'}")
            return "\n".join(out)

        elif name == 'detail_karyawan':
            import datetime as _dt
            nama = f"%{inp.get('nama', '').strip()}%"
            row = db.execute('''
                SELECT name, jabatan, divisi, level, employment_type,
                       contract_start, contract_end, email, phone, notes
                FROM employees
                WHERE name LIKE ? AND is_active=1
                ORDER BY name LIMIT 1
            ''', (nama,)).fetchone()
            if not row:
                return f"Karyawan '{inp.get('nama','')}' tidak ditemukan."
            tipe = (row['employment_type'] or 'tetap').lower()
            cs   = row['contract_start'] or ''
            ce   = row['contract_end'] or ''
            # Hitung masa kerja dan status probasi
            masa_kerja = '—'
            status_probasi = '—'
            if cs:
                try:
                    tgl_masuk = _dt.date.fromisoformat(cs[:10])
                    hari_ini  = _dt.date.today()
                    delta     = hari_ini - tgl_masuk
                    bulan     = delta.days // 30
                    tahun     = bulan // 12
                    sisa      = bulan % 12
                    if tahun > 0:
                        masa_kerja = f"{tahun} tahun {sisa} bulan"
                    else:
                        masa_kerja = f"{bulan} bulan"
                    # Probasi dihitung 3 bulan (atau 6 bulan jika kontrak/probasi)
                    batas_probasi = 6 if tipe in ('kontrak', 'probasi') else 3
                    if bulan >= batas_probasi:
                        status_probasi = f"Selesai (mulai kerja {cs[:10]}, sudah {masa_kerja})"
                    else:
                        total_m = tgl_masuk.month + batas_probasi - 1
                        end_y   = tgl_masuk.year + total_m // 12
                        end_m   = total_m % 12 + 1
                        end_probasi = tgl_masuk.replace(year=end_y, month=end_m)
                        sisa_hari = (end_probasi - hari_ini).days
                        status_probasi = f"Masih probasi — sisa ±{max(sisa_hari,0)} hari"
                except Exception:
                    pass
            out = [
                f"Nama        : {row['name']}",
                f"Jabatan     : {row['jabatan'] or '—'}",
                f"Divisi      : {row['divisi'] or '—'}",
                f"Level       : {row['level'] or '—'}",
                f"Tipe kontrak: {tipe}",
                f"Mulai kerja : {cs or '—'}",
                f"Akhir kontrak: {ce or '—'}",
                f"Masa kerja  : {masa_kerja}",
                f"Status probasi: {status_probasi}",
                f"Email       : {row['email'] or '—'}",
                f"Phone       : {row['phone'] or '—'}",
            ]
            if row['notes']:
                out.append(f"Catatan     : {row['notes']}")
            return "\n".join(out)

        elif name == 'cari_aset':
            kw = f"%{inp.get('keyword','')}%"
            cat = (inp.get('category','') or '').lower()
            out = []

            # 1. Search ac_assets (Laptops/PCs)
            if not cat or 'laptop' in cat or 'pc' in cat or 'komputer' in cat or 'asset' in cat:
                assets = db.execute('''
                    SELECT a.device_type, a.brand, a.asset_tag, a.serial_number, a.status, a.condition, a.notes,
                           e.name as employee_name, a.manual_employee_name
                    FROM ac_assets a
                    LEFT JOIN employees e ON e.id = a.employee_id
                    WHERE (a.asset_tag LIKE ? OR a.brand LIKE ? OR a.device_type LIKE ? OR a.serial_number LIKE ? OR a.notes LIKE ? OR e.name LIKE ? OR a.manual_employee_name LIKE ?)
                    ORDER BY a.id DESC LIMIT 10
                ''', (kw, kw, kw, kw, kw, kw, kw)).fetchall()
                for a in assets:
                    holder = a['employee_name'] or a['manual_employee_name'] or 'Belum ditugaskan'
                    out.append(f"[Asset] {a['brand']} {a['device_type']} ({a['asset_tag'] or 'Tanpa Tag'}) | S/N: {a['serial_number'] or '—'} | Status: {a['status']} | Kondisi: {a['condition']} | Pemegang: {holder}")

            # 2. Search ac_infrastructure (Servers, Networks, etc.)
            if not cat or 'infra' in cat or 'server' in cat or 'router' in cat or 'switch' in cat or 'jaringan' in cat:
                infra = db.execute('''
                    SELECT device_type, brand, model, serial_number, location, status, nickname, description
                    FROM ac_infrastructure
                    WHERE (device_type LIKE ? OR brand LIKE ? OR model LIKE ? OR serial_number LIKE ? OR nickname LIKE ? OR location LIKE ? OR description LIKE ?)
                    ORDER BY id DESC LIMIT 10
                ''', (kw, kw, kw, kw, kw, kw, kw)).fetchall()
                for i in infra:
                    name_infra = f"{i['brand']} {i['model']}" if i['brand'] else i['device_type']
                    nick = f" ({i['nickname']})" if i['nickname'] else ""
                    out.append(f"[Infra] {i['device_type']}: {name_infra}{nick} | S/N: {i['serial_number'] or '—'} | Status: {i['status']} | Lokasi: {i['location'] or '—'} | Deskripsi: {i['description'] or '—'}")

            # 3. Search ac_licenses (Software Licenses)
            if not cat or 'license' in cat or 'lisensi' in cat or 'software' in cat or 'aplikasi' in cat:
                licenses = db.execute('''
                    SELECT software_name, license_type, version, max_seats, is_active, notes
                    FROM ac_licenses
                    WHERE (software_name LIKE ? OR notes LIKE ?)
                    ORDER BY id DESC LIMIT 10
                ''', (kw, kw)).fetchall()
                for l in licenses:
                    status_lic = "Aktif" if l['is_active'] else "Nonaktif"
                    out.append(f"[Lisensi] {l['software_name']} v{l['version'] or '—'} | Tipe: {l['license_type']} | Max Seats: {l['max_seats']} | Status: {status_lic}")

            if not out:
                return "Tidak ditemukan aset, infrastruktur, atau lisensi yang cocok."
            return "\n".join(out)

        elif name == 'statistik_aplikasi':
            mod = inp.get('module','all')
            out = []
            if mod in ('support','all'):
                t = db.execute("SELECT COUNT(*) as n FROM sc_tickets").fetchone()['n']
                o = db.execute("SELECT COUNT(*) as n FROM sc_tickets WHERE status NOT IN ('resolved','closed')").fetchone()['n']
                out.append(f"Support: {t} total tiket, {o} open")
            if mod in ('project','all'):
                p = db.execute("SELECT COUNT(*) as n FROM pc_projects WHERE status='active'").fetchone()['n']
                tk = db.execute("SELECT COUNT(*) as n FROM pc_tasks WHERE status NOT IN ('done','cancelled')").fetchone()['n']
                out.append(f"Project: {p} project aktif, {tk} task open")
            if mod in ('talent','all'):
                e = db.execute("SELECT COUNT(*) as n FROM employees WHERE is_active=1").fetchone()['n']
                out.append(f"Talent: {e} karyawan aktif")
            if mod in ('asset','all'):
                a = db.execute("SELECT COUNT(*) as n FROM ac_assets WHERE status='active'").fetchone()['n']
                out.append(f"Aset: {a} aset aktif")
            return "\n".join(out) if out else "Tidak ada data statistik."

    except Exception as ex:
        return f"Error mengambil data: {ex}"
    return "Tool tidak dikenal."

CHATBOT_SYSTEM = """You are an AI assistant for Hive — an internal IT management application.
Hive modules: TalentCore (employees & evaluations), SupportCore (support tickets), ProjectCore (projects & tasks), AssetCore (IT assets), BookingCore (rooms & vehicles).

STUDY HIVE DATABASE SCHEMA & BUSINESS LOGIC:
- Always use the 'execute_sql_query' tool to query the PostgreSQL database for any application data questions.
- Understand the tables and their relations:
  * users (system users with logins): id, username, full_name, role, is_active, email
  * employees (company employees): id, name, jabatan, divisi, level, employment_type, contract_start, contract_end, is_active, email, phone, telegram_id
    - NOTE: When searching or listing employees, always filter by `is_active = 1` unless inactive employees are explicitly requested.
  * sc_tickets (support tickets): id, ticket_no, subject, description, customer_id (references sc_customers.id), support_type_id (references sc_support_types.id), priority, status, reported_at, resolved_at, solution_note
    - status values: 'open', 'in_progress', 'resolved', 'closed'
  * sc_customers: id, name, code, is_active
  * sc_support_types: id, name, is_active
  * pc_projects: id, code, name, description, status, start_date, end_date
    - status values: 'active', 'completed', 'on_hold'
  * pc_tasks: id, project_id (references pc_projects.id), title, description, status, priority, assigned_to (references employees.id), due_date
  * ac_assets (company hardware assets like laptops): id, device_type, brand, asset_tag, serial_number, status, condition, employee_id (references employees.id), manual_employee_name, notes
    - Join with `employees` to find who holds which asset (`employee_id = employees.id`). If `employee_id` is null, check `manual_employee_name`.
  * ac_infrastructure (IT infrastructure): id, device_type, brand, model, serial_number, location, status, nickname, description
  * ac_licenses (software licenses): id, software_name, license_type, version, max_seats, is_active, notes
  * bk_resources (booking resources): id, name, type ('room' or 'vehicle'), subtype, capacity, description, location, color, is_active
  * bk_bookings (resource bookings): id, resource_id (references bk_resources.id), title, purpose, booked_by (references users.id), start_dt, end_dt, status ('confirmed' or 'cancelled')
    - NOTE: `booked_by` in `bk_bookings` references `users.id`, not `employees.id`! Join with `users` to find who booked it.

RULES:
- USER POINT OF VIEW: Never show raw SQL queries, query execution logs, table names, or technical schema details in your final response. The user wants the information, not the code. Translate all database outputs into a clean, polite, human-friendly, and natural Indonesian response from the user's perspective.
- LANGUAGE: Always reply in the SAME language the user used. If they write in Indonesian → reply in Indonesian. Detect from each message.
- CONFIDENTIAL — NEVER show: salary, compensation components, rate_mandays, password hashes, or personal evaluation scores. Queries containing 'employee_salary' or 'salary' or 'rate_mandays' or 'password_hash' will be blocked by the system.
- For questions outside app scope, use your general knowledge.
- Keep answers concise and direct; use bullet points when listing multiple items.
"""

@app.route('/chatbot')
@login_required
def chatbot():
    db = get_db()
    settings = get_settings(db)
    if settings.get('chatbot_enabled','0') != '1':
        flash('Fitur chatbot belum diaktifkan. Hubungi administrator.', 'warning')
        return redirect(url_for('portal'))
    return render_template('chatbot.html')

AI_PROVIDER_MODELS = {
    'anthropic':    ['claude-sonnet-4-6', 'claude-haiku-4-5-20251001', 'claude-opus-4-8'],
    'openai':       ['gpt-4o', 'gpt-4o-mini', 'gpt-4-turbo', 'gpt-3.5-turbo'],
    'openai_compat': [],  # user isi manual
}

AI_PROVIDER_DEFAULTS = {
    'anthropic':    'claude-sonnet-4-6',
    'openai':       'gpt-4o',
    'openai_compat': 'gpt-4o',
}

# Tools dalam format OpenAI (function calling)
def _tools_openai():
    result = []
    for t in CHATBOT_TOOLS:
        result.append({
            'type': 'function',
            'function': {
                'name': t['name'],
                'description': t['description'],
                'parameters': t['input_schema'],
            }
        })
    return result

def _chatbot_call_anthropic(api_key, model, messages, system, tools):
    import anthropic as _ant
    client = _ant.Anthropic(api_key=api_key)
    for _ in range(5):
        resp = client.messages.create(
            model=model,
            max_tokens=2048,
            system=system,
            tools=tools,
            messages=messages,
        )
        if resp.stop_reason == 'end_turn':
            return next((b.text for b in resp.content if hasattr(b,'text')), '')
        if resp.stop_reason == 'tool_use':
            messages.append({'role': 'assistant', 'content': resp.content})
            tool_results = []
            for blk in resp.content:
                if blk.type == 'tool_use':
                    from flask import g as _g
                    db = get_db()
                    result = _chatbot_exec_tool(db, blk.name, blk.input)
                    tool_results.append({'type': 'tool_result', 'tool_use_id': blk.id, 'content': result})
            messages.append({'role': 'user', 'content': tool_results})
            continue
        break
    return 'Tidak ada respons dari AI.'

_SIMPLE_QUESTION_PATTERNS = (
    'halo', 'hai', 'hi', 'selamat', 'terima kasih', 'makasih', 'oke', 'ok',
    'siapa kamu', 'apa itu', 'bantu', 'help', 'tolong',
)

def _is_simple_question(text):
    """True jika pertanyaan tidak perlu tool calling (salam/umum/singkat)."""
    t = text.lower().strip()
    if len(t) < 30:
        return any(p in t for p in _SIMPLE_QUESTION_PATTERNS)
    return False

def _chatbot_call_openai(api_key, model, messages, system, tools_oa, base_url=None):
    import openai as _oai
    kwargs = {'api_key': api_key}
    if base_url:
        kwargs['base_url'] = base_url
    client = _oai.OpenAI(**kwargs)
    oai_msgs = [{'role': 'system', 'content': system}] + messages

    # Deteksi apakah perlu tool calling — skip jika pertanyaan sederhana
    last_user = next((m['content'] for m in reversed(messages) if m.get('role') == 'user'), '')
    use_tools = tools_oa and not _is_simple_question(last_user)

    def _call_once(with_tools):
        kwargs = dict(model=model, max_tokens=1024, messages=oai_msgs)
        if with_tools and tools_oa:
            kwargs['tools'] = tools_oa
            kwargs['tool_choice'] = 'auto'
        return client.chat.completions.create(**kwargs)

    for _ in range(5):
        try:
            resp = _call_once(use_tools)
        except Exception as e:
            # Jika gagal karena tool calling tidak didukung, retry tanpa tools
            err = str(e).lower()
            if use_tools and ('tool' in err or 'function' in err or 'unsupported' in err or '400' in err):
                use_tools = False
                resp = _call_once(False)
            else:
                raise
        choice = resp.choices[0]
        if choice.finish_reason == 'stop':
            return choice.message.content or ''
        if choice.finish_reason == 'tool_calls' and use_tools:
            import json as _json
            msg = choice.message
            oai_msgs.append(msg)
            db = get_db()
            for tc in msg.tool_calls:
                inp = _json.loads(tc.function.arguments)
                result = _chatbot_exec_tool(db, tc.function.name, inp)
                oai_msgs.append({'role': 'tool', 'tool_call_id': tc.id, 'content': result})
            continue
        break
    return 'Tidak ada respons dari AI.'

# Rate limiter sederhana: max N request per menit per user
_chatbot_rate: dict = {}  # {user_id: [timestamp, ...]}
_CHATBOT_MAX_RPM = 8       # max 8 request/menit per user

def _chatbot_check_rate(user_id):
    """Return True jika masih dalam batas, False jika over limit."""
    now = datetime.now().timestamp()
    window = 60.0
    hits = [t for t in _chatbot_rate.get(user_id, []) if now - t < window]
    if len(hits) >= _CHATBOT_MAX_RPM:
        return False
    hits.append(now)
    _chatbot_rate[user_id] = hits
    return True

def _friendly_ai_error(ex, provider='gemini', base_url=''):
    """Ubah error API menjadi pesan yang mudah dipahami."""
    msg = str(ex)
    # Normalise provider label
    _labels = {
        'gemini': 'Gemini', 'openai': 'OpenAI', 'claude': 'Claude (Anthropic)',
        'ollama': 'Ollama', 'openwebui': 'Open WebUI',
    }
    # Fallback deteksi dari base_url jika provider tidak diset
    if provider not in _labels:
        if 'googleapis' in base_url or not base_url:
            provider = 'gemini'
        elif '11434' in base_url or 'ollama' in base_url:
            provider = 'ollama'
        elif '3000' in base_url or 'openwebui' in base_url.lower():
            provider = 'openwebui'
        else:
            provider = 'openai'
    provider_label = _labels.get(provider, 'AI')
    is_gemini    = provider == 'gemini'
    is_ollama    = provider == 'ollama'
    is_claude    = provider == 'claude'
    is_openai    = provider == 'openai'

    if '429' in msg or 'RESOURCE_EXHAUSTED' in msg or 'quota' in msg.lower():
        if is_gemini and ('limit: 0' in msg or 'free_tier' in msg):
            return ('Quota Gemini free tier habis atau API key tidak punya akses free tier. '
                    'Buat API key baru dari aistudio.google.com/apikey → "Create API key in new project".')
        return f'Quota {provider_label} tercapai. Coba lagi dalam beberapa menit.'
    if '401' in msg or 'UNAUTHENTICATED' in msg or 'unauthorized' in msg.lower() or 'authentication_error' in msg.lower():
        if is_ollama:
            return 'Ollama: tidak perlu API key. Kosongkan field API key di Pengaturan Sistem.'
        if is_claude:
            return 'API key Anthropic tidak valid. Buat API key di console.anthropic.com → API Keys.'
        if is_openai:
            return 'API key OpenAI tidak valid. Periksa di platform.openai.com/api-keys.'
        return f'API key {provider_label} tidak valid. Periksa konfigurasi di Pengaturan Sistem → AI Assistant.'
    if '403' in msg or 'PERMISSION_DENIED' in msg:
        return f'API key tidak punya izin menggunakan model ini ({provider_label}).'
    if 'credit' in msg.lower() or 'billing' in msg.lower() or 'insufficient_quota' in msg.lower():
        if is_openai:
            return 'Saldo OpenAI habis. Top up di platform.openai.com/settings/billing.'
        if is_claude:
            return 'Saldo Anthropic habis. Top up di console.anthropic.com/settings/billing.'
        return f'Saldo {provider_label} habis atau quota terlampaui.'
    if 'model' in msg.lower() and ('not found' in msg.lower() or 'does not exist' in msg.lower() or 'pull' in msg.lower()):
        if is_ollama:
            return 'Model Ollama tidak ditemukan. Jalankan: ollama pull <nama_model> di server, lalu coba lagi.'
        if provider == 'openwebui':
            return 'Model tidak ditemukan di Open WebUI. Pastikan model sudah di-pull via Ollama atau tersedia di Open WebUI.'
        if is_claude:
            return 'Model Claude tidak ditemukan. Contoh model valid: claude-sonnet-4-6, claude-haiku-4-5-20251001, claude-opus-4-8.'
        if is_openai:
            return 'Model OpenAI tidak ditemukan. Contoh: gpt-4o, gpt-4o-mini, gpt-4-turbo.'
        return 'Model tidak ditemukan. Ganti model di Pengaturan Sistem → AI Assistant.'
    if 'tool' in msg.lower() and ('not support' in msg.lower() or 'unsupported' in msg.lower()):
        return 'Model ini tidak mendukung tool calling. Coba model lain yang support function calling.'
    if 'connect' in msg.lower() or 'timeout' in msg.lower() or 'connection' in msg.lower():
        if is_ollama:
            return f'Tidak bisa terhubung ke Ollama ({base_url}). Pastikan Ollama berjalan: systemctl status ollama'
        if provider == 'openwebui':
            return f'Tidak bisa terhubung ke Open WebUI ({base_url}). Pastikan container/service berjalan.'
        return 'Tidak bisa terhubung ke server AI. Periksa koneksi internet server.'
    first_line = msg.split('\n')[0][:250]
    return f'Error {provider_label}: {first_line}'

@app.route('/api/chatbot/send', methods=['POST'])
@login_required
def chatbot_send():
    db = get_db()
    settings = get_settings(db)
    if settings.get('chatbot_enabled','0') != '1':
        return jsonify({'error': 'Chatbot tidak aktif'}), 403

    provider    = settings.get('ai_provider', 'gemini').strip() or 'gemini'
    ai_base_url = settings.get('ai_base_url', '').strip()
    api_key     = settings.get('ai_api_key', '').strip()

    # Tentukan base_url dan default_model berdasarkan provider
    if provider == 'openai':
        ai_base_url   = 'https://api.openai.com/v1'
        default_model = 'gpt-4o'
        if not api_key:
            return jsonify({'error': 'API key OpenAI belum dikonfigurasi. Isi di Pengaturan Sistem → AI Assistant.'}), 503
    elif provider == 'claude':
        default_model = 'claude-sonnet-4-6'
        if not api_key:
            return jsonify({'error': 'API key Anthropic (Claude) belum dikonfigurasi. Isi di Pengaturan Sistem → AI Assistant.'}), 503
    elif provider in ('ollama', 'openwebui'):
        if not ai_base_url:
            ai_base_url = 'http://localhost:11434/v1' if provider == 'ollama' else 'http://localhost:3000/api'
        default_model = 'llama3'
        api_key = api_key or 'ollama'
    else:
        # Default: Gemini Google AI Studio
        provider      = 'gemini'
        ai_base_url   = 'https://generativelanguage.googleapis.com/v1beta/openai/'
        default_model = 'gemini-2.0-flash'
        if not api_key:
            return jsonify({'error': 'API key Google AI Studio belum dikonfigurasi. Isi di Pengaturan Sistem → AI Assistant.'}), 503

    model = settings.get('ai_model', '').strip() or default_model

    # Rate limit per user
    if not _chatbot_check_rate(session.get('user_id', 0)):
        return jsonify({'error': f'Terlalu banyak permintaan. Maksimal {_CHATBOT_MAX_RPM} pesan per menit. Tunggu sebentar.'}), 429

    data = request.get_json()
    messages = data.get('messages', [])
    if not messages:
        return jsonify({'error': 'Pesan kosong'}), 400
    messages = messages[-8:]

    try:
        if provider == 'claude':
            # Gunakan Anthropic SDK (native tool calling)
            reply = _chatbot_call_anthropic(api_key, model, messages, CHATBOT_SYSTEM, CHATBOT_TOOLS)
        else:
            # OpenAI-compatible: OpenAI / Gemini / Ollama / Open WebUI
            reply = _chatbot_call_openai(api_key, model, messages, CHATBOT_SYSTEM, _tools_openai(),
                                         base_url=ai_base_url)
        return jsonify({'reply': reply})
    except Exception as ex:
        return jsonify({'error': _friendly_ai_error(ex, provider, ai_base_url)}), 500

# ─── Portal: Notification Settings ───────────────────────────────────────────

@app.route('/portal/notifications', methods=['GET', 'POST'])
@login_required
def portal_notifications():
    if not is_portal_admin():
        flash('Akses ditolak.', 'danger')
        return redirect(url_for('portal'))
    import json as _j
    db = get_db()
    _init_notif_types(db)
    _migrate_notif_from_settings(db)

    if request.method == 'POST':
        action = request.form.get('action', '')

        if action == 'save_types':
            active_slugs = set(request.form.getlist('active_notif_types'))
            for slug, _, _ in _NOTIF_TYPES:
                db.execute('UPDATE notif_type_settings SET is_active=? WHERE slug=?',
                           (1 if slug in active_slugs else 0, slug))
            db.commit()
            flash('Pengaturan tipe notifikasi disimpan.', 'success')

        elif action == 'add_recipient':
            name    = request.form.get('recipient_name', '').strip()
            channel = request.form.get('recipient_channel', '').strip()
            address = request.form.get('recipient_address', '').strip()
            types   = request.form.getlist('recipient_notif_types')
            if not name or not channel or not address:
                flash('Nama, channel, dan address wajib diisi.', 'warning')
            elif channel not in ('email', 'telegram', 'wa'):
                flash('Channel tidak valid.', 'danger')
            else:
                if channel == 'telegram':
                    address = normalize_telegram_id(address)
                db.execute(
                    'INSERT INTO notif_recipients(name,channel,address,notif_types,is_active) VALUES(?,?,?,?,1)',
                    (name, channel, address, _j.dumps(types if types else ['*']))
                )
                db.commit()
                flash(f'Penerima "{name}" ditambahkan.', 'success')

        elif action == 'toggle_recipient':
            rid = int(request.form.get('recipient_id', 0))
            row = db.execute('SELECT is_active FROM notif_recipients WHERE id=?', (rid,)).fetchone()
            if row:
                db.execute('UPDATE notif_recipients SET is_active=? WHERE id=?',
                           (0 if row['is_active'] else 1, rid))
                db.commit()
                flash('Status penerima diperbarui.', 'success')

        elif action == 'edit_recipient':
            rid     = int(request.form.get('recipient_id', 0))
            name    = request.form.get('recipient_name', '').strip()
            address = request.form.get('recipient_address', '').strip()
            types   = request.form.getlist('recipient_notif_types')
            channel = request.form.get('recipient_channel', '').strip()
            if channel == 'telegram':
                address = normalize_telegram_id(address)
            if name and address:
                db.execute(
                    'UPDATE notif_recipients SET name=?,address=?,notif_types=? WHERE id=?',
                    (name, address, _j.dumps(types if types else ['*']), rid)
                )
                db.commit()
                flash('Penerima diperbarui.', 'success')

        elif action == 'delete_recipient':
            rid = int(request.form.get('recipient_id', 0))
            db.execute('DELETE FROM notif_recipients WHERE id=?', (rid,))
            db.commit()
            flash('Penerima dihapus.', 'success')

        elif action == 'save_talentcore':
            for k in TALENTCORE_SETTINGS_KEYS:
                v = request.form.get(k, '').strip()
                if k == 'reminder_enabled':
                    v = '1' if request.form.get('reminder_enabled') else '0'
                save_setting(db, k, v)
            db.commit()
            flash('Pengaturan TalentCore disimpan.', 'success')

        elif action == 'save_assetcore':
            for k in AC_SETTINGS_KEYS:
                if k == 'ac_sub_reminder_enabled':
                    v = '1' if request.form.get(k) else '0'
                else:
                    v = request.form.get(k, '').strip()
                save_setting(db, k, v)
            db.commit()
            flash('Pengaturan AssetCore disimpan.', 'success')

        tab = {'save_talentcore': 'talentcore', 'save_assetcore': 'assetcore'}.get(action, 'sistem')
        return redirect(url_for('portal_notifications', tab=tab))

    # GET
    tab = request.args.get('tab', 'sistem')
    notif_types = db.execute(
        'SELECT slug, is_active, label, description FROM notif_type_settings ORDER BY slug'
    ).fetchall()
    recipients = db.execute(
        'SELECT * FROM notif_recipients ORDER BY channel, name'
    ).fetchall()
    by_channel = {'email': [], 'telegram': [], 'wa': []}
    for r in recipients:
        ch = r['channel']
        if ch in by_channel:
            by_channel[ch].append(r)
    cfg = get_settings(db)
    return render_template('portal_notifications.html',
                           notif_types=notif_types,
                           by_channel=by_channel,
                           all_notif_types=_NOTIF_TYPES,
                           cfg=cfg,
                           active_tab=tab)

# ─── Portal: Audit Trail ───────────────────────────────────────────────────────

@app.route('/portal/audit')
@login_required
def portal_audit():
    if not is_portal_admin():
        flash('Akses ditolak', 'danger')
        return redirect(url_for('portal'))
    db = get_db()
    app_filter   = request.args.get('app', '')
    tab          = request.args.get('tab', 'activity')
    user_filter  = request.args.get('user', '')
    date_from    = request.args.get('from', '')
    date_to      = request.args.get('to', '')
    page         = max(1, int(request.args.get('page', 1)))
    per_page     = 50
    offset       = (page - 1) * per_page

    apps_list = db.execute('SELECT slug, name FROM superapp_apps WHERE is_active=1 ORDER BY sort_order').fetchall()

    def build_where(base_conds, params):
        where = ' AND '.join(base_conds) if base_conds else '1=1'
        return f'WHERE {where}', params

    if tab == 'activity':
        conds, params = [], []
        if app_filter: conds.append('app_slug=?'); params.append(app_filter)
        if user_filter: conds.append('username LIKE ?'); params.append(f'%{user_filter}%')
        if date_from: conds.append('created_at >= ?'); params.append(date_from)
        if date_to: conds.append('created_at <= ?'); params.append(date_to + ' 23:59:59')
        where, params = build_where(conds, params)
        total = db.execute(f'SELECT COUNT(*) FROM audit_activity {where}', params).fetchone()[0]
        rows  = db.execute(f'SELECT * FROM audit_activity {where} ORDER BY id DESC LIMIT ? OFFSET ?',
                           params + [per_page, offset]).fetchall()

    elif tab == 'errors':
        conds, params = [], []
        if app_filter: conds.append('app_slug=?'); params.append(app_filter)
        if date_from: conds.append('created_at >= ?'); params.append(date_from)
        if date_to: conds.append('created_at <= ?'); params.append(date_to + ' 23:59:59')
        where, params = build_where(conds, params)
        total = db.execute(f'SELECT COUNT(*) FROM audit_errors {where}', params).fetchone()[0]
        rows  = db.execute(f'SELECT * FROM audit_errors {where} ORDER BY id DESC LIMIT ? OFFSET ?',
                           params + [per_page, offset]).fetchall()

    else:  # notifications
        conds, params = [], []
        if app_filter: conds.append('app_slug=?'); params.append(app_filter)
        if date_from: conds.append('created_at >= ?'); params.append(date_from)
        if date_to: conds.append('created_at <= ?'); params.append(date_to + ' 23:59:59')
        where, params = build_where(conds, params)
        total = db.execute(f'SELECT COUNT(*) FROM audit_notifications {where}', params).fetchone()[0]
        rows  = db.execute(f'SELECT * FROM audit_notifications {where} ORDER BY id DESC LIMIT ? OFFSET ?',
                           params + [per_page, offset]).fetchall()

    total_pages = max(1, (total + per_page - 1) // per_page)
    return render_template('portal_audit.html',
                           rows=rows, tab=tab, apps_list=apps_list,
                           app_filter=app_filter, user_filter=user_filter,
                           date_from=date_from, date_to=date_to,
                           page=page, total_pages=total_pages, total=total)

@app.route('/portal/audit/clear', methods=['POST'])
@login_required
def portal_audit_clear():
    if not is_portal_admin():
        flash('Akses ditolak', 'danger')
        return redirect(url_for('portal'))
    db  = get_db()
    tab = request.form.get('tab', 'activity')
    tbl = {'activity': 'audit_activity', 'errors': 'audit_errors', 'notifications': 'audit_notifications'}.get(tab)
    if tbl:
        app_filter = request.form.get('app', '')
        if app_filter:
            db.execute(f'DELETE FROM {tbl} WHERE app_slug=?', (app_filter,))
        else:
            db.execute(f'DELETE FROM {tbl}')
        db.commit()
        flash('Log berhasil dihapus', 'success')
    return redirect(url_for('portal_audit', tab=tab))

# ─── BookingCore ───────────────────────────────────────────────────────────────

def _bk_require_access():
    if not session.get('user_id'):
        return redirect(url_for('login'))
    acc = get_db().execute(
        'SELECT app_role FROM user_app_access WHERE user_id=? AND app_slug=? AND is_active=1',
        (session['user_id'], 'booking')
    ).fetchone()
    if session.get('user_role') == 'superadmin' or acc:
        return None
    flash('Anda tidak memiliki akses ke BookingCore.', 'danger')
    return redirect(url_for('portal'))

def _bk_generate_recurring(base, resource_id, title, purpose, booked_by, attendees,
                             attendee_count, notes, rec_type, rec_days, rec_until,
                             destination, start_dt, end_dt):
    from datetime import datetime, timedelta
    db = get_db()
    dt_start = datetime.strptime(start_dt, '%Y-%m-%d %H:%M')
    dt_end   = datetime.strptime(end_dt,   '%Y-%m-%d %H:%M')
    until    = datetime.strptime(rec_until, '%Y-%m-%d')
    delta    = dt_end - dt_start
    days_map = {'sen':0,'sel':1,'rab':2,'kam':3,'jum':4,'sab':5,'min':6}
    chosen   = [days_map[d] for d in (rec_days.split(',') if rec_days else []) if d in days_map]

    cur = dt_start + timedelta(days=1)
    count = 0
    while cur.date() <= until.date() and count < 365:
        if rec_type == 'weekly' and chosen:
            if cur.weekday() not in chosen:
                cur += timedelta(days=1)
                continue
        elif rec_type == 'daily':
            pass
        elif rec_type == 'weekly' and not chosen:
            if cur.weekday() != dt_start.weekday():
                cur += timedelta(days=1)
                continue
        s = cur.strftime('%Y-%m-%d %H:%M')
        e = (cur + delta).strftime('%Y-%m-%d %H:%M')
        db.execute('''INSERT INTO bk_bookings(resource_id,title,purpose,booked_by,attendees,
            attendee_count,start_dt,end_dt,status,notes,is_recurring,recurring_type,
            recurring_days,recurring_until,parent_id,destination)
            VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)''',
            (resource_id,title,purpose,booked_by,attendees,attendee_count,
             s,e,'confirmed',notes,1,rec_type,rec_days,rec_until,base,destination))
        cur += timedelta(days=1)
        count += 1
    db.commit()


_holidays_cache = {}  # {year: {date_str: holiday_name}}

def get_indonesian_holidays(year):
    global _holidays_cache
    if year in _holidays_cache:
        return _holidays_cache[year]
    
    import urllib.request
    import json

    year_holidays = {}
    # Try upset.dev API first
    url = f"https://tanggalmerah.upset.dev/api/holidays?year={year}"
    try:
        req = urllib.request.Request(url, headers={'User-Agent': 'Mozilla/5.0'})
        with urllib.request.urlopen(req, timeout=3) as response:
            data = json.loads(response.read().decode())
            if isinstance(data, dict) and 'data' in data:
                for item in data['data']:
                    d_str = item.get('date')
                    name = item.get('name')
                    if d_str and name:
                        year_holidays[d_str] = name
                _holidays_cache[year] = year_holidays
                return year_holidays
    except Exception as e:
        print(f"Failed to fetch holidays from upset.dev: {e}")

    # Try libur.deno.dev fallback
    url_fallback = f"https://libur.deno.dev/api?year={year}"
    try:
        req = urllib.request.Request(url_fallback, headers={'User-Agent': 'Mozilla/5.0'})
        with urllib.request.urlopen(req, timeout=3) as response:
            data = json.loads(response.read().decode())
            if isinstance(data, list):
                for item in data:
                    d_str = item.get('date')
                    name = item.get('name') or item.get('holiday')
                    if d_str and name:
                        year_holidays[d_str] = name
                _holidays_cache[year] = year_holidays
                return year_holidays
    except Exception as e:
        print(f"Failed to fetch holidays from libur.deno.dev: {e}")

    return {}


@app.route('/booking/public')
def booking_public():
    db = get_db()
    resources = db.execute('SELECT * FROM bk_resources WHERE is_active=1 ORDER BY sort_order').fetchall()
    resource_id = request.args.get('resource', type=int)
    view = request.args.get('view', 'dashboard')
    date_str = request.args.get('date', '')
    
    from datetime import datetime, timedelta
    today = datetime.now().date()
    
    if date_str:
        try: ref_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        except: ref_date = today
    else:
        ref_date = today

    # Month calculations
    month_start = ref_date.replace(day=1)
    cal_start = month_start - timedelta(days=month_start.weekday())
    cal_dates = [cal_start + timedelta(days=i) for i in range(42)]
    cal_end = cal_start + timedelta(days=41)

    # Fetch holidays
    start_year = cal_start.year
    end_year = cal_end.year
    holidays = {}
    for y in range(start_year, end_year + 1):
        y_holidays = get_indonesian_holidays(y)
        holidays.update(y_holidays)

    # Fetch bookings for current range
    booking_query = '''
        SELECT b.*, r.name res_name, r.color res_color, r.icon res_icon, u.full_name booker_name
        FROM bk_bookings b
        JOIN bk_resources r ON b.resource_id = r.id
        LEFT JOIN users u ON u.id = b.booked_by
        WHERE b.start_dt >= ? AND b.start_dt <= ? AND r.is_active = 1
    '''
    params = [cal_start.strftime('%Y-%m-%d 00:00:00'), cal_end.strftime('%Y-%m-%d 23:59:59')]
    if resource_id:
        booking_query += ' AND b.resource_id = ?'
        params.append(resource_id)
        
    db_bookings = db.execute(booking_query, params).fetchall()

    # Map bookings by date
    bookings_by_date = {}
    for b in db_bookings:
        try:
            b_date = datetime.strptime(b['start_dt'][:10], '%Y-%m-%d').date()
        except Exception:
            continue
        if b_date not in bookings_by_date:
            bookings_by_date[b_date] = []
        
        try:
            start_t = datetime.strptime(b['start_dt'], '%Y-%m-%d %H:%M:%S').strftime('%H:%M')
            end_t = datetime.strptime(b['end_dt'], '%Y-%m-%d %H:%M:%S').strftime('%H:%M')
        except Exception:
            try:
                start_t = datetime.strptime(b['start_dt'], '%Y-%m-%d %H:%M').strftime('%H:%M')
                end_t = datetime.strptime(b['end_dt'], '%Y-%m-%d %H:%M').strftime('%H:%M')
            except Exception:
                start_t = b['start_dt']
                end_t = b['end_dt']
        
        bookings_by_date[b_date].append({
            'id': b['id'],
            'resource_id': b['resource_id'],
            'res_name': b['res_name'],
            'res_color': b['res_color'],
            'res_icon': b['res_icon'],
            'time_range': f"{start_t} - {end_t}",
            'title': b['title'],
            'booked_by': b['booker_name'] or b['booked_by']
        })

    # Fetch today's bookings for display
    today_query = '''
        SELECT b.*, r.name res_name, r.color res_color, r.icon res_icon, u.full_name booker_name
        FROM bk_bookings b
        JOIN bk_resources r ON b.resource_id = r.id
        LEFT JOIN users u ON u.id = b.booked_by
        WHERE DATE(b.start_dt) = ? AND r.is_active = 1
        ORDER BY b.start_dt ASC
    '''
    today_bookings = db.execute(today_query, (ref_date.strftime('%Y-%m-%d'),)).fetchall()
    
    # Format today's bookings
    formatted_today_bookings = []
    for tb in today_bookings:
        try:
            st = datetime.strptime(tb['start_dt'], '%Y-%m-%d %H:%M:%S').strftime('%H:%M')
            et = datetime.strptime(tb['end_dt'], '%Y-%m-%d %H:%M:%S').strftime('%H:%M')
        except Exception:
            try:
                st = datetime.strptime(tb['start_dt'], '%Y-%m-%d %H:%M').strftime('%H:%M')
                et = datetime.strptime(tb['end_dt'], '%Y-%m-%d %H:%M').strftime('%H:%M')
            except Exception:
                st = tb['start_dt']
                et = tb['end_dt']
        formatted_today_bookings.append({
            'id': tb['id'],
            'resource_id': tb['resource_id'],
            'res_name': tb['res_name'],
            'res_color': tb['res_color'],
            'res_icon': tb['res_icon'],
            'time_range': f"{st} - {et}",
            'title': tb['title'],
            'booked_by': tb['booker_name'] or tb['booked_by'],
            'notes': tb.get('notes', '')
        })

    if month_start.month == 1:
        prev_month = month_start.replace(year=month_start.year - 1, month=12, day=1).isoformat()
    else:
        prev_month = month_start.replace(month=month_start.month - 1, day=1).isoformat()
        
    if month_start.month == 12:
        next_month = month_start.replace(year=month_start.year + 1, month=1, day=1).isoformat()
    else:
        next_month = month_start.replace(month=month_start.month + 1, day=1).isoformat()

    return render_template('booking_public_dashboard.html',
                           resources=resources,
                           resource_id=resource_id,
                           cal_dates=cal_dates,
                           month_start=month_start,
                           bookings_by_date=bookings_by_date,
                           holidays=holidays,
                           view=view,
                           ref_date=ref_date,
                           today_bookings=formatted_today_bookings,
                           today=today,
                           prev_month=prev_month,
                           next_month=next_month)


@app.route('/booking/')
@app.route('/booking')
def booking_index():
    redir = _bk_require_access()
    if redir: return redir
    db = get_db()
    resources = db.execute('SELECT * FROM bk_resources WHERE is_active=1 ORDER BY sort_order').fetchall()
    resource_id = request.args.get('resource', type=int)
    view = request.args.get('view', 'month')
    date_str = request.args.get('date', '')

    from datetime import datetime, timedelta
    today = datetime.now().date()
    if date_str:
        try: ref_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        except: ref_date = today
    else:
        ref_date = today

    # Month calculations
    month_start = ref_date.replace(day=1)
    cal_start = month_start - timedelta(days=month_start.weekday())
    cal_dates = [cal_start + timedelta(days=i) for i in range(42)]
    cal_end = cal_start + timedelta(days=41)

    if month_start.month == 1:
        prev_month = month_start.replace(year=month_start.year - 1, month=12, day=1).isoformat()
    else:
        prev_month = month_start.replace(month=month_start.month - 1, day=1).isoformat()
        
    if month_start.month == 12:
        next_month = month_start.replace(year=month_start.year + 1, month=1, day=1).isoformat()
    else:
        next_month = month_start.replace(month=month_start.month + 1, day=1).isoformat()

    # Week calculations
    week_start = ref_date - timedelta(days=ref_date.weekday())
    week_dates = [week_start + timedelta(days=i) for i in range(7)]

    q = 'SELECT b.*,r.name res_name,r.type res_type,r.color res_color,r.icon res_icon,u.full_name booker_name FROM bk_bookings b JOIN bk_resources r ON r.id=b.resource_id JOIN users u ON u.id=b.booked_by WHERE b.status!=? '
    params = ['cancelled']
    if resource_id:
        q += ' AND b.resource_id=?'
        params.append(resource_id)
    if view == 'month':
        q += ' AND date(b.start_dt)>=? AND date(b.start_dt)<=?'
        params += [cal_start.isoformat(), cal_end.isoformat()]
    elif view == 'week':
        q += ' AND date(b.start_dt)>=? AND date(b.start_dt)<=?'
        params += [week_start.isoformat(), (week_start + timedelta(days=6)).isoformat()]
    else:
        q += ' AND date(b.start_dt)>=?'
        params.append(today.isoformat())
    q += ' ORDER BY b.start_dt'
    bookings = db.execute(q, params).fetchall()

    # Fetch holidays for all years in cal_dates
    cal_years = {d.year for d in cal_dates}
    holidays = {}
    for y in cal_years:
        holidays.update(get_indonesian_holidays(y))

    return render_template('booking_index.html',
        resources=resources, bookings=bookings,
        selected_resource=resource_id, view=view,
        ref_date=ref_date, today=today,
        week_dates=week_dates,
        prev_week=(week_start - timedelta(days=7)).isoformat(),
        next_week=(week_start + timedelta(days=7)).isoformat(),
        cal_dates=cal_dates, month_start=month_start,
        prev_month=prev_month, next_month=next_month,
        holidays=holidays)


@app.route('/booking/new', methods=['GET','POST'])
def booking_new():
    redir = _bk_require_access()
    if redir: return redir
    db = get_db()
    resources = db.execute('SELECT * FROM bk_resources WHERE is_active=1 ORDER BY sort_order').fetchall()

    if request.method == 'POST':
        resource_id    = int(request.form['resource_id'])
        title          = request.form['title'].strip()
        purpose        = request.form.get('purpose','').strip()
        attendees      = request.form.get('attendees','').strip()
        attendee_count = int(request.form.get('attendee_count') or 0)
        notes          = request.form.get('notes','').strip()
        destination    = request.form.get('destination','').strip()
        start_dt       = request.form['start_dt']
        end_dt         = request.form['end_dt']
        is_recurring   = 1 if request.form.get('is_recurring') else 0
        rec_type       = request.form.get('recurring_type','')
        rec_days       = ','.join(request.form.getlist('recurring_days'))
        rec_until      = request.form.get('recurring_until','')

        if not title or not start_dt or not end_dt:
            flash('Judul, waktu mulai dan selesai wajib diisi.', 'danger')
        elif start_dt >= end_dt:
            flash('Waktu selesai harus setelah waktu mulai.', 'danger')
        else:
            # check conflict
            conflict = db.execute('''SELECT id FROM bk_bookings
                WHERE resource_id=? AND status!='cancelled'
                AND start_dt < ? AND end_dt > ?''',
                (resource_id, end_dt, start_dt)).fetchone()
            if conflict:
                flash('Waktu yang dipilih bertabrakan dengan booking lain.', 'danger')
            else:
                db.execute('''INSERT INTO bk_bookings(resource_id,title,purpose,booked_by,attendees,
                    attendee_count,start_dt,end_dt,status,notes,is_recurring,recurring_type,
                    recurring_days,recurring_until,destination)
                    VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)''',
                    (resource_id,title,purpose,session['user_id'],attendees,attendee_count,
                     start_dt,end_dt,'confirmed',notes,is_recurring,rec_type,rec_days,rec_until,destination))
                db.commit()
                parent_id = db.execute('SELECT last_insert_rowid()').fetchone()[0]

                if is_recurring and rec_until:
                    _bk_generate_recurring(parent_id, resource_id, title, purpose,
                        session['user_id'], attendees, attendee_count, notes,
                        rec_type, rec_days, rec_until, destination, start_dt, end_dt)

                # Simpan booking items
                for item_id_str in request.form.getlist('item_ids'):
                    try:
                        item_id = int(item_id_str)
                        qty = int(request.form.get(f'item_qty_{item_id}') or 1)
                        item_notes = request.form.get(f'item_notes_{item_id}', '')
                        db.execute('INSERT INTO bk_booking_items(booking_id,item_id,qty,notes) VALUES(?,?,?,?)',
                                   (parent_id, item_id, qty, item_notes))
                    except (ValueError, TypeError):
                        pass
                db.commit()
                flash('Booking berhasil dibuat!', 'success')
                return redirect(url_for('booking_detail', bid=parent_id))

    pre_resource = request.args.get('resource', type=int)
    bk_items = db.execute('SELECT * FROM bk_items WHERE is_active=1 ORDER BY sort_order').fetchall()
    # load images per resource untuk preview
    res_images = {}
    for r in resources:
        imgs = db.execute('SELECT image FROM bk_resource_images WHERE resource_id=? ORDER BY sort_order LIMIT 5', (r['id'],)).fetchall()
        all_imgs = []
        if r['image']:
            all_imgs.append(r['image'])
        all_imgs += [i['image'] for i in imgs]
        res_images[r['id']] = all_imgs
    return render_template('booking_form.html', resources=resources, pre_resource=pre_resource,
                           bk_items=bk_items, res_images=res_images)


@app.route('/booking/resource/<int:rid>')
def booking_resource_detail(rid):
    redir = _bk_require_access()
    if redir: return redir
    db = get_db()
    resource = db.execute('SELECT * FROM bk_resources WHERE id=? AND is_active=1', (rid,)).fetchone()
    if not resource:
        flash('Resource tidak ditemukan.', 'danger')
        return redirect(url_for('booking_index'))
    from datetime import datetime as _dt
    today = _dt.now().date().isoformat()
    upcoming = db.execute('''
        SELECT b.*,u.full_name booker_name FROM bk_bookings b
        JOIN users u ON u.id=b.booked_by
        WHERE b.resource_id=? AND b.status!='cancelled' AND date(b.start_dt)>=?
        ORDER BY b.start_dt LIMIT 30''', (rid, today)).fetchall()
    past = db.execute('''
        SELECT b.*,u.full_name booker_name FROM bk_bookings b
        JOIN users u ON u.id=b.booked_by
        WHERE b.resource_id=? AND b.status!='cancelled' AND date(b.start_dt)<?
        ORDER BY b.start_dt DESC LIMIT 10''', (rid, today)).fetchall()
    images = db.execute('SELECT * FROM bk_resource_images WHERE resource_id=? ORDER BY sort_order', (rid,)).fetchall()
    return render_template('booking_resource.html', resource=resource,
                           images=images, upcoming=upcoming, past=past, today=today)


@app.route('/booking/resource/<int:rid>/edit', methods=['GET', 'POST'])
@login_required
def booking_resource_edit(rid):
    redir = _bk_require_access()
    if redir: return redir
    db = get_db()
    resource = db.execute('SELECT * FROM bk_resources WHERE id=?', (rid,)).fetchone()
    if not resource:
        flash('Resource tidak ditemukan.', 'danger')
        return redirect(url_for('booking_index'))
    if request.method == 'POST':
        name       = request.form.get('name', '').strip()
        rtype      = request.form.get('type', 'room')
        subtype    = request.form.get('subtype', '').strip()
        capacity   = request.form.get('capacity', 0, type=int)
        location   = request.form.get('location', '').strip()
        description= request.form.get('description', '').strip()
        facilities = request.form.get('facilities', '').strip()
        notes      = request.form.get('notes', '').strip()
        color      = request.form.get('color', '#d97706').strip()
        icon       = request.form.get('icon', 'door-open').strip()
        sort_order = request.form.get('sort_order', 0, type=int)
        is_active  = 1 if request.form.get('is_active') else 0
        image      = resource['image'] or ''
        
        # Handle library selected main image or uploaded file
        selected_main = request.form.get('selected_main_image')
        if selected_main:
            image = selected_main
        else:
            f = request.files.get('image')
            if f and f.filename:
                saved = _save_upload(f, 'bookingcore')
                if saved:
                    image = saved
                else:
                    flash('Format gambar tidak didukung. Gunakan JPG, PNG, atau WEBP.', 'warning')
        
        if not name:
            flash('Nama resource wajib diisi.', 'danger')
        else:
            db.execute('''UPDATE bk_resources SET name=?,type=?,subtype=?,capacity=?,
                location=?,description=?,facilities=?,notes=?,color=?,icon=?,
                sort_order=?,is_active=?,image=? WHERE id=?''',
                (name, rtype, subtype, capacity, location, description,
                 facilities, notes, color, icon, sort_order, is_active, image, rid))
            
            # Handle additional uploaded gallery images
            gallery_files = request.files.getlist('images')
            sort_base = db.execute('SELECT COALESCE(MAX(sort_order),0) FROM bk_resource_images WHERE resource_id=?', (rid,)).fetchone()[0]
            for i, gf in enumerate(gallery_files):
                if gf and gf.filename:
                    saved_g = _save_upload(gf, 'bookingcore')
                    if saved_g:
                        caption = request.form.get(f'caption_{i}', '')
                        db.execute('INSERT INTO bk_resource_images(resource_id,image,caption,sort_order) VALUES(?,?,?,?)',
                                   (rid, saved_g, caption, sort_base + i + 1))
                        sort_base += 1
            
            # Handle additional selected gallery images from library
            selected_galleries = request.form.getlist('selected_gallery_images')
            for sg in selected_galleries:
                if sg:
                    db.execute('INSERT INTO bk_resource_images(resource_id,image,caption,sort_order) VALUES(?,?,?,?)',
                               (rid, sg, '', sort_base + 1))
                    sort_base += 1
            
            db.commit()
            flash('Resource berhasil diperbarui.', 'success')
            return redirect(url_for('booking_resource_detail', rid=rid))
    images = db.execute('SELECT * FROM bk_resource_images WHERE resource_id=? ORDER BY sort_order', (rid,)).fetchall()
    return render_template('booking_resource_edit.html', resource=resource, images=images)


@app.route('/booking/resource/add', methods=['GET', 'POST'])
@login_required
def booking_resource_add():
    redir = _bk_require_access()
    if redir: return redir
    db = get_db()
    if request.method == 'POST':
        name       = request.form.get('name', '').strip()
        rtype      = request.form.get('type', 'room')
        subtype    = request.form.get('subtype', '').strip()
        capacity   = request.form.get('capacity', 0, type=int)
        location   = request.form.get('location', '').strip()
        description= request.form.get('description', '').strip()
        facilities = request.form.get('facilities', '').strip()
        notes      = request.form.get('notes', '').strip()
        color      = request.form.get('color', '#d97706').strip()
        icon       = request.form.get('icon', 'door-open').strip()
        sort_order = request.form.get('sort_order', 0, type=int)
        is_active  = 1 if request.form.get('is_active') else 0
        image      = ''
        
        # Handle library selected main image or uploaded file
        selected_main = request.form.get('selected_main_image')
        if selected_main:
            image = selected_main
        else:
            f = request.files.get('image')
            if f and f.filename:
                saved = _save_upload(f, 'bookingcore')
                if saved:
                    image = saved
                    
        if not name:
            flash('Nama resource wajib diisi.', 'danger')
        else:
            cur = db.execute('''INSERT INTO bk_resources(name,type,subtype,capacity,location,
                description,facilities,notes,color,icon,sort_order,image,is_active)
                VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?)''',
                (name, rtype, subtype, capacity, location, description,
                 facilities, notes, color, icon, sort_order, image, is_active))
            rid = cur.lastrowid
            
            # Handle additional uploaded gallery images
            gallery_files = request.files.getlist('images')
            sort_base = 0
            for i, gf in enumerate(gallery_files):
                if gf and gf.filename:
                    saved_g = _save_upload(gf, 'bookingcore')
                    if saved_g:
                        caption = request.form.get(f'caption_{i}', '')
                        db.execute('INSERT INTO bk_resource_images(resource_id,image,caption,sort_order) VALUES(?,?,?,?)',
                                   (rid, saved_g, caption, sort_base + i + 1))
                        sort_base += 1
            
            # Handle additional selected gallery images from library
            selected_galleries = request.form.getlist('selected_gallery_images')
            for sg in selected_galleries:
                if sg:
                    db.execute('INSERT INTO bk_resource_images(resource_id,image,caption,sort_order) VALUES(?,?,?,?)',
                               (rid, sg, '', sort_base + 1))
                    sort_base += 1
            
            db.commit()
            flash(f'Resource "{name}" berhasil ditambahkan.', 'success')
            return redirect(url_for('booking_index'))
    return render_template('booking_resource_edit.html', resource=None)


@app.route('/booking/resource/<int:rid>/images', methods=['POST'])
@login_required
def booking_resource_images(rid):
    """Upload atau hapus gambar tambahan resource."""
    redir = _bk_require_access()
    if redir: return redir
    db = get_db()
    action = request.form.get('action', 'upload')
    if action == 'delete':
        img_id = request.form.get('img_id', type=int)
        if img_id:
            db.execute('DELETE FROM bk_resource_images WHERE id=? AND resource_id=?', (img_id, rid))
            db.commit()
        return jsonify({'ok': True, 'msg': 'Foto galeri berhasil dihapus.'})
    else:
        files = request.files.getlist('images')
        sort_base = db.execute('SELECT COALESCE(MAX(sort_order),0) FROM bk_resource_images WHERE resource_id=?', (rid,)).fetchone()[0]
        for i, f in enumerate(files):
            if f and f.filename:
                saved = _save_upload(f, 'bookingcore')
                if saved:
                    caption = request.form.get(f'caption_{i}', '')
                    db.execute('INSERT INTO bk_resource_images(resource_id,image,caption,sort_order) VALUES(?,?,?,?)',
                               (rid, saved, caption, sort_base + i + 1))
        db.commit()
    return redirect(url_for('booking_resource_edit', rid=rid))

@app.route('/booking/resource/<int:rid>/delete-main-image', methods=['POST'])
@login_required
def booking_resource_delete_main_image(rid):
    redir = _bk_require_access()
    if redir: return redir
    db = get_db()
    db.execute('UPDATE bk_resources SET image=? WHERE id=?', ('', rid))
    db.commit()
    return jsonify({'ok': True})


@app.route('/booking/media/list')
@login_required
def booking_media_list():
    redir = _bk_require_access()
    if redir: return redir
    
    db = get_db()
    cfg = get_settings(db)
    storage_type = cfg.get('media_storage_type', 'local')
    
    urls = set()
    
    # 1. Query unique images from DB tables
    rows_res = db.execute("SELECT DISTINCT image FROM bk_resources WHERE image IS NOT NULL AND image != ''").fetchall()
    rows_gal = db.execute("SELECT DISTINCT image FROM bk_resource_images WHERE image IS NOT NULL AND image != ''").fetchall()
    for r in rows_res:
        urls.add(r['image'])
    for r in rows_gal:
        urls.add(r['image'])
        
    # 2. List from local static uploads folder
    import os
    local_dir = os.path.join('static', 'uploads', 'bookingcore')
    if os.path.exists(local_dir):
        for f in os.listdir(local_dir):
            if f.lower().endswith(('.png', '.jpg', '.jpeg', '.webp', '.gif')):
                urls.add(f'/static/uploads/bookingcore/{f}')
                
    # 3. List from S3 folder
    if storage_type == 's3':
        try:
            endpoint = cfg.get('backup_dest_s3_endpoint', '').strip()
            access_key = cfg.get('backup_dest_s3_access_key', '').strip()
            secret_key = cfg.get('backup_dest_s3_secret_key', '').strip()
            bucket = cfg.get('backup_dest_s3_bucket', '').strip()
            region = cfg.get('backup_dest_s3_region', '').strip()
            
            if access_key and secret_key and bucket:
                import boto3
                from botocore.config import Config
                config = Config(
                    region_name=region or 'us-east-1',
                    signature_version='s3v4',
                    connect_timeout=3,
                    read_timeout=5,
                    retries={'max_attempts': 1}
                )
                s3 = boto3.client(
                    's3',
                    endpoint_url=endpoint or None,
                    aws_access_key_id=access_key,
                    aws_secret_access_key=secret_key,
                    config=config
                )
                paginator = s3.get_paginator('list_objects_v2')
                for page in paginator.paginate(Bucket=bucket, Prefix='upload/media/bookingcore/'):
                    for obj in page.get('Contents', []):
                        key = obj['Key']
                        if key.lower().endswith(('.png', '.jpg', '.jpeg', '.webp', '.gif')):
                            urls.add(f"/media/proxy/{key}")
        except Exception as e:
            print(f"Error listing S3 media: {e}")
            
    # Sort and return unique images list
    return jsonify({'ok': True, 'images': sorted(list(urls))})


@app.route('/booking/resource/<int:rid>/delete', methods=['POST'])
@login_required
def booking_resource_delete(rid):
    redir = _bk_require_access()
    if redir: return redir
    db = get_db()
    
    # Cascade delete bookings and resource images
    db.execute('DELETE FROM bk_bookings WHERE resource_id=?', (rid,))
    db.execute('DELETE FROM bk_resource_images WHERE resource_id=?', (rid,))
    db.execute('DELETE FROM bk_resources WHERE id=?', (rid,))
    db.commit()
    
    flash('Resource berhasil dihapus secara permanen.', 'success')
    return redirect(url_for('booking_index'))


@app.route('/booking/<int:bid>')
def booking_detail(bid):
    redir = _bk_require_access()
    if redir: return redir
    db = get_db()
    b = db.execute('''SELECT b.*,r.name res_name,r.type res_type,r.color res_color,r.icon res_icon,
        r.capacity res_cap,r.location res_loc,u.full_name booker_name,u.username booker_username
        FROM bk_bookings b JOIN bk_resources r ON r.id=b.resource_id
        JOIN users u ON u.id=b.booked_by WHERE b.id=?''', (bid,)).fetchone()
    if not b:
        flash('Booking tidak ditemukan.', 'danger')
        return redirect(url_for('booking_index'))
    children = []
    if b['is_recurring'] and not b['parent_id']:
        children = db.execute('''SELECT * FROM bk_bookings WHERE parent_id=? AND status!='cancelled'
            ORDER BY start_dt LIMIT 50''', (bid,)).fetchall()
    booking_items = db.execute('''SELECT bi.*,i.name item_name,i.icon item_icon,i.category item_cat
        FROM bk_booking_items bi JOIN bk_items i ON i.id=bi.item_id
        WHERE bi.booking_id=? ORDER BY i.sort_order''', (bid,)).fetchall()
    return render_template('booking_detail.html', b=b, children=children, booking_items=booking_items)


@app.route('/booking/<int:bid>/cancel', methods=['POST'])
def booking_cancel(bid):
    redir = _bk_require_access()
    if redir: return redir
    db = get_db()
    b = db.execute('SELECT * FROM bk_bookings WHERE id=?', (bid,)).fetchone()
    if not b:
        flash('Booking tidak ditemukan.', 'danger')
        return redirect(url_for('booking_index'))
    is_owner = b['booked_by'] == session['user_id']
    is_admin = session.get('user_role') in ('superadmin', 'admin')
    if not is_owner and not is_admin:
        flash('Hanya pemesan atau admin yang dapat membatalkan booking ini.', 'danger')
        return redirect(url_for('booking_detail', bid=bid))
    cancel_children = request.form.get('cancel_children') == '1'
    from datetime import datetime
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    db.execute('UPDATE bk_bookings SET status=?,cancelled_at=?,cancelled_by=? WHERE id=?',
               ('cancelled', now, session['user_id'], bid))
    if cancel_children and b['is_recurring'] and not b['parent_id']:
        db.execute('UPDATE bk_bookings SET status=?,cancelled_at=?,cancelled_by=? WHERE parent_id=? AND status!=?',
                   ('cancelled', now, session['user_id'], bid, 'cancelled'))
    db.commit()
    flash('Booking berhasil dibatalkan.', 'success')
    return redirect(url_for('booking_index'))


@app.route('/booking/api/slots')
def booking_api_slots():
    redir = _bk_require_access()
    if redir: return ('', 403)
    resource_id = request.args.get('resource', type=int)
    date_str    = request.args.get('date', '')
    if not resource_id or not date_str:
        return app.response_class('[]', mimetype='application/json')
    db = get_db()
    bookings = db.execute('''SELECT id,title,start_dt,end_dt,status FROM bk_bookings
        WHERE resource_id=? AND date(start_dt)=? AND status!='cancelled'
        ORDER BY start_dt''', (resource_id, date_str)).fetchall()
    import json
    data = [dict(id=r['id'],title=r['title'],start=r['start_dt'],end=r['end_dt']) for r in bookings]
    return app.response_class(json.dumps(data), mimetype='application/json')


# ─── AssetCore ─────────────────────────────────────────────────────────────────

def ac_require(perm):
    db = get_db()
    if not has_permission(session.get('user_role', ''), perm, db):
        flash(f'Akses ditolak — permission "{perm}" diperlukan', 'danger')
        return False
    return True

def _record_asset_history(db, asset_id, employee_id, manual_name, started_at, reason='', notes=''):
    """Simpan satu baris history penggunaan asset. Dipanggil sebelum employee diganti."""
    if not employee_id and not (manual_name or '').strip():
        return  # tidak ada user sebelumnya, tidak perlu catat
    today = date.today().isoformat()
    db.execute(
        '''INSERT INTO ac_asset_history
           (asset_id, employee_id, manual_employee_name, started_at, ended_at, reason, notes)
           VALUES (?,?,?,?,?,?,?)''',
        (asset_id, employee_id, manual_name or '', started_at or '', today, reason, notes)
    )

AC_ASSET_DEVICE_TYPES = {'Laptop', 'PC Desktop', 'All-in-One', 'Mini PC'}

def _clean_asset_device_type(category='', item_name=''):
    category = (category or '').strip()
    if category in AC_ASSET_DEVICE_TYPES:
        return category
    hay = f'{category} {item_name or ""}'.lower()
    if 'all-in-one' in hay or 'all in one' in hay or 'aio' in hay:
        return 'All-in-One'
    if 'mini pc' in hay:
        return 'Mini PC'
    if 'desktop' in hay or 'pc' in hay or 'komputer' in hay or 'computer' in hay:
        return 'PC Desktop'
    if 'laptop' in hay or 'notebook' in hay:
        return 'Laptop'
    return ''

def _tool_request_targets_asset(req):
    return bool(_clean_asset_device_type(req['item_category'], req['item_name']))

def _format_rupiah(value):
    try:
        amount = float(value or 0)
    except (TypeError, ValueError):
        return ''
    if amount <= 0:
        return ''
    return 'Rp {:,.0f}'.format(amount).replace(',', '.')

def _parse_float_value(value, default=0.0):
    try:
        return float(value or default)
    except (TypeError, ValueError):
        return default

AC_TOOL_ATTACHMENT_CONFIG = {
    'request_capture': {
        'field': 'attach_request_capture',
        'allowed': ALLOWED_ATTACHMENT_EXT,
    },
    'unit_photo': {
        'field': 'attach_unit_photo',
        'allowed': ALLOWED_IMAGE_EXT,
    },
    'approval_proof': {
        'field': 'attach_approval_proof',
        'allowed': ALLOWED_ATTACHMENT_EXT,
    },
    'handover_proof': {
        'field': 'attach_handover_proof',
        'allowed': ALLOWED_ATTACHMENT_EXT,
    },
}

def _tool_request_channel_label(req):
    channel = (req['request_channel'] or '').strip()
    other = (req['request_channel_other'] or '').strip()
    if channel == 'Other' and other:
        return other
    return channel or 'Email'

def _save_tool_request_attachments(db, request_id, section):
    cfg = AC_TOOL_ATTACHMENT_CONFIG.get(section)
    if not cfg:
        return 0
    saved = 0
    for file_obj in request.files.getlist(cfg['field']):
        if not file_obj or not file_obj.filename:
            continue
        url = _save_upload_file(file_obj, 'assetcore/tool_requests', cfg['allowed'])
        if not url:
            continue
        db.execute(
            '''INSERT INTO ac_tool_request_attachments
               (request_id,section,filename,original_name,uploaded_by,uploaded_by_name)
               VALUES(?,?,?,?,?,?)''',
            (request_id, section, url, file_obj.filename,
             session.get('user_id'), session.get('user_name',''))
        )
        saved += 1
    return saved

def _create_asset_from_tool_request(db, req):
    """Create Laptop/PC asset from a completed tool request once."""
    existing_asset_id = req['asset_id']
    if existing_asset_id:
        existing = db.execute('SELECT id FROM ac_assets WHERE id=?', (existing_asset_id,)).fetchone()
        if existing:
            return existing_asset_id, False

    device_type = _clean_asset_device_type(req['item_category'], req['item_name'])
    if not device_type:
        return None, False

    emp_id = req['employee_id'] or None
    manual_name = (req['manual_user_name'] or '').strip() if not emp_id else ''
    started = ''
    if emp_id or manual_name:
        started = req['receipt_date'] or req['received_date'] or date.today().isoformat()

    notes = []
    channel_label = _tool_request_channel_label(req)
    if channel_label:
        notes.append(f"Requested by: {channel_label}")
    if req['reason']:
        notes.append(f"Kebutuhan: {req['reason']}")
    if req['admin_specs']:
        notes.append(f"Spek rekomendasi: {req['admin_specs']}")
    if req['spec_gpu']:
        notes.append(f"GPU: {req['spec_gpu']}")
    if req['spec_screen']:
        notes.append(f"Layar: {req['spec_screen']}")
    if req['pic_support']:
        notes.append(f"PIC Support: {req['pic_support']}")
    if req['ket']:
        notes.append(f"Ket: {req['ket']}")
    price_text = _format_rupiah(req['admin_price'])
    if price_text:
        notes.append(f"Harga pembelian: {price_text}")
    if req['admin_url']:
        notes.append(f"Link pembelian: {req['admin_url']}")

    cur = db.execute(
        '''INSERT INTO ac_assets
           (employee_id,manual_employee_name,device_type,brand,os,processor,ram,disk,
            office_version,asset_tag,serial_number,purchase_date,condition,notes,started_using)
           VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)''',
        (emp_id, manual_name, device_type,
         req['admin_item_type'] or req['item_name'],
         req['spec_os'] or '',
         req['spec_cpu_type'] or '',
         req['spec_ram'] or '',
         req['spec_disk'] or '',
         req['spec_office'] or '',
         req['asset_tag'] or '',
         req['serial_number'] or '',
         req['purchase_date'] or '',
         'Baik',
         ' | '.join(notes),
         started)
    )
    asset_id = cur.lastrowid
    db.execute(
        'UPDATE ac_tool_requests SET asset_id=?, updated_at=datetime("now","localtime") WHERE id=?',
        (asset_id, req['id'])
    )
    return asset_id, True

def _get_ac_masters_dict(db):
    rows = db.execute('SELECT category, name FROM ac_masters ORDER BY name').fetchall()
    res = {c: [] for c in ['cpu', 'ram', 'disk', 'gpu', 'screen', 'os', 'office', 'software']}
    for r in rows:
        if r['category'] in res:
            res[r['category']].append(r['name'])
    return res

# ── Masters Data Spec ──────────────────────────────────────────────────────────
@app.route('/aset/masters', methods=['GET', 'POST'])
@login_required
def ac_masters():
    if not ac_require('ac_manage_assets'): return redirect(url_for('ac_index'))
    db = get_db()
    
    categories = {
        'cpu': 'CPU Type',
        'ram': 'RAM',
        'disk': 'Storage / Disk',
        'gpu': 'GPU',
        'screen': 'Layar / Screen',
        'os': 'Sistem Operasi / OS',
        'office': 'Microsoft Office',
        'software': 'Software / Aplikasi'
    }
    
    if request.method == 'POST':
        category = request.form.get('category', '').strip()
        name = request.form.get('name', '').strip()
        if category in categories and name:
            try:
                db.execute('INSERT INTO ac_masters(category, name) VALUES(?, ?)', (category, name))
                db.commit()
                flash(f'Master {categories[category]} "{name}" berhasil ditambahkan.', 'success')
            except Exception:
                flash(f'Master {categories[category]} "{name}" sudah ada.', 'warning')
        else:
            flash('Kategori dan Nama harus diisi.', 'danger')
        return redirect(url_for('ac_masters'))
        
    masters_by_cat = {c: [] for c in categories}
    rows = db.execute('SELECT * FROM ac_masters ORDER BY category, name').fetchall()
    for r in rows:
        if r['category'] in masters_by_cat:
            masters_by_cat[r['category']].append(r)
            
    return render_template('ac_masters.html', categories=categories, masters=masters_by_cat)

@app.route('/aset/masters/<int:mid>/delete', methods=['POST'])
@login_required
def ac_master_delete(mid):
    if not ac_require('ac_manage_assets'): return redirect(url_for('ac_index'))
    db = get_db()
    m = db.execute('SELECT * FROM ac_masters WHERE id=?', (mid,)).fetchone()
    if m:
        db.execute('DELETE FROM ac_masters WHERE id=?', (mid,))
        db.commit()
        flash('Master berhasil dihapus.', 'success')
    return redirect(url_for('ac_masters'))

@app.route('/aset/')
@login_required
def ac_index():
    db = get_db()
    if not ac_require('ac_view'): return redirect(url_for('portal'))
    from datetime import date as _date, timedelta
    today = _date.today().isoformat()
    limit30 = (_date.today() + timedelta(days=30)).isoformat()
    stats = {
        'assets':    db.execute('SELECT COUNT(*) FROM ac_assets').fetchone()[0],
        'infra':     db.execute('SELECT COUNT(*) FROM ac_infrastructure').fetchone()[0],
        'licenses':  db.execute('SELECT COUNT(*) FROM ac_licenses WHERE is_active=1').fetchone()[0],
        'subs':      db.execute('SELECT COUNT(*) FROM ac_subscriptions WHERE is_active=1').fetchone()[0],
        'requests':  db.execute("SELECT COUNT(*) FROM ac_software_requests WHERE status='Pending'").fetchone()[0],
        'tool_requests': db.execute("SELECT COUNT(*) FROM ac_tool_requests WHERE status='Pending'").fetchone()[0],
    }
    expiring = db.execute(
        "SELECT * FROM ac_subscriptions WHERE is_active=1 AND end_date!='' AND end_date BETWEEN ? AND ? ORDER BY end_date",
        (today, limit30)).fetchall()
    recent_requests = db.execute(
        "SELECT r.*, e.name as emp_name FROM ac_software_requests r LEFT JOIN employees e ON r.employee_id=e.id ORDER BY r.requested_at DESC LIMIT 5"
    ).fetchall()
    recent_tool_requests = db.execute(
        "SELECT r.*, e.name as emp_name FROM ac_tool_requests r LEFT JOIN employees e ON r.employee_id=e.id ORDER BY r.requested_at DESC LIMIT 5"
    ).fetchall()
    recent_history = db.execute(
        """SELECT h.*, e.name as emp_name,
                  a.brand, a.device_type, a.asset_tag, a.id as asset_id
           FROM ac_asset_history h
           LEFT JOIN employees e ON h.employee_id=e.id
           LEFT JOIN ac_assets a ON h.asset_id=a.id
           ORDER BY h.id DESC LIMIT 10"""
    ).fetchall()
    limit30 = (_date.today() + timedelta(days=30)).isoformat()
    maintenance_alert = db.execute(
        """SELECT * FROM ac_maintenance WHERE is_active=1
           AND next_maintenance!='' AND next_maintenance<=?
           ORDER BY next_maintenance ASC LIMIT 10""", (limit30,)
    ).fetchall()
    return render_template('ac_index.html', stats=stats, expiring=expiring,
                           recent_requests=recent_requests, recent_tool_requests=recent_tool_requests,
                           recent_history=recent_history, maintenance_alert=maintenance_alert, today=today)

# ── Assets ────────────────────────────────────────────────────────────────────
@app.route('/aset/assets')
@login_required
def ac_assets():
    if not ac_require('ac_view'): return redirect(url_for('ac_index'))
    db = get_db()
    q             = request.args.get('q', '')
    divisi        = request.args.get('divisi', '')
    status_filter = request.args.get('status', '')       # 'linked' | 'unlinked'
    asset_status  = request.args.get('asset_status', '') # 'aktif' | 'end'
    sort          = request.args.get('sort', 'updated')  # 'updated' | 'name' | 'divisi'
    sql = """SELECT a.*, e.name as emp_name, e.divisi,
             CASE WHEN a.employee_id IS NOT NULL THEN 'linked'
                  WHEN a.manual_employee_name!='' THEN 'unlinked'
                  ELSE 'no_user' END as link_status,
             (SELECT id FROM ac_tool_requests WHERE asset_id=a.id LIMIT 1) as tool_request_id
             FROM ac_assets a LEFT JOIN employees e ON a.employee_id=e.id WHERE 1=1"""
    params = []
    if q:
        sql += ' AND (e.name LIKE ? OR a.manual_employee_name LIKE ? OR a.device_type LIKE ? OR a.asset_tag LIKE ?)'
        params += [f'%{q}%'] * 4
    if divisi:
        sql += ' AND e.divisi=?'; params.append(divisi)
    if status_filter == 'unlinked':
        sql += ' AND a.employee_id IS NULL AND a.manual_employee_name!=""'
    elif status_filter == 'linked':
        sql += ' AND a.employee_id IS NOT NULL'
    if asset_status == 'end':
        sql += " AND a.status='End'"
    elif asset_status == 'aktif':
        sql += " AND (a.status IS NULL OR a.status='Aktif')"
    ORDER_MAP = {
        'updated': 'a.updated_at DESC, a.id DESC',
        'name':    'COALESCE(e.name, a.manual_employee_name) ASC',
        'divisi':  'COALESCE(e.divisi,"") ASC, COALESCE(e.name, a.manual_employee_name) ASC',
    }
    sql += ' ORDER BY a.status ASC, ' + ORDER_MAP.get(sort, ORDER_MAP['updated'])
    assets = db.execute(sql, params).fetchall()
    divisis       = [r[0] for r in db.execute("SELECT DISTINCT divisi FROM employees WHERE divisi!='' ORDER BY divisi").fetchall()]
    unlinked_count = db.execute("SELECT COUNT(*) FROM ac_assets WHERE (status IS NULL OR status='Aktif') AND employee_id IS NULL AND manual_employee_name!=''").fetchone()[0]
    end_count      = db.execute("SELECT COUNT(*) FROM ac_assets WHERE status='End'").fetchone()[0]
    return render_template('ac_assets.html', assets=assets, q=q, divisi=divisi,
                           divisis=divisis, status_filter=status_filter,
                           asset_status=asset_status, sort=sort,
                           unlinked_count=unlinked_count, end_count=end_count)

@app.route('/aset/assets/new', methods=['GET','POST'])
@login_required
def ac_asset_new():
    if not ac_require('ac_manage_assets'): return redirect(url_for('ac_assets'))
    db = get_db()
    employees = db.execute('SELECT id,name,divisi FROM employees WHERE is_active=1 ORDER BY divisi,name').fetchall()
    if request.method == 'POST':
        mode = request.form.get('emp_mode', 'linked')
        emp_id = request.form.get('employee_id') or None if mode == 'linked' else None
        manual_name = request.form.get('manual_employee_name', '').strip() if mode == 'manual' else ''
        today = date.today().isoformat()
        started = today if (emp_id or manual_name) else ''
        cur = db.execute(
            'INSERT INTO ac_assets(employee_id,manual_employee_name,device_type,brand,os,os_license_type,processor,ram,disk,office_version,asset_tag,serial_number,purchase_date,condition,notes,started_using) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)',
            (emp_id, manual_name, request.form.get('device_type','Laptop'),
             request.form.get('brand',''), request.form.get('os',''), request.form.get('os_license_type',''),
             request.form.get('processor',''), request.form.get('ram',''), request.form.get('disk',''),
             request.form.get('office_version',''), request.form.get('asset_tag',''),
             request.form.get('serial_number',''), request.form.get('purchase_date',''),
             request.form.get('condition','Baik'), request.form.get('notes',''), started))
        aid = cur.lastrowid
        for s in [x.strip() for x in request.form.get('softwares','').split('\n') if x.strip()]:
            db.execute('INSERT INTO ac_asset_software(asset_id,software_name) VALUES(?,?)', (aid, s))
        db.commit()
        label = request.form.get('employee_id','') if mode == 'linked' else manual_name
        audit_log('create', 'ac_assets', aid, f"Asset baru: {label}", 'aset')
        flash('Asset berhasil ditambahkan.', 'success')
        return redirect(url_for('ac_asset_detail', aid=aid))
    return render_template('ac_asset_form.html', asset=None, employees=employees, sw_text='', masters=_get_ac_masters_dict(db))

@app.route('/aset/assets/<int:aid>')
@login_required
def ac_asset_detail(aid):
    if not ac_require('ac_view'): return redirect(url_for('ac_assets'))
    db = get_db()
    asset = db.execute('''SELECT a.*, e.name as emp_name, e.divisi, e.jabatan, e.phone, e.email,
                          CASE WHEN a.employee_id IS NOT NULL THEN 'linked'
                               WHEN a.manual_employee_name!='' THEN 'unlinked'
                               ELSE 'no_user' END as link_status
                          FROM ac_assets a LEFT JOIN employees e ON a.employee_id=e.id WHERE a.id=?''', (aid,)).fetchone()
    if not asset: flash('Asset tidak ditemukan.', 'danger'); return redirect(url_for('ac_assets'))
    softwares = db.execute('SELECT * FROM ac_asset_software WHERE asset_id=? ORDER BY software_name', (aid,)).fetchall()
    employees = db.execute('SELECT id,name,divisi FROM employees WHERE is_active=1 ORDER BY divisi,name').fetchall()
    history = db.execute(
        '''SELECT h.*, e.name as emp_name, e.divisi
           FROM ac_asset_history h LEFT JOIN employees e ON h.employee_id=e.id
           WHERE h.asset_id=? ORDER BY h.ended_at DESC, h.id DESC''', (aid,)).fetchall()
    tool_request = db.execute('SELECT id, item_name, status FROM ac_tool_requests WHERE asset_id=?', (aid,)).fetchone()
    return render_template('ac_asset_detail.html', asset=asset, softwares=softwares,
                           employees=employees, history=history, tool_request=tool_request)

@app.route('/aset/assets/<int:aid>/link', methods=['POST'])
@login_required
def ac_asset_link(aid):
    if not ac_require('ac_manage_assets'): return redirect(url_for('ac_asset_detail', aid=aid))
    db = get_db()
    cur_asset = db.execute('SELECT * FROM ac_assets WHERE id=?', (aid,)).fetchone()
    if not cur_asset: return redirect(url_for('ac_assets'))
    action = request.form.get('action', 'link')
    reason = request.form.get('reason', '').strip() or ('Putuskan koneksi' if action == 'unlink' else 'Hubungkan ke karyawan')
    today  = date.today().isoformat()

    if action == 'unlink':
        old_user = _asset_user_display(db, cur_asset)
        _record_asset_history(db, aid, cur_asset['employee_id'], cur_asset['manual_employee_name'],
                              cur_asset['started_using'], reason)
        db.execute('UPDATE ac_assets SET employee_id=NULL, started_using=?, updated_at=datetime("now","localtime") WHERE id=?',
                   ('', aid))
        db.commit()
        _notify_asset_change(db, cur_asset, 'released', old_user, '', reason)
        flash('Link karyawan dilepas. Riwayat penggunaan tersimpan.', 'success')
    else:
        emp_id = request.form.get('employee_id') or None
        if not emp_id:
            flash('Pilih karyawan terlebih dahulu.', 'warning')
            return redirect(url_for('ac_asset_detail', aid=aid))
        old_user = _asset_user_display(db, cur_asset)
        _record_asset_history(db, aid, cur_asset['employee_id'], cur_asset['manual_employee_name'],
                              cur_asset['started_using'], reason)
        db.execute('UPDATE ac_assets SET employee_id=?, manual_employee_name=?, started_using=?, updated_at=datetime("now","localtime") WHERE id=?',
                   (emp_id, '', today, aid))
        db.commit()
        emp = db.execute('SELECT name FROM employees WHERE id=?', (emp_id,)).fetchone()
        new_user = emp['name'] if emp else ''
        _notify_asset_change(db, cur_asset, 'assigned', old_user, new_user, reason)
        flash(f'Asset dihubungkan ke {new_user or "karyawan"}. Riwayat tersimpan.', 'success')
    return redirect(url_for('ac_asset_detail', aid=aid))

@app.route('/aset/assets/<int:aid>/edit', methods=['GET','POST'])
@login_required
def ac_asset_edit(aid):
    if not ac_require('ac_manage_assets'): return redirect(url_for('ac_assets'))
    db = get_db()
    asset = db.execute('SELECT * FROM ac_assets WHERE id=?', (aid,)).fetchone()
    if not asset: flash('Asset tidak ditemukan.', 'danger'); return redirect(url_for('ac_assets'))
    employees = db.execute('SELECT id,name,divisi FROM employees WHERE is_active=1 ORDER BY divisi,name').fetchall()
    if request.method == 'POST':
        mode = request.form.get('emp_mode', 'linked')
        emp_id      = request.form.get('employee_id') or None if mode == 'linked' else None
        manual_name = request.form.get('manual_employee_name', '').strip() if mode == 'manual' else ''
        new_status  = request.form.get('status', 'Aktif')
        today = date.today().isoformat()

        # Alasan yang membuat asset kembali available (stok)
        FREEING_REASONS = {'Resign', 'Pindah ke Laptop Lain', 'Upgrade Laptop'}
        change_reason = request.form.get('change_reason', '').strip()
        has_current_user = bool(asset['employee_id'] or (asset['manual_employee_name'] or '').strip())

        notif_event = notif_old = notif_new = notif_reason = None

        if change_reason in FREEING_REASONS and has_current_user:
            # Paksa asset jadi available — simpan history lalu hapus user
            notif_old = _asset_user_display(db, asset)
            _record_asset_history(db, aid, asset['employee_id'], asset['manual_employee_name'],
                                  asset['started_using'], change_reason)
            emp_id      = None
            manual_name = ''
            new_started = ''
            notif_event = 'released'; notif_reason = change_reason
        else:
            # Deteksi perubahan user biasa → simpan history jika berbeda
            emp_changed = (str(emp_id or '') != str(asset['employee_id'] or '') or
                           manual_name != (asset['manual_employee_name'] or ''))
            if emp_changed:
                reason = change_reason or 'Edit asset'
                notif_old = _asset_user_display(db, asset)
                _record_asset_history(db, aid, asset['employee_id'], asset['manual_employee_name'],
                                      asset['started_using'], reason)
                new_started = today if (emp_id or manual_name) else ''
                notif_event  = 'assigned' if (emp_id or manual_name) else 'released'
                notif_reason = reason
            else:
                new_started = asset['started_using'] or ''

        db.execute(
            'UPDATE ac_assets SET employee_id=?,manual_employee_name=?,started_using=?,status=?,'
            'device_type=?,brand=?,os=?,os_license_type=?,processor=?,ram=?,disk=?,'
            'office_version=?,asset_tag=?,serial_number=?,purchase_date=?,condition=?,notes=?,'
            'updated_at=datetime("now","localtime") WHERE id=?',
            (emp_id, manual_name, new_started, new_status,
             request.form.get('device_type','Laptop'),
             request.form.get('brand',''), request.form.get('os',''), request.form.get('os_license_type',''),
             request.form.get('processor',''), request.form.get('ram',''), request.form.get('disk',''),
             request.form.get('office_version',''), request.form.get('asset_tag',''),
             request.form.get('serial_number',''), request.form.get('purchase_date',''),
             request.form.get('condition','Baik'), request.form.get('notes',''), aid))
        db.execute('DELETE FROM ac_asset_software WHERE asset_id=?', (aid,))
        for s in [x.strip() for x in request.form.get('softwares','').split('\n') if x.strip()]:
            db.execute('INSERT INTO ac_asset_software(asset_id,software_name) VALUES(?,?)', (aid, s))
        db.commit()

        if notif_event:
            if notif_event == 'assigned':
                if emp_id:
                    emp_row = db.execute('SELECT name FROM employees WHERE id=?', (emp_id,)).fetchone()
                    notif_new = emp_row['name'] if emp_row else manual_name
                else:
                    notif_new = manual_name or ''
            _notify_asset_change(db, asset, notif_event, notif_old or '', notif_new or '', notif_reason or '')

        flash('Asset diperbarui.', 'success')
        return redirect(url_for('ac_asset_detail', aid=aid))
    sw_text = '\n'.join(r['software_name'] for r in db.execute('SELECT software_name FROM ac_asset_software WHERE asset_id=? ORDER BY software_name', (aid,)).fetchall())
    return render_template('ac_asset_form.html', asset=asset, employees=employees, sw_text=sw_text, masters=_get_ac_masters_dict(db))

@app.route('/aset/assets/<int:aid>/end', methods=['POST'])
@login_required
def ac_asset_end(aid):
    """Tandai asset sebagai End (rusak total / hilang). Simpan history user terakhir."""
    if not ac_require('ac_manage_assets'): return redirect(url_for('ac_asset_detail', aid=aid))
    db = get_db()
    asset = db.execute('SELECT * FROM ac_assets WHERE id=?', (aid,)).fetchone()
    if not asset: return redirect(url_for('ac_assets'))
    reason = request.form.get('end_reason', 'End').strip() or 'End'
    old_user = _asset_user_display(db, asset)
    _record_asset_history(db, aid, asset['employee_id'], asset['manual_employee_name'],
                          asset['started_using'], reason)
    db.execute(
        'UPDATE ac_assets SET status=?,employee_id=NULL,manual_employee_name=?,started_using=?,'
        'updated_at=datetime("now","localtime") WHERE id=?',
        ('End', asset['manual_employee_name'] or '', '', aid))
    db.commit()
    _notify_asset_change(db, asset, 'end', old_user, '', reason)
    audit_log('end', 'ac_assets', aid, f'Asset ditandai End: {reason}', 'aset')
    flash(f'Asset ditandai sebagai End ({reason}). Riwayat pengguna terakhir tersimpan.', 'warning')
    return redirect(url_for('ac_asset_detail', aid=aid))

@app.route('/aset/assets/<int:aid>/reactivate', methods=['POST'])
@login_required
def ac_asset_reactivate(aid):
    """Reaktifkan asset yang sebelumnya End."""
    if not ac_require('ac_manage_assets'): return redirect(url_for('ac_asset_detail', aid=aid))
    db = get_db()
    db.execute('UPDATE ac_assets SET status=?,updated_at=datetime("now","localtime") WHERE id=?',
               ('Aktif', aid))
    db.commit()
    flash('Asset diaktifkan kembali.', 'success')
    return redirect(url_for('ac_asset_detail', aid=aid))

@app.route('/aset/assets/<int:aid>/delete', methods=['POST'])
@login_required
def ac_asset_delete(aid):
    if not ac_require('ac_manage_assets'): return redirect(url_for('ac_assets'))
    get_db().execute('DELETE FROM ac_assets WHERE id=?', (aid,)); get_db().commit()
    flash('Asset dihapus.', 'success'); return redirect(url_for('ac_assets'))

# ── Maintenance ───────────────────────────────────────────────────────────────

MAINTENANCE_FREQ_DAYS = {
    'Harian': 1, 'Mingguan': 7, 'Bulanan': 30,
    '3 Bulan': 90, '6 Bulan': 180, 'Tahunan': 365,
}

def _maintenance_next(last_date_str, frequency):
    """Hitung next_maintenance dari last_maintenance + frequency."""
    try:
        from datetime import date as _d, timedelta
        last = _d.fromisoformat(last_date_str)
        delta = timedelta(days=MAINTENANCE_FREQ_DAYS.get(frequency, 30))
        return (last + delta).isoformat()
    except Exception:
        return ''

def _maintenance_status(item, today_str):
    nxt = (item['next_maintenance'] or '').strip()
    if not nxt:
        return 'no_schedule'
    if nxt < today_str:
        return 'overdue'
    from datetime import date as _d, timedelta
    soon = (_d.today() + timedelta(days=30)).isoformat()
    return 'upcoming' if nxt <= soon else 'ok'

@app.route('/aset/maintenance')
@login_required
def ac_maintenance():
    if not ac_require('ac_view'): return redirect(url_for('ac_index'))
    db = get_db()
    from datetime import date as _d
    today = _d.today().isoformat()
    q = request.args.get('q', '')
    cat = request.args.get('cat', '')
    sort = request.args.get('sort', 'schedule')
    sql = "SELECT * FROM ac_maintenance WHERE is_active=1"
    params = []
    if q:
        sql += " AND (title LIKE ? OR location LIKE ? OR vendor LIKE ? OR category LIKE ?)"
        params += [f'%{q}%'] * 4
    if cat:
        sql += " AND category=?"
        params.append(cat)
    if sort == 'updated':
        sql += " ORDER BY COALESCE(NULLIF(updated_at,''),'1970') DESC, id DESC"
    else:
        sql += " ORDER BY CASE WHEN next_maintenance='' THEN '9999' ELSE next_maintenance END ASC"
    items = db.execute(sql, params).fetchall()
    items_with_status = [(r, _maintenance_status(r, today)) for r in items]
    categories = [r[0] for r in db.execute("SELECT DISTINCT category FROM ac_maintenance WHERE category!='' ORDER BY category").fetchall()]
    overdue_count = sum(1 for _, s in items_with_status if s == 'overdue')
    return render_template('ac_maintenance.html', items=items_with_status, q=q, cat=cat, sort=sort,
                           categories=categories, today=today, overdue_count=overdue_count)

@app.route('/aset/maintenance/new', methods=['GET', 'POST'])
@app.route('/aset/maintenance/<int:mid>/edit', methods=['GET', 'POST'])
@login_required
def ac_maintenance_form(mid=None):
    if not ac_require('ac_manage_assets'): return redirect(url_for('ac_maintenance'))
    db = get_db()
    item = db.execute('SELECT * FROM ac_maintenance WHERE id=?', (mid,)).fetchone() if mid else None
    if request.method == 'POST':
        title       = request.form.get('title', '').strip()
        category    = request.form.get('category', 'Lainnya').strip()
        location    = request.form.get('location', '').strip()
        description = request.form.get('description', '').strip()
        frequency   = request.form.get('frequency', 'Bulanan').strip()
        last_maint  = request.form.get('last_maintenance', '').strip()
        next_maint  = request.form.get('next_maintenance', '').strip()
        vendor      = request.form.get('vendor', '').strip()
        pic         = request.form.get('pic', '').strip()
        cost_raw    = request.form.get('cost_estimate', '').strip()
        cost_est    = float(cost_raw) if cost_raw else None
        notes       = request.form.get('notes', '').strip()
        if not title:
            flash('Judul wajib diisi.', 'warning')
            return render_template('ac_maintenance_form.html', item=item)
        if not next_maint and last_maint:
            next_maint = _maintenance_next(last_maint, frequency)
        if mid:
            db.execute(
                'UPDATE ac_maintenance SET title=?,category=?,location=?,description=?,frequency=?,'
                'last_maintenance=?,next_maintenance=?,vendor=?,pic=?,cost_estimate=?,notes=?,'
                'updated_at=datetime("now","localtime") WHERE id=?',
                (title, category, location, description, frequency, last_maint, next_maint,
                 vendor, pic, cost_est, notes, mid))
            db.commit()
            flash('Jadwal maintenance diperbarui.', 'success')
        else:
            db.execute(
                'INSERT INTO ac_maintenance(title,category,location,description,frequency,'
                'last_maintenance,next_maintenance,vendor,pic,cost_estimate,notes) VALUES(?,?,?,?,?,?,?,?,?,?,?)',
                (title, category, location, description, frequency, last_maint, next_maint,
                 vendor, pic, cost_est, notes))
            db.commit()
            flash('Jadwal maintenance ditambahkan.', 'success')
        return redirect(url_for('ac_maintenance'))
    return render_template('ac_maintenance_form.html', item=item)

@app.route('/aset/maintenance/<int:mid>')
@login_required
def ac_maintenance_detail(mid):
    if not ac_require('ac_view'): return redirect(url_for('ac_index'))
    db = get_db()
    from datetime import date as _d
    item = db.execute('SELECT * FROM ac_maintenance WHERE id=?', (mid,)).fetchone()
    if not item: flash('Data tidak ditemukan.', 'danger'); return redirect(url_for('ac_maintenance'))
    logs = db.execute('SELECT * FROM ac_maintenance_log WHERE maintenance_id=? ORDER BY done_at DESC, id DESC', (mid,)).fetchall()
    today = _d.today().isoformat()
    status = _maintenance_status(item, today)
    return render_template('ac_maintenance_detail.html', item=item, logs=logs, status=status, today=today)

@app.route('/aset/maintenance/<int:mid>/done', methods=['POST'])
@login_required
def ac_maintenance_done(mid):
    if not ac_require('ac_manage_assets'): return redirect(url_for('ac_maintenance_detail', mid=mid))
    db = get_db()
    item = db.execute('SELECT * FROM ac_maintenance WHERE id=?', (mid,)).fetchone()
    if not item: return redirect(url_for('ac_maintenance'))
    done_at  = request.form.get('done_at', '').strip() or date.today().isoformat()
    done_by  = request.form.get('done_by', '').strip()
    result   = request.form.get('result', 'OK').strip()
    cost_raw = request.form.get('cost', '').strip()
    cost     = float(cost_raw) if cost_raw else None
    notes    = request.form.get('notes', '').strip()
    db.execute(
        'INSERT INTO ac_maintenance_log(maintenance_id,done_at,done_by,cost,result,notes) VALUES(?,?,?,?,?,?)',
        (mid, done_at, done_by, cost, result, notes))
    next_maint = _maintenance_next(done_at, item['frequency'])
    db.execute(
        'UPDATE ac_maintenance SET last_maintenance=?,next_maintenance=?,updated_at=datetime("now","localtime") WHERE id=?',
        (done_at, next_maint, mid))
    db.commit()
    flash(f'Maintenance dicatat. Jadwal berikutnya: {next_maint or "—"}.', 'success')
    return redirect(url_for('ac_maintenance_detail', mid=mid))

@app.route('/aset/maintenance/<int:mid>/delete', methods=['POST'])
@login_required
def ac_maintenance_delete(mid):
    if not ac_require('ac_manage_assets'): return redirect(url_for('ac_maintenance'))
    db = get_db()
    db.execute('DELETE FROM ac_maintenance WHERE id=?', (mid,))
    db.commit()
    flash('Jadwal maintenance dihapus.', 'success')
    return redirect(url_for('ac_maintenance'))

# ── Infrastruktur ─────────────────────────────────────────────────────────────
@app.route('/aset/infra')
@login_required
def ac_infra():
    if not ac_require('ac_view'): return redirect(url_for('ac_index'))
    db = get_db()
    q = request.args.get('q',''); dtype = request.args.get('dtype','')
    sort = request.args.get('sort', 'updated')
    sql = 'SELECT * FROM ac_infrastructure WHERE 1=1'; params = []
    if q:
        sql += ' AND (device_type LIKE ? OR model LIKE ? OR nickname LIKE ? OR serial_number LIKE ?)'; params += [f'%{q}%']*4
    if dtype:
        sql += ' AND device_type=?'; params.append(dtype)
    order = ("COALESCE(NULLIF(updated_at,''),'1970') DESC, id DESC" if sort == 'updated'
             else 'device_type ASC, nickname ASC')
    items = db.execute(sql + ' ORDER BY ' + order, params).fetchall()
    dtypes = [r[0] for r in db.execute("SELECT DISTINCT device_type FROM ac_infrastructure ORDER BY device_type").fetchall()]
    return render_template('ac_infra.html', items=items, q=q, dtype=dtype, dtypes=dtypes, sort=sort)

@app.route('/aset/infra/new', methods=['GET','POST'])
@login_required
def ac_infra_new():
    if not ac_require('ac_manage_infra'): return redirect(url_for('ac_infra'))
    db = get_db()
    if request.method == 'POST':
        db.execute('INSERT INTO ac_infrastructure(device_type,brand,model,description,serial_number,nickname,ups_group,location,status,condition_notes) VALUES(?,?,?,?,?,?,?,?,?,?)',
                   (request.form.get('device_type',''), request.form.get('brand',''), request.form.get('model',''),
                    request.form.get('description',''), request.form.get('serial_number',''), request.form.get('nickname',''),
                    request.form.get('ups_group',''), request.form.get('location',''),
                    request.form.get('status','Aktif'), request.form.get('condition_notes','')))
        db.commit(); flash('Perangkat ditambahkan.', 'success')
        return redirect(url_for('ac_infra'))
    return render_template('ac_infra_form.html', item=None)

@app.route('/aset/infra/<int:iid>/edit', methods=['GET','POST'])
@login_required
def ac_infra_edit(iid):
    if not ac_require('ac_manage_infra'): return redirect(url_for('ac_infra'))
    db = get_db()
    item = db.execute('SELECT * FROM ac_infrastructure WHERE id=?', (iid,)).fetchone()
    if not item: flash('Tidak ditemukan.', 'danger'); return redirect(url_for('ac_infra'))
    if request.method == 'POST':
        db.execute('UPDATE ac_infrastructure SET device_type=?,brand=?,model=?,description=?,serial_number=?,nickname=?,ups_group=?,location=?,status=?,condition_notes=?,updated_at=datetime("now","localtime") WHERE id=?',
                   (request.form.get('device_type',''), request.form.get('brand',''), request.form.get('model',''),
                    request.form.get('description',''), request.form.get('serial_number',''), request.form.get('nickname',''),
                    request.form.get('ups_group',''), request.form.get('location',''),
                    request.form.get('status','Aktif'), request.form.get('condition_notes',''), iid))
        db.commit(); flash('Perangkat diperbarui.', 'success')
        return redirect(url_for('ac_infra'))
    return render_template('ac_infra_form.html', item=item)

@app.route('/aset/infra/<int:iid>/delete', methods=['POST'])
@login_required
def ac_infra_delete(iid):
    if not ac_require('ac_manage_infra'): return redirect(url_for('ac_infra'))
    get_db().execute('DELETE FROM ac_infrastructure WHERE id=?', (iid,)); get_db().commit()
    flash('Perangkat dihapus.', 'success'); return redirect(url_for('ac_infra'))

# ── Lisensi ───────────────────────────────────────────────────────────────────
@app.route('/aset/licenses')
@login_required
def ac_licenses():
    if not ac_require('ac_view'): return redirect(url_for('ac_index'))
    db = get_db()
    q = request.args.get('q','')
    sort = request.args.get('sort', 'updated')
    sql = "SELECT l.*, COUNT(la.id) as assigned FROM ac_licenses l LEFT JOIN ac_license_assignments la ON la.license_id=l.id"
    params = []
    if q:
        sql += ' WHERE l.software_name LIKE ?'; params.append(f'%{q}%')
    order = ("COALESCE(NULLIF(l.updated_at,''),'1970') DESC, l.id DESC" if sort == 'updated'
             else 'l.software_name ASC')
    lics = db.execute(sql + ' GROUP BY l.id ORDER BY ' + order, params).fetchall()
    return render_template('ac_licenses.html', licenses=lics, q=q, sort=sort)

@app.route('/aset/licenses/new', methods=['GET','POST'])
@app.route('/aset/licenses/<int:lid>/edit', methods=['GET','POST'])
@login_required
def ac_license_form(lid=None):
    if not ac_require('ac_manage_licenses'): return redirect(url_for('ac_licenses'))
    db = get_db()
    lic = db.execute('SELECT * FROM ac_licenses WHERE id=?', (lid,)).fetchone() if lid else None
    employees = db.execute('SELECT id,name,divisi FROM employees WHERE is_active=1 ORDER BY name').fetchall()
    assignments = [r['employee_id'] for r in db.execute('SELECT employee_id FROM ac_license_assignments WHERE license_id=?', (lid,)).fetchall()] if lid else []
    if request.method == 'POST':
        if lid:
            db.execute('UPDATE ac_licenses SET software_name=?,license_key=?,license_type=?,version=?,year=?,max_seats=?,notes=?,is_active=?,updated_at=datetime("now","localtime") WHERE id=?',
                       (request.form['software_name'].strip(), request.form.get('license_key','').strip(),
                        request.form.get('license_type','Perpetual'), request.form.get('version','').strip(),
                        request.form.get('year') or None, request.form.get('max_seats',1),
                        request.form.get('notes','').strip(), 1 if request.form.get('is_active') else 0, lid))
            db.execute('DELETE FROM ac_license_assignments WHERE license_id=?', (lid,))
        else:
            cur = db.execute('INSERT INTO ac_licenses(software_name,license_key,license_type,version,year,max_seats,notes) VALUES(?,?,?,?,?,?,?)',
                       (request.form['software_name'].strip(), request.form.get('license_key','').strip(),
                        request.form.get('license_type','Perpetual'), request.form.get('version','').strip(),
                        request.form.get('year') or None, request.form.get('max_seats',1), request.form.get('notes','').strip()))
            lid = cur.lastrowid
        for emp_id in request.form.getlist('assigned_employees'):
            if emp_id: db.execute('INSERT INTO ac_license_assignments(license_id,employee_id) VALUES(?,?)', (lid, emp_id))
        db.commit(); flash('Lisensi disimpan.', 'success')
        return redirect(url_for('ac_licenses'))
    return render_template('ac_license_form.html', lic=lic, employees=employees, assignments=assignments)

@app.route('/aset/licenses/<int:lid>/delete', methods=['POST'])
@login_required
def ac_license_delete(lid):
    if not ac_require('ac_manage_licenses'): return redirect(url_for('ac_licenses'))
    get_db().execute('DELETE FROM ac_licenses WHERE id=?', (lid,)); get_db().commit()
    flash('Lisensi dihapus.', 'success'); return redirect(url_for('ac_licenses'))

# ── Subscription ──────────────────────────────────────────────────────────────
@app.route('/aset/subscriptions')
@login_required
def ac_subscriptions():
    if not ac_require('ac_view'): return redirect(url_for('ac_index'))
    from datetime import date as _date, timedelta
    today = _date.today()
    settings = get_settings(get_db())
    reminder_days_raw = settings.get('ac_sub_reminder_days', '30,14,7,1')
    reminder_days = [int(d.strip()) for d in reminder_days_raw.split(',') if d.strip().isdigit()]
    sort = request.args.get('sort', 'updated')
    db = get_db()
    order = ("COALESCE(NULLIF(updated_at,''),'1970') DESC, id DESC" if sort == 'updated'
             else 'is_active DESC, end_date ASC')
    subs = db.execute(f'SELECT * FROM ac_subscriptions ORDER BY {order}').fetchall()
    return render_template('ac_subscriptions.html',
                           subscriptions=subs, sort=sort,
                           today=today.isoformat(),
                           today_plus_7=(today + timedelta(days=7)).isoformat(),
                           today_plus_30=(today + timedelta(days=30)).isoformat(),
                           reminder_days=sorted(reminder_days, reverse=True))

@app.route('/aset/subscriptions/new', methods=['GET','POST'])
@app.route('/aset/subscriptions/<int:sid>/edit', methods=['GET','POST'])
@login_required
def ac_subscription_form(sid=None):
    if not ac_require('ac_manage_subs'): return redirect(url_for('ac_subscriptions'))
    db = get_db()
    sub = db.execute('SELECT * FROM ac_subscriptions WHERE id=?', (sid,)).fetchone() if sid else None
    if request.method == 'POST':
        vals = (request.form['provider'].strip(), request.form.get('category','SaaS'),
                request.form.get('billing_cycle','Monthly'), request.form.get('start_date',''),
                request.form.get('end_date',''), request.form.get('username','').strip(),
                request.form.get('password','').strip(), request.form.get('access_url','').strip(),
                request.form.get('notes','').strip())
        if sid:
            db.execute('UPDATE ac_subscriptions SET provider=?,category=?,billing_cycle=?,start_date=?,end_date=?,username=?,password=?,access_url=?,notes=?,is_active=?,updated_at=datetime("now","localtime") WHERE id=?',
                       vals + (1 if request.form.get('is_active') else 0, sid))
        else:
            db.execute('INSERT INTO ac_subscriptions(provider,category,billing_cycle,start_date,end_date,username,password,access_url,notes) VALUES(?,?,?,?,?,?,?,?,?)', vals)
        db.commit(); flash('Subscription disimpan.', 'success')
        return redirect(url_for('ac_subscriptions'))
    return render_template('ac_subscription_form.html', sub=sub)

@app.route('/aset/subscriptions/<int:sid>/delete', methods=['POST'])
@login_required
def ac_subscription_delete(sid):
    if not ac_require('ac_manage_subs'): return redirect(url_for('ac_subscriptions'))
    get_db().execute('DELETE FROM ac_subscriptions WHERE id=?', (sid,)); get_db().commit()
    flash('Subscription dihapus.', 'success'); return redirect(url_for('ac_subscriptions'))

@app.route('/aset/subscriptions/remind', methods=['POST'])
@login_required
def ac_subscription_remind():
    """Trigger manual reminder untuk semua subscription yang mendekati expired."""
    if not ac_require('ac_manage_subs'): return redirect(url_for('ac_subscriptions'))
    sent, failed = run_subscription_reminders(triggered_by='manual')
    if sent == 0 and failed == 0:
        flash('Tidak ada subscription yang perlu diingatkan (belum masuk periode reminder).', 'info')
    else:
        flash(f'Reminder terkirim: {sent} sukses, {failed} gagal.', 'success' if failed == 0 else 'warning')
    return redirect(url_for('ac_subscriptions'))

@app.route('/aset/subscriptions/<int:sid>/remind', methods=['POST'])
@login_required
def ac_subscription_remind_one(sid):
    """Trigger reminder manual untuk satu subscription tertentu."""
    if not ac_require('ac_manage_subs'): return redirect(url_for('ac_subscriptions'))
    db = get_db()
    sub = db.execute('SELECT * FROM ac_subscriptions WHERE id=?', (sid,)).fetchone()
    if not sub:
        flash('Subscription tidak ditemukan.', 'danger')
        return redirect(url_for('ac_subscriptions'))
    settings = get_settings(db)
    try:
        days_left = (date.fromisoformat(sub['end_date']) - date.today()).days if sub['end_date'] else 999
    except Exception:
        days_left = 999
    sent = failed = 0
    subj   = f"[AssetCore] Reminder {sub['provider']} — {'berakhir dalam '+str(days_left)+' hari' if days_left > 0 else 'BERAKHIR HARI INI'}"
    html   = compose_sub_email(sub, days_left)
    tg_msg = compose_sub_telegram(sub, days_left)
    wa_msg = compose_sub_wa(sub, days_left)
    bot_token  = settings.get('telegram_bot_token', '').strip()
    wa_url     = settings.get('openwa_url', '').strip()
    wa_key     = settings.get('openwa_api_key', '').strip()
    wa_session = get_openwa_session(settings, 'aset')
    wa_enabled = settings.get('openwa_enabled', '0') == '1'
    ac_emails  = get_ac_notification_emails(settings)
    ac_tg_ids  = get_ac_notification_telegram_ids(settings)
    ac_phones  = get_ac_notification_wa_phones(settings)
    if settings.get('smtp_host', '').strip() and ac_emails:
        for to_email in ac_emails:
            ok, err = send_email(settings, to_email, subj, html)
            audit_notif('email', to_email, subj, html, ok, err, 'manual', app_slug='aset')
            if ok: sent += 1
            else:  failed += 1
    if bot_token and ac_tg_ids:
        for chat_id in ac_tg_ids:
            ok, err = send_telegram(bot_token, chat_id, tg_msg, _log_subject=subj, _app_slug='aset')
            if ok: sent += 1
            else:  failed += 1
    if wa_enabled and wa_url and ac_phones:
        for phone in ac_phones:
            ok, err = send_whatsapp(wa_url, wa_key, wa_session, phone, wa_msg)
            audit_notif('whatsapp', phone, subj, wa_msg, ok, err, 'manual', app_slug='aset')
            if ok: sent += 1
            else:  failed += 1
    db.execute("UPDATE ac_subscriptions SET last_reminder_sent=? WHERE id=?",
               (date.today().isoformat(), sid))
    db.commit()
    if sent == 0 and failed == 0:
        flash(f'Tidak ada penerima notifikasi yang dikonfigurasi. Isi dulu di Pengaturan → Notifikasi.', 'warning')
    else:
        flash(f'Reminder "{sub["provider"]}": {sent} terkirim, {failed} gagal.', 'success' if failed == 0 else 'warning')
    return redirect(url_for('ac_subscriptions'))

# ── AssetCore Settings ────────────────────────────────────────────────────────
AC_SETTINGS_KEYS = [
    'ac_sub_reminder_enabled', 'ac_sub_reminder_days',
    'ac_notification_emails', 'ac_notification_telegram_ids', 'ac_notification_wa_phones',
]

@app.route('/aset/settings', methods=['GET', 'POST'])
@login_required
def ac_settings():
    if not ac_require('ac_manage_subs'): return redirect(url_for('ac_index'))
    db = get_db()
    if request.method == 'POST':
        for k in AC_SETTINGS_KEYS:
            if k == 'ac_sub_reminder_enabled':
                v = '1' if request.form.get(k) else '0'
            else:
                v = request.form.get(k, '').strip()
            save_setting(db, k, v)
        db.commit()
        flash('Pengaturan notifikasi AssetCore disimpan.', 'success')
        return redirect(url_for('ac_settings'))
    cfg = get_settings(db)
    return render_template('ac_settings.html', cfg=cfg)

@app.route('/aset/settings/test-email', methods=['POST'])
@login_required
def ac_settings_test_email():
    if not ac_require('ac_manage_subs'): return jsonify({'ok': False, 'msg': 'Akses ditolak'})
    db = get_db()
    settings = get_settings(db)
    to_email = request.form.get('test_email', '').strip()
    if not to_email:
        return jsonify({'ok': False, 'msg': 'Masukkan alamat email tujuan'})
    if not settings.get('smtp_host', '').strip():
        return jsonify({'ok': False, 'msg': 'SMTP belum dikonfigurasi di Pengaturan Sistem Portal'})
    ok, err = send_email(settings, to_email, '[AssetCore] Test Email Notifikasi',
                         '<h3>✅ Test berhasil!</h3><p>Konfigurasi notifikasi email AssetCore sudah benar.</p>')
    return jsonify({'ok': ok, 'msg': 'Email berhasil dikirim' if ok else str(err)})

@app.route('/aset/settings/test-telegram', methods=['POST'])
@login_required
def ac_settings_test_telegram():
    if not ac_require('ac_manage_subs'): return jsonify({'ok': False, 'msg': 'Akses ditolak'})
    db = get_db()
    settings = get_settings(db)
    bot_token = settings.get('telegram_bot_token', '').strip()
    chat_id   = request.form.get('test_chat_id', '').strip()
    if not bot_token:
        return jsonify({'ok': False, 'msg': 'Bot Token belum dikonfigurasi di Pengaturan Sistem Portal'})
    if not chat_id:
        return jsonify({'ok': False, 'msg': 'Masukkan Chat ID tujuan test'})
    ok, err = send_telegram(bot_token, normalize_telegram_id(chat_id),
                            '✅ <b>Test berhasil!</b>\n\nNotifikasi Telegram AssetCore sudah terhubung.',
                            _app_slug='aset')
    return jsonify({'ok': ok, 'msg': 'Pesan Telegram berhasil dikirim' if ok else str(err)})

@app.route('/aset/settings/test-whatsapp', methods=['POST'])
@login_required
def ac_settings_test_whatsapp():
    if not ac_require('ac_manage_subs'): return jsonify({'ok': False, 'msg': 'Akses ditolak'})
    db = get_db()
    settings = get_settings(db)
    wa_url     = settings.get('openwa_url', '').strip()
    wa_key     = settings.get('openwa_api_key', '').strip()
    wa_session = get_openwa_session(settings, 'aset')
    phone      = request.form.get('test_wa_phone', '').strip()
    if not settings.get('openwa_enabled', '0') == '1' or not wa_url:
        return jsonify({'ok': False, 'msg': 'WhatsApp (OpenWA) belum diaktifkan di Pengaturan Sistem Portal'})
    if not phone:
        return jsonify({'ok': False, 'msg': 'Masukkan nomor HP tujuan test'})
    ok, err = send_whatsapp(wa_url, wa_key, wa_session, phone,
                            '✅ *Test berhasil!*\n\nNotifikasi WhatsApp AssetCore sudah terhubung.')
    return jsonify({'ok': ok, 'msg': f'Pesan terkirim ke {normalize_phone_wa(phone)}' if ok else str(err)})

# ── Software Requests ─────────────────────────────────────────────────────────
@app.route('/aset/requests')
@login_required
def ac_requests():
    if not ac_require('ac_view'): return redirect(url_for('ac_index'))
    db = get_db()
    status_filter = request.args.get('status','')
    sort = request.args.get('sort', 'updated')
    sql = "SELECT r.*, e.name as emp_name, e.divisi FROM ac_software_requests r LEFT JOIN employees e ON r.employee_id=e.id WHERE 1=1"
    params = []
    if status_filter:
        sql += ' AND r.status=?'; params.append(status_filter)
    order = ("COALESCE(NULLIF(r.updated_at,''),r.requested_at) DESC, r.id DESC" if sort == 'updated'
             else 'r.requested_at DESC')
    reqs = db.execute(sql + ' ORDER BY ' + order, params).fetchall()
    return render_template('ac_requests.html', requests=reqs, status_filter=status_filter, sort=sort)

@app.route('/aset/requests/new', methods=['GET','POST'])
@login_required
def ac_request_new():
    if not ac_require('ac_manage_requests'): return redirect(url_for('ac_requests'))
    db = get_db()
    employees = db.execute('SELECT id,name,divisi FROM employees WHERE is_active=1 ORDER BY name').fetchall()
    if request.method == 'POST':
        db.execute('INSERT INTO ac_software_requests(employee_id,software_name,version,reason) VALUES(?,?,?,?)',
                   (request.form.get('employee_id') or None, request.form['software_name'].strip(),
                    request.form.get('version','').strip(), request.form.get('reason','').strip()))
        db.commit(); flash('Request ditambahkan.', 'success')
        return redirect(url_for('ac_requests'))
    return render_template('ac_request_form.html', employees=employees)

@app.route('/aset/requests/<int:rid>/status', methods=['POST'])
@login_required
def ac_request_status(rid):
    if not ac_require('ac_manage_requests'): return redirect(url_for('ac_requests'))
    from datetime import datetime as _dt
    db = get_db()
    new_status = request.form.get('status')
    resolved_at = _dt.now().strftime('%Y-%m-%d %H:%M:%S') if new_status in ('Approved','Rejected','Installed') else ''
    db.execute('UPDATE ac_software_requests SET status=?,notes=?,resolved_at=?,resolved_by=?,updated_at=datetime("now","localtime") WHERE id=?',
               (new_status, request.form.get('notes','').strip(), resolved_at, session.get('user_name',''), rid))
    db.commit(); flash(f'Status diubah ke {new_status}.', 'success')
    return redirect(url_for('ac_requests'))

@app.route('/aset/requests/<int:rid>/delete', methods=['POST'])
@login_required
def ac_request_delete(rid):
    if not ac_require('ac_manage_requests'): return redirect(url_for('ac_requests'))
    get_db().execute('DELETE FROM ac_software_requests WHERE id=?', (rid,)); get_db().commit()
    flash('Request dihapus.', 'success'); return redirect(url_for('ac_requests'))

# ── Work Tool Requests ───────────────────────────────────────────────────────
@app.route('/aset/tool-requests')
@login_required
def ac_tool_requests():
    if not ac_require('ac_view'): return redirect(url_for('ac_index'))
    db = get_db()
    status_filter = request.args.get('status','')
    sort = request.args.get('sort', 'updated')
    if DB_TYPE == 'postgresql':
        completion_days_expr = (
            "CASE WHEN r.receipt_date!='' AND r.received_date!='' "
            "THEN (NULLIF(r.receipt_date,'')::date - NULLIF(r.received_date,'')::date + 1) "
            "ELSE NULL END"
        )
    else:
        completion_days_expr = (
            "CASE WHEN r.receipt_date!='' AND r.received_date!='' "
            "THEN CAST((julianday(r.receipt_date) - julianday(r.received_date) + 1) AS INTEGER) "
            "ELSE NULL END"
        )
    sql = """SELECT r.*, e.name as emp_name, e.divisi,
             {completion_days_expr} as completion_days
             FROM ac_tool_requests r LEFT JOIN employees e ON r.employee_id=e.id WHERE 1=1"""
    sql = sql.format(completion_days_expr=completion_days_expr)
    params = []
    if status_filter:
        sql += ' AND r.status=?'; params.append(status_filter)
    order = ("COALESCE(NULLIF(r.updated_at,''),r.requested_at) DESC, r.id DESC" if sort == 'updated'
             else 'r.requested_at DESC')
    reqs = db.execute(sql + ' ORDER BY ' + order, params).fetchall()
    attachments_by_request = {}
    req_ids = [r['id'] for r in reqs]
    if req_ids:
        placeholders = ','.join(['?'] * len(req_ids))
        att_rows = db.execute(
            f'''SELECT * FROM ac_tool_request_attachments
                WHERE request_id IN ({placeholders})
                ORDER BY section, id''',
            req_ids
        ).fetchall()
        for att in att_rows:
            grouped = attachments_by_request.setdefault(
                att['request_id'],
                {'request_capture': [], 'unit_photo': []}
            )
            grouped.setdefault(att['section'], []).append(att)
    employees = db.execute('SELECT id,name,divisi FROM employees WHERE is_active=1 ORDER BY divisi,name').fetchall()
    return render_template('ac_tool_requests.html', requests=reqs, employees=employees, masters=_get_ac_masters_dict(db),
                           attachments_by_request=attachments_by_request,
                           status_filter=status_filter, sort=sort)

@app.route('/aset/tool-requests/new', methods=['GET','POST'])
@login_required
def ac_tool_request_new():
    if not ac_require('ac_manage_requests'): return redirect(url_for('ac_tool_requests'))
    db = get_db()
    employees = db.execute('SELECT id,name,divisi FROM employees WHERE is_active=1 ORDER BY divisi,name').fetchall()
    if request.method == 'POST':
        item_name = request.form['item_name'].strip()
        cur = db.execute(
            '''INSERT INTO ac_tool_requests
               (employee_id,manual_user_name,requestor_name,item_name,item_category,request_channel,request_channel_other,
                reason,admin_item_type,admin_specs,
                admin_url,admin_price,request_date,purchase_date,received_date,receipt_date,pic_support,
                ket,spec_cpu_type,spec_ram,spec_disk,spec_gpu,spec_screen,spec_os,spec_office,
                asset_tag,serial_number,notes)
               VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)''',
            (request.form.get('employee_id') or None,
             request.form.get('manual_user_name','').strip(),
             request.form.get('requestor_name','').strip(),
             item_name,
             request.form.get('item_category','Laptop').strip() or 'Laptop',
             request.form.get('request_channel','Email').strip() or 'Email',
             request.form.get('request_channel_other','').strip(),
             request.form.get('reason','').strip(),
             request.form.get('admin_item_type','').strip(),
             request.form.get('admin_specs','').strip(),
             request.form.get('admin_url','').strip(),
             _parse_float_value(request.form.get('admin_price')),
             request.form.get('request_date','').strip(),
             request.form.get('purchase_date','').strip(),
             request.form.get('received_date','').strip(),
             request.form.get('receipt_date','').strip(),
             request.form.get('pic_support','').strip(),
             request.form.get('ket','').strip(),
             request.form.get('spec_cpu_type','').strip(),
             request.form.get('spec_ram','').strip(),
             request.form.get('spec_disk','').strip(),
             request.form.get('spec_gpu','').strip(),
             request.form.get('spec_screen','').strip(),
             request.form.get('spec_os','').strip(),
             request.form.get('spec_office','').strip(),
             request.form.get('asset_tag','').strip(),
             request.form.get('serial_number','').strip(),
             request.form.get('notes','').strip()))
        rid = cur.lastrowid
        capture_saved = _save_tool_request_attachments(db, rid, 'request_capture')
        photo_saved = _save_tool_request_attachments(db, rid, 'unit_photo')
        db.commit()
        msg = 'Request alat kerja ditambahkan.'
        if capture_saved or photo_saved:
            msg += f' Attachment tersimpan: {capture_saved + photo_saved}.'
        flash(msg, 'success')
        return redirect(url_for('ac_tool_requests'))
    return render_template('ac_tool_request_form.html', employees=employees, today=date.today().isoformat(), masters=_get_ac_masters_dict(db))

@app.route('/aset/tool-requests/<int:rid>/status', methods=['POST'])
@login_required
def ac_tool_request_status(rid):
    if not ac_require('ac_manage_requests'): return redirect(url_for('ac_tool_requests'))
    from datetime import datetime as _dt
    db = get_db()
    new_status = request.form.get('status')
    resolved_at = _dt.now().strftime('%Y-%m-%d %H:%M:%S') if new_status in ('Approved','Rejected','Completed') else ''

    # Mandatory attachment proof for Approved / Rejected
    if new_status in ('Approved', 'Rejected'):
        proof_files = [f for f in request.files.getlist('attach_approval_proof')
                       if f and f.filename]
        if not proof_files:
            flash(
                f'Status "{new_status}" memerlukan bukti attachment (capture approval/rejection). '
                'Harap upload setidaknya satu file sebagai bukti.',
                'danger'
            )
            return redirect(url_for('ac_tool_requests'))

    # Mandatory handover proof for Completed
    if new_status == 'Completed':
        handover_files = [f for f in request.files.getlist('attach_handover_proof')
                          if f and f.filename]
        if not handover_files:
            flash(
                'Status "Completed" memerlukan bukti serah terima perangkat ke user. '
                'Harap upload setidaknya satu file (foto/dokumen serah terima).',
                'danger'
            )
            return redirect(url_for('ac_tool_requests'))

    admin_item_type = request.form.get('admin_item_type', '').strip()
    admin_specs = request.form.get('admin_specs', '').strip()
    admin_url = request.form.get('admin_url', '').strip()
    admin_price = _parse_float_value(request.form.get('admin_price'))

    db.execute('''UPDATE ac_tool_requests
                  SET employee_id=?, manual_user_name=?, requestor_name=?, item_name=?, item_category=?,
                      request_channel=?, request_channel_other=?, reason=?,
                      status=?, notes=?, admin_item_type=?, admin_specs=?, admin_url=?, admin_price=?,
                      request_date=?, purchase_date=?, received_date=?, receipt_date=?, pic_support=?,
                      ket=?, spec_cpu_type=?, spec_ram=?, spec_disk=?, spec_gpu=?, spec_screen=?,
                      spec_os=?, spec_office=?, asset_tag=?, serial_number=?,
                      resolved_at=?, resolved_by=?, updated_at=datetime("now","localtime")
                  WHERE id=?''',
               (request.form.get('employee_id') or None,
                request.form.get('manual_user_name','').strip(),
                request.form.get('requestor_name','').strip(),
                request.form.get('item_name','').strip(),
                request.form.get('item_category','Laptop').strip() or 'Laptop',
                request.form.get('request_channel','Email').strip() or 'Email',
                request.form.get('request_channel_other','').strip(),
                request.form.get('reason','').strip(),
                new_status, request.form.get('notes','').strip(), admin_item_type, admin_specs, admin_url, admin_price,
                request.form.get('request_date','').strip(),
                request.form.get('purchase_date','').strip(),
                request.form.get('received_date','').strip(),
                request.form.get('receipt_date','').strip(),
                request.form.get('pic_support','').strip(),
                request.form.get('ket','').strip(),
                request.form.get('spec_cpu_type','').strip(),
                request.form.get('spec_ram','').strip(),
                request.form.get('spec_disk','').strip(),
                request.form.get('spec_gpu','').strip(),
                request.form.get('spec_screen','').strip(),
                request.form.get('spec_os','').strip(),
                request.form.get('spec_office','').strip(),
                request.form.get('asset_tag','').strip(),
                request.form.get('serial_number','').strip(),
                resolved_at, session.get('user_name',''), rid))
    capture_saved = _save_tool_request_attachments(db, rid, 'request_capture')
    photo_saved = _save_tool_request_attachments(db, rid, 'unit_photo')
    approval_saved = _save_tool_request_attachments(db, rid, 'approval_proof')
    handover_saved = _save_tool_request_attachments(db, rid, 'handover_proof')

    asset_id = None
    asset_created = False
    if new_status == 'Completed' and request.form.get('create_asset') == '1':
        req = db.execute('SELECT * FROM ac_tool_requests WHERE id=?', (rid,)).fetchone()
        if req and _tool_request_targets_asset(req):
            asset_id, asset_created = _create_asset_from_tool_request(db, req)

    db.commit()
    if asset_created:
        audit_log('create', 'ac_assets', asset_id, f"Asset dari request alat kerja #{rid}", 'aset')
        msg = f'Status request alat kerja diubah ke {new_status}. Asset Laptop/PC berhasil dibuat.'
        total_att = capture_saved + photo_saved + approval_saved + handover_saved
        if total_att:
            msg += f' Attachment baru tersimpan: {total_att}.'
        flash(msg, 'success')
    elif new_status == 'Completed' and request.form.get('create_asset') == '1' and not asset_id:
        msg = f'Status request alat kerja diubah ke {new_status}. Asset tidak dibuat karena kategori bukan Laptop/PC.'
        total_att = capture_saved + photo_saved + approval_saved + handover_saved
        if total_att:
            msg += f' Attachment baru tersimpan: {total_att}.'
        flash(msg, 'warning')
    else:
        msg = f'Status request alat kerja diubah ke {new_status}.'
        total_att = capture_saved + photo_saved + approval_saved + handover_saved
        if total_att:
            msg += f' Attachment baru tersimpan: {total_att}.'
        flash(msg, 'success')
    return redirect(url_for('ac_tool_requests'))

@app.route('/aset/tool-requests/<int:rid>/attachments/<int:att_id>/delete', methods=['POST'])
@login_required
def ac_tool_request_attachment_delete(rid, att_id):
    if not ac_require('ac_manage_requests'): return redirect(url_for('ac_tool_requests'))
    db = get_db()
    att = db.execute(
        'SELECT * FROM ac_tool_request_attachments WHERE id=? AND request_id=?',
        (att_id, rid)
    ).fetchone()
    if att:
        filename = att['filename'] or ''
        if filename.startswith('/static/uploads/'):
            try:
                local_path = os.path.join(os.path.dirname(__file__), filename.lstrip('/'))
                if os.path.exists(local_path):
                    os.remove(local_path)
            except Exception:
                pass
        db.execute('DELETE FROM ac_tool_request_attachments WHERE id=?', (att_id,))
        db.commit()
        flash('Attachment request dihapus.', 'success')
    return redirect(url_for('ac_tool_requests'))

@app.route('/aset/tool-requests/<int:rid>/delete', methods=['POST'])
@login_required
def ac_tool_request_delete(rid):
    if not ac_require('ac_manage_requests'): return redirect(url_for('ac_tool_requests'))
    get_db().execute('DELETE FROM ac_tool_requests WHERE id=?', (rid,)); get_db().commit()
    flash('Request alat kerja dihapus.', 'success'); return redirect(url_for('ac_tool_requests'))

# ─── ProjectCore ───────────────────────────────────────────────────────────────

def _pc_next_issue_no(db, project_id):
    last = db.execute(
        "SELECT issue_no FROM pc_issues WHERE project_id=? ORDER BY id DESC LIMIT 1", (project_id,)
    ).fetchone()
    proj = db.execute("SELECT code FROM pc_projects WHERE id=?", (project_id,)).fetchone()
    prefix = proj['code'].upper() if proj else 'PC'
    if not last:
        return f"{prefix}-001"
    try:
        n = int(last['issue_no'].rsplit('-', 1)[-1]) + 1
    except Exception:
        n = 1
    return f"{prefix}-{n:03d}"

def _pc_log_issue(db, issue_id, action, old_val, new_val, notes=''):
    uid = session.get('user_id')
    db.execute(
        "INSERT INTO pc_issue_history(issue_id,action,old_value,new_value,changed_by,notes) VALUES(?,?,?,?,?,?)",
        (issue_id, action, old_val or '', new_val or '', uid, notes)
    )

def _pc_members(db, project_id):
    rows = db.execute(
        '''SELECT pm.id, pm.role, pm.name_ext,
                  e.id as emp_id, e.name, e.jabatan, e.divisi
           FROM pc_members pm
           LEFT JOIN employees e ON e.id=pm.employee_id
           WHERE pm.project_id=? ORDER BY pm.role, e.name, pm.name_ext''', (project_id,)
    ).fetchall()
    result = []
    for r in rows:
        d = dict(r)
        if not d['name'] and d['name_ext']:
            d['name'] = d['name_ext']
            d['jabatan'] = 'Eksternal'
        result.append(d)
    return result

@app.route('/project/')
@app.route('/project')
@login_required
def pc_index():
    db = get_db()
    projects = db.execute(
        '''SELECT p.*, e.name as pic_name,
                  (SELECT COUNT(*) FROM pc_issues WHERE project_id=p.id) as total_issues,
                  (SELECT COUNT(*) FROM pc_issues WHERE project_id=p.id AND status_programmer='Done') as done_issues,
                  (SELECT COUNT(*) FROM pc_tasks WHERE project_id=p.id) as total_tasks,
                  (SELECT COUNT(*) FROM pc_tasks WHERE project_id=p.id AND status='done') as done_tasks
           FROM pc_projects p
           LEFT JOIN employees e ON e.id=p.pic_id
           WHERE p.status != 'archived' AND p.deleted_at IS NULL
           ORDER BY p.created_at DESC'''
    ).fetchall()
    counts = {
        'projects': len(projects),
        'open_issues': db.execute(
            "SELECT COUNT(*) FROM pc_issues WHERE status_programmer NOT IN ('Done','Hold')"
        ).fetchone()[0],
        'pending_tasks': db.execute(
            "SELECT COUNT(*) FROM pc_tasks WHERE status NOT IN ('done')"
        ).fetchone()[0],
        'milestones': db.execute(
            "SELECT COUNT(*) FROM pc_milestones WHERE status='in_progress'"
        ).fetchone()[0],
    }
    return render_template('pc_index.html', projects=projects, counts=counts)

@app.route('/project/projects')
@login_required
def pc_projects():
    db       = get_db()
    show_del = request.args.get('show') == 'deleted'
    _proj_cols = '''SELECT p.*, e.name as pic_name, c.name as customer_name,
                      (SELECT COUNT(*) FROM pc_issues WHERE project_id=p.id) as total_issues,
                      (SELECT COUNT(*) FROM pc_issues WHERE project_id=p.id AND status_programmer='Done') as done_issues,
                      (SELECT COUNT(*) FROM pc_issues WHERE project_id=p.id AND status_programmer NOT IN ('Done','Hold')) as open_issues,
                      (SELECT COUNT(*) FROM pc_tasks WHERE project_id=p.id) as total_tasks,
                      (SELECT COUNT(*) FROM pc_tasks WHERE project_id=p.id AND status='done') as done_tasks
               FROM pc_projects p
               LEFT JOIN employees e ON e.id=p.pic_id
               LEFT JOIN sc_customers c ON c.id=p.customer_id'''
    if show_del:
        rows = db.execute(_proj_cols + ' WHERE p.deleted_at IS NOT NULL ORDER BY p.deleted_at DESC').fetchall()
    else:
        rows = db.execute(_proj_cols + ' WHERE p.deleted_at IS NULL ORDER BY p.status, p.created_at DESC').fetchall()
    deleted_count = db.execute(
        "SELECT COUNT(*) FROM pc_projects WHERE deleted_at IS NOT NULL"
    ).fetchone()[0]
    return render_template('pc_projects.html', rows=rows, show_deleted=show_del, deleted_count=deleted_count)

def _pc_save_team_members(db, pid, programmers, prog_exts, testers, test_exts):
    """Replace programmer dan QC/tester members untuk project pid.
    programmers/testers: list of employee_id (int string) or '' if external
    prog_exts/test_exts: list of external name strings
    """
    db.execute("DELETE FROM pc_members WHERE project_id=? AND role IN ('programmer','qc_tester')", (pid,))
    for eid, ext in zip(programmers, prog_exts):
        eid = eid.strip(); ext = ext.strip()
        if eid:
            try:
                db.execute("INSERT INTO pc_members(project_id,employee_id,name_ext,role) VALUES(?,?,?,'programmer')",
                           (pid, int(eid), ''))
            except Exception:
                pass
        elif ext:
            db.execute("INSERT INTO pc_members(project_id,employee_id,name_ext,role) VALUES(?,NULL,?,'programmer')",
                       (pid, ext))
    for eid, ext in zip(testers, test_exts):
        eid = eid.strip(); ext = ext.strip()
        if eid:
            try:
                db.execute("INSERT INTO pc_members(project_id,employee_id,name_ext,role) VALUES(?,?,?,'qc_tester')",
                           (pid, int(eid), ''))
            except Exception:
                pass
        elif ext:
            db.execute("INSERT INTO pc_members(project_id,employee_id,name_ext,role) VALUES(?,NULL,?,'qc_tester')",
                       (pid, ext))

@app.route('/project/projects/add', methods=['GET','POST'])
@login_required
def pc_project_add():
    db        = get_db()
    emps      = db.execute("SELECT id, name, jabatan FROM employees WHERE is_active=1 ORDER BY name").fetchall()
    customers = db.execute("SELECT id, code, name FROM sc_customers WHERE is_active=1 ORDER BY name").fetchall()
    if request.method == 'POST':
        code = request.form.get('code','').strip().upper()
        name = request.form.get('name','').strip()
        if not code or not name:
            flash('Kode dan nama wajib diisi', 'danger')
            return render_template('pc_project_form.html', emps=emps, customers=customers, colors=PC_PROJECT_COLORS, r={}, members_prog=[], members_test=[])
        if db.execute("SELECT id FROM pc_projects WHERE code=?", (code,)).fetchone():
            flash(f'Kode proyek {code} sudah dipakai', 'danger')
            return render_template('pc_project_form.html', emps=emps, customers=customers, colors=PC_PROJECT_COLORS, r={}, members_prog=[], members_test=[])
        cur = db.execute(
            '''INSERT INTO pc_projects(code,name,customer_id,description,status,start_date,end_date,
               pic_id,pic_ext,implementor_id,implementor_ext,co_leader_id,co_leader_ext,color,created_by)
               VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)''',
            (code, name,
             request.form.get('customer_id') or None,
             request.form.get('description','').strip(),
             request.form.get('status','active'),
             request.form.get('start_date') or None,
             request.form.get('end_date') or None,
             request.form.get('pic_id') or None,
             request.form.get('pic_ext','').strip(),
             request.form.get('implementor_id') or None,
             request.form.get('implementor_ext','').strip(),
             request.form.get('co_leader_id') or None,
             request.form.get('co_leader_ext','').strip(),
             request.form.get('color','#0ea5e9'),
             session.get('user_id'))
        )
        pid = cur.lastrowid
        _pc_save_team_members(db, pid,
                              request.form.getlist('programmers'),
                              request.form.getlist('prog_exts'),
                              request.form.getlist('testers'),
                              request.form.getlist('test_exts'))
        db.commit()
        flash(f'Proyek {name} berhasil dibuat', 'success')
        return redirect(url_for('pc_projects'))
    return render_template('pc_project_form.html', emps=emps, customers=customers,
                           colors=PC_PROJECT_COLORS, r={}, members_prog=[], members_test=[])

@app.route('/project/projects/<int:pid>/edit', methods=['GET','POST'])
@login_required
def pc_project_edit(pid):
    db        = get_db()
    proj      = db.execute("SELECT * FROM pc_projects WHERE id=?", (pid,)).fetchone()
    if not proj: abort(404)
    emps      = db.execute("SELECT id, name, jabatan FROM employees WHERE is_active=1 ORDER BY name").fetchall()
    customers = db.execute("SELECT id, code, name FROM sc_customers WHERE is_active=1 ORDER BY name").fetchall()
    if request.method == 'POST':
        db.execute(
            '''UPDATE pc_projects SET name=?,customer_id=?,description=?,status=?,
               start_date=?,end_date=?,pic_id=?,pic_ext=?,
               implementor_id=?,implementor_ext=?,co_leader_id=?,co_leader_ext=?,color=? WHERE id=?''',
            (request.form.get('name','').strip(),
             request.form.get('customer_id') or None,
             request.form.get('description','').strip(),
             request.form.get('status','active'),
             request.form.get('start_date') or None,
             request.form.get('end_date') or None,
             request.form.get('pic_id') or None,
             request.form.get('pic_ext','').strip(),
             request.form.get('implementor_id') or None,
             request.form.get('implementor_ext','').strip(),
             request.form.get('co_leader_id') or None,
             request.form.get('co_leader_ext','').strip(),
             request.form.get('color','#0ea5e9'),
             pid)
        )
        _pc_save_team_members(db, pid,
                              request.form.getlist('programmers'),
                              request.form.getlist('prog_exts'),
                              request.form.getlist('testers'),
                              request.form.getlist('test_exts'))
        db.commit()
        flash('Proyek diperbarui', 'success')
        return redirect(url_for('pc_project_detail', pid=pid))
    members_prog = db.execute(
        "SELECT employee_id, name_ext FROM pc_members WHERE project_id=? AND role='programmer'", (pid,)).fetchall()
    members_test = db.execute(
        "SELECT employee_id, name_ext FROM pc_members WHERE project_id=? AND role='qc_tester'", (pid,)).fetchall()
    return render_template('pc_project_form.html', emps=emps, customers=customers,
                           colors=PC_PROJECT_COLORS, r=dict(proj),
                           members_prog=members_prog, members_test=members_test)

@app.route('/project/projects/<int:pid>')
@login_required
def pc_project_detail(pid):
    db   = get_db()
    proj_row = db.execute(
        '''SELECT p.*, c.name as customer_name,
                  COALESCE(pm.name, p.pic_ext)        as pic_name,
                  COALESCE(im.name, p.implementor_ext) as implementor_name,
                  COALESCE(cl.name, p.co_leader_ext)   as co_leader_name
           FROM pc_projects p
           LEFT JOIN sc_customers c ON c.id=p.customer_id
           LEFT JOIN employees pm ON pm.id=p.pic_id
           LEFT JOIN employees im ON im.id=p.implementor_id
           LEFT JOIN employees cl ON cl.id=p.co_leader_id
           WHERE p.id=?''', (pid,)
    ).fetchone()
    proj = dict(proj_row) if proj_row else None
    if not proj: abort(404)
    issues = db.execute(
        '''SELECT i.*, ep.name as pic_prog_name, et.name as pic_test_name
           FROM pc_issues i
           LEFT JOIN employees ep ON ep.id=i.pic_programmer_id
           LEFT JOIN employees et ON et.id=i.pic_tester_id
           WHERE i.project_id=? ORDER BY i.id DESC''', (pid,)
    ).fetchall()
    _raw_tasks = db.execute(
        '''SELECT t.*, GROUP_CONCAT(e.name, ', ') as assignees
           FROM pc_tasks t
           LEFT JOIN pc_task_assignees ta ON ta.task_id=t.id
           LEFT JOIN employees e ON e.id=ta.employee_id
           WHERE t.project_id=? GROUP BY t.id ORDER BY t.sort_order, t.id''', (pid,)
    ).fetchall()
    _ext_map = {}
    for row in db.execute(
        '''SELECT ea.task_id, GROUP_CONCAT(ea.name, ', ') as ext_names
           FROM pc_task_external_assignees ea
           JOIN pc_tasks t ON t.id=ea.task_id
           WHERE t.project_id=? GROUP BY ea.task_id''', (pid,)
    ).fetchall():
        _ext_map[row['task_id']] = row['ext_names']
    tasks = []
    for t in _raw_tasks:
        t = dict(t)
        parts = [p for p in [t.get('assignees'), _ext_map.get(t['id'])] if p]
        t['assignees'] = ', '.join(parts) if parts else ''
        tasks.append(t)
    milestones = db.execute(
        "SELECT * FROM pc_milestones WHERE project_id=? ORDER BY due_date, sort_order", (pid,)
    ).fetchall()
    proposed = db.execute(
        "SELECT * FROM pc_proposed_changes WHERE project_id=? ORDER BY id", (pid,)
    ).fetchall()
    phases = db.execute(
        '''SELECT ph.*, COALESCE(e.name, ph.pic_ext) as pic_name,
                  a.name as app_name, m.name as module_name
           FROM pc_phases ph
           LEFT JOIN employees e   ON e.id=ph.pic_id
           LEFT JOIN sc_apps a     ON a.id=ph.app_id
           LEFT JOIN sc_modules m  ON m.id=ph.module_id
           WHERE ph.project_id=? ORDER BY ph.sort_order, ph.id''', (pid,)
    ).fetchall()
    sc_apps_list = db.execute(
        "SELECT id, name FROM sc_apps WHERE is_active=1 ORDER BY name"
    ).fetchall()
    sc_modules_list = db.execute(
        "SELECT id, app_id, name FROM sc_modules WHERE is_active=1 ORDER BY app_id, name"
    ).fetchall()
    members  = _pc_members(db, pid)
    emps     = db.execute("SELECT id, name, jabatan FROM employees WHERE is_active=1 ORDER BY name").fetchall()
    stats = {
        'total_issues': len(issues),
        'open_issues':  sum(1 for i in issues if i['status_programmer'] not in ('Done','Hold')),
        'done_issues':  sum(1 for i in issues if i['status_programmer'] == 'Done'),
        'total_tasks':  len(tasks),
        'done_tasks':   sum(1 for t in tasks if t['status'] == 'done'),
        'total_phases': len(phases),
        'done_phases':  sum(1 for p in phases if p['status'] == 'done'),
    }
    kanban = {s: [t for t in tasks if t['status'] == s] for s, _, _ in PC_TASK_STATUSES}
    return render_template('pc_project_detail.html',
        proj=proj, issues=issues, tasks=tasks, milestones=milestones,
        proposed=proposed, phases=phases, members=members, emps=emps,
        stats=stats, kanban=kanban,
        task_statuses=PC_TASK_STATUSES, milestone_statuses=PC_MILESTONE_STATUSES,
        proposed_statuses=PC_PROPOSED_STATUSES,
        phase_types=PC_PHASE_TYPES, phase_statuses=PC_PHASE_STATUSES,
        sc_apps_list=sc_apps_list, sc_modules_list=sc_modules_list,
        priorities=PC_PRIORITIES, difficulties=PC_DIFFICULTIES)

@app.route('/project/projects/<int:pid>/delete', methods=['POST'])
@login_required
def pc_project_delete(pid):
    db = get_db()
    p  = db.execute("SELECT id FROM pc_projects WHERE id=? AND deleted_at IS NULL", (pid,)).fetchone()
    if not p: abort(404)
    db.execute("UPDATE pc_projects SET deleted_at=datetime('now','localtime') WHERE id=?", (pid,))
    db.commit()
    flash('Proyek dipindahkan ke tempat sampah. Bisa dipulihkan dari daftar proyek dihapus.', 'warning')
    return redirect(url_for('pc_projects'))

@app.route('/project/projects/<int:pid>/restore', methods=['POST'])
@login_required
def pc_project_restore(pid):
    db = get_db()
    p  = db.execute("SELECT id FROM pc_projects WHERE id=? AND deleted_at IS NOT NULL", (pid,)).fetchone()
    if not p: abort(404)
    db.execute("UPDATE pc_projects SET deleted_at=NULL WHERE id=?", (pid,))
    db.commit()
    flash('Proyek berhasil dipulihkan.', 'success')
    return redirect(url_for('pc_project_detail', pid=pid))

@app.route('/project/projects/<int:pid>/members', methods=['POST'])
@login_required
def pc_member_add(pid):
    db  = get_db()
    eid = request.form.get('employee_id')
    role= request.form.get('role','developer')
    if eid:
        try:
            db.execute("INSERT OR IGNORE INTO pc_members(project_id,employee_id,role) VALUES(?,?,?)", (pid, eid, role))
            db.commit()
        except Exception:
            pass
    return redirect(url_for('pc_project_detail', pid=pid) + '#members')

@app.route('/project/projects/<int:pid>/members/<int:eid>/delete', methods=['POST'])
@login_required
def pc_member_delete(pid, eid):
    db = get_db()
    db.execute("DELETE FROM pc_members WHERE project_id=? AND employee_id=?", (pid, eid))
    db.commit()
    return redirect(url_for('pc_project_detail', pid=pid) + '#members')

# ── Issues ─────────────────────────────────────────────────────────────────────

@app.route('/project/projects/<int:pid>/issues/add', methods=['GET','POST'])
@login_required
def pc_issue_add(pid):
    db   = get_db()
    proj = db.execute("SELECT * FROM pc_projects WHERE id=?", (pid,)).fetchone()
    if not proj: abort(404)
    emps = db.execute("SELECT id, name, jabatan FROM employees WHERE is_active=1 ORDER BY name").fetchall()
    if request.method == 'POST':
        issue_no = _pc_next_issue_no(db, pid)
        db.execute(
            '''INSERT INTO pc_issues
               (project_id,issue_no,title,description,role,menu,stage,solution_type,
                priority,severity,difficulty,issued_type,issued_date,issued_by,
                pic_programmer_id,pic_tester_id,md_days,plan_hours,
                status_programmer,notes,redmine)
               VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)''',
            (pid, issue_no,
             request.form.get('title','').strip(),
             request.form.get('description','').strip(),
             request.form.get('role','').strip(),
             request.form.get('menu','').strip(),
             request.form.get('stage','').strip(),
             request.form.get('solution_type',''),
             request.form.get('priority','Medium'),
             request.form.get('severity','Medium'),
             request.form.get('difficulty','Normal'),
             request.form.get('issued_type','Bugs'),
             request.form.get('issued_date') or None,
             request.form.get('issued_by','').strip(),
             request.form.get('pic_programmer_id') or None,
             request.form.get('pic_tester_id') or None,
             request.form.get('md_days') or None,
             request.form.get('plan_hours') or None,
             'New',
             request.form.get('notes','').strip(),
             request.form.get('redmine','').strip())
        )
        iid = db.execute("SELECT last_insert_rowid()").fetchone()[0]
        _pc_log_issue(db, iid, 'created', '', 'New', f'Issue {issue_no} dibuat')
        db.commit()
        flash(f'Issue {issue_no} berhasil ditambahkan', 'success')
        return redirect(url_for('pc_project_detail', pid=pid) + '#issues')
    return render_template('pc_issue_form.html', proj=proj, emps=emps,
        priorities=PC_PRIORITIES, severities=PC_SEVERITIES,
        difficulties=PC_DIFFICULTIES, issued_types=PC_ISSUED_TYPES,
        solution_types=PC_SOLUTION_TYPES, r={})

@app.route('/project/issues/<int:iid>')
@login_required
def pc_issue_detail(iid):
    db    = get_db()
    issue = db.execute(
        '''SELECT i.*, p.name as project_name, p.id as project_id, p.code as project_code,
                  ep.name as pic_prog_name, et.name as pic_test_name
           FROM pc_issues i
           JOIN pc_projects p ON p.id=i.project_id
           LEFT JOIN employees ep ON ep.id=i.pic_programmer_id
           LEFT JOIN employees et ON et.id=i.pic_tester_id
           WHERE i.id=?''', (iid,)
    ).fetchone()
    if not issue: abort(404)
    history = db.execute(
        '''SELECT h.*, u.username as changed_by_name
           FROM pc_issue_history h LEFT JOIN users u ON u.id=h.changed_by
           WHERE h.issue_id=? ORDER BY h.id DESC''', (iid,)
    ).fetchall()
    emps = db.execute("SELECT id, name, jabatan FROM employees WHERE is_active=1 ORDER BY name").fetchall()
    return render_template('pc_issue_detail.html', issue=issue, history=history, emps=emps,
        prg_statuses=PC_ISSUE_STATUSES_PRG, test_statuses=PC_ISSUE_STATUSES_TEST,
        priorities=PC_PRIORITIES, severities=PC_SEVERITIES,
        difficulties=PC_DIFFICULTIES, solution_types=PC_SOLUTION_TYPES)

@app.route('/project/issues/<int:iid>/edit', methods=['GET','POST'])
@login_required
def pc_issue_edit(iid):
    db    = get_db()
    issue = db.execute("SELECT * FROM pc_issues WHERE id=?", (iid,)).fetchone()
    if not issue: abort(404)
    proj  = db.execute("SELECT * FROM pc_projects WHERE id=?", (issue['project_id'],)).fetchone()
    emps  = db.execute("SELECT id, name, jabatan FROM employees WHERE is_active=1 ORDER BY name").fetchall()
    if request.method == 'POST':
        old_status = issue['status_programmer']
        new_status = request.form.get('status_programmer', old_status)
        db.execute(
            '''UPDATE pc_issues SET title=?,description=?,role=?,menu=?,stage=?,solution_type=?,
               priority=?,severity=?,difficulty=?,issued_type=?,issued_date=?,issued_by=?,
               pic_programmer_id=?,pic_tester_id=?,md_days=?,plan_hours=?,
               status_programmer=?,status_testing=?,testing_date=?,resolved_date=?,
               notes=?,redmine=? WHERE id=?''',
            (request.form.get('title','').strip(),
             request.form.get('description','').strip(),
             request.form.get('role','').strip(),
             request.form.get('menu','').strip(),
             request.form.get('stage','').strip(),
             request.form.get('solution_type',''),
             request.form.get('priority','Medium'),
             request.form.get('severity','Medium'),
             request.form.get('difficulty','Normal'),
             request.form.get('issued_type','Bugs'),
             request.form.get('issued_date') or None,
             request.form.get('issued_by','').strip(),
             request.form.get('pic_programmer_id') or None,
             request.form.get('pic_tester_id') or None,
             request.form.get('md_days') or None,
             request.form.get('plan_hours') or None,
             new_status,
             request.form.get('status_testing',''),
             request.form.get('testing_date') or None,
             request.form.get('resolved_date') or None,
             request.form.get('notes','').strip(),
             request.form.get('redmine','').strip(),
             iid)
        )
        if old_status != new_status:
            _pc_log_issue(db, iid, 'status_change', old_status, new_status)
        else:
            _pc_log_issue(db, iid, 'update', '', '', 'Data issue diperbarui')
        db.commit()
        flash('Issue diperbarui', 'success')
        return redirect(url_for('pc_issue_detail', iid=iid))
    return render_template('pc_issue_form.html', proj=proj, emps=emps,
        priorities=PC_PRIORITIES, severities=PC_SEVERITIES,
        difficulties=PC_DIFFICULTIES, issued_types=PC_ISSUED_TYPES,
        solution_types=PC_SOLUTION_TYPES, r=issue, is_edit=True)

@app.route('/project/issues/<int:iid>/status', methods=['POST'])
@login_required
def pc_issue_status(iid):
    db    = get_db()
    issue = db.execute("SELECT * FROM pc_issues WHERE id=?", (iid,)).fetchone()
    if not issue: abort(404)
    new_prg  = request.form.get('status_programmer', issue['status_programmer'])
    new_test = request.form.get('status_testing', issue['status_testing'] or '')
    notes    = request.form.get('notes','').strip()
    if new_prg != issue['status_programmer']:
        _pc_log_issue(db, iid, 'status_change', issue['status_programmer'], new_prg, notes)
    resolved_date = issue['resolved_date']
    if new_prg == 'Done' and not resolved_date:
        resolved_date = datetime.now().strftime('%Y-%m-%d')
    db.execute(
        "UPDATE pc_issues SET status_programmer=?,status_testing=?,resolved_date=? WHERE id=?",
        (new_prg, new_test, resolved_date, iid)
    )
    db.commit()
    flash('Status diperbarui', 'success')
    return redirect(url_for('pc_issue_detail', iid=iid))

@app.route('/project/issues/<int:iid>/delete', methods=['POST'])
@login_required
def pc_issue_delete(iid):
    db    = get_db()
    issue = db.execute("SELECT * FROM pc_issues WHERE id=?", (iid,)).fetchone()
    if not issue: abort(404)
    pid   = issue['project_id']
    db.execute("DELETE FROM pc_issues WHERE id=?", (iid,))
    db.commit()
    flash('Issue dihapus', 'warning')
    return redirect(url_for('pc_project_detail', pid=pid) + '#issues')

# ── Tasks / Kanban ─────────────────────────────────────────────────────────────

@app.route('/project/projects/<int:pid>/tasks/add', methods=['POST'])
@login_required
def pc_task_add(pid):
    db = get_db()
    title = request.form.get('title','').strip()
    if title:
        status = request.form.get('status','backlog')
        max_order = db.execute(
            "SELECT COALESCE(MAX(sort_order),0) FROM pc_tasks WHERE project_id=? AND status=?", (pid, status)
        ).fetchone()[0]
        db.execute(
            '''INSERT INTO pc_tasks(project_id,milestone_id,title,description,status,priority,due_date,sort_order)
               VALUES(?,?,?,?,?,?,?,?)''',
            (pid, request.form.get('milestone_id') or None,
             title, request.form.get('description','').strip(),
             status, request.form.get('priority','Medium'),
             request.form.get('due_date') or None, max_order + 1)
        )
        tid = db.execute("SELECT last_insert_rowid()").fetchone()[0]
        for eid in request.form.getlist('assignee_ids'):
            db.execute("INSERT OR IGNORE INTO pc_task_assignees(task_id,employee_id) VALUES(?,?)", (tid, eid))
        ext_names = [n for n in request.form.get('external_assignees','').split('\n') if n.strip()]
        _sync_external_assignees(db, 'pc_task_external_assignees', 'task_id', tid, ext_names)
        db.commit()
        flash('Task ditambahkan', 'success')
    return redirect(url_for('pc_project_detail', pid=pid) + '#kanban')

@app.route('/project/tasks/<int:tid>/move', methods=['POST'])
@login_required
def pc_task_move(tid):
    db  = get_db()
    new_status = request.json.get('status') if request.is_json else request.form.get('status')
    if new_status:
        db.execute("UPDATE pc_tasks SET status=? WHERE id=?", (new_status, tid))
        db.commit()
    return jsonify({'ok': True})

@app.route('/project/tasks/<int:tid>/delete', methods=['POST'])
@login_required
def pc_task_delete(tid):
    db   = get_db()
    task = db.execute("SELECT project_id FROM pc_tasks WHERE id=?", (tid,)).fetchone()
    if not task: abort(404)
    pid  = task['project_id']
    db.execute("DELETE FROM pc_tasks WHERE id=?", (tid,))
    db.commit()
    flash('Task dihapus', 'warning')
    return redirect(url_for('pc_project_detail', pid=pid) + '#kanban')

# ── Milestones ─────────────────────────────────────────────────────────────────

@app.route('/project/projects/<int:pid>/milestones/add', methods=['POST'])
@login_required
def pc_milestone_add(pid):
    db    = get_db()
    title = request.form.get('title','').strip()
    due   = request.form.get('due_date','').strip()
    if title and due:
        db.execute(
            "INSERT INTO pc_milestones(project_id,title,description,due_date,status) VALUES(?,?,?,?,?)",
            (pid, title, request.form.get('description','').strip(), due, 'upcoming')
        )
        db.commit()
        flash('Milestone ditambahkan', 'success')
    return redirect(url_for('pc_project_detail', pid=pid) + '#milestones')

@app.route('/project/milestones/<int:mid>/status', methods=['POST'])
@login_required
def pc_milestone_status(mid):
    db = get_db()
    ms = db.execute("SELECT project_id FROM pc_milestones WHERE id=?", (mid,)).fetchone()
    if not ms: abort(404)
    db.execute("UPDATE pc_milestones SET status=? WHERE id=?", (request.form.get('status','upcoming'), mid))
    db.commit()
    flash('Status milestone diperbarui', 'success')
    return redirect(url_for('pc_project_detail', pid=ms['project_id']) + '#milestones')

@app.route('/project/milestones/<int:mid>/delete', methods=['POST'])
@login_required
def pc_milestone_delete(mid):
    db = get_db()
    ms = db.execute("SELECT project_id FROM pc_milestones WHERE id=?", (mid,)).fetchone()
    if not ms: abort(404)
    pid = ms['project_id']
    db.execute("DELETE FROM pc_milestones WHERE id=?", (mid,))
    db.commit()
    flash('Milestone dihapus', 'warning')
    return redirect(url_for('pc_project_detail', pid=pid) + '#milestones')

# ── Phases / Timeline ──────────────────────────────────────────────────────────

@app.route('/project/projects/<int:pid>/phases/add', methods=['POST'])
@login_required
def pc_phase_add(pid):
    db   = get_db()
    name = request.form.get('name','').strip()
    ptype= request.form.get('phase_type','custom')
    if not name:
        ptype_labels = {k: v for k, v, _ in PC_PHASE_TYPES}
        name = ptype_labels.get(ptype, 'Fase Baru')
    start = request.form.get('start_date','').strip() or None
    end   = request.form.get('end_date','').strip() or None
    pic_id    = request.form.get('pic_id','').strip() or None
    pic_ext   = request.form.get('pic_ext','').strip()
    app_id    = request.form.get('app_id','').strip() or None
    module_id = request.form.get('module_id','').strip() or None
    notes = request.form.get('notes','').strip()
    max_order = db.execute(
        "SELECT COALESCE(MAX(sort_order),0) FROM pc_phases WHERE project_id=?", (pid,)
    ).fetchone()[0]
    db.execute(
        '''INSERT INTO pc_phases(project_id,phase_type,name,start_date,end_date,
           status,sort_order,pic_id,pic_ext,app_id,module_id,notes)
           VALUES(?,?,?,?,?,'planned',?,?,?,?,?,?)''',
        (pid, ptype, name, start, end, max_order + 1, pic_id, pic_ext, app_id, module_id, notes)
    )
    db.commit()
    flash('Fase ditambahkan', 'success')
    return redirect(url_for('pc_project_detail', pid=pid) + '#timeline')

@app.route('/project/phases/<int:phid>/status', methods=['POST'])
@login_required
def pc_phase_status(phid):
    db = get_db()
    ph = db.execute("SELECT project_id FROM pc_phases WHERE id=?", (phid,)).fetchone()
    if not ph: abort(404)
    new_status = request.form.get('status', 'planned')
    sign_off   = request.form.get('sign_off_date','').strip() or None
    db.execute(
        "UPDATE pc_phases SET status=?, sign_off_date=COALESCE(?,sign_off_date) WHERE id=?",
        (new_status, sign_off, phid)
    )
    db.commit()
    flash('Status fase diperbarui', 'success')
    return redirect(url_for('pc_project_detail', pid=ph['project_id']) + '#timeline')

@app.route('/project/phases/<int:phid>/edit', methods=['POST'])
@login_required
def pc_phase_edit(phid):
    db = get_db()
    ph = db.execute("SELECT project_id FROM pc_phases WHERE id=?", (phid,)).fetchone()
    if not ph: abort(404)
    name      = request.form.get('name','').strip()
    start     = request.form.get('start_date','').strip() or None
    end       = request.form.get('end_date','').strip() or None
    pic_id    = request.form.get('pic_id','').strip() or None
    pic_ext   = request.form.get('pic_ext','').strip()
    app_id    = request.form.get('app_id','').strip() or None
    module_id = request.form.get('module_id','').strip() or None
    sign_off  = request.form.get('sign_off_date','').strip() or None
    notes     = request.form.get('notes','').strip()
    if name:
        db.execute(
            '''UPDATE pc_phases SET name=?,start_date=?,end_date=?,
               pic_id=?,pic_ext=?,app_id=?,module_id=?,sign_off_date=?,notes=? WHERE id=?''',
            (name, start, end, pic_id, pic_ext, app_id, module_id, sign_off, notes, phid)
        )
        db.commit()
        flash('Fase diperbarui', 'success')
    return redirect(url_for('pc_project_detail', pid=ph['project_id']) + '#timeline')

@app.route('/project/phases/<int:phid>/delete', methods=['POST'])
@login_required
def pc_phase_delete(phid):
    db = get_db()
    ph = db.execute("SELECT project_id FROM pc_phases WHERE id=?", (phid,)).fetchone()
    if not ph: abort(404)
    pid = ph['project_id']
    db.execute("DELETE FROM pc_phases WHERE id=?", (phid,))
    db.commit()
    flash('Fase dihapus', 'warning')
    return redirect(url_for('pc_project_detail', pid=pid) + '#timeline')

# ── Proposed Changes ───────────────────────────────────────────────────────────

@app.route('/project/projects/<int:pid>/proposed/add', methods=['POST'])
@login_required
def pc_proposed_add(pid):
    db = get_db()
    title = request.form.get('title','').strip()
    if title:
        db.execute(
            '''INSERT INTO pc_proposed_changes
               (project_id,title,description,module,pekerjaan,impact,difficulty,status,tester,notes)
               VALUES(?,?,?,?,?,?,?,?,?,?)''',
            (pid, title,
             request.form.get('description','').strip(),
             request.form.get('module','').strip(),
             request.form.get('pekerjaan','').strip(),
             request.form.get('impact','Medium'),
             request.form.get('difficulty','Normal'),
             'proposed',
             request.form.get('tester','').strip(),
             request.form.get('notes','').strip())
        )
        db.commit()
        flash('Proposed change ditambahkan', 'success')
    return redirect(url_for('pc_project_detail', pid=pid) + '#proposed')

@app.route('/project/proposed/<int:cid>/status', methods=['POST'])
@login_required
def pc_proposed_status(cid):
    db = get_db()
    pc = db.execute("SELECT project_id FROM pc_proposed_changes WHERE id=?", (cid,)).fetchone()
    if not pc: abort(404)
    db.execute("UPDATE pc_proposed_changes SET status=? WHERE id=?",
               (request.form.get('status','proposed'), cid))
    db.commit()
    return redirect(url_for('pc_project_detail', pid=pc['project_id']) + '#proposed')

@app.route('/project/proposed/<int:cid>/delete', methods=['POST'])
@login_required
def pc_proposed_delete(cid):
    db = get_db()
    pc = db.execute("SELECT project_id FROM pc_proposed_changes WHERE id=?", (cid,)).fetchone()
    if not pc: abort(404)
    pid = pc['project_id']
    db.execute("DELETE FROM pc_proposed_changes WHERE id=?", (cid,))
    db.commit()
    flash('Proposed change dihapus', 'warning')
    return redirect(url_for('pc_project_detail', pid=pid) + '#proposed')

# ─── Global Error Handler ──────────────────────────────────────────────────────

@app.errorhandler(500)
def handle_500(e):
    import traceback as _tb
    tb_str = _tb.format_exc()
    try:
        db  = get_db()
        uid = session.get('user_id')
        uname = session.get('username', '')
        db.execute('''INSERT INTO audit_errors
                      (app_slug, user_id, username, url, method, error_code,
                       error_type, error_msg, traceback, ip)
                      VALUES(?,?,?,?,?,500,?,?,?,?)''',
                   ('portal', uid, uname,
                    request.url, request.method,
                    type(e).__name__, str(e), tb_str,
                    request.remote_addr or ''))
        db.commit()
    except Exception:
        pass  # jangan sampai error handler ikut error

    # Tampilkan halaman error yang informatif (traceback visible untuk superadmin)
    show_tb = session.get('user_role') == 'superadmin'
    return render_template('error_500.html',
                           error_type=type(e).__name__,
                           error_msg=str(e),
                           traceback=tb_str if show_tb else '',
                           url=request.url), 500


@app.errorhandler(404)
def handle_404(e):
    return render_template('error_404.html' if False else 'base.html'), 404


# ─── AttendanceCore: Helper ───────────────────────────────────────────────────
def at_require(perm):
    """Pastikan user memiliki permission atau superadmin."""
    if session.get('user_role') == 'superadmin':
        return True
    return has_permission(session.get('user_role', ''), perm, get_db())

# ─── AttendanceCore: Views ────────────────────────────────────────────────────
@app.route('/attendance/')
@login_required
def at_index():
    if not at_require('at_view'):
        flash('Anda tidak memiliki akses ke AttendanceCore.', 'danger')
        return redirect(url_for('portal'))
    
    session['active_app'] = 'attendance'
    db = get_db()
    uid = session['user_id']
    today = datetime.now().strftime('%Y-%m-%d')
    
    today_att = db.execute(
        'SELECT * FROM attendance WHERE user_id=? AND date=?',
        (uid, today)
    ).fetchone()
    
    history = db.execute(
        'SELECT * FROM attendance WHERE user_id=? ORDER BY date DESC LIMIT 30',
        (uid,)
    ).fetchall()
    
    # Kueri semua kehadiran user hari ini
    rows = db.execute('''
        SELECT u.id as user_id, u.full_name as employee_name, 
               a.clock_in, a.clock_out, a.location_in, a.location_out, a.status, a.plan, a.progress,
               a.notes_in, a.notes_out
        FROM users u
        LEFT JOIN attendance a ON a.user_id = u.id AND a.date = ?
        WHERE u.is_active = 1
        ORDER BY CASE WHEN a.clock_in IS NULL THEN 1 ELSE 0 END, a.clock_in ASC, u.full_name ASC
    ''', (today,)).fetchall()
    
    all_today_att = []
    for row in rows:
        r = dict(row)
        # Hitung Durasi Kerja (Hours)
        if r['clock_in'] and r['clock_out']:
            try:
                t1_str = r['clock_in']
                t2_str = r['clock_out']
                t1 = datetime.strptime(t1_str, '%H:%M') if len(t1_str.split(':')) == 2 else datetime.strptime(t1_str, '%H:%M:%S')
                t2 = datetime.strptime(t2_str, '%H:%M') if len(t2_str.split(':')) == 2 else datetime.strptime(t2_str, '%H:%M:%S')
                diff = t2 - t1
                hours = diff.total_seconds() / 3600.0
                r['work_hours'] = f"{hours:.2f} jam"
            except Exception:
                r['work_hours'] = '-'
        else:
            r['work_hours'] = '-'
            
        # Group Name / Sumber Absensi
        source_in = r['notes_in'] or ''
        group_name = "-"
        if "Telegram Group:" in source_in:
            group_name = source_in.replace("Telegram Group:", "").strip()
        elif "Telegram Japri" in source_in:
            group_name = "Japri"
        elif "Web" in source_in or "Web Attendance" in source_in:
            group_name = "Web Portal"
            
        r['group_name'] = group_name
        all_today_att.append(r)
        
    unmapped_employees = []
    mappable_employees = []
    if at_require('at_manage'):
        unmapped_employees = db.execute(
            "SELECT id, name, telegram_id FROM employees WHERE divisi = 'Telegram Core' AND telegram_id IS NOT NULL AND telegram_id != '' AND (name LIKE 'Telegram User%' OR name LIKE 'tg_%') AND is_active = 1 ORDER BY name"
        ).fetchall()
        mappable_employees = db.execute(
            "SELECT id, name FROM employees WHERE (telegram_id IS NULL OR telegram_id = '') AND is_active = 1 ORDER BY name"
        ).fetchall()

    return render_template('at_index.html', today_att=today_att, history=history, today=today, 
                           all_today_att=all_today_att, unmapped_employees=unmapped_employees, 
                           mappable_employees=mappable_employees)

@app.route('/attendance/map-telegram', methods=['POST'])
@login_required
def at_map_telegram():
    if not at_require('at_manage'):
        flash('Anda tidak memiliki izin untuk memetakan Telegram ID.', 'danger')
        return redirect(url_for('at_index'))
    
    temp_emp_id = request.form.get('temp_employee_id')
    target_emp_id = request.form.get('target_employee_id')
    
    if not temp_emp_id or not target_emp_id:
        flash('Data pemetaan tidak lengkap.', 'danger')
        return redirect(url_for('at_index'))
    
    db = get_db()
    temp_emp = db.execute('SELECT * FROM employees WHERE id=?', (temp_emp_id,)).fetchone()
    target_emp = db.execute('SELECT * FROM employees WHERE id=?', (target_emp_id,)).fetchone()
    
    if not temp_emp or not target_emp:
        flash('Data karyawan tidak ditemukan.', 'danger')
        return redirect(url_for('at_index'))
        
    try:
        telegram_id = temp_emp['telegram_id']
        
        # 1. Update target employee's telegram_id
        db.execute('UPDATE employees SET telegram_id=? WHERE id=?', (telegram_id, target_emp_id))
        
        # 2. Update user_id for past attendance records
        if temp_emp['user_id'] and target_emp['user_id']:
            db.execute('UPDATE attendance SET user_id=? WHERE user_id=?', (target_emp['user_id'], temp_emp['user_id']))
        
        # 3. Delete temporary employee and user
        db.execute('DELETE FROM employees WHERE id=?', (temp_emp_id,))
        if temp_emp['user_id']:
            db.execute('DELETE FROM users WHERE id=?', (temp_emp['user_id'],))
            
        db.commit()
        flash(f"Sukses! Telegram ID '{telegram_id}' berhasil dipetakan ke Karyawan '{target_emp['name']}'.", 'success')
    except Exception as ex:
        db.rollback()
        flash(f"Gagal memetakan Telegram ID: {str(ex)}", 'danger')
        
    return redirect(url_for('at_index'))

@app.route('/attendance/clock', methods=['POST'])
@login_required
def at_clock():
    if not at_require('at_view'):
        return jsonify({'ok': False, 'msg': 'Akses ditolak'})
    
    db = get_db()
    uid = session['user_id']
    action = request.form.get('action')
    notes = request.form.get('notes', '').strip()
    lat = request.form.get('lat', '').strip()
    lng = request.form.get('lng', '').strip()
    loc = f"{lat},{lng}" if (lat and lng) else "Web Browser"
    
    now_time = datetime.now().strftime('%H:%M:%S')
    today = datetime.now().strftime('%Y-%m-%d')
    
    today_att = db.execute(
        'SELECT * FROM attendance WHERE user_id=? AND date=?',
        (uid, today)
    ).fetchone()
    
    if action == 'in':
        if today_att:
            return jsonify({'ok': False, 'msg': 'Anda sudah clock-in hari ini.'})
        status = 'present'
        if now_time > '09:15:00':
            status = 'late'
        db.execute(
            'INSERT INTO attendance (user_id, date, clock_in, location_in, notes_in, status) VALUES (?, ?, ?, ?, ?, ?)',
            (uid, today, now_time, loc, notes, status)
        )
    elif action == 'out':
        if not today_att:
            return jsonify({'ok': False, 'msg': 'Anda belum clock-in hari ini.'})
        if today_att['clock_out']:
            return jsonify({'ok': False, 'msg': 'Anda sudah clock-out hari ini.'})
        
        db.execute(
            'UPDATE attendance SET clock_out=?, location_out=?, notes_out=? WHERE id=?',
            (now_time, loc, notes, today_att['id'])
        )
    db.commit()
    return jsonify({'ok': True})

@app.route('/attendance/leave', methods=['GET', 'POST'])
@login_required
def at_leave():
    if not at_require('at_view'):
        flash('Akses ditolak', 'danger')
        return redirect(url_for('portal'))
    
    session['active_app'] = 'attendance'
    db = get_db()
    uid = session['user_id']
    
    if request.method == 'POST':
        leave_type = request.form.get('leave_type')
        start_date = request.form.get('start_date')
        end_date = request.form.get('end_date')
        reason = request.form.get('reason', '').strip()
        
        if not leave_type or not start_date or not end_date:
            flash('Harap isi semua kolom.', 'danger')
        else:
            db.execute(
                'INSERT INTO attendance_leaves (user_id, leave_type, start_date, end_date, reason) VALUES (?, ?, ?, ?, ?)',
                (uid, leave_type, start_date, end_date, reason)
            )
            db.commit()
            flash('Pengajuan cuti berhasil dikirim.', 'success')
            return redirect(url_for('at_leave'))
            
    history = db.execute(
        'SELECT * FROM attendance_leaves WHERE user_id=? ORDER BY created_at DESC',
        (uid,)
    ).fetchall()
    return render_template('at_leave.html', history=history)

@app.route('/attendance/overtime', methods=['GET', 'POST'])
@login_required
def at_overtime():
    if not at_require('at_view'):
        flash('Akses ditolak', 'danger')
        return redirect(url_for('portal'))
    
    session['active_app'] = 'attendance'
    db = get_db()
    uid = session['user_id']
    
    if request.method == 'POST':
        date = request.form.get('date')
        start_time = request.form.get('start_time')
        end_time = request.form.get('end_time')
        hours = request.form.get('hours', 0, type=float)
        reason = request.form.get('reason', '').strip()
        
        if not date or not start_time or not end_time or hours <= 0:
            flash('Harap isi semua kolom dengan benar.', 'danger')
        else:
            db.execute(
                'INSERT INTO attendance_overtime (user_id, date, start_time, end_time, hours, reason) VALUES (?, ?, ?, ?, ?, ?)',
                (uid, date, start_time, end_time, hours, reason)
            )
            db.commit()
            flash('Pengajuan lembur berhasil dikirim.', 'success')
            return redirect(url_for('at_overtime'))
            
    history = db.execute(
        'SELECT * FROM attendance_overtime WHERE user_id=? ORDER BY created_at DESC',
        (uid,)
    ).fetchall()
    return render_template('at_overtime.html', history=history)

@app.route('/attendance/correction', methods=['GET', 'POST'])
@login_required
def at_correction():
    if not at_require('at_view'):
        flash('Akses ditolak', 'danger')
        return redirect(url_for('portal'))
    
    session['active_app'] = 'attendance'
    db = get_db()
    uid = session['user_id']
    
    if request.method == 'POST':
        date = request.form.get('date')
        requested_clock_in = request.form.get('requested_clock_in') or None
        requested_clock_out = request.form.get('requested_clock_out') or None
        reason = request.form.get('reason', '').strip()
        
        if not date or (not requested_clock_in and not requested_clock_out):
            flash('Harap isi tanggal dan minimal salah satu jam koreksi.', 'danger')
        else:
            db.execute(
                'INSERT INTO attendance_corrections (user_id, date, requested_clock_in, requested_clock_out, reason) VALUES (?, ?, ?, ?, ?)',
                (uid, date, requested_clock_in, requested_clock_out, reason)
            )
            db.commit()
            flash('Pengajuan koreksi kehadiran berhasil dikirim.', 'success')
            return redirect(url_for('at_correction'))
            
    history = db.execute(
        'SELECT * FROM attendance_corrections WHERE user_id=? ORDER BY created_at DESC',
        (uid,)
    ).fetchall()
    return render_template('at_correction.html', history=history)

@app.route('/attendance/admin/approvals')
@login_required
def at_approvals():
    if not at_require('at_manage'):
        flash('Akses ditolak — Menu khusus administrator.', 'danger')
        return redirect(url_for('at_index'))
    
    session['active_app'] = 'attendance'
    db = get_db()
    
    leaves = db.execute('''
        SELECT l.*, u.full_name as employee_name 
        FROM attendance_leaves l 
        JOIN users u ON u.id = l.user_id 
        WHERE l.status=\'pending\' 
        ORDER BY l.created_at ASC
    ''').fetchall()
    
    overtimes = db.execute('''
        SELECT o.*, u.full_name as employee_name 
        FROM attendance_overtime o 
        JOIN users u ON u.id = o.user_id 
        WHERE o.status=\'pending\' 
        ORDER BY o.created_at ASC
    ''').fetchall()
    
    corrections = db.execute('''
        SELECT c.*, u.full_name as employee_name 
        FROM attendance_corrections c 
        JOIN users u ON u.id = c.user_id 
        WHERE c.status=\'pending\' 
        ORDER BY c.created_at ASC
    ''').fetchall()
    
    return render_template('at_approvals.html', leaves=leaves, overtimes=overtimes, corrections=corrections)

@app.route('/attendance/admin/approve/<string:type>/<int:id>', methods=['POST'])
@login_required
def at_approve_action(type, id):
    if not at_require('at_manage'):
        return jsonify({'ok': False, 'msg': 'Akses ditolak'})
        
    db = get_db()
    action = request.form.get('action')
    status = 'approved' if action == 'approve' else 'rejected'
    approver = session['user_id']
    
    if type == 'leave':
        db.execute('UPDATE attendance_leaves SET status=?, approved_by=? WHERE id=?', (status, approver, id))
    elif type == 'overtime':
        db.execute('UPDATE attendance_overtime SET status=?, approved_by=? WHERE id=?', (status, approver, id))
    elif type == 'correction':
        if status == 'approved':
            corr = db.execute('SELECT * FROM attendance_corrections WHERE id=?', (id,)).fetchone()
            if corr:
                att = db.execute('SELECT id FROM attendance WHERE user_id=? AND date=?', (corr['user_id'], corr['date'])).fetchone()
                if att:
                    db.execute(
                        'UPDATE attendance SET clock_in=COALESCE(?, clock_in), clock_out=COALESCE(?, clock_out), status=? WHERE id=?',
                        (corr['requested_clock_in'], corr['requested_clock_out'], 'present', att['id'])
                    )
                else:
                    db.execute(
                        'INSERT INTO attendance (user_id, date, clock_in, clock_out, status) VALUES (?, ?, ?, ?, ?)',
                        (corr['user_id'], corr['date'], corr['requested_clock_in'], corr['requested_clock_out'], 'present')
                    )
        db.execute('UPDATE attendance_corrections SET status=?, approved_by=? WHERE id=?', (status, approver, id))
    
    db.commit()
    return jsonify({'ok': True, 'msg': 'Persetujuan berhasil diproses'})
    
@app.route('/attendance/report')
@login_required
def at_report():
    if not at_require('at_view'):
        flash('Anda tidak memiliki akses ke AttendanceCore.', 'danger')
        return redirect(url_for('portal'))
    
    session['active_app'] = 'attendance'
    db = get_db()
    
    # Range parameters
    from datetime import datetime, timedelta
    today_str = datetime.now().strftime('%Y-%m-%d')
    default_start = (datetime.now() - timedelta(days=30)).strftime('%Y-%m-%d')
    
    start_date = request.args.get('start_date', default_start)
    end_date = request.args.get('end_date', today_str)
    
    # Permission checks
    is_mgr = at_require('at_manage')
    
    # Filter employee
    selected_user_id = request.args.get('user_id', '')
    if not is_mgr:
        # Normal user can only see their own logs
        selected_user_id = str(session['user_id'])
    
    target_emp_name = "Semua Karyawan"
    if selected_user_id:
        user_row = db.execute('SELECT full_name FROM users WHERE id=?', (int(selected_user_id),)).fetchone()
        if user_row:
            target_emp_name = user_row['full_name']

    # Fetch SupportCore tickets assigned and worked on by this employee or all employees in date range
    assigned_tickets = []
    if selected_user_id:
        emp_row = db.execute('SELECT id FROM employees WHERE user_id = ?', (int(selected_user_id),)).fetchone()
        if emp_row:
            emp_id = emp_row['id']
            assigned_tickets = db.execute('''
                SELECT t.id, t.ticket_no, t.subject, t.status, t.priority, t.reported_at, t.resolved_at, t.due_date,
                       c.name as customer_name, u.full_name as assignee_name
                FROM sc_ticket_assignees ta
                JOIN sc_tickets t ON t.id = ta.ticket_id
                JOIN sc_customers c ON c.id = t.customer_id
                JOIN employees e ON e.id = ta.employee_id
                JOIN users u ON u.id = e.user_id
                WHERE ta.employee_id = ? 
                  AND (
                      (date(t.reported_at) >= ? AND date(t.reported_at) <= ?)
                      OR (t.resolved_at IS NOT NULL AND date(t.resolved_at) >= ? AND date(t.resolved_at) <= ?)
                      OR (t.status IN ('open', 'progress'))
                  )
                  AND t.status != 'cancelled'
                ORDER BY t.reported_at DESC
            ''', (emp_id, start_date, end_date, start_date, end_date)).fetchall()
    else:
        assigned_tickets = db.execute('''
            SELECT t.id, t.ticket_no, t.subject, t.status, t.priority, t.reported_at, t.resolved_at, t.due_date,
                   c.name as customer_name, u.full_name as assignee_name
            FROM sc_ticket_assignees ta
            JOIN sc_tickets t ON t.id = ta.ticket_id
            JOIN sc_customers c ON c.id = t.customer_id
            JOIN employees e ON e.id = ta.employee_id
            JOIN users u ON u.id = e.user_id
            WHERE (
                  (date(t.reported_at) >= ? AND date(t.reported_at) <= ?)
                  OR (t.resolved_at IS NOT NULL AND date(t.resolved_at) >= ? AND date(t.resolved_at) <= ?)
                  OR (t.status IN ('open', 'progress'))
              )
              AND t.status != 'cancelled'
            ORDER BY t.reported_at DESC
        ''', (start_date, end_date, start_date, end_date)).fetchall()

    # Fetch active employees for dropdown filter if admin/manager
    employees = []
    if is_mgr:
        employees = db.execute(
            'SELECT id, full_name, username FROM users WHERE is_active = 1 ORDER BY full_name'
        ).fetchall()
        
    # Fetch attendance logs
    query = '''
        SELECT a.*, u.full_name as employee_name, u.username as telegram_username
        FROM attendance a
        JOIN users u ON a.user_id = u.id
        WHERE a.date >= ? AND a.date <= ?
    '''
    params = [start_date, end_date]
    if selected_user_id:
        query += ' AND a.user_id = ?'
        params.append(int(selected_user_id))
        
    query += ' ORDER BY a.date DESC, a.clock_in DESC'
    rows = db.execute(query, params).fetchall()
    
    # Format rows
    attendance_list = []
    for row in rows:
        r = dict(row)
        # Work hours duration
        if r['clock_in'] and r['clock_out']:
            try:
                t1_str = r['clock_in']
                t2_str = r['clock_out']
                t1 = datetime.strptime(t1_str, '%H:%M') if len(t1_str.split(':')) == 2 else datetime.strptime(t1_str, '%H:%M:%S')
                t2 = datetime.strptime(t2_str, '%H:%M') if len(t2_str.split(':')) == 2 else datetime.strptime(t2_str, '%H:%M:%S')
                diff = t2 - t1
                hours = diff.total_seconds() / 3600.0
                r['work_hours'] = f"{hours:.2f} jam"
            except Exception:
                r['work_hours'] = '-'
        else:
            r['work_hours'] = '-'
            
        # Group Name / Sumber Absensi
        source_in = r['notes_in'] or ''
        group_name = "-"
        if "Telegram Group:" in source_in:
            parts = source_in.split("Telegram Group:")
            if len(parts) > 1:
                group_name = parts[1].split("(ID:")[0].strip()
        elif "Telegram Japri" in source_in:
            group_name = "Telegram Japri"
        elif "WhatsApp Group:" in source_in:
            parts = source_in.split("WhatsApp Group:")
            if len(parts) > 1:
                group_name = "WA Group: " + parts[1].split("(ID:")[0].strip()
        elif "WhatsApp Japri" in source_in:
            group_name = "WA Japri"
        elif "Web" in source_in or "Web Attendance" in source_in:
            group_name = "Web Portal"
            
        r['group_name'] = group_name

        # Fetch ticket progress updates for this user on this date
        progress_updates = db.execute('''
            SELECT h.old_value, h.new_value, h.created_at, t.ticket_no, t.subject
            FROM sc_ticket_history h
            JOIN sc_tickets t ON t.id = h.ticket_id
            WHERE h.changed_by = ? 
              AND h.field_name = 'pct_done' 
              AND date(h.created_at) = ?
            ORDER BY h.created_at ASC
        ''', (r['user_id'], r['date'])).fetchall()
        
        r['progress_updates'] = [dict(pu) for pu in progress_updates]
        r['progress_count'] = len(progress_updates)
        attendance_list.append(r)
        
    return render_template(
        'at_report.html',
        attendance=attendance_list,
        start_date=start_date,
        end_date=end_date,
        employees=employees,
        selected_user_id=selected_user_id,
        is_mgr=is_mgr,
        assigned_tickets=assigned_tickets,
        target_emp_name=target_emp_name
    )

@app.route('/attendance/report/export/excel')
@login_required
def at_report_export_excel():
    if not at_require('at_view'):
        flash('Anda tidak memiliki akses ke AttendanceCore.', 'danger')
        return redirect(url_for('portal'))
        
    db = get_db()
    from datetime import datetime, timedelta
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    import io
    
    today_str = datetime.now().strftime('%Y-%m-%d')
    default_start = (datetime.now() - timedelta(days=30)).strftime('%Y-%m-%d')
    
    start_date = request.args.get('start_date', default_start)
    end_date = request.args.get('end_date', today_str)
    
    is_mgr = at_require('at_manage')
    selected_user_id = request.args.get('user_id', '')
    if not is_mgr:
        selected_user_id = str(session['user_id'])
        
    # Fetch user details if specific filter applied
    target_emp_name = "Semua Karyawan"
    if selected_user_id:
        user_row = db.execute('SELECT full_name FROM users WHERE id=?', (int(selected_user_id),)).fetchone()
        if user_row:
            target_emp_name = user_row['full_name']

    # Fetch SupportCore tickets assigned and worked on by this employee or all employees in date range
    assigned_tickets = []
    if selected_user_id:
        emp_row = db.execute('SELECT id FROM employees WHERE user_id = ?', (int(selected_user_id),)).fetchone()
        if emp_row:
            emp_id = emp_row['id']
            assigned_tickets = db.execute('''
                SELECT t.id, t.ticket_no, t.subject, t.status, t.priority, t.reported_at, t.resolved_at, t.due_date,
                       c.name as customer_name, u.full_name as assignee_name
                FROM sc_ticket_assignees ta
                JOIN sc_tickets t ON t.id = ta.ticket_id
                JOIN sc_customers c ON c.id = t.customer_id
                JOIN employees e ON e.id = ta.employee_id
                JOIN users u ON u.id = e.user_id
                WHERE ta.employee_id = ? 
                  AND (
                      (date(t.reported_at) >= ? AND date(t.reported_at) <= ?)
                      OR (t.resolved_at IS NOT NULL AND date(t.resolved_at) >= ? AND date(t.resolved_at) <= ?)
                      OR (t.status IN ('open', 'progress'))
                  )
                  AND t.status != 'cancelled'
                ORDER BY t.reported_at DESC
            ''', (emp_id, start_date, end_date, start_date, end_date)).fetchall()
    else:
        assigned_tickets = db.execute('''
            SELECT t.id, t.ticket_no, t.subject, t.status, t.priority, t.reported_at, t.resolved_at, t.due_date,
                   c.name as customer_name, u.full_name as assignee_name
            FROM sc_ticket_assignees ta
            JOIN sc_tickets t ON t.id = ta.ticket_id
            JOIN sc_customers c ON c.id = t.customer_id
            JOIN employees e ON e.id = ta.employee_id
            JOIN users u ON u.id = e.user_id
            WHERE (
                  (date(t.reported_at) >= ? AND date(t.reported_at) <= ?)
                  OR (t.resolved_at IS NOT NULL AND date(t.resolved_at) >= ? AND date(t.resolved_at) <= ?)
                  OR (t.status IN ('open', 'progress'))
              )
              AND t.status != 'cancelled'
            ORDER BY t.reported_at DESC
        ''', (start_date, end_date, start_date, end_date)).fetchall()
            
    # Fetch logs
    query = '''
        SELECT a.*, u.full_name as employee_name, u.username as telegram_username
        FROM attendance a
        JOIN users u ON a.user_id = u.id
        WHERE a.date >= ? AND a.date <= ?
    '''
    params = [start_date, end_date]
    if selected_user_id:
        query += ' AND a.user_id = ?'
        params.append(int(selected_user_id))
        
    query += ' ORDER BY a.date DESC, a.clock_in DESC'
    rows = db.execute(query, params).fetchall()
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Log Kehadiran"
    
    # Styling
    BLUE_DARK = 'FF1E293B' # Slate 800
    BLUE_LIGHT = 'FFF1F5F9' # Slate 100
    WHITE = 'FFFFFFFF'
    BORDER_COLOR = 'FFCBD5E1' # Slate 300
    
    thin_border = Border(
        left=Side(style='thin', color=BORDER_COLOR),
        right=Side(style='thin', color=BORDER_COLOR),
        top=Side(style='thin', color=BORDER_COLOR),
        bottom=Side(style='thin', color=BORDER_COLOR)
    )
    
    # Title Block
    ws.merge_cells('A1:K1')
    ws['A1'] = "LAPORAN KEHADIRAN KARYAWAN"
    ws['A1'].font = Font(size=14, bold=True, color='FF0F172A')
    ws['A1'].alignment = Alignment(horizontal='center')
    
    ws['A2'] = "Karyawan:"
    ws['B2'] = target_emp_name
    ws['A2'].font = Font(bold=True)
    
    ws['A3'] = "Periode:"
    ws['B3'] = f"{start_date} s/d {end_date}"
    ws['A3'].font = Font(bold=True)
    
    # Headers
    headers = [
        "No", "Tanggal", "Nama Karyawan", "Clock In", "Clock Out", 
        "Durasi Kerja", "Status", "Lokasi Clock In", "Lokasi Clock Out", "Sumber", "Progress Tiket"
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col_num)
        cell.value = header
        cell.font = Font(bold=True, color=WHITE)
        cell.fill = PatternFill('solid', fgColor=BLUE_DARK)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
        
    row_num = 6
    for idx, row in enumerate(rows, 1):
        r = dict(row)
        # Format Work hours
        work_hours = '-'
        if r['clock_in'] and r['clock_out']:
            try:
                t1_str = r['clock_in']
                t2_str = r['clock_out']
                t1 = datetime.strptime(t1_str, '%H:%M') if len(t1_str.split(':')) == 2 else datetime.strptime(t1_str, '%H:%M:%S')
                t2 = datetime.strptime(t2_str, '%H:%M') if len(t2_str.split(':')) == 2 else datetime.strptime(t2_str, '%H:%M:%S')
                diff = t2 - t1
                hours = diff.total_seconds() / 3600.0
                work_hours = f"{hours:.2f} jam"
            except Exception:
                pass
                
        # Format Source
        source_in = r['notes_in'] or ''
        group_name = "-"
        if "Telegram Group:" in source_in:
            parts = source_in.split("Telegram Group:")
            if len(parts) > 1:
                group_name = parts[1].split("(ID:")[0].strip()
        elif "Telegram Japri" in source_in:
            group_name = "Telegram Japri"
        elif "WhatsApp Group:" in source_in:
            parts = source_in.split("WhatsApp Group:")
            if len(parts) > 1:
                group_name = "WA Group: " + parts[1].split("(ID:")[0].strip()
        elif "WhatsApp Japri" in source_in:
            group_name = "WA Japri"
        elif "Web" in source_in or "Web Attendance" in source_in:
            group_name = "Web Portal"
            
        # Fetch ticket progress updates for this user on this date
        progress_updates = db.execute('''
            SELECT h.old_value, h.new_value, t.ticket_no
            FROM sc_ticket_history h
            JOIN sc_tickets t ON t.id = h.ticket_id
            WHERE h.changed_by = ? 
              AND h.field_name = 'pct_done' 
              AND date(h.created_at) = ?
            ORDER BY h.created_at ASC
        ''', (r['user_id'], r['date'])).fetchall()
        
        progress_str = "-"
        if progress_updates:
            progress_parts = []
            for pu in progress_updates:
                progress_parts.append(f"{pu['ticket_no']} ({pu['old_value']}%->{pu['new_value']}%)")
            progress_str = f"{len(progress_updates)} Update: " + ", ".join(progress_parts)

        vals = [
            idx,
            r['date'],
            r['employee_name'],
            r['clock_in'] or '-',
            r['clock_out'] or '-',
            work_hours,
            (r['status'] or '').upper(),
            r['location_in'] or '-',
            r['location_out'] or '-',
            group_name,
            progress_str
        ]
        
        for col_num, val in enumerate(vals, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = val
            cell.border = thin_border
            if col_num in (1, 2, 4, 5, 6, 7):
                cell.alignment = Alignment(horizontal='center')
            else:
                cell.alignment = Alignment(horizontal='left')
                
            # Alternate row background
            if row_num % 2 == 1:
                cell.fill = PatternFill('solid', fgColor=BLUE_LIGHT)
        row_num += 1
        
    # ── Section 2: Support Tickets ──
    if assigned_tickets:
        row_num += 2
        ws.cell(row=row_num, column=1, value="DAFTAR TIKET & TASK SUPPORTCORE (PERIODE INI)").font = Font(size=11, bold=True, color='FF0F172A')
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=8)
        
        row_num += 1
        t_headers = ["No", "No Tiket", "Waktu Lapor", "Pelanggan", "Subjek", "Status", "Prioritas", "Penerima Tugas"]
        for col_num, h in enumerate(t_headers, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = h
            cell.font = Font(bold=True, color=WHITE)
            cell.fill = PatternFill('solid', fgColor='FF475569') # Slate 600
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
            
        row_num += 1
        for idx, t in enumerate(assigned_tickets, 1):
            assignee_display = t.get('assignee_name') or target_emp_name
            vals = [
                idx,
                t['ticket_no'],
                t['reported_at'],
                t['customer_name'],
                t['subject'],
                t['status'].upper(),
                t['priority'] or 'Medium',
                assignee_display
            ]
            for col_num, val in enumerate(vals, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = val
                cell.border = thin_border
                if col_num in (1, 2, 3, 6, 7):
                    cell.alignment = Alignment(horizontal='center')
                else:
                    cell.alignment = Alignment(horizontal='left')
                if row_num % 2 == 1:
                    cell.fill = PatternFill('solid', fgColor=BLUE_LIGHT)
            row_num += 1
        
    # Auto-adjust column widths
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.row >= 5 and cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(max_len + 3, 10)
        
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    
    target_name_slug = target_emp_name.replace(" ", "_")
    fname = f"Log_Kehadiran_{target_name_slug}_{start_date}_to_{end_date}.xlsx"
    return send_file(
        out,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=fname
    )

@app.route('/attendance/report/export/pdf')
@login_required
def at_report_export_pdf():
    if not at_require('at_view'):
        flash('Anda tidak memiliki akses ke AttendanceCore.', 'danger')
        return redirect(url_for('portal'))
        
    db = get_db()
    from datetime import datetime, timedelta
    
    today_str = datetime.now().strftime('%Y-%m-%d')
    default_start = (datetime.now() - timedelta(days=30)).strftime('%Y-%m-%d')
    
    start_date = request.args.get('start_date', default_start)
    end_date = request.args.get('end_date', today_str)
    
    is_mgr = at_require('at_manage')
    selected_user_id = request.args.get('user_id', '')
    if not is_mgr:
        selected_user_id = str(session['user_id'])
        
    target_emp_name = "Semua Karyawan"
    if selected_user_id:
        user_row = db.execute('SELECT full_name FROM users WHERE id=?', (int(selected_user_id),)).fetchone()
        if user_row:
            target_emp_name = user_row['full_name']

    # Fetch SupportCore tickets assigned and worked on by this employee or all employees in date range
    assigned_tickets = []
    if selected_user_id:
        emp_row = db.execute('SELECT id FROM employees WHERE user_id = ?', (int(selected_user_id),)).fetchone()
        if emp_row:
            emp_id = emp_row['id']
            assigned_tickets = db.execute('''
                SELECT t.id, t.ticket_no, t.subject, t.status, t.priority, t.reported_at, t.resolved_at, t.due_date,
                       c.name as customer_name, u.full_name as assignee_name
                FROM sc_ticket_assignees ta
                JOIN sc_tickets t ON t.id = ta.ticket_id
                JOIN sc_customers c ON c.id = t.customer_id
                JOIN employees e ON e.id = ta.employee_id
                JOIN users u ON u.id = e.user_id
                WHERE ta.employee_id = ? 
                  AND (
                      (date(t.reported_at) >= ? AND date(t.reported_at) <= ?)
                      OR (t.resolved_at IS NOT NULL AND date(t.resolved_at) >= ? AND date(t.resolved_at) <= ?)
                      OR (t.status IN ('open', 'progress'))
                  )
                  AND t.status != 'cancelled'
                ORDER BY t.reported_at DESC
            ''', (emp_id, start_date, end_date, start_date, end_date)).fetchall()
    else:
        assigned_tickets = db.execute('''
            SELECT t.id, t.ticket_no, t.subject, t.status, t.priority, t.reported_at, t.resolved_at, t.due_date,
                   c.name as customer_name, u.full_name as assignee_name
            FROM sc_ticket_assignees ta
            JOIN sc_tickets t ON t.id = ta.ticket_id
            JOIN sc_customers c ON c.id = t.customer_id
            JOIN employees e ON e.id = ta.employee_id
            JOIN users u ON u.id = e.user_id
            WHERE (
                  (date(t.reported_at) >= ? AND date(t.reported_at) <= ?)
                  OR (t.resolved_at IS NOT NULL AND date(t.resolved_at) >= ? AND date(t.resolved_at) <= ?)
                  OR (t.status IN ('open', 'progress'))
              )
              AND t.status != 'cancelled'
            ORDER BY t.reported_at DESC
        ''', (start_date, end_date, start_date, end_date)).fetchall()
            
    # Fetch logs
    query = '''
        SELECT a.*, u.full_name as employee_name, u.username as telegram_username
        FROM attendance a
        JOIN users u ON a.user_id = u.id
        WHERE a.date >= ? AND a.date <= ?
    '''
    params = [start_date, end_date]
    if selected_user_id:
        query += ' AND a.user_id = ?'
        params.append(int(selected_user_id))
        
    query += ' ORDER BY a.date DESC, a.clock_in DESC'
    rows = db.execute(query, params).fetchall()
    
    # Format rows
    attendance_list = []
    for row in rows:
        r = dict(row)
        # Work hours duration
        if r['clock_in'] and r['clock_out']:
            try:
                t1_str = r['clock_in']
                t2_str = r['clock_out']
                t1 = datetime.strptime(t1_str, '%H:%M') if len(t1_str.split(':')) == 2 else datetime.strptime(t1_str, '%H:%M:%S')
                t2 = datetime.strptime(t2_str, '%H:%M') if len(t2_str.split(':')) == 2 else datetime.strptime(t2_str, '%H:%M:%S')
                diff = t2 - t1
                hours = diff.total_seconds() / 3600.0
                r['work_hours'] = f"{hours:.2f} jam"
            except Exception:
                r['work_hours'] = '-'
        else:
            r['work_hours'] = '-'
            
        # Group Name / Sumber Absensi
        source_in = r['notes_in'] or ''
        group_name = "-"
        if "Telegram Group:" in source_in:
            parts = source_in.split("Telegram Group:")
            if len(parts) > 1:
                group_name = parts[1].split("(ID:")[0].strip()
        elif "Telegram Japri" in source_in:
            group_name = "Telegram Japri"
        elif "WhatsApp Group:" in source_in:
            parts = source_in.split("WhatsApp Group:")
            if len(parts) > 1:
                group_name = "WA Group: " + parts[1].split("(ID:")[0].strip()
        elif "WhatsApp Japri" in source_in:
            group_name = "WA Japri"
        elif "Web" in source_in or "Web Attendance" in source_in:
            group_name = "Web Portal"
            
        r['group_name'] = group_name

        # Fetch ticket progress updates for this user on this date
        progress_updates = db.execute('''
            SELECT h.old_value, h.new_value, h.created_at, t.ticket_no, t.subject
            FROM sc_ticket_history h
            JOIN sc_tickets t ON t.id = h.ticket_id
            WHERE h.changed_by = ? 
              AND h.field_name = 'pct_done' 
              AND date(h.created_at) = ?
            ORDER BY h.created_at ASC
        ''', (r['user_id'], r['date'])).fetchall()
        
        r['progress_updates'] = [dict(pu) for pu in progress_updates]
        r['progress_count'] = len(progress_updates)
        attendance_list.append(r)
        
    now_str = datetime.now().strftime('%d %b %Y %H:%M')
    return render_template(
        'at_report_print.html',
        attendance=attendance_list,
        start_date=start_date,
        end_date=end_date,
        target_emp_name=target_emp_name,
        now=now_str,
        assigned_tickets=assigned_tickets
    )

@app.route('/telegram/webhook', methods=['POST'])
def telegram_webhook():
    db = get_db()
    settings = get_settings(db)
    bot_token = settings.get('telegram_bot_token', '').strip()
    if not bot_token:
        bot_token = os.environ.get('TELEGRAM_BOT_TOKEN', '1095530966:AAFkSV9puxmT2z7cvpsbBQy_TWqj9-MCvbM').strip()
    if not bot_token:
        return 'OK', 200

    try:
        data = request.get_json()
        print(f"[Telegram Webhook] Payload: {json.dumps(data)}")
    except Exception as e:
        print(f"[Telegram Webhook] Failed to parse JSON: {e}")
        return 'OK', 200

    if not data:
        return 'OK', 200

    # Detect custom Node.js forwarder payload vs standard Telegram webhook update
    is_custom = 'telegram_user_id' in data and 'latitude' in data

    if is_custom:
        from_id = data.get('telegram_user_id')
        username = data.get('username', '') or ''
        chat_id = data.get('chat_id') or data.get('chatId') or data.get('group_id') or data.get('groupId') or data.get('telegram_chat_id') or from_id
        is_location = True
        lat = data.get('latitude')
        lng = data.get('longitude')
        loc = f"{lat},{lng}"
        text = ""
    else:
        message = data.get('message') or data.get('edited_message')
        if not message or 'chat' not in message:
            return 'OK', 200

        chat_id = message['chat']['id']
        from_user = message.get('from', {})
        from_id = from_user.get('id')
        username = from_user.get('username', '') or ''
        
        is_location = 'location' in message
        lat = None
        lng = None
        loc = ""
        if is_location:
            lat = message['location']['latitude']
            lng = message['location']['longitude']
            loc = f"{lat},{lng}"
        
        text = message.get('text', '') or ''
        if text.strip().startswith('/location'):
            is_location = True
            # Try to parse coordinates from text if present (e.g. "/location -6.2,106.8")
            parts = text.strip().split()
            if len(parts) > 1:
                coord_part = parts[1]
                if ',' in coord_part:
                    try:
                        lat_str, lng_str = coord_part.split(',', 1)
                        lat = float(lat_str.strip())
                        lng = float(lng_str.strip())
                        loc = f"{lat},{lng}"
                    except Exception:
                        pass

    def reply(text_msg, auto_delete=True):
        if not chat_id:
            print("[Telegram Webhook] Reply skipped: chat_id is empty or None")
            return
        
        # Only auto-delete for /birthday, /lapar, and /haus commands
        is_target_command = False
        if text:
            clean_txt = text.strip().lower()
            if clean_txt.startswith('/birthday') or clean_txt.startswith('/lapar') or clean_txt.startswith('/haus'):
                is_target_command = True
        if not is_target_command:
            auto_delete = False

        url = f"https://api.telegram.org/bot{bot_token}/sendMessage"
        payload = {
            'chat_id': chat_id,
            'text': text_msg,
            'parse_mode': 'HTML'
        }
        try:
            print(f"[Telegram Webhook] Sending reply to {chat_id}: {text_msg}")
            r = req_lib.post(url, json=payload, timeout=5)
            print(f"[Telegram Webhook] Telegram API Response (Status {r.status_code}): {r.text}")
            
            if auto_delete and r.status_code == 200:
                resp_json = r.json()
                if resp_json.get('ok'):
                    bot_msg_id = resp_json['result']['message_id']
                    
                    def delete_messages():
                        # Delete bot reply
                        try:
                            req_lib.post(f"https://api.telegram.org/bot{bot_token}/deleteMessage", json={
                                'chat_id': chat_id,
                                'message_id': bot_msg_id
                            }, timeout=5)
                        except Exception:
                            pass
                        # Delete user message
                        if not is_custom and 'message' in data and 'message_id' in data['message']:
                            try:
                                req_lib.post(f"https://api.telegram.org/bot{bot_token}/deleteMessage", json={
                                    'chat_id': chat_id,
                                    'message_id': data['message']['message_id']
                                }, timeout=5)
                            except Exception:
                                pass
                                
                    import threading
                    threading.Timer(30.0, delete_messages).start()
        except Exception as ex:
            print(f"[Telegram Webhook Error] Exception in reply helper: {str(ex)}")

    def format_indo_datetime(dt):
        months = ["", "Januari", "Februari", "Maret", "April", "Mei", "Juni", "Juli", "Agustus", "September", "Oktober", "November", "Desember"]
        day = dt.strftime('%d')
        month = months[dt.month]
        year = dt.strftime('%Y')
        time_str = dt.strftime('%H:%M')
        return f"{day} {month} {year} {time_str} WIB"

    try:
        tg_id_variants = []
        if from_id:
            tg_id_variants.append(str(from_id))
        if username:
            clean_username = username.lstrip('@')
            tg_id_variants.append(clean_username)
            tg_id_variants.append(f"@{clean_username}")

        # Extract telegram user's first/last name if available from Telegram payload
        tg_first_name = ""
        tg_last_name = ""
        if not is_custom and message:
            tg_first_name = from_user.get('first_name', '') or ''
            tg_last_name = from_user.get('last_name', '') or ''
        tg_full_name = f"{tg_first_name} {tg_last_name}".strip()

        # 1. Check in employees table first
        employee = None
        for variant in tg_id_variants:
            employee = db.execute('SELECT * FROM employees WHERE LOWER(telegram_id) = LOWER(?) AND is_active = 1', (variant,)).fetchone()
            if employee:
                break

        user = None
        if employee and employee['user_id']:
            user = db.execute('SELECT * FROM users WHERE id = ? AND is_active = 1', (employee['user_id'],)).fetchone()

        if not user:
            for variant in tg_id_variants:
                user = db.execute('SELECT * FROM users WHERE LOWER(telegram_id) = LOWER(?) AND is_active = 1', (variant,)).fetchone()
                if user:
                    break

        # Fallback: if not registered in HIVE, auto-register as user & employee
        if not user:
            import time
            from werkzeug.security import generate_password_hash

            fallback_username = f"tg_{from_id}" if from_id else f"tg_anon_{int(time.time())}"
            existing_user = db.execute('SELECT * FROM users WHERE username = ?', (fallback_username,)).fetchone()
            if existing_user:
                user = existing_user
            else:
                dummy_pwd = generate_password_hash("TelegramAttendanceFallback@2026")
                full_name = employee['name'] if employee else (tg_full_name or username or f"Telegram User {from_id}")
                
                db.execute('''
                    INSERT INTO users (username, password_hash, full_name, role, is_active, telegram_id)
                    VALUES (?, ?, ?, ?, ?, ?)
                ''', (fallback_username, dummy_pwd, full_name, 'karyawan', 1, str(from_id) if from_id else ''))
                db.commit()
                user = db.execute('SELECT * FROM users WHERE username = ?', (fallback_username,)).fetchone()

            if employee:
                db.execute('UPDATE employees SET user_id = ? WHERE id = ?', (user['id'], employee['id']))
                db.commit()
            else:
                emp_name = tg_full_name or username or f"Telegram User {from_id}"
                db.execute('''
                    INSERT INTO employees (name, divisi, telegram_id, user_id, is_active)
                    VALUES (?, ?, ?, ?, ?)
                ''', (emp_name, 'Telegram Core', str(from_id) if from_id else '', user['id'], 1))
                db.commit()

        uid = user['id']
        full_name = user['full_name']
        
        # Resolve employee/username display with HTML escaping
        import html
        
        display_name = ""
        if employee and employee['name']:
            display_name = employee['name']
        elif user and user['full_name']:
            display_name = user['full_name']
        elif tg_full_name:
            display_name = tg_full_name
            
        telegram_identifier = f"@{username}" if username else (str(from_id) if from_id else "")
        
        # Resolve display string: (@telegram_identifier (display_name))
        if telegram_identifier and display_name:
            user_display = f"({html.escape(telegram_identifier)} ({html.escape(display_name)}))"
        elif telegram_identifier:
            user_display = f"({html.escape(telegram_identifier)})"
        elif display_name:
            user_display = f"({html.escape(display_name)})"
        else:
            user_display = "(Unknown)"

        if is_location:
            # Get group title / name
            group_title = ""
            if is_custom:
                group_title = data.get('chat_title', '') or ''
            else:
                chat_obj = message.get('chat', {})
                if chat_obj.get('type') in ('group', 'supergroup'):
                    group_title = chat_obj.get('title', '') or ''

            if group_title:
                source_label = f"Telegram Group: {group_title} (ID: {chat_id})"
            else:
                source_label = f"Telegram Japri (ID: {chat_id})"
            
            today = datetime.now().strftime('%Y-%m-%d')
            now_time = datetime.now().strftime('%H:%M:%S')
            
            today_att = db.execute(
                'SELECT * FROM attendance WHERE user_id=? AND date=?',
                (uid, today)
            ).fetchone()
            
            # Check GPS location validity
            if not lat or not lng:
                reply(
                    f"❌ {user_display}\n"
                    "Lokasi GPS tidak valid. Harap gunakan fitur Kirim Lokasi pada Telegram."
                )
                return 'OK', 200

            # Reverse geocode using OpenStreetMap Nominatim API
            resolved_address = ""
            try:
                osm_url = f"https://nominatim.openstreetmap.org/reverse?format=json&lat={lat}&lon={lng}&zoom=18&addressdetails=1"
                osm_headers = {"User-Agent": "HIVE-Attendance-Bot/2.0 (md@workspace)"}
                osm_resp = req_lib.get(osm_url, headers=osm_headers, timeout=5)
                if osm_resp.status_code == 200:
                    resolved_address = osm_resp.json().get('display_name', '')
            except Exception as osm_ex:
                print(f"[OSM Nominatim Error] {osm_ex}")
            
            loc = f"{resolved_address} ({lat},{lng})" if resolved_address else f"{lat},{lng}"

            if not today_att:
                status = 'present'
                if now_time > '09:15:00':
                    status = 'late'
                
                # Save to database
                db.execute(
                    'INSERT INTO attendance (user_id, date, clock_in, location_in, notes_in, status) VALUES (?, ?, ?, ?, ?, ?)',
                    (uid, today, now_time, loc, source_label, status)
                )
                db.commit()

                # Send reply
                now_dt = datetime.now()
                time_str = format_indo_datetime(now_dt)
                reply(
                    f"✅ {user_display}\n"
                    f"Tagging berhasil dicatat pada {time_str}\n"
                    f"Lokasi: {html.escape(loc)}\n\n"
                    f"Silakan segera membuat PLAN dan PROGRESS sebelum clock out."
                )
            else:
                # User is sending location again (Clock Out / update Clock Out)
                # Strict check: must checkout in the same group/chat as checkin
                import re
                m_chat = re.search(r'\(ID:\s*(-?\d+)\)', today_att['notes_in'] or '')
                if m_chat:
                    checkin_chat_id = m_chat.group(1)
                    if str(chat_id) != checkin_chat_id:
                        m_name = re.match(r'^Telegram Group:\s*(.*?)\s*\(ID:', today_att['notes_in'] or '')
                        group_name = m_name.group(1) if m_name else "Japri"
                        if group_name != "Japri":
                            reply(
                                f"❌ {user_display}\n"
                                f"Checkout gagal!\nAnda checkin di group <b>{group_name}</b>, silahkan checkout di group yang sama."
                            )
                        else:
                            reply(
                                f"❌ {user_display}\n"
                                f"Checkout gagal!\nAnda checkin via Japri, silahkan checkout via Japri kembali."
                            )
                        return 'OK', 200

                # First, check 9 hours limit!
                try:
                    clk_in = today_att['clock_in']
                    if len(clk_in.split(':')) == 2:
                        in_dt = datetime.strptime(f"{today} {clk_in}", "%Y-%m-%d %H:%M")
                    else:
                        in_dt = datetime.strptime(f"{today} {clk_in}", "%Y-%m-%d %H:%M:%S")
                    curr_dt = datetime.now()
                    diff = curr_dt - in_dt
                    total_seconds = diff.total_seconds()
                    # 9 hours is 9 * 3600 = 32400 seconds
                    if total_seconds < 9 * 3600:
                        needed_seconds = 9 * 3600 - total_seconds
                        needed_hours = int(needed_seconds // 3600)
                        needed_minutes = int((needed_seconds % 3600) // 60)
                        
                        reply(
                            f"❌ {user_display}\n"
                            f"Clock Out GAGAL! Total waktu kerja Anda baru berjalan selama {total_seconds / 3600:.2f} jam.\n"
                            f"Anda baru dapat melakukan Clock Out setelah bekerja minimal 9 jam (kurang {needed_hours} jam {needed_minutes} menit lagi)."
                        )
                        return 'OK', 200
                except Exception as ex:
                    print(f"[Clock Out Calculation Error] {ex}")
                plan = (today_att['plan'] or '').strip()
                progress = (today_att['progress'] or '').strip()
                
                # Check character lengths
                import re
                clean_plan = re.sub(r'(?i)#plan', '', plan).strip()
                clean_progress = re.sub(r'(?i)#progress', '', progress).strip()
                
                plan_ok = len(clean_plan) >= 10
                progress_ok = len(clean_progress) >= 10
                
                if not plan_ok or not progress_ok:
                    status_msg = []
                    if plan_ok:
                        status_msg.append("✅ <b>#PLAN:</b> Terisi (≥ 10 karakter)")
                    else:
                        status_msg.append("❌ <b>#PLAN:</b> Belum terisi atau kurang dari 10 karakter")
                        
                    if progress_ok:
                        status_msg.append("✅ <b>#PROGRESS:</b> Terisi (≥ 10 karakter)")
                    else:
                        status_msg.append("❌ <b>#PROGRESS:</b> Belum terisi atau kurang dari 10 karakter")
                    
                    # Send failure reply
                    reply(
                        f"❌ {user_display}\n"
                        "Clock Out GAGAL! Anda wajib mengisi #PLAN dan #PROGRESS (minimal 10 karakter per item) sebelum melakukan Clock Out.\n\n"
                        "<b>Status Pengisian Hari Ini:</b>\n" + "\n".join(status_msg)
                    )
                else:
                    # Save to database
                    db.execute(
                        'UPDATE attendance SET clock_out=?, location_out=?, notes_out=? WHERE id=?',
                        (now_time, loc, source_label, today_att['id'])
                    )
                    db.commit()

                    # Send success reply
                    now_dt = datetime.now()
                    time_str = format_indo_datetime(now_dt)
                    reply(
                        f"✅ {user_display}\n"
                        f"Tagging berhasil dicatat pada {time_str}"
                    )
        else:
            text = message.get('text', '').strip()
            lower_text = text.lower()
            is_plan = any(x in lower_text for x in ['#plan', '#addplan', '#changeplan'])
            is_progress = any(x in lower_text for x in ['#progress', '#addprogress', '#changeprogress'])

            if is_plan and is_progress:
                reply(
                    f"❌ {user_display}\n"
                    "Gagal! Rencana kerja (#PLAN) dan kemajuan kerja (#PROGRESS) harus dikirimkan dalam pesan terpisah secara berurutan:\n\n"
                    "1. Lakukan Clock In (Kirim Lokasi)\n"
                    "2. Kirim rencana kerja: #PLAN [isi rencana]\n"
                    "3. Kirim kemajuan kerja: #PROGRESS [isi laporan]\n"
                    "4. Lakukan Clock Out (Kirim Lokasi)"
                )
                return 'OK', 200

            if is_plan or is_progress:
                import re
                today = datetime.now().strftime('%Y-%m-%d')
                today_att = db.execute(
                    'SELECT * FROM attendance WHERE user_id=? AND date=?',
                    (uid, today)
                ).fetchone()

                is_clocked_in = today_att is not None and bool(today_att['clock_in'])

                if is_plan:
                    if not is_clocked_in:
                        reply(
                            f"❌ {user_display}\n"
                            "Gagal mencatat PLAN! Anda wajib melakukan Clock In (Kirim Lokasi) terlebih dahulu."
                        )
                        return 'OK', 200

                    cleaned_plan = re.sub(r'(?i)#(add|change)?plan', '', text).strip()
                    if len(cleaned_plan) < 10:
                        reply(
                            f"❌ {user_display}\n"
                            "PLAN Gagal: Rencana kerja minimal harus 10 karakter (tidak termasuk tag)."
                        )
                    else:
                        # Save to database
                        db.execute(
                            'UPDATE attendance SET plan=? WHERE id=?',
                            (cleaned_plan, today_att['id'])
                        )
                        db.commit()

                        # Reply success
                        now_dt = datetime.now()
                        time_str = format_indo_datetime(now_dt)
                        reply(
                            f"✅ {user_display}\n"
                            f"PLAN berhasil dicatat pada {time_str}"
                        )

                elif is_progress:
                    if not is_clocked_in:
                        reply(
                            f"❌ {user_display}\n"
                            "Gagal mencatat PROGRESS! Anda wajib melakukan Clock In (Kirim Lokasi) terlebih dahulu."
                        )
                        return 'OK', 200

                    # Check if #PLAN is already filled with at least 10 characters
                    db_plan = (today_att['plan'] or '').strip()
                    clean_db_plan = re.sub(r'(?i)#(add|change)?plan', '', db_plan).strip()
                    if len(clean_db_plan) < 10:
                        reply(
                            f"❌ {user_display}\n"
                            "Gagal mencatat PROGRESS! Anda wajib mengisi #PLAN (minimal 10 karakter) terlebih dahulu."
                        )
                        return 'OK', 200

                    cleaned_prog = re.sub(r'(?i)#(add|change)?progress', '', text).strip()
                    if len(cleaned_prog) < 10:
                        reply(
                            f"❌ {user_display}\n"
                            "PROGRESS Gagal: Kemajuan kerja minimal harus 10 karakter (tidak termasuk tag)."
                        )
                    else:
                        # Save to database
                        db.execute(
                            'UPDATE attendance SET progress=? WHERE id=?',
                            (cleaned_prog, today_att['id'])
                        )
                        db.commit()

                        # Reply success
                        now_dt = datetime.now()
                        time_str = format_indo_datetime(now_dt)
                        reply(
                            f"✅ {user_display}\n"
                            f"PROGRESS berhasil dicatat pada {time_str}"
                        )

            elif text.startswith('/start') or text.startswith('/help') or text.startswith('/absen'):
                reply(
                    "ℹ️ <b>Informasi</b>\n\n"
                    "Halo, saya adalah Hive😊\n\n"
                    "Saya akan mencatat daily aktifitas kamu secara automatis\n\n"
                    "Untuk melakukan checkin dan checkout share location kamu ke group\n\n"
                    "Untuk membuat, menambah dan merubah plan atau progress, kamu dapat mengirim pesan dengan menambahkan salah satu kata kunci <code>#plan</code>, <code>#addplan</code>, <code>#changeplan</code>, <code>#progress</code>, <code>#addprogress</code> atau <code>#changeprogress</code>\n\n"
                    "Kamu dapat melihat status kamu dengan mengetikan perintah <code>/myinfo</code>\n\n"
                    "Ketik <code>/checkin</code> di group untuk melihat user yg sudah checkin\n"
                    "Ketik <code>/birthday</code> untuk melihat birthday user yang terdaftar\n"
                    "Ketik <code>/lapar</code> untuk melihat user yang sedang ulang tahun\n"
                    "Ketik <code>/haus</code> untuk melihat user yang sedang ulang tahun\n\n"
                    "Kirim pesan pribadi ke saya dan ketik <code>/mylink</code> untuk mendapatkan link login kamu ke website"
                )

            elif text.startswith('/myinfo'):
                today = datetime.now().strftime('%Y-%m-%d')
                today_att = db.execute('SELECT * FROM attendance WHERE user_id=? AND date=?', (uid, today)).fetchone()
                
                status_txt = "Belum Check In"
                time_in = "—"
                time_out = "—"
                
                if today_att:
                    if today_att['clock_in']:
                        status_txt = "Sudah Check In"
                        time_in = today_att['clock_in']
                    if today_att['clock_out']:
                        status_txt = "Sudah Check Out"
                        time_out = today_att['clock_out']
                
                info_msg = (
                    f"ℹ️ <b>Status Absensi Anda Hari Ini:</b>\n"
                    f"👤 <b>Karyawan:</b> {user_display}\n"
                    f"📅 <b>Tanggal:</b> {today}\n"
                    f"📊 <b>Status:</b> {status_txt}\n"
                    f"📥 <b>Check In:</b> {time_in}\n"
                    f"📤 <b>Check Out:</b> {time_out}\n"
                )
                reply(info_msg)

            elif text.startswith('/checkin'):
                today = datetime.now().strftime('%Y-%m-%d')
                checked_in = db.execute('''
                    SELECT u.full_name, a.clock_in FROM attendance a 
                    JOIN users u ON a.user_id = u.id 
                    WHERE a.date = ? AND a.clock_in IS NOT NULL 
                    ORDER BY a.clock_in ASC
                ''', (today,)).fetchall()
                
                if not checked_in:
                    reply("⚠️ Belum ada karyawan yang check-in hari ini.")
                else:
                    user_list = "\n".join([f"• {r['full_name']} ({r['clock_in']})" for r in checked_in])
                    reply(f"👥 <b>Sudah checkin hari ini ({today}):</b>\n{user_list}")

            elif text.startswith('/birthday'):
                rows = db.execute("SELECT name, birthday FROM employees WHERE is_active=1 AND birthday IS NOT NULL AND birthday != ''").fetchall()
                
                if not rows:
                    reply("🎂 Belum ada data ulang tahun user.")
                else:
                    bday_list = []
                    months_names_en = ["", "July", "August", "September", "October", "November", "December", "January", "February", "March", "April", "May", "June"]
                    # Let's map month index to english names
                    months_names = ["", "January", "February", "March", "April", "May", "June", "July", "August", "September", "October", "November", "December"]
                    
                    for r in rows:
                        try:
                            b_date = datetime.strptime(r['birthday'], "%Y-%m-%d").date()
                            bday_list.append({
                                'name': r['name'],
                                'date_str': f"{b_date.day:02d} {months_names[b_date.month]}",
                                'month': b_date.month,
                                'day': b_date.day
                            })
                        except Exception:
                            pass
                    
                    # Sort chronologically by month and day (Jan to Dec)
                    bday_list.sort(key=lambda x: (x['month'], x['day']))
                    bday_txt = "\n".join([f"• {item['name']} , {item['date_str']}" for item in bday_list])
                    reply(f"🎂 <b>Upcoming Birthdays (sorted):</b>\n\n{bday_txt}")

            elif text.startswith('/lapar') or text.startswith('/haus'):
                curr_month = datetime.now().month
                months_names = ["", "Januari", "Februari", "Maret", "April", "Mei", "Juni", "Juli", "Agustus", "September", "Oktober", "November", "Desember"]
                months_names_en = ["", "January", "February", "March", "April", "May", "June", "July", "August", "September", "October", "November", "December"]
                
                rows = db.execute("SELECT name, birthday FROM employees WHERE is_active=1 AND birthday IS NOT NULL AND birthday != ''").fetchall()
                
                bday_this_month = []
                for r in rows:
                    try:
                        b_date = datetime.strptime(r['birthday'], "%Y-%m-%d").date()
                        if b_date.month == curr_month:
                            bday_this_month.append({
                                'name': r['name'],
                                'day': b_date.day,
                                'date_str': f"{b_date.day:02d} {months_names_en[b_date.month]}"
                            })
                    except Exception:
                        pass
                
                bday_this_month.sort(key=lambda x: x['day'])
                m_name = months_names[curr_month]
                if not bday_this_month:
                    reply(f"🎈 Tidak ada karyawan yang ulang tahun di bulan {m_name}.")
                else:
                    user_list = "\n".join([f"• {item['name']} , {item['date_str']}" for item in bday_this_month])
                    reply(f"🎂 <b>Karyawan yang ulang tahun di bulan {m_name}:</b>\n\n{user_list}")

            elif text.startswith('/mylink'):
                is_group = False
                if not is_custom:
                    chat_obj = message.get('chat', {})
                    if chat_obj.get('type') in ('group', 'supergroup'):
                        is_group = True
                
                if is_group:
                    reply(f"❌ {user_display}\nKirim pesan pribadi (Japri) ke saya dan ketik <code>/mylink</code> untuk mendapatkan link login Anda.")
                else:
                    portal_url = request.url_root.rstrip('/')
                    reply(
                        f"🔑 <b>Login Portal HIVE:</b>\n\n"
                        f"Silakan login ke web portal HIVE menggunakan tautan berikut:\n"
                        f"🔗 <a href='{portal_url}'>{portal_url}</a>\n\n"
                        f"Gunakan Username Anda: <code>{user['username']}</code>"
                    )
            else:
                # Ignore unrecognized messages in group chats to avoid spam
                is_group = False
                if not is_custom:
                    chat_obj = message.get('chat', {})
                    if chat_obj.get('type') in ('group', 'supergroup'):
                        is_group = True
                if not is_group and text:
                    reply("Perintah tidak dikenali. Silakan kirimkan lokasi Anda untuk melakukan Clock In / Clock Out, atau kirim pesan berisi #PLAN / #PROGRESS untuk memperbarui rencana/kemajuan kerja Anda.")
    except Exception as e:
        import traceback
        print(f"[Telegram Webhook Error] {str(e)}")
        traceback.print_exc()

    return 'OK', 200


@app.route('/whatsapp/webhook', methods=['POST'])
def whatsapp_webhook():
    db = get_db()
    settings = get_settings(db)
    wa_url = settings.get('openwa_url', '').strip()
    wa_key = settings.get('openwa_api_key', '').strip()
    wa_session = get_openwa_session(settings, 'evaluasi')
    wa_enabled = settings.get('openwa_enabled', '0') == '1'
    
    if not wa_enabled or not wa_url:
        return 'OK', 200

    try:
        data = request.get_json()
        print(f"[WhatsApp Webhook] Payload: {json.dumps(data)}")
    except Exception as e:
        print(f"[WhatsApp Webhook] Failed to parse JSON: {e}")
        return 'OK', 200

    if not data:
        return 'OK', 200

    msg = data.get('data') if data.get('event') == 'onMessage' else data
    if not isinstance(msg, dict):
        return 'OK', 200

    sender_jid = msg.get('from', '') or ''
    author_jid = msg.get('author', '') or sender_jid
    
    if not sender_jid:
        return 'OK', 200

    is_group = msg.get('isGroupMsg', False) or '@g.us' in sender_jid
    chat_id = sender_jid
    
    phone_number = author_jid.split('@')[0] if '@' in author_jid else author_jid
    sender_name = msg.get('sender', {}).get('pushname', '') or msg.get('sender', {}).get('name', '') or phone_number

    msg_type = msg.get('type', '')
    is_location = msg_type == 'location'
    lat = msg.get('lat') or msg.get('latitude')
    lng = msg.get('lng') or msg.get('longitude')
    
    text = msg.get('body', '') or ''
    
    def reply(text_msg):
        clean_msg = text_msg.replace('<b>', '*').replace('</b>', '*').replace('<code>', '`').replace('</code>', '`').replace('<i>', '_').replace('</i>', '_')
        try:
            print(f"[WhatsApp Webhook] Replying to {chat_id}: {clean_msg}")
            send_whatsapp(wa_url, wa_key, wa_session, chat_id, clean_msg)
        except Exception as ex:
            print(f"[WhatsApp Webhook Reply Error] {ex}")

    def format_indo_datetime(dt):
        months = ["", "Januari", "Februari", "Maret", "April", "Mei", "Juni", "Juli", "Agustus", "September", "Oktober", "November", "Desember"]
        day = dt.strftime('%d')
        month = months[dt.month]
        year = dt.strftime('%Y')
        time_str = dt.strftime('%H:%M')
        return f"{day} {month} {year} {time_str} WIB"

    try:
        tg_id_variants = [phone_number, f"+{phone_number}"]
        
        employee = None
        for variant in tg_id_variants:
            employee = db.execute('SELECT * FROM employees WHERE LOWER(telegram_id) = LOWER(?) AND is_active = 1', (variant,)).fetchone()
            if employee:
                break
            employee = db.execute('SELECT * FROM employees WHERE phone = ? AND is_active = 1', (variant,)).fetchone()
            if employee:
                break

        user = None
        if employee and employee['user_id']:
            user = db.execute('SELECT * FROM users WHERE id = ? AND is_active = 1', (employee['user_id'],)).fetchone()

        if not user:
            for variant in tg_id_variants:
                user = db.execute('SELECT * FROM users WHERE LOWER(telegram_id) = LOWER(?) AND is_active = 1', (variant,)).fetchone()
                if user:
                    break
                user = db.execute('SELECT * FROM users WHERE phone = ? AND is_active = 1', (variant,)).fetchone()
                if user:
                    break

        if not user:
            import time
            from werkzeug.security import generate_password_hash
            fallback_username = f"wa_{phone_number}"
            existing_user = db.execute('SELECT * FROM users WHERE username = ?', (fallback_username,)).fetchone()
            if existing_user:
                user = existing_user
            else:
                dummy_pwd = generate_password_hash("WhatsAppAttendanceFallback@2026")
                full_name = employee['name'] if employee else sender_name
                db.execute('''
                    INSERT INTO users (username, password_hash, full_name, role, is_active, telegram_id)
                    VALUES (?, ?, ?, ?, ?, ?)
                ''', (fallback_username, dummy_pwd, full_name, 'karyawan', 1, phone_number))
                db.commit()
                user = db.execute('SELECT * FROM users WHERE username = ?', (fallback_username,)).fetchone()

            if employee:
                db.execute('UPDATE employees SET user_id = ? WHERE id = ?', (user['id'], employee['id']))
                db.commit()
            else:
                db.execute('''
                    INSERT INTO employees (name, divisi, telegram_id, user_id, is_active)
                    VALUES (?, ?, ?, ?, ?)
                ''', (sender_name, 'Telegram Core', phone_number, user['id'], 1))
                db.commit()

        uid = user['id']
        full_name = user['full_name']
        user_display = f"{full_name} ({phone_number})"

        if is_location:
            group_title = msg.get('chat', {}).get('name', '') or ''
            if group_title:
                source_label = f"WhatsApp Group: {group_title} (ID: {chat_id})"
            else:
                source_label = f"WhatsApp Japri (ID: {chat_id})"

            today = datetime.now().strftime('%Y-%m-%d')
            now_time = datetime.now().strftime('%H:%M:%S')
            
            today_att = db.execute(
                'SELECT * FROM attendance WHERE user_id=? AND date=?',
                (uid, today)
            ).fetchone()

            if not lat or not lng:
                reply(f"❌ {user_display}\nLokasi GPS tidak valid.")
                return 'OK', 200

            resolved_address = ""
            try:
                osm_url = f"https://nominatim.openstreetmap.org/reverse?format=json&lat={lat}&lon={lng}&zoom=18&addressdetails=1"
                osm_headers = {"User-Agent": "HIVE-Attendance-Bot/2.0 (md@workspace)"}
                osm_resp = req_lib.get(osm_url, headers=osm_headers, timeout=5)
                if osm_resp.status_code == 200:
                    resolved_address = osm_resp.json().get('display_name', '')
            except Exception as osm_ex:
                print(f"[OSM Nominatim WA Error] {osm_ex}")
            loc = f"{resolved_address} ({lat},{lng})" if resolved_address else f"{lat},{lng}"

            if not today_att:
                status = 'present'
                if now_time > '09:15:00':
                    status = 'late'
                db.execute(
                    'INSERT INTO attendance (user_id, date, clock_in, location_in, notes_in, status) VALUES (?, ?, ?, ?, ?, ?)',
                    (uid, today, now_time, loc, source_label, status)
                )
                db.commit()
                now_dt = datetime.now()
                time_str = format_indo_datetime(now_dt)
                reply(
                    f"✅ {user_display}\n"
                    f"Tagging berhasil dicatat pada {time_str}\n"
                    f"Lokasi: {loc}\n\n"
                    f"Silakan segera membuat PLAN dan PROGRESS sebelum clock out."
                )
            else:
                import re
                m_chat = re.search(r'\(ID:\s*(.*?)\)', today_att['notes_in'] or '')
                if m_chat:
                    checkin_chat_id = m_chat.group(1)
                    if str(chat_id) != checkin_chat_id:
                        m_name = re.match(r'^WhatsApp Group:\s*(.*?)\s*\(ID:', today_att['notes_in'] or '')
                        group_name = m_name.group(1) if m_name else "Japri"
                        if group_name != "Japri":
                            reply(f"❌ {user_display}\nCheckout gagal! Anda checkin di group {group_name}, silakan checkout di group yang sama.")
                        else:
                            reply(f"❌ {user_display}\nCheckout gagal! Anda checkin via Japri, silakan checkout via Japri kembali.")
                        return 'OK', 200

                try:
                    clk_in = today_att['clock_in']
                    if len(clk_in.split(':')) == 2:
                        in_dt = datetime.strptime(f"{today} {clk_in}", "%Y-%m-%d %H:%M")
                    else:
                        in_dt = datetime.strptime(f"{today} {clk_in}", "%Y-%m-%d %H:%M:%S")
                    curr_dt = datetime.now()
                    diff = curr_dt - in_dt
                    total_seconds = diff.total_seconds()
                    if total_seconds < 9 * 3600:
                        needed_seconds = 9 * 3600 - total_seconds
                        needed_hours = int(needed_seconds // 3600)
                        needed_minutes = int((needed_seconds % 3600) // 60)
                        reply(f"❌ {user_display}\nClock Out GAGAL! Total waktu kerja baru berjalan selama {total_seconds / 3600:.2f} jam. Anda baru dapat Clock Out setelah bekerja minimal 9 jam (kurang {needed_hours} jam {needed_minutes} menit).")
                        return 'OK', 200
                except Exception as ex:
                    print(f"[WA Clock Out Calculation Error] {ex}")

                plan = (today_att['plan'] or '').strip()
                progress = (today_att['progress'] or '').strip()
                
                clean_plan = re.sub(r'(?i)#(add|change)?plan', '', plan).strip()
                clean_progress = re.sub(r'(?i)#(add|change)?progress', '', progress).strip()
                
                plan_ok = len(clean_plan) >= 10
                progress_ok = len(clean_progress) >= 10
                
                if not plan_ok or not progress_ok:
                    status_msg = []
                    status_msg.append(f"• PLAN: {'Terisi' if plan_ok else 'Kosong atau < 10 karakter'}")
                    status_msg.append(f"• PROGRESS: {'Terisi' if progress_ok else 'Kosong atau < 10 karakter'}")
                    reply(f"❌ {user_display}\nClock Out GAGAL! Anda wajib mengisi #PLAN dan #PROGRESS (min 10 karakter) sebelum Clock Out.\n\nStatus:\n" + "\n".join(status_msg))
                else:
                    db.execute(
                        'UPDATE attendance SET clock_out=?, location_out=?, notes_out=? WHERE id=?',
                        (now_time, loc, source_label, today_att['id'])
                    )
                    db.commit()
                    now_dt = datetime.now()
                    time_str = format_indo_datetime(now_dt)
                    reply(f"✅ {user_display}\nTagging berhasil dicatat pada {time_str}")
        else:
            text = text.strip()
            lower_text = text.lower()
            is_plan = any(x in lower_text for x in ['#plan', '#addplan', '#changeplan'])
            is_progress = any(x in lower_text for x in ['#progress', '#addprogress', '#changeprogress'])

            if is_plan and is_progress:
                reply(f"❌ {user_display}\nPLAN dan PROGRESS harus dikirimkan dalam pesan terpisah secara berurutan.")
                return 'OK', 200

            if is_plan or is_progress:
                import re
                today = datetime.now().strftime('%Y-%m-%d')
                today_att = db.execute('SELECT * FROM attendance WHERE user_id=? AND date=?', (uid, today)).fetchone()
                is_clocked_in = today_att is not None and bool(today_att['clock_in'])

                if is_plan:
                    if not is_clocked_in:
                        reply(f"❌ {user_display}\nGagal mencatat PLAN! Anda wajib Clock In (Kirim Lokasi) terlebih dahulu.")
                        return 'OK', 200

                    cleaned_plan = re.sub(r'(?i)#(add|change)?plan', '', text).strip()
                    if len(cleaned_plan) < 10:
                        reply(f"❌ {user_display}\nPLAN Gagal: Rencana kerja minimal harus 10 karakter.")
                    else:
                        db.execute('UPDATE attendance SET plan=? WHERE id=?', (cleaned_plan, today_att['id']))
                        db.commit()
                        reply(f"✅ {user_display}\nPLAN berhasil dicatat.")

                elif is_progress:
                    if not is_clocked_in:
                        reply(f"❌ {user_display}\nGagal mencatat PROGRESS! Anda wajib Clock In terlebih dahulu.")
                        return 'OK', 200

                    db_plan = (today_att['plan'] or '').strip()
                    clean_db_plan = re.sub(r'(?i)#(add|change)?plan', '', db_plan).strip()
                    if len(clean_db_plan) < 10:
                        reply(f"❌ {user_display}\nGagal mencatat PROGRESS! Anda wajib mengisi #PLAN (min 10 karakter) terlebih dahulu.")
                        return 'OK', 200

                    cleaned_prog = re.sub(r'(?i)#(add|change)?progress', '', text).strip()
                    if len(cleaned_prog) < 10:
                        reply(f"❌ {user_display}\nPROGRESS Gagal: Laporan kerja minimal harus 10 karakter.")
                    else:
                        db.execute('UPDATE attendance SET progress=? WHERE id=?', (cleaned_prog, today_att['id']))
                        db.commit()
                        reply(f"✅ {user_display}\nPROGRESS berhasil dicatat.")

            elif text.startswith('/start') or text.startswith('/help') or text.startswith('/absen'):
                reply(
                    "Halo! Saya adalah Bot WA HIVE 😊\n\n"
                    "• Kirim share lokasi Anda untuk *Clock In* / *Clock Out*.\n"
                    "• Kirim `#plan [rencana]` untuk mengisi plan.\n"
                    "• Kirim `#progress [laporan]` untuk mengisi progress.\n\n"
                    "Daftar Perintah:\n"
                    "- `/myinfo` : Status absensi Anda hari ini\n"
                    "- `/checkin` : Daftar karyawan sudah checkin hari ini\n"
                    "- `/birthday` : Daftar ulang tahun semua karyawan\n"
                    "- `/lapar` atau `/haus` : Karyawan ultah bulan ini\n"
                    "- `/mylink` : Tautan login web"
                )

            elif text.startswith('/myinfo'):
                today = datetime.now().strftime('%Y-%m-%d')
                today_att = db.execute('SELECT * FROM attendance WHERE user_id=? AND date=?', (uid, today)).fetchone()
                status_txt = "Belum Check In"
                time_in = "—"
                time_out = "—"
                if today_att:
                    if today_att['clock_in']: status_txt = "Sudah Check In"; time_in = today_att['clock_in']
                    if today_att['clock_out']: status_txt = "Sudah Check Out"; time_out = today_att['clock_out']
                reply(f"👤 Karyawan: {user_display}\n📅 Tanggal: {today}\n📊 Status: {status_txt}\n📥 Check In: {time_in}\n📤 Check Out: {time_out}")

            elif text.startswith('/checkin'):
                today = datetime.now().strftime('%Y-%m-%d')
                checked_in = db.execute('''
                    SELECT u.full_name, a.clock_in FROM attendance a 
                    JOIN users u ON a.user_id = u.id 
                    WHERE a.date = ? AND a.clock_in IS NOT NULL 
                    ORDER BY a.clock_in ASC
                ''', (today,)).fetchall()
                if not checked_in:
                    reply("⚠️ Belum ada karyawan yang check-in hari ini.")
                else:
                    user_list = "\n".join([f"• {r['full_name']} ({r['clock_in']})" for r in checked_in])
                    reply(f"👥 Sudah checkin hari ini ({today}):\n{user_list}")

            elif text.startswith('/birthday'):
                rows = db.execute("SELECT name, birthday FROM employees WHERE is_active=1 AND birthday IS NOT NULL AND birthday != ''").fetchall()
                if not rows:
                    reply("🎂 Belum ada data ulang tahun.")
                else:
                    from datetime import datetime
                    bday_list = []
                    months_names = ["", "January", "February", "March", "April", "May", "June", "July", "August", "September", "October", "November", "December"]
                    for r in rows:
                        try:
                            b_date = datetime.strptime(r['birthday'], "%Y-%m-%d").date()
                            bday_list.append({
                                'name': r['name'],
                                'date_str': f"{b_date.day:02d} {months_names[b_date.month]}",
                                'month': b_date.month,
                                'day': b_date.day
                            })
                        except Exception: pass
                    bday_list.sort(key=lambda x: (x['month'], x['day']))
                    bday_txt = "\n".join([f"• {item['name']} , {item['date_str']}" for item in bday_list])
                    reply(f"🎂 Daftar Ulang Tahun Karyawan:\n\n{bday_txt}")

            elif text.startswith('/lapar') or text.startswith('/haus'):
                curr_month = datetime.now().month
                months_names = ["", "Januari", "Februari", "Maret", "April", "Mei", "Juni", "Juli", "Agustus", "September", "Oktober", "November", "Desember"]
                months_names_en = ["", "January", "February", "March", "April", "May", "June", "July", "August", "September", "October", "November", "December"]
                rows = db.execute("SELECT name, birthday FROM employees WHERE is_active=1 AND birthday IS NOT NULL AND birthday != ''").fetchall()
                bday_this_month = []
                for r in rows:
                    try:
                        b_date = datetime.strptime(r['birthday'], "%Y-%m-%d").date()
                        if b_date.month == curr_month:
                            bday_this_month.append({
                                'name': r['name'],
                                'day': b_date.day,
                                'date_str': f"{b_date.day:02d} {months_names_en[b_date.month]}"
                            })
                    except Exception: pass
                bday_this_month.sort(key=lambda x: x['day'])
                m_name = months_names[curr_month]
                if not bday_this_month:
                    reply(f"🎈 Tidak ada yang ulang tahun di bulan {m_name}.")
                else:
                    user_list = "\n".join([f"• {item['name']} , {item['date_str']}" for item in bday_this_month])
                    reply(f"🎂 Karyawan ulang tahun bulan {m_name}:\n\n{user_list}")

            elif text.startswith('/mylink'):
                if is_group:
                    reply("❌ Silakan Japri saya dan ketik /mylink untuk mendapatkan tautan login Anda.")
                else:
                    portal_url = request.url_root.rstrip('/')
                    reply(f"🔑 Login Portal HIVE:\n\nTautan: {portal_url}\nUsername: {user['username']}")
            else:
                if not is_group and text:
                    reply("Perintah tidak dikenali. Silakan kirimkan lokasi untuk Clock In/Out, atau pesan dengan #PLAN / #PROGRESS.")
    except Exception as e:
        import traceback
        print(f"[WhatsApp Webhook Error] {str(e)}")
        traceback.print_exc()

    return 'OK', 200




# ─── Backup and Restore Helpers ────────────────────────────────────────────────

def dump_postgres(output_path):
    import subprocess
    base_dir = os.path.dirname(os.path.abspath(__file__))
    script_path = os.path.join(base_dir, 'backup_db.sh')
    
    # Ensure script is executable if running on unix
    if os.path.exists(script_path) and os.name != 'nt':
        try:
            os.chmod(script_path, 0o755)
        except Exception:
            pass
            
    # Fallback to direct pg_dump if script_path doesn't exist
    if not os.path.exists(script_path):
        host = os.environ.get('PG_HOST', 'localhost')
        port = os.environ.get('PG_PORT', 5432)
        dbname = os.environ.get('PG_NAME', 'hive_db')
        user = os.environ.get('PG_USER', 'hive')
        password = os.environ.get('PG_PASS', '')
        pg_dump_bin = os.environ.get('PG_DUMP_PATH', 'pg_dump')
        cmd = [pg_dump_bin, '-h', host, '-p', str(port), '-U', user, '-F', 'p', '-f', output_path, dbname]
        env = os.environ.copy()
        if password:
            env['PGPASSWORD'] = password
        res = subprocess.run(cmd, env=env, capture_output=True, text=True)
    else:
        # Run using bash shell script
        if os.name == 'nt':
            cmd = ['bash', script_path, output_path]
        else:
            cmd = [script_path, output_path]
        res = subprocess.run(cmd, capture_output=True, text=True)
        
    if res.returncode != 0:
        raise Exception(f"pg_dump / backup_db.sh failed: {res.stderr}")

def run_backup_housekeeping(retention_days):
    import time
    if not retention_days:
        return
    try:
        days = int(retention_days)
        if days <= 0:
            return
        retention_sec = days * 24 * 60 * 60
        now = time.time()
        base_dir = os.path.dirname(os.path.abspath(__file__))
        backups_dir = os.path.join(base_dir, 'backups')
        if os.path.exists(backups_dir):
            for f in os.listdir(backups_dir):
                if f.endswith('.zip') and f.startswith('backup_hive_'):
                    fpath = os.path.join(backups_dir, f)
                    stat = os.stat(fpath)
                    if (now - stat.st_mtime) > retention_sec:
                        try:
                            os.remove(fpath)
                        except Exception:
                            pass
    except Exception:
        pass

def create_backup_zip(backup_app=True, backup_uploads=True, backup_db=True):
    import zipfile
    import shutil
    base_dir = os.path.dirname(os.path.abspath(__file__))
    backups_dir = os.path.join(base_dir, 'backups')
    os.makedirs(backups_dir, exist_ok=True)
    
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    zip_filename = f"backup_hive_{timestamp}.zip"
    zip_filepath = os.path.join(backups_dir, zip_filename)
    
    temp_dir = os.path.join(backups_dir, f"temp_{timestamp}")
    os.makedirs(temp_dir, exist_ok=True)
    
    try:
        # 1. Backup DB
        if backup_db:
            if DB_TYPE == 'postgresql':
                db_dump_path = os.path.join(temp_dir, 'database_dump.sql')
                dump_postgres(db_dump_path)
            else:
                # SQLite
                db_dump_path = os.path.join(temp_dir, 'evaluasi.db')
                if os.path.exists(DB_PATH):
                    shutil.copy2(DB_PATH, db_dump_path)
                else:
                    with open(db_dump_path, 'w') as f:
                        pass
        
        # 2. Backup Uploads
        if backup_uploads:
            uploads_src = os.path.join(base_dir, 'static', 'uploads')
            if os.path.exists(uploads_src):
                uploads_dest = os.path.join(temp_dir, 'uploads')
                shutil.copytree(uploads_src, uploads_dest, dirs_exist_ok=True)
        
        # 3. Backup App (source code)
        if backup_app:
            app_dest = os.path.join(temp_dir, 'app_source')
            os.makedirs(app_dest, exist_ok=True)
            exclude_folders = {'.git', 'venv', '.venv', '__pycache__', 'backups', 'static/uploads', '.agents', '.claude', 'node_modules'}
            exclude_files = {'evaluasi.db', '.env'}
            
            for root, dirs, files in os.walk(base_dir):
                rel_root = os.path.relpath(root, base_dir)
                if rel_root == '.':
                    rel_root = ''
                
                parts = rel_root.replace('\\', '/').split('/')
                skip = False
                for part in parts:
                    if part in exclude_folders:
                        skip = True
                        break
                if skip:
                    dirs[:] = []
                    continue
                
                if rel_root:
                    os.makedirs(os.path.join(app_dest, rel_root), exist_ok=True)
                
                for file in files:
                    file_rel_path = os.path.join(rel_root, file) if rel_root else file
                    file_norm = file_rel_path.replace('\\', '/')
                    if file_norm in exclude_files:
                        continue
                    if file.endswith('.zip') and 'backup' in file:
                        continue
                    shutil.copy2(os.path.join(root, file), os.path.join(app_dest, file_rel_path))
                    
        # Compress
        with zipfile.ZipFile(zip_filepath, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            for root, dirs, files in os.walk(temp_dir):
                for file in files:
                    file_path = os.path.join(root, file)
                    arcname = os.path.relpath(file_path, temp_dir)
                    zip_file.write(file_path, arcname)
                    
        # Housekeeping
        try:
            db = _get_raw_db()
            cfg = get_settings(db)
            db.close()
            retention = cfg.get('backup_retention_days', '30')
            run_backup_housekeeping(retention)
        except Exception:
            pass
            
        return zip_filepath
    finally:
        if os.path.exists(temp_dir):
            shutil.rmtree(temp_dir)

def send_email_with_attachment(settings, to_email, subject, html_body, filepath):
    from email.mime.base import MIMEBase
    from email import encoders
    try:
        msg = MIMEMultipart()
        msg['Subject'] = subject
        msg['From']    = settings.get('smtp_from') or settings.get('smtp_user', '')
        msg['To']      = to_email
        
        msg.attach(MIMEText(html_body, 'html', 'utf-8'))
        
        filename = os.path.basename(filepath)
        with open(filepath, 'rb') as f:
            part = MIMEBase('application', 'octet-stream')
            part.set_payload(f.read())
            encoders.encode_base64(part)
            part.add_header('Content-Disposition', f'attachment; filename="{filename}"')
            msg.attach(part)
            
        use_ssl = settings.get('smtp_ssl', '0') == '1'
        host, port = settings.get('smtp_host',''), int(settings.get('smtp_port', 587))
        server = smtplib.SMTP_SSL(host, port) if use_ssl else smtplib.SMTP(host, port)
        if not use_ssl:
            server.starttls()
        server.login(settings.get('smtp_user',''), settings.get('smtp_password',''))
        server.send_message(msg)
        server.quit()
        return True, None
    except Exception as e:
        return False, str(e)

def upload_to_s3(endpoint, access_key, secret_key, bucket, region, filepath):
    import boto3
    from botocore.config import Config
    
    filename = os.path.basename(filepath)
    s3_key = f"backups/{filename}"
    config = Config(
        region_name=region or 'us-east-1',
        signature_version='s3v4'
    )
    s3 = boto3.client(
        's3',
        endpoint_url=endpoint or None,
        aws_access_key_id=access_key,
        aws_secret_access_key=secret_key,
        config=config
    )
    s3.upload_file(filepath, bucket, s3_key)

def test_s3_connection(endpoint, access_key, secret_key, bucket, region):
    import boto3
    from botocore.config import Config
    config = Config(
        region_name=region or 'us-east-1',
        signature_version='s3v4'
    )
    s3 = boto3.client(
        's3',
        endpoint_url=endpoint or None,
        aws_access_key_id=access_key,
        aws_secret_access_key=secret_key,
        config=config
    )
    s3.list_objects_v2(Bucket=bucket, MaxKeys=1)
    return True

# Scheduler callback
def check_and_run_scheduled_backup():
    db = _get_raw_db()
    try:
        cfg = get_settings(db)
        if cfg.get('backup_sched_enabled') != '1':
            return
            
        interval = cfg.get('backup_sched_interval', 'daily')
        time_str = cfg.get('backup_sched_time', '02:00')
        
        try:
            hour, minute = map(int, time_str.split(':'))
        except Exception:
            hour, minute = 2, 0
            
        now = datetime.now()
        
        if now.hour == hour and abs(now.minute - minute) < 5:
            should_run = False
            last_run_str = cfg.get('backup_sched_last_run', '')
            if not last_run_str:
                should_run = True
            else:
                last_run = datetime.strptime(last_run_str, '%Y-%m-%d %H:%M:%S')
                if interval == 'daily':
                    should_run = (now.date() > last_run.date())
                elif interval == 'weekly':
                    should_run = (now - last_run).days >= 7
                elif interval == 'monthly':
                    should_run = (now.year > last_run.year or now.month > last_run.month)
            
            if should_run:
                lock_key = f"backup_lock_{now.strftime('%Y%m%d_%H%M')}"
                try:
                    res = db.execute("SELECT value FROM app_settings WHERE key=?", (lock_key,)).fetchone()
                    if res:
                        return
                    save_setting(db, lock_key, '1')
                    db.commit()
                except Exception:
                    return
                
                save_setting(db, 'backup_sched_last_run', now.strftime('%Y-%m-%d %H:%M:%S'))
                db.commit()
                
                import threading
                threading.Thread(target=execute_scheduled_backup, args=(cfg,)).start()
    finally:
        db.close()

def execute_scheduled_backup(cfg):
    try:
        backup_app = cfg.get('backup_target_app', '1') == '1'
        backup_uploads = cfg.get('backup_target_uploads', '1') == '1'
        backup_db = cfg.get('backup_target_db', '1') == '1'
        
        filepath = create_backup_zip(backup_app=backup_app, backup_uploads=backup_uploads, backup_db=backup_db)
        
        # Dest 1: Email
        if cfg.get('backup_dest_email_enabled') == '1' and cfg.get('backup_dest_email_recipient'):
            subject = f"Scheduled Backup HIVE - {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            html_body = "<p>Terlampir adalah file backup otomatis HIVE.</p>"
            send_email_with_attachment(cfg, cfg.get('backup_dest_email_recipient'), subject, html_body, filepath)
            
        # Dest 2: S3
        if cfg.get('backup_dest_s3_enabled') == '1':
            endpoint = cfg.get('backup_dest_s3_endpoint', '').strip()
            access_key = cfg.get('backup_dest_s3_access_key', '').strip()
            secret_key = cfg.get('backup_dest_s3_secret_key', '').strip()
            bucket = cfg.get('backup_dest_s3_bucket', '').strip()
            region = cfg.get('backup_dest_s3_region', '').strip()
            if endpoint or access_key:
                upload_to_s3(endpoint, access_key, secret_key, bucket, region, filepath)
                
        # Status success
        db = _get_raw_db()
        try:
            log_msg = f"Backup otomatis berhasil pada {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            save_setting(db, 'backup_last_status', 'Sukses')
            save_setting(db, 'backup_last_log', log_msg)
            db.commit()
        finally:
            db.close()
            
    except Exception as e:
        db = _get_raw_db()
        try:
            log_msg = f"Backup otomatis gagal pada {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}: {str(e)}"
            save_setting(db, 'backup_last_status', 'Gagal')
            save_setting(db, 'backup_last_log', log_msg)
            db.commit()
        finally:
            db.close()

# Routes
@app.route('/portal/backup', methods=['GET'])
@login_required
def portal_backup():
    if not is_portal_admin():
        flash('Akses ditolak.', 'danger')
        return redirect(url_for('portal'))
    db = get_db()
    cfg = get_settings(db)
    
    # Get last 10 backup files in backups/ directory
    base_dir = os.path.dirname(os.path.abspath(__file__))
    backups_dir = os.path.join(base_dir, 'backups')
    backup_files = []
    if os.path.exists(backups_dir):
        for f in os.listdir(backups_dir):
            if f.endswith('.zip') and f.startswith('backup_hive_'):
                fpath = os.path.join(backups_dir, f)
                stat = os.stat(fpath)
                backup_files.append({
                    'filename': f,
                    'size': f"{stat.st_size / (1024*1024):.2f} MB",
                    'created_at': datetime.fromtimestamp(stat.st_mtime).strftime('%Y-%m-%d %H:%M:%S')
                })
        backup_files.sort(key=lambda x: x['created_at'], reverse=True)
        backup_files = backup_files[:10]
        
    return render_template('portal_backup.html', cfg=cfg, backup_files=backup_files, DB_TYPE=DB_TYPE)

@app.route('/portal/backup/download/<filename>')
@login_required
def download_backup_file(filename):
    if not is_portal_admin():
        abort(403)
    from flask import send_from_directory
    base_dir = os.path.dirname(os.path.abspath(__file__))
    backups_dir = os.path.join(base_dir, 'backups')
    return send_from_directory(backups_dir, filename, as_attachment=True)

@app.route('/portal/backup/run', methods=['POST'])
@login_required
def run_backup_manual():
    if not is_portal_admin():
        return jsonify({'ok': False, 'msg': 'Akses ditolak.'})
        
    backup_app = request.form.get('backup_app') == '1'
    backup_uploads = request.form.get('backup_uploads') == '1'
    backup_db = request.form.get('backup_db') == '1'
    dest_download = request.form.get('dest_download') == '1'
    dest_email = request.form.get('dest_email') == '1'
    dest_s3 = request.form.get('dest_s3') == '1'
    
    db = get_db()
    cfg = get_settings(db)
    
    try:
        filepath = create_backup_zip(backup_app=backup_app, backup_uploads=backup_uploads, backup_db=backup_db)
        log_msgs = []
        
        if dest_email:
            recipient = request.form.get('email_recipient', '').strip()
            if not recipient:
                recipient = cfg.get('backup_dest_email_recipient')
            if recipient:
                subject = f"Manual Backup HIVE - {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
                html_body = "<p>Terlampir adalah file backup manual HIVE.</p>"
                ok, err = send_email_with_attachment(cfg, recipient, subject, html_body, filepath)
                if ok:
                    log_msgs.append("Kirim email berhasil.")
                else:
                    log_msgs.append(f"Kirim email gagal: {err}")
            else:
                log_msgs.append("Kirim email gagal: Alamat email penerima kosong.")
                
        if dest_s3:
            endpoint = request.form.get('s3_endpoint', '').strip() or cfg.get('backup_dest_s3_endpoint')
            access_key = request.form.get('s3_access_key', '').strip() or cfg.get('backup_dest_s3_access_key')
            secret_key = request.form.get('s3_secret_key', '').strip() or cfg.get('backup_dest_s3_secret_key')
            bucket = request.form.get('s3_bucket', '').strip() or cfg.get('backup_dest_s3_bucket')
            region = request.form.get('s3_region', '').strip() or cfg.get('backup_dest_s3_region')
            
            try:
                upload_to_s3(endpoint, access_key, secret_key, bucket, region, filepath)
                log_msgs.append("Upload ke Object Storage S3 berhasil.")
            except Exception as e:
                log_msgs.append(f"Upload ke S3 gagal: {str(e)}")
                
        # Simpan status manual backup
        status_msg = f"Manual backup sukses pada {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}."
        if log_msgs:
            status_msg += " Info: " + ", ".join(log_msgs)
            
        save_setting(db, 'backup_last_status', 'Sukses')
        save_setting(db, 'backup_last_log', status_msg)
        db.commit()
        
        filename = os.path.basename(filepath)
        return jsonify({
            'ok': True,
            'msg': status_msg,
            'filename': filename if dest_download else None
        })
    except Exception as e:
        status_msg = f"Manual backup gagal pada {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}: {str(e)}"
        save_setting(db, 'backup_last_status', 'Gagal')
        save_setting(db, 'backup_last_log', status_msg)
        db.commit()
        return jsonify({'ok': False, 'msg': status_msg})

@app.route('/portal/backup/settings', methods=['POST'])
@login_required
def save_backup_settings():
    if not is_portal_admin():
        flash('Akses ditolak.', 'danger')
        return redirect(url_for('portal'))
    db = get_db()
    
    keys_to_save = [
        'backup_sched_enabled', 'backup_sched_interval', 'backup_sched_time',
        'backup_target_app', 'backup_target_uploads', 'backup_target_db',
        'backup_dest_email_enabled', 'backup_dest_email_recipient',
        'backup_dest_s3_enabled', 'backup_dest_s3_endpoint',
        'backup_dest_s3_access_key', 'backup_dest_s3_secret_key',
        'backup_dest_s3_bucket', 'backup_dest_s3_region',
        'backup_retention_days',
        'media_storage_type'
    ]
    
    for k in keys_to_save:
        if k in ('backup_sched_enabled', 'backup_target_app', 'backup_target_uploads',
                 'backup_target_db', 'backup_dest_email_enabled', 'backup_dest_s3_enabled'):
            v = '1' if request.form.get(k) else '0'
        else:
            v = request.form.get(k, '').strip()
        save_setting(db, k, v)
        
    db.commit()
    flash('Pengaturan backup berhasil disimpan.', 'success')
    return redirect(url_for('portal_backup'))

@app.route('/portal/backup/delete/<filename>', methods=['POST'])
@login_required
def delete_backup_file(filename):
    if not is_portal_admin():
        return jsonify({'ok': False, 'msg': 'Akses ditolak.'})
    # Secure filename
    import werkzeug
    filename = werkzeug.utils.secure_filename(filename)
    base_dir = os.path.dirname(os.path.abspath(__file__))
    backups_dir = os.path.join(base_dir, 'backups')
    filepath = os.path.join(backups_dir, filename)
    if os.path.exists(filepath):
        try:
            os.remove(filepath)
            return jsonify({'ok': True, 'msg': f'File backup {filename} berhasil dihapus.'})
        except Exception as e:
            return jsonify({'ok': False, 'msg': str(e)})
    return jsonify({'ok': False, 'msg': 'File tidak ditemukan.'})

@app.route('/portal/backup/test-s3', methods=['POST'])
@login_required
def portal_test_s3():
    if not is_portal_admin():
        return jsonify({'ok': False, 'msg': 'Akses ditolak.'})
    endpoint = request.form.get('s3_endpoint', '').strip()
    access_key = request.form.get('s3_access_key', '').strip()
    secret_key = request.form.get('s3_secret_key', '').strip()
    bucket = request.form.get('s3_bucket', '').strip()
    region = request.form.get('s3_region', '').strip()
    
    try:
        test_s3_connection(endpoint, access_key, secret_key, bucket, region)
        return jsonify({'ok': True, 'msg': 'Koneksi ke Object Storage S3 berhasil!'})
    except Exception as e:
        return jsonify({'ok': False, 'msg': f'Koneksi S3 gagal: {str(e)}'})

# ─── Main ──────────────────────────────────────────────────────────────────────

if __name__ == '__main__':
    init_db()
    if not app.debug or os.environ.get('WERKZEUG_RUN_MAIN') == 'true':
        start_scheduler()
    port = int(os.environ.get('PORT', 5000))
    debug = os.environ.get('FLASK_DEBUG', '0') == '1'
    print("=" * 55)
    print(" Aplikasi TalentCore")
    print(f" Buka browser: http://127.0.0.1:{port}")
    print("=" * 55)
    app.run(debug=debug, host='0.0.0.0', port=port)
